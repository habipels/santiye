from django.shortcuts import render,redirect,get_object_or_404
from django.http import HttpResponse
from users.models import *
from django.contrib.auth.models import User
from django.core.paginator import Paginator , EmptyPage, PageNotAnInteger
from django.db.models.query_utils import Q
from site_settings.models import *
from django.utils.translation  import gettext as _
from django.utils.translation import get_language, activate, gettext
from site_info.models import *
from main.views import super_admin_kontrolu,dil_bilgisi,translate,sozluk_yapisi,yetki
from .models import *
from hashids import Hashids
from asgiref.sync import sync_to_async, async_to_sync
# Salt değeri ve minimum hash uzunluğu belirleyin
HASHIDS_SALT = "habip_elis_12345"
HASHIDS_MIN_LENGTH = 32

hashids = Hashids(salt=HASHIDS_SALT, min_length=HASHIDS_MIN_LENGTH)
def toplam_tutar_cikarmai(id):
    a = gider_urun_bilgisi.objects.filter(gider_bilgis = id)
    topla = 0
    for i in a:
        topla = topla + ((i.urun_fiyati*i.urun_adeti)-i.urun_indirimi)
    return topla
def kalan_tutuari(id):
    a = Gider_odemesi.objects.filter(gelir_kime_ait_oldugu = id)
    toplam = 0
    indirim = 0
    for i in a:
        toplam = toplam+i.tutar
    a = gider_urun_bilgisi.objects.filter(gider_bilgis = id)
    genel_toplam = 0
    for i in a:
        genel_toplam = genel_toplam+(i.urun_fiyati*i.urun_adeti)
        indirim = indirim+ i.urun_indirimi
    return round(float(genel_toplam - toplam -indirim),2)



def get_fatura_gelir(request, fatura_id):
    print(fatura_id)
    if True:
        fatura = get_object_or_none(Gelir_Bilgisi , id = fatura_id)
        kalemler = gelir_urun_bilgisi.objects.filter(gider_bilgis = fatura)
        makbuzlar = Gelir_odemesi.objects.filter(gelir_kime_ait_oldugu = fatura)
        kasalar = Kasa.objects.filter(kasa_kart_ait_bilgisi = fatura.gelir_kime_ait_oldugu,silinme_bilgisi = False)
        toplam_fiyat  = 0
        sonuc = 0
        try:
           sonuc =  float(fatura.doviz) * float(fatura.toplam_tutar)
        except:
            sonuc = fatura.toplam_tutar
        toplam_genel = 0
        for j in kalemler:
            toplam_fiyat = toplam_fiyat + j.urun_fiyati
            toplam_genel = toplam_genel + (j.urun_adeti*j.urun_fiyati)
        if fatura.doviz:
            fatura_doviz = fatura.doviz
        else:
            fatura_doviz = ""
        fatura_data = {
            'cari':fatura.cari_bilgisi.cari_adi,
        'fatura_no': fatura.fatura_no,
        'doviz': fatura_doviz,
        'genel_doviz_tutari': sonuc,
        'aciklama': fatura.aciklama,
        "kategori" : fatura.gelir_kategorisii.gelir_kategori_adi if fatura.gelir_kategorisii else "Kategori Belirtilmemiş",
        "fatura_tarihi" : fatura.fatura_tarihi.strftime("%d.%m.%Y"),
        "vade_tarihi" : fatura.vade_tarihi.strftime("%d.%m.%Y"),
        "kalemler": [
            {
                "urun_adi": kalem.urun_bilgisi.urun_adi,
                "urun_adeti": kalem.urun_adeti,
                "urun_fiyati": kalem.urun_fiyati,
                "toplam": kalem.urun_adeti * kalem.urun_fiyati
            } for kalem in kalemler
        ],
        "makbuzlar": [
            {
                "id": makbuz.id,
                "tarihi": makbuz.tarihi.strftime("%d.%m.%Y"),
                "kasa_adi": makbuz.kasa_bilgisi.kasa_adi,
                "kayit_tarihi": makbuz.kayit_tarihi.strftime("%d.%m.%Y"),
                "islemi_yapan": makbuz.islemi_yapan.last_name if makbuz.islemi_yapan else "",
                "tutar": makbuz.tutar,
                "resim":makbuz.gelir_makbuzu.url if makbuz.gelir_makbuzu else ""

            } for makbuz in makbuzlar
        ],"kasalar": [
            {
                "id": kasa.id,
                "kasa_adi": kasa.kasa_adi

            } for kasa in kasalar
        ],
        
        "toplam_fiyat" : round(float(toplam_fiyat),2),
        "genel_toplam" : round(float(toplam_genel),2),
        "id" : fatura.id,
        "kalan_tutar":str(kalan_tutuar(fatura))
        }
        print(fatura_data)
        return JsonResponse(fatura_data)


def get_fatura_gider(request, fatura_id):
    print(fatura_id)
    if True:
        fatura = get_object_or_none(Gider_Bilgisi , id = fatura_id)
        kalemler = gider_urun_bilgisi.objects.filter(gider_bilgis = fatura)
        makbuzlar = Gider_odemesi.objects.filter(gelir_kime_ait_oldugu = fatura)
        kasalar = Kasa.objects.filter(kasa_kart_ait_bilgisi = fatura.gelir_kime_ait_oldugu,silinme_bilgisi = False)
        toplam_fiyat  = 0
        toplam_genel = 0
        sonuc = 0
        try:
           sonuc =  float(fatura.doviz) * float(fatura.toplam_tutar)
        except:
            sonuc = float(fatura.toplam_tutar)
        if fatura.doviz:
            fatura_doviz = fatura.doviz
        else:
            fatura_doviz = ""
        for j in kalemler:
            toplam_fiyat = toplam_fiyat + j.urun_fiyati
            toplam_genel = toplam_genel + (j.urun_adeti*j.urun_fiyati)
        fatura_data = {
            'cari': fatura.cari_bilgisi.cari_adi if fatura.cari_bilgisi.cari_adi else "",
        'fatura_no': fatura.fatura_no,
        'doviz': fatura_doviz,
        'genel_doviz_tutari': sonuc,
        'aciklama': fatura.aciklama,
        "kategori" : fatura.gelir_kategorisii.gider_kategori_adi if fatura.gelir_kategorisii else "Kategori Belirtilmemiş",
        "fatura_tarihi" : fatura.fatura_tarihi.strftime("%d.%m.%Y"),
        "vade_tarihi" : fatura.vade_tarihi.strftime("%d.%m.%Y"),
        "kalemler": [
            {
                "urun_adi": kalem.urun_bilgisi.urun_adi,
                "urun_adeti": kalem.urun_adeti,
                "urun_fiyati": kalem.urun_fiyati,
                "toplam": kalem.urun_adeti * kalem.urun_fiyati
            } for kalem in kalemler
        ],
        "makbuzlar": [
            {
                "id": makbuz.id,
                "tarihi": makbuz.tarihi.strftime("%d.%m.%Y"),
                "kasa_adi": makbuz.kasa_bilgisi.kasa_adi,
                "kayit_tarihi": makbuz.kayit_tarihi.strftime("%d.%m.%Y"),
                "islemi_yapan": makbuz.islemi_yapan.last_name if makbuz.islemi_yapan else "",
                "tutar": makbuz.tutar,
                "resim":makbuz.gelir_makbuzu.url if makbuz.gelir_makbuzu else ""

            } for makbuz in makbuzlar
        ],"kasalar": [
            {
                "id": kasa.id,
                "kasa_adi": kasa.kasa_adi

            } for kasa in kasalar
        ],
        
        "toplam_fiyat" : round(float(toplam_fiyat),2),
        "genel_toplam" : round(float(toplam_genel),2),
        "id" : fatura.id,
        "kalan_tutar":str(kalan_tutuari(fatura))
        }
        print(fatura_data)
        return JsonResponse(fatura_data)
import time

def jhson_gonder(a):
    start_time = time.time()
    data = []
    for i in a:
        if i.silinme_bilgisi:
            pass
        else:
            s = i.gelir_etiketi_sec.all()
            
            try:
                j = s[0].gider_etiketi_adi
                
            except:
                j = ""
            try:
                l = s[1].gider_etiketi_adi
                
            except:
                l = ""
            try:
                v = s[2].gider_etiketi_adi
                
            except:
                v = ""
            if i.silinme_bilgisi:
                b = "İPTAL"
            tutar = toplam_tutar_cikarmai(i)
            odeme = toplam_odenme_tutari(i)
            if odeme == tutar :
                b = "Ödendi"
            elif odeme > 0:
                b  = "Parçalı Ödendi"
            elif odeme == 0:
                b = "Ödenmedi"
            id = i.id
            y =   {
                "incele":f'<button class="faturabilgisi bg-sucsses" id="{id}" onclick="loadFaturaDetails({id})">İncele</button>',
            "fatura_no": str(i.fatura_no),
            "cari": i.cari_bilgisi.cari_adi if i.cari_bilgisi.cari_adi else "",
            "aciklama": f'<span class="monospace-bold" title="{str(i.aciklama)}">{str(i.aciklama)[:15]}</span>',
            "etiket1": j ,
            "etiket2": l,       
            "etiket3": v ,
            "duzenleme_tarihi": str(i.fatura_tarihi.strftime("%d.%m.%Y")),
            "vade_tarihi": str(i.vade_tarihi.strftime("%d.%m.%Y")),
            "fatura_bedeli": "$"+str(toplam_tutar_cikarmai(i)),
            "kalan_tutar": "$"+str(kalan_tutuari(i)),
            "durum": b
            }
            data.append(y)
        
    return data

def toplam_tutar_cikarma(id):
    a = gelir_urun_bilgisi.objects.filter(gider_bilgis = id)
    topla = 0
    for i in a:
        topla = topla + ((i.urun_fiyati*i.urun_adeti)-i.urun_indirimi)
    return topla
def toplam_odenme_tutar(id):
    a = Gelir_odemesi.objects.filter(gelir_kime_ait_oldugu = id)
    topla = 0
    for i in a:
        topla = topla + i.tutar
    return topla
def kalan_tutuar(id):
    a = Gelir_odemesi.objects.filter(gelir_kime_ait_oldugu = id)
    toplam = 0
    for i in a:
        toplam = toplam+i.tutar
    a = gelir_urun_bilgisi.objects.filter(gider_bilgis = id)
    genel_toplam = 0
    indirim = 0
    for i in a:
        genel_toplam = genel_toplam+(i.urun_fiyati*i.urun_adeti)
        indirim = indirim+ i.urun_indirimi
    return round(float(genel_toplam - toplam-indirim),2)
def jhson_gonder_2(a):
    from django.shortcuts import render
    from django.http import JsonResponse
    data = []
    for i in a:
        if i.silinme_bilgisi:
            pass
        else:
            s = i.gelir_etiketi_sec.all()
            
            try:
                j = s[0].gelir_etiketi_adi
                
            except:
                j = ""
            try:
                l = s[1].gelir_etiketi_adi
                
            except:
                l = ""
            try:
                v = s[2].gelir_etiketi_adi
                
            except:
                v = ""
            if i.silinme_bilgisi:
                b = "İPTAL"
            tutar = toplam_tutar_cikarma(i)
            odeme = toplam_odenme_tutar(i)
            id = i.id
            if odeme == tutar :
                b = "Ödendi"
            elif odeme > 0:
                b  = "Parçalı Ödendi"
            elif odeme == 0:
                b = "Ödenmedi"
            if i.cari_bilgisi:
                cari_bilgi = i.cari_bilgisi.cari_adi
            else:
                cari_bilgi = ""
            y =   {
            "incele":f'<button class="faturabilgisi bg-sucsses" id="{id}" onclick="loadFaturaDetails({id})">İncele</button>',
            "fatura_no": str(i.fatura_no),
            "cari": cari_bilgi,
            "aciklama": f'<span class="monospace-bold" title="{str(i.aciklama)}">{str(i.aciklama)[:15]}</span>',
            "etiket1": j ,
            "etiket2": l,       
            "etiket3": v ,
            "duzenleme_tarihi": str(i.fatura_tarihi.strftime("%d.%m.%Y")),
            "vade_tarihi": str(i.vade_tarihi.strftime("%d.%m.%Y")),
            "fatura_bedeli": "$"+str(toplam_tutar_cikarma(i)),
            "kalan_tutar": "$"+str(kalan_tutuar(i)),
            "durum": b
            }
            data.append(y)
    
    return data

def encode_id(id):
    return hashids.encode(id)

def decode_id(hash_id):
    ids = hashids.decode(hash_id)
    return ids[0] if ids else None
def get_object_or_none(model, *args, **kwargs):
    try:
        return model.objects.get(*args, **kwargs)
    except :
        return None
# Create your views here.
def kasa_viev(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.kasa_gosterme_izni:
                    profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =Kasa.objects.filter(Q(kasa_kart_ait_bilgisi__first_name__icontains = search)|Q(kasa_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kasa_gosterme_izni:
                        profile = Kasa.objects.filter(Q(kasa_kart_ait_bilgisi = request.user.kullanicilar_db) & Q(kasa_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = Kasa.objects.filter(Q(kasa_kart_ait_bilgisi = request.user) & Q(kasa_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/muhasebe_index.html",content)
def a_kasa_viev(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        profile = Kasa.objects.filter(kasa_kart_ait_bilgisi =  users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =Kasa.objects.filter(Q(kasa_kart_ait_bilgisi__first_name__icontains = search)|Q(kasa_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = Kasa.objects.filter(Q(kasa_kart_ait_bilgisi = request.user) & Q(kasa_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/muhasebe_index.html",content)
def kasa_tekli(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        k_gonder = get_object_or_404(Kasa,id = id)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.kasa_detay_izni:
                    profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                    k_gonder = get_object_or_404(Kasa,id = id)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
            k_gonder = get_object_or_404(Kasa,id = id)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =Kasa.objects.filter(Q(kasa_kart_ait_bilgisi__first_name__icontains = search)|Q(kasa_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kasa_detay_izni:
                        profile = Kasa.objects.filter(Q(kasa_kart_ait_bilgisi = request.user.kullanicilar_db) & Q(kasa_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = Kasa.objects.filter(Q(kasa_kart_ait_bilgisi = request.user) & Q(kasa_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["kasabilgiis_getirme"] = k_gonder
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/kasa_hareketleri.html",content)

#kasa ekleme
def kasa_ekle(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            kasa_Adi   = request.POST.get("kasaadi")
            bakiye = request.POST.get("bakiye")
            konumu = request.POST.get("konumu")
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
            Kasa.objects.create(kasa_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi )
                                ,kasa_adi = kasa_Adi,aciklama = konumu,bakiye = bakiye,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan
                                )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kasa_olusturma_izni:
                        kasa_Adi   = request.POST.get("kasaadi")
                        bakiye = request.POST.get("bakiye")
                        konumu = request.POST.get("konumu")
                        maaslarda_kullan = request.POST.get("maaslarda_kullan")
                        avanslarda_kullan = request.POST.get("avanslarda_kullan")
                        if maaslarda_kullan == "1":
                            maaslarda_kullan = False
                        else:
                            maaslarda_kullan = True
                        if avanslarda_kullan == "1":
                            avanslarda_kullan = False
                        else:
                            avanslarda_kullan = True
                        Kasa.objects.create(kasa_kart_ait_bilgisi = request.user.kullanicilar_db
                            ,kasa_adi = kasa_Adi,aciklama = konumu,bakiye = bakiye,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                kasa_Adi   = request.POST.get("kasaadi")
                bakiye = request.POST.get("bakiye")
                konumu = request.POST.get("konumu")
                maaslarda_kullan = request.POST.get("maaslarda_kullan")
                avanslarda_kullan = request.POST.get("avanslarda_kullan")
                if maaslarda_kullan == "1":
                    maaslarda_kullan = False
                else:
                    maaslarda_kullan = True
                if avanslarda_kullan == "1":
                    avanslarda_kullan = False
                else:
                    avanslarda_kullan = True
                Kasa.objects.create(kasa_kart_ait_bilgisi = request.user
                    ,kasa_adi = kasa_Adi,aciklama = konumu,bakiye = bakiye,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)

    return redirect("accounting:kasa")

#kasa silme
def kasa_sil(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        Kasa.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.Kasa_silme_izni:
                    Kasa.objects.filter(kasa_kart_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            Kasa.objects.filter(kasa_kart_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect("accounting:kasa")


#kasa düzenle
def kasa_duzenle(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        silinmedurumu = request.POST.get("silinmedurumu")
        konumu = request.POST.get("konumu")
        maaslarda_kullan = request.POST.get("maaslarda_kullan")
        avanslarda_kullan = request.POST.get("avanslarda_kullan")
        if maaslarda_kullan == "1":
            maaslarda_kullan = False
        else:
            maaslarda_kullan = True
        if avanslarda_kullan == "1":
            avanslarda_kullan = False
        else:
            avanslarda_kullan = True
        if silinmedurumu == "1":
            silinmedurumu = False
            Kasa.objects.filter(id = id).update(kasa_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,kasa_adi = proje_tip_adi,aciklama = konumu,silinme_bilgisi = silinmedurumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        elif silinmedurumu == "2":
            silinmedurumu = True
            Kasa.objects.filter(id = id).update(kasa_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,kasa_adi = proje_tip_adi,aciklama = konumu,silinme_bilgisi = silinmedurumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        else:
            Kasa.objects.filter(id = id).update(kasa_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,kasa_adi = proje_tip_adi,aciklama = konumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.kasa_guncelleme_izni:
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    konumu = request.POST.get("konumu")
                    maaslarda_kullan = request.POST.get("maaslarda_kullan")
                    avanslarda_kullan = request.POST.get("avanslarda_kullan")
                    if maaslarda_kullan == "1":
                        maaslarda_kullan = False
                    else:
                        maaslarda_kullan = True
                    if avanslarda_kullan == "1":
                        avanslarda_kullan = False
                    else:
                        avanslarda_kullan = True
                    Kasa.objects.filter(kasa_kart_ait_bilgisi = request.user.kullanicilar_db,id = id).update(kasa_adi = proje_tip_adi
                            ,aciklama = konumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("yetkili_adi")
            konumu = request.POST.get("konumu")
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
            Kasa.objects.filter(kasa_kart_ait_bilgisi = request.user,id = id).update(kasa_adi = proje_tip_adi
                    ,aciklama = konumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    return redirect("accounting:kasa")

def kasa_tekli_2(request,id,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = users)
        k_gonder = get_object_or_404(Kasa,id = id)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.kasa_detay_izni:
                    profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                    k_gonder = get_object_or_404(Kasa,id = id)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
            k_gonder = get_object_or_404(Kasa,id = id)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =Kasa.objects.filter(Q(kasa_kart_ait_bilgisi__first_name__icontains = search)|Q(kasa_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kasa_detay_izni:
                        profile = Kasa.objects.filter(Q(kasa_kart_ait_bilgisi = request.user.kullanicilar_db) & Q(kasa_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = Kasa.objects.filter(Q(kasa_kart_ait_bilgisi = request.user) & Q(kasa_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["kasabilgiis_getirme"] = k_gonder
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/kasa_hareketleri.html",content)

#kasa ekleme
def kasa_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            kasa_Adi   = request.POST.get("kasaadi")
            bakiye = request.POST.get("bakiye")
            konumu = request.POST.get("konumu")
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
            Kasa.objects.create(kasa_kart_ait_bilgisi = users
                                ,kasa_adi = kasa_Adi,aciklama = konumu,bakiye = bakiye,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan
                                )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kasa_olusturma_izni:
                        kasa_Adi   = request.POST.get("kasaadi")
                        bakiye = request.POST.get("bakiye")
                        konumu = request.POST.get("konumu")
                        maaslarda_kullan = request.POST.get("maaslarda_kullan")
                        avanslarda_kullan = request.POST.get("avanslarda_kullan")
                        if maaslarda_kullan == "1":
                            maaslarda_kullan = False
                        else:
                            maaslarda_kullan = True
                        if avanslarda_kullan == "1":
                            avanslarda_kullan = False
                        else:
                            avanslarda_kullan = True
                        Kasa.objects.create(kasa_kart_ait_bilgisi = request.user.kullanicilar_db
                            ,kasa_adi = kasa_Adi,aciklama = konumu,bakiye = bakiye,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                kasa_Adi   = request.POST.get("kasaadi")
                bakiye = request.POST.get("bakiye")
                konumu = request.POST.get("konumu")
                maaslarda_kullan = request.POST.get("maaslarda_kullan")
                avanslarda_kullan = request.POST.get("avanslarda_kullan")
                if maaslarda_kullan == "1":
                    maaslarda_kullan = False
                else:
                    maaslarda_kullan = True
                if avanslarda_kullan == "1":
                    avanslarda_kullan = False
                else:
                    avanslarda_kullan = True
                Kasa.objects.create(kasa_kart_ait_bilgisi = request.user
                    ,kasa_adi = kasa_Adi,aciklama = konumu,bakiye = bakiye,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)

    return redirect("accounting:a_kasa_viev",hash)

#kasa silme
def kasa_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        Kasa.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.Kasa_silme_izni:
                    Kasa.objects.filter(kasa_kart_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            Kasa.objects.filter(kasa_kart_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect("accounting:a_kasa_viev",hash)


#kasa düzenle
def kasa_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        silinmedurumu = request.POST.get("silinmedurumu")
        konumu = request.POST.get("konumu")
        maaslarda_kullan = request.POST.get("maaslarda_kullan")
        avanslarda_kullan = request.POST.get("avanslarda_kullan")
        if maaslarda_kullan == "1":
            maaslarda_kullan = False
        else:
            maaslarda_kullan = True
        if avanslarda_kullan == "1":
            avanslarda_kullan = False
        else:
            avanslarda_kullan = True
        if silinmedurumu == "1":
            silinmedurumu = False
            Kasa.objects.filter(id = id).update(kasa_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,kasa_adi = proje_tip_adi,aciklama = konumu,silinme_bilgisi = silinmedurumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        elif silinmedurumu == "2":
            silinmedurumu = True
            Kasa.objects.filter(id = id).update(kasa_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,kasa_adi = proje_tip_adi,aciklama = konumu,silinme_bilgisi = silinmedurumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        else:
            Kasa.objects.filter(id = id).update(kasa_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,kasa_adi = proje_tip_adi,aciklama = konumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.kasa_guncelleme_izni:
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    konumu = request.POST.get("konumu")
                    maaslarda_kullan = request.POST.get("maaslarda_kullan")
                    avanslarda_kullan = request.POST.get("avanslarda_kullan")
                    if maaslarda_kullan == "1":
                        maaslarda_kullan = False
                    else:
                        maaslarda_kullan = True
                    if avanslarda_kullan == "1":
                        avanslarda_kullan = False
                    else:
                        avanslarda_kullan = True
                    Kasa.objects.filter(kasa_kart_ait_bilgisi = request.user.kullanicilar_db,id = id).update(kasa_adi = proje_tip_adi
                            ,aciklama = konumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("yetkili_adi")
            konumu = request.POST.get("konumu")
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
            Kasa.objects.filter(kasa_kart_ait_bilgisi = request.user,id = id).update(kasa_adi = proje_tip_adi
                    ,aciklama = konumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    return redirect("accounting:a_kasa_viev",hash)



#Gelir Kategorisi
def gelir_kategorisi_tipleri(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =gelir_kategorisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_kategorisi_gorme:
                    profile = gelir_kategorisi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = gelir_kategorisi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =gelir_kategorisi.objects.filter(Q(gelir_kategoris_ait_bilgisi__last_name__icontains = search)|Q(gelir_kategori_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gelir_kategorisi_gorme:
                        profile = gelir_kategorisi.objects.filter(Q(gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db) & Q(gelir_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = gelir_kategorisi.objects.filter(Q(gelir_kategoris_ait_bilgisi = request.user) & Q(gelir_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/gelir_kategorisi.html",content)
def gelir_kategorisi_tipleri_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        #profile = Kasa.objects.filter(kasa_kart_ait_bilgisi =  get_object_or_404(CustomUser,id = post_id))
        profile = gelir_kategorisi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = gelir_kategorisi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =gelir_kategorisi.objects.filter(Q(gelir_kategoris_ait_bilgisi__last_name__icontains = search)|Q(gelir_kategori_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = gelir_kategorisi.objects.filter(Q(gelir_kategoris_ait_bilgisi = request.user) & Q(gelir_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/gelir_kategorisi.html",content)
#gelir KAtegorisi Ekleme
#gelir KAtegorisi Ekleme


def gelir_kategoisi_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        gelir_kategorisi.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_kategorisi_silme:
                    gelir_kategorisi.objects.filter(gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            gelir_kategorisi.objects.filter(gelir_kategoris_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect("accounting:gelir_kategorisi_tipleri_2",hash)
#gelir düzenleme
def gelir_kategorisi_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        proje_tip_adi   = request.POST.get("yetkili_adi")
        aciklama = request.POST.get("aciklama")
        renk = request.POST.get("renk")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        gelir_kategorisi.objects.filter(gelir_kategoris_ait_bilgisi = users,id = id).update(gelir_kategori_adi = proje_tip_adi,gelir_kategorisi_renk = renk,aciklama = aciklama)
    return redirect("accounting:gelir_kategorisi_tipleri_2",hash)
def gelir_kategorisi_ekleme_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            aciklama = request.POST.get("aciklama")
            renk = request.POST.get("renk")

            gelir_kategorisi.objects.create(gelir_kategoris_ait_bilgisi = users ,gelir_kategori_adi = proje_tip_adi,gelir_kategorisi_renk = renk,aciklama = aciklama)
        
    return redirect("accounting:gelir_kategorisi_tipleri_2",hash)

#gelir Kategorisi Silme
def gelir_kategorisi_ekleme(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            aciklama = request.POST.get("aciklama")
            renk = request.POST.get("renk")

            gelir_kategorisi.objects.create(gelir_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gelir_kategori_adi = proje_tip_adi,gelir_kategorisi_renk = renk,aciklama = aciklama)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gelir_kategorisi_olusturma:
                        proje_tip_adi   = request.POST.get("yetkili_adi")
                        aciklama = request.POST.get("aciklama")
                        renk = request.POST.get("renk")
                        gelir_kategorisi.objects.create(gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db,gelir_kategori_adi = proje_tip_adi,gelir_kategorisi_renk = renk,aciklama = aciklama)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                proje_tip_adi   = request.POST.get("yetkili_adi")
                aciklama = request.POST.get("aciklama")
                renk = request.POST.get("renk")
                gelir_kategorisi.objects.create(gelir_kategoris_ait_bilgisi = request.user,gelir_kategori_adi = proje_tip_adi,gelir_kategorisi_renk = renk,aciklama = aciklama)
    return redirect("accounting:gelir_kategorisi_tipleri")

#gelir Kate
def gelir_kategoisi_sil(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        gelir_kategorisi.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_kategorisi_silme:
                    gelir_kategorisi.objects.filter(gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            gelir_kategorisi.objects.filter(gelir_kategoris_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect("accounting:gelir_kategorisi_tipleri")
#gelir düzenleme
def gelir_kategorisi_duzenle(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        aciklama = request.POST.get("aciklama")
        renk = request.POST.get("renk")
        silinmedurumu = request.POST.get("silinmedurumu")
        if silinmedurumu == "1":
            silinmedurumu = False
            gelir_kategorisi.objects.filter(id = id).update(gelir_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gelir_kategori_adi = proje_tip_adi,gelir_kategorisi_renk = renk,aciklama = aciklama,silinme_bilgisi = silinmedurumu)
        elif silinmedurumu == "2":
            silinmedurumu = True
            gelir_kategorisi.objects.filter(id = id).update(gelir_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gelir_kategori_adi = proje_tip_adi,gelir_kategorisi_renk = renk,aciklama = aciklama,silinme_bilgisi = silinmedurumu)
        else:
            gelir_kategorisi.objects.filter(id = id).update(gelir_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gelir_kategori_adi = proje_tip_adi,gelir_kategorisi_renk = renk,aciklama = aciklama)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_kategorisi_guncelleme:
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    aciklama = request.POST.get("aciklama")
                    renk = request.POST.get("renk")
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    gelir_kategorisi.objects.filter(gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db,id = id).update(gelir_kategori_adi = proje_tip_adi,gelir_kategorisi_renk = renk,aciklama = aciklama)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("yetkili_adi")
            aciklama = request.POST.get("aciklama")
            renk = request.POST.get("renk")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            gelir_kategorisi.objects.filter(gelir_kategoris_ait_bilgisi = request.user,id = id).update(gelir_kategori_adi = proje_tip_adi,gelir_kategorisi_renk = renk,aciklama = aciklama)
    return redirect("accounting:gelir_kategorisi_tipleri")



#gider_kategorisi.html

#gider Kategorisi
def gider_kategorisi_tipleri(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =gider_kategorisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_kategorisi_gorme:
                    profile = gider_kategorisi.objects.filter(silinme_bilgisi = False,gider_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = gider_kategorisi.objects.filter(silinme_bilgisi = False,gider_kategoris_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =gider_kategorisi.objects.filter(Q(gider_kategoris_ait_bilgisi__last_name__icontains = search)|Q(gider_kategori_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gider_kategorisi_gorme:
                        profile = gider_kategorisi.objects.filter(Q(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db) & Q(gider_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = gider_kategorisi.objects.filter(Q(gider_kategoris_ait_bilgisi = request.user) & Q(gider_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/gider_kategorisi.html",content)
#gider Kategorisi
def gider_kategorisi_tipleri_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        profile = gider_kategorisi.objects.filter(silinme_bilgisi = False,gider_kategoris_ait_bilgisi = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = gider_kategorisi.objects.filter(silinme_bilgisi = False,gider_kategoris_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =gider_kategorisi.objects.filter(Q(gider_kategoris_ait_bilgisi__last_name__icontains = search)|Q(gider_kategori_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = gider_kategorisi.objects.filter(Q(gider_kategoris_ait_bilgisi = request.user) & Q(gider_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/gider_kategorisi.html",content)
#gider KAtegorisi Ekleme
#gider KAtegorisi Ekleme
def gider_kategorisi_ekleme(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            aciklama = request.POST.get("aciklama")
            renk = request.POST.get("renk")
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
            gider_kategorisi.objects.create(gider_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gider_kategori_adi = proje_tip_adi,gider_kategorisi_renk = renk,aciklama = aciklama,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gider_kategorisi_olusturma:
                        proje_tip_adi   = request.POST.get("yetkili_adi")
                        aciklama = request.POST.get("aciklama")
                        renk = request.POST.get("renk")
                        maaslarda_kullan = request.POST.get("maaslarda_kullan")
                        avanslarda_kullan = request.POST.get("avanslarda_kullan")
                        if maaslarda_kullan == "1":
                            maaslarda_kullan = False
                        else:
                            maaslarda_kullan = True
                        if avanslarda_kullan == "1":
                            avanslarda_kullan = False
                        else:
                            avanslarda_kullan = True
                        gider_kategorisi.objects.create(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db,gider_kategori_adi = proje_tip_adi,gider_kategorisi_renk = renk,aciklama = aciklama,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                proje_tip_adi   = request.POST.get("yetkili_adi")
                aciklama = request.POST.get("aciklama")
                renk = request.POST.get("renk")
                maaslarda_kullan = request.POST.get("maaslarda_kullan")
                avanslarda_kullan = request.POST.get("avanslarda_kullan")
                if maaslarda_kullan == "1":
                    maaslarda_kullan = False
                else:
                    maaslarda_kullan = True
                if avanslarda_kullan == "1":
                    avanslarda_kullan = False
                else:
                    avanslarda_kullan = True
                gider_kategorisi.objects.create(gider_kategoris_ait_bilgisi = request.user,gider_kategori_adi = proje_tip_adi,gider_kategorisi_renk = renk,aciklama = aciklama,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    return redirect("accounting:gider_kategorisi_tipleri")

#gider Kategorisi Silme

def gider_kategoisi_sil(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        gider_kategorisi.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_kategorisi_silme:
                    gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect("accounting:gider_kategorisi_tipleri")
#gider düzenleme
def gider_kategorisi_duzenle(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        aciklama = request.POST.get("aciklama")
        renk = request.POST.get("renk")
        if True:
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
        silinmedurumu = request.POST.get("silinmedurumu")
        if silinmedurumu == "1":
            silinmedurumu = False
            gider_kategorisi.objects.filter(id = id).update(gider_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gider_kategori_adi = proje_tip_adi,gider_kategorisi_renk = renk,aciklama = aciklama,silinme_bilgisi = silinmedurumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        elif silinmedurumu == "2":
            silinmedurumu = True
            gider_kategorisi.objects.filter(id = id).update(gider_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gider_kategori_adi = proje_tip_adi,gider_kategorisi_renk = renk,aciklama = aciklama,silinme_bilgisi = silinmedurumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        else:
            gider_kategorisi.objects.filter(id = id).update(gider_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gider_kategori_adi = proje_tip_adi,gider_kategorisi_renk = renk,aciklama = aciklama,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    else:
        if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gider_kategorisi_guncelleme:
                        proje_tip_adi   = request.POST.get("yetkili_adi")
                        aciklama = request.POST.get("aciklama")
                        renk = request.POST.get("renk")
                        proje_tip_adi   = request.POST.get("yetkili_adi")
                        maaslarda_kullan = request.POST.get("maaslarda_kullan")
                        avanslarda_kullan = request.POST.get("avanslarda_kullan")
                        if maaslarda_kullan == "1":
                            maaslarda_kullan = False
                        else:
                            maaslarda_kullan = True
                        if avanslarda_kullan == "1":
                            avanslarda_kullan = False
                        else:
                            avanslarda_kullan = True
                        gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db,id = id).update(gider_kategori_adi = proje_tip_adi,gider_kategorisi_renk = renk,aciklama = aciklama,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
        else:
                proje_tip_adi   = request.POST.get("yetkili_adi")
                aciklama = request.POST.get("aciklama")
                renk = request.POST.get("renk")
                proje_tip_adi   = request.POST.get("yetkili_adi")
                maaslarda_kullan = request.POST.get("maaslarda_kullan")
                avanslarda_kullan = request.POST.get("avanslarda_kullan")
                if maaslarda_kullan == "1":
                    maaslarda_kullan = False
                else:
                    maaslarda_kullan = True
                if avanslarda_kullan == "1":
                    avanslarda_kullan = False
                else:
                    avanslarda_kullan = True
                gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = request.user,id = id).update(gider_kategori_adi = proje_tip_adi,gider_kategorisi_renk = renk,aciklama = aciklama,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    return redirect("accounting:gider_kategorisi_tipleri")
#gider KAtegorisi Ekleme
def gider_kategorisi_ekleme_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            aciklama = request.POST.get("aciklama")
            renk = request.POST.get("renk")
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
            gider_kategorisi.objects.create(gider_kategoris_ait_bilgisi = users ,gider_kategori_adi = proje_tip_adi,gider_kategorisi_renk = renk,aciklama = aciklama,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    return redirect("accounting:gider_kategorisi_tipleri_2",hash)

#gider Kategorisi Silme

def gider_kategoisi_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        gider_kategorisi.objects.filter(id = id).update(silinme_bilgisi = True)
    
    return redirect("accounting:gider_kategorisi_tipleri_2",hash)
#gider düzenleme
def gider_kategorisi_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        aciklama = request.POST.get("aciklama")
        renk = request.POST.get("renk")
        if True:
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
        
            gider_kategorisi.objects.filter(id = id).update(gider_kategoris_ait_bilgisi = users ,gider_kategori_adi = proje_tip_adi,gider_kategorisi_renk = renk,aciklama = aciklama,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    
    return redirect("accounting:gider_kategorisi_tipleri_2",hash)


#cari işlemler
def cari_viev(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =cari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.cari_gosterme_izni:
                    profile = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =cari.objects.filter(Q(cari_kart_ait_bilgisi__first_name__icontains = search)|Q(cari_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.cari_gosterme_izni:
                        profile = cari.objects.filter(Q(cari_kart_ait_bilgisi = request.user.kullanicilar_db) & Q(cari_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = cari.objects.filter(Q(cari_kart_ait_bilgisi = request.user) & Q(cari_adi__icontains = search)& Q(silinme_bilgisi = False))

    content["santiyeler"] = profile
    return render(request,"muhasebe_page/cariler.html",content)
def cari_viev_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        profile = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =cari.objects.filter(Q(cari_kart_ait_bilgisi__first_name__icontains = search)|Q(cari_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = cari.objects.filter(Q(cari_kart_ait_bilgisi = request.user) & Q(cari_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/cariler.html",content)
#cari işlemler
#cari işlemler
def cari_views_details_2(request,id,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =cari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = users)
        k_gonder = get_object_or_404(cari,id = id)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.cari_detay_izni:
                    profile = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user.kullanicilar_db)
                    k_gonder = get_object_or_404(cari,id = id)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user)
            k_gonder = get_object_or_404(cari,id = id)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =cari.objects.filter(Q(cari_kart_ait_bilgisi__first_name__icontains = search)|Q(cari_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.cari_detay_izni:
                        profile = cari.objects.filter(Q(cari_kart_ait_bilgisi = request.user.kullanicilar_db) & Q(cari_adi__icontains = search)& Q(silinme_bilgisi = False))
                        k_gonder = get_object_or_404(cari,id = id)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = cari.objects.filter(Q(cari_kart_ait_bilgisi = request.user) & Q(cari_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["cari"] = k_gonder
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/cari_detay.html",content)

#cari ekleme
def cari_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            cari_Adi   = request.POST.get("cariadi")
            bakiye = request.POST.get("bakiye")
            konumu = request.POST.get("konumu")
            cari.objects.create(cari_kart_ait_bilgisi = users
                                ,cari_adi = cari_Adi,aciklama = konumu,telefon_numarasi = bakiye

                                )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.cari_olusturma:
                        cari_Adi   = request.POST.get("cariadi")
                        bakiye = request.POST.get("bakiye")
                        konumu = request.POST.get("konumu")
                        cari.objects.create(cari_kart_ait_bilgisi = request.user.kullanicilar_db
                            ,cari_adi = cari_Adi,aciklama = konumu,telefon_numarasi = bakiye)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                cari_Adi   = request.POST.get("cariadi")
                bakiye = request.POST.get("bakiye")
                konumu = request.POST.get("konumu")
                cari.objects.create(cari_kart_ait_bilgisi = request.user
                    ,cari_adi = cari_Adi,aciklama = konumu,telefon_numarasi = bakiye)

    return redirect("accounting:cari_viev_2",hash)

#cari silme
def cari_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        cari.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.cari_silme_izni:
                    cari.objects.filter(cari_kart_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            cari.objects.filter(cari_kart_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect("accounting:cari_viev_2",hash)


#cari düzenle
def cari_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        silinmedurumu = request.POST.get("silinmedurumu")
        konumu = request.POST.get("konumu")
        if silinmedurumu == "1":
            silinmedurumu = False
            cari.objects.filter(id = id).update(cari_kart_ait_bilgisi = users ,cari_adi = proje_tip_adi,aciklama = konumu,silinme_bilgisi = silinmedurumu)
        elif silinmedurumu == "2":
            silinmedurumu = True
            cari.objects.filter(id = id).update(cari_kart_ait_bilgisi = users ,cari_adi = proje_tip_adi,aciklama = konumu,silinme_bilgisi = silinmedurumu)
        else:
            cari.objects.filter(id = id).update(cari_kart_ait_bilgisi = users ,cari_adi = proje_tip_adi,aciklama = konumu)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.cari_guncelleme_izni:
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    konumu = request.POST.get("konumu")
                    cari.objects.filter(cari_kart_ait_bilgisi = request.user.kullanicilar_db,id = id).update(cari_adi = proje_tip_adi
                            ,aciklama = konumu)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("yetkili_adi")
            konumu = request.POST.get("konumu")
            cari.objects.filter(cari_kart_ait_bilgisi = request.user,id = id).update(cari_adi = proje_tip_adi
                    ,aciklama = konumu)
    return redirect("accounting:cari_viev_2",hash)
################################################3
#cari işlemler
def cari_views_details(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =cari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        k_gonder = get_object_or_404(cari,id = id)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.cari_detay_izni:
                    profile = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user.kullanicilar_db)
                    k_gonder = get_object_or_404(cari,id = id)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user)
            k_gonder = get_object_or_404(cari,id = id)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =cari.objects.filter(Q(cari_kart_ait_bilgisi__first_name__icontains = search)|Q(cari_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.cari_detay_izni:
                        profile = cari.objects.filter(Q(cari_kart_ait_bilgisi = request.user.kullanicilar_db) & Q(cari_adi__icontains = search)& Q(silinme_bilgisi = False))
                        k_gonder = get_object_or_404(cari,id = id)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = cari.objects.filter(Q(cari_kart_ait_bilgisi = request.user) & Q(cari_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["cari"] = k_gonder
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/cari_detay.html",content)

#cari ekleme
def cari_ekle(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            cari_Adi   = request.POST.get("cariadi")
            bakiye = request.POST.get("bakiye")
            konumu = request.POST.get("konumu")
            cari.objects.create(cari_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi )
                                ,cari_adi = cari_Adi,aciklama = konumu,telefon_numarasi = bakiye

                                )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.cari_olusturma:
                        cari_Adi   = request.POST.get("cariadi")
                        bakiye = request.POST.get("bakiye")
                        konumu = request.POST.get("konumu")
                        cari.objects.create(cari_kart_ait_bilgisi = request.user.kullanicilar_db
                            ,cari_adi = cari_Adi,aciklama = konumu,telefon_numarasi = bakiye)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                cari_Adi   = request.POST.get("cariadi")
                bakiye = request.POST.get("bakiye")
                konumu = request.POST.get("konumu")
                cari.objects.create(cari_kart_ait_bilgisi = request.user
                    ,cari_adi = cari_Adi,aciklama = konumu,telefon_numarasi = bakiye)

    return redirect("accounting:cari")

#cari silme
def cari_sil(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        cari.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.cari_silme_izni:
                    cari.objects.filter(cari_kart_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            cari.objects.filter(cari_kart_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect("accounting:cari")


#cari düzenle
def cari_duzenle(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        silinmedurumu = request.POST.get("silinmedurumu")
        konumu = request.POST.get("konumu")
        if silinmedurumu == "1":
            silinmedurumu = False
            cari.objects.filter(id = id).update(cari_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,cari_adi = proje_tip_adi,aciklama = konumu,silinme_bilgisi = silinmedurumu)
        elif silinmedurumu == "2":
            silinmedurumu = True
            cari.objects.filter(id = id).update(cari_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,cari_adi = proje_tip_adi,aciklama = konumu,silinme_bilgisi = silinmedurumu)
        else:
            cari.objects.filter(id = id).update(cari_kart_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,cari_adi = proje_tip_adi,aciklama = konumu)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.cari_guncelleme_izni:
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    konumu = request.POST.get("konumu")
                    cari.objects.filter(cari_kart_ait_bilgisi = request.user.kullanicilar_db,id = id).update(cari_adi = proje_tip_adi
                            ,aciklama = konumu)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("yetkili_adi")
            konumu = request.POST.get("konumu")
            cari.objects.filter(cari_kart_ait_bilgisi = request.user,id = id).update(cari_adi = proje_tip_adi
                    ,aciklama = konumu)
    return redirect("accounting:cari")
#cari işlemler

#gelir Etiketleri
def gelir_etiketi_tipleri(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =gelir_etiketi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_etiketi_gorme:
                    profile = gelir_etiketi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = gelir_etiketi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =gelir_etiketi.objects.filter(Q(gelir_kategoris_ait_bilgisi__last_name__icontains = search)|Q(gelir_kategori_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gelir_etiketi_gorme:
                        profile = gelir_etiketi.objects.filter(Q(gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db) & Q(gelir_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = gelir_etiketi.objects.filter(Q(gelir_kategoris_ait_bilgisi = request.user) & Q(gelir_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/gelir_etiketi.html",content)
def gelir_etiketi_tipleri_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        #profile =gelir_etiketi.objects.all()
        profile = gelir_etiketi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi =users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = gelir_etiketi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =gelir_etiketi.objects.filter(Q(gelir_kategoris_ait_bilgisi__last_name__icontains = search)|Q(gelir_kategori_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = gelir_etiketi.objects.filter(Q(gelir_kategoris_ait_bilgisi = request.user) & Q(gelir_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/gelir_etiketi.html",content)
#

def gelir_etiketi_ekleme(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            proje_tip_adi   = request.POST.get("yetkili_adi")

            gelir_etiketi.objects.create(gelir_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gelir_etiketi_adi = proje_tip_adi)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gelir_etiketi_olusturma:
                        proje_tip_adi   = request.POST.get("yetkili_adi")
                        gelir_etiketi.objects.create(gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db,gelir_etiketi_adi = proje_tip_adi)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                proje_tip_adi   = request.POST.get("yetkili_adi")
                gelir_etiketi.objects.create(gelir_kategoris_ait_bilgisi = request.user,gelir_etiketi_adi = proje_tip_adi)
    return redirect("accounting:gelir_etiketi_tipleri")

def gelir_etiketi_sil(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        gelir_etiketi.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_etiketi_silme:
                    gelir_etiketi.objects.filter(gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            gelir_etiketi.objects.filter(gelir_kategoris_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect("accounting:gelir_etiketi_tipleri")
def gelir_etiketi_duzenle(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")

        silinmedurumu = request.POST.get("silinmedurumu")
        if silinmedurumu == "1":
            silinmedurumu = False
            gelir_etiketi.objects.filter(id = id).update(gelir_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gelir_etiketi_adi = proje_tip_adi,silinme_bilgisi = silinmedurumu)
        elif silinmedurumu == "2":
            silinmedurumu = True
            gelir_etiketi.objects.filter(id = id).update(gelir_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gelir_etiketi_adi = proje_tip_adi,silinme_bilgisi = silinmedurumu)
        else:
           gelir_etiketi.objects.filter(id = id).update(gelir_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gelir_etiketi_adi = proje_tip_adi)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_etiketi_guncelleme:
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    gelir_etiketi.objects.filter(gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db,id = id).update(gelir_etiketi_adi = proje_tip_adi)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("yetkili_adi")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            gelir_etiketi.objects.filter(gelir_kategoris_ait_bilgisi = request.user,id = id).update(gelir_etiketi_adi = proje_tip_adi)
    return redirect("accounting:gelir_etiketi_tipleri")
#gelir Etiketleri

def gelir_etiketi_ekleme_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            proje_tip_adi   = request.POST.get("yetkili_adi")

            gelir_etiketi.objects.create(gelir_kategoris_ait_bilgisi = users ,gelir_etiketi_adi = proje_tip_adi)
        
    return redirect("accounting:gelir_etiketi_tipleri_2",hash)

def gelir_etiketi_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        gelir_etiketi.objects.filter(id = id).update(silinme_bilgisi = True)
    
    return redirect("accounting:gelir_etiketi_tipleri_2",hash)
def gelir_etiketi_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        gelir_etiketi.objects.filter(gelir_kategoris_ait_bilgisi = users,id = id).update(gelir_etiketi_adi = proje_tip_adi)
    return redirect("accounting:gelir_etiketi_tipleri_2",hash)
#gelir Etiketleri
#gider etiketi


#gider Etiketleri
def gider_etiketi_tipleri(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =gider_etiketi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_etiketi_gorme:
                    profile = gider_etiketi.objects.filter(silinme_bilgisi = False,gider_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = gider_etiketi.objects.filter(silinme_bilgisi = False,gider_kategoris_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =gider_etiketi.objects.filter(Q(gider_kategoris_ait_bilgisi__last_name__icontains = search)|Q(gider_kategori_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gider_etiketi_gorme:
                        profile = gider_etiketi.objects.filter(Q(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db) & Q(gider_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = gider_etiketi.objects.filter(Q(gider_kategoris_ait_bilgisi = request.user) & Q(gider_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/gider_etiketi.html",content)
def gider_etiketi_tipleri_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        #profile =gelir_etiketi.objects.all()
        profile = gider_etiketi.objects.filter(silinme_bilgisi = False,gider_kategoris_ait_bilgisi = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = gider_etiketi.objects.filter(silinme_bilgisi = False,gider_kategoris_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =gider_etiketi.objects.filter(Q(gider_kategoris_ait_bilgisi__last_name__icontains = search)|Q(gider_kategori_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = gider_etiketi.objects.filter(Q(gider_kategoris_ait_bilgisi = request.user) & Q(gider_kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/gider_etiketi.html",content)
#
def gider_etiketi_ekleme(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
            gider_etiketi.objects.create(gider_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gider_etiketi_adi = proje_tip_adi,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gider_etiketi_olusturma:
                        proje_tip_adi   = request.POST.get("yetkili_adi")
                        maaslarda_kullan = request.POST.get("maaslarda_kullan")
                        avanslarda_kullan = request.POST.get("avanslarda_kullan")
                        if maaslarda_kullan == "1":
                            maaslarda_kullan = False
                        else:
                            maaslarda_kullan = True
                        if avanslarda_kullan == "1":
                            avanslarda_kullan = False
                        else:
                            avanslarda_kullan = True
                        gider_etiketi.objects.create(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db,gider_etiketi_adi = proje_tip_adi,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                proje_tip_adi   = request.POST.get("yetkili_adi")
                maaslarda_kullan = request.POST.get("maaslarda_kullan")
                avanslarda_kullan = request.POST.get("avanslarda_kullan")
                if maaslarda_kullan == "1":
                    maaslarda_kullan = False
                else:
                    maaslarda_kullan = True
                if avanslarda_kullan == "1":
                    avanslarda_kullan = False
                else:
                    avanslarda_kullan = True
                gider_etiketi.objects.create(gider_kategoris_ait_bilgisi = request.user,gider_etiketi_adi = proje_tip_adi,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    return redirect("accounting:gider_etiketi_tipleri")

def gider_etiketi_sil(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        gider_etiketi.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_etiketi_silme:
                    gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect("accounting:gider_etiketi_tipleri")
def gider_etiketi_duzenle(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        if True:
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
        silinmedurumu = request.POST.get("silinmedurumu")
        if silinmedurumu == "1":
            silinmedurumu = False
            gider_etiketi.objects.filter(id = id).update(gider_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gider_etiketi_adi = proje_tip_adi,silinme_bilgisi = silinmedurumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        elif silinmedurumu == "2":
            silinmedurumu = True
            gider_etiketi.objects.filter(id = id).update(gider_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gider_etiketi_adi = proje_tip_adi,silinme_bilgisi = silinmedurumu,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        else:
            gider_etiketi.objects.filter(id = id).update(gider_kategoris_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,gider_etiketi_adi = proje_tip_adi,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_etiketi_guncelleme:
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    maaslarda_kullan = request.POST.get("maaslarda_kullan")
                    avanslarda_kullan = request.POST.get("avanslarda_kullan")
                    if maaslarda_kullan == "1":
                        maaslarda_kullan = False
                    else:
                        maaslarda_kullan = True
                    if avanslarda_kullan == "1":
                        avanslarda_kullan = False
                    else:
                        avanslarda_kullan = True
                    gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db,id = id).update(gider_etiketi_adi = proje_tip_adi,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("yetkili_adi")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
            gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = request.user,id = id).update(gider_etiketi_adi = proje_tip_adi,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    return redirect("accounting:gider_etiketi_tipleri")
#gider Etiketleri
#gider etiketi

#virman olayları
def virman_yapma(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            z = "/accounting/superadmintransfer/"+kullanici_bilgisi
            return redirect(z)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kasa_virman_olusturma_izni:
                        gonderen = request.POST.get("gonderen")
                        alici = request.POST.get("alici")
                        islemtarihi = request.POST.get("islemtarihi")
                        tutar = float(str(request.POST.get("tutar")).replace(",","."))
                        aciklama = request.POST.get("aciklama")
                        virman.objects.create(virman_ait_oldugu = request.user.kullanicilar_db,
                            virman_tarihi = islemtarihi,gonderen_kasa = get_object_or_404(Kasa,id = gonderen)
                            ,alici_kasa = get_object_or_404(Kasa,id = alici),tutar = tutar,
                            aciklama = aciklama
                            )
                        bakiye_dusme = get_object_or_404(Kasa,id = gonderen).bakiye
                        bakiye_yukseltme = get_object_or_404(Kasa,id = alici).bakiye
                        bakiye_dusme = bakiye_dusme - float(tutar)
                        bakiye_yukseltme = bakiye_yukseltme + float(tutar)
                        Kasa.objects.filter(id = gonderen).update(bakiye = bakiye_dusme)
                        Kasa.objects.filter(id = alici).update(bakiye = bakiye_yukseltme)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                gonderen = request.POST.get("gonderen")
                alici = request.POST.get("alici")
                islemtarihi = request.POST.get("islemtarihi")
                tutar = float(str(request.POST.get("tutar")).replace(",","."))
                aciklama = request.POST.get("aciklama")
                virman.objects.create(virman_ait_oldugu = request.user,
                    virman_tarihi = islemtarihi,gonderen_kasa = get_object_or_404(Kasa,id = gonderen)
                    ,alici_kasa = get_object_or_404(Kasa,id = alici),tutar = tutar,
                    aciklama = aciklama
                    )
                bakiye_dusme = get_object_or_404(Kasa,id = gonderen).bakiye
                bakiye_yukseltme = get_object_or_404(Kasa,id = alici).bakiye
                bakiye_dusme = bakiye_dusme - float(tutar)
                bakiye_yukseltme = bakiye_yukseltme + float(tutar)
                Kasa.objects.filter(id = gonderen).update(bakiye = bakiye_dusme)
                Kasa.objects.filter(id = alici).update(bakiye = bakiye_yukseltme)

    return redirect("accounting:kasa")
#virman olayları
def virman_yapma_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            if True:
                if True:
                    if True:
                        gonderen = request.POST.get("gonderen")
                        alici = request.POST.get("alici")
                        islemtarihi = request.POST.get("islemtarihi")
                        tutar = float(str(request.POST.get("tutar")).replace(",","."))
                        aciklama = request.POST.get("aciklama")
                        virman.objects.create(virman_ait_oldugu = users,
                            virman_tarihi = islemtarihi,gonderen_kasa = get_object_or_404(Kasa,id = gonderen)
                            ,alici_kasa = get_object_or_404(Kasa,id = alici),tutar = tutar,
                            aciklama = aciklama
                            )
                        bakiye_dusme = get_object_or_404(Kasa,id = gonderen).bakiye
                        bakiye_yukseltme = get_object_or_404(Kasa,id = alici).bakiye
                        bakiye_dusme = bakiye_dusme - float(tutar)
                        bakiye_yukseltme = bakiye_yukseltme + float(tutar)
                        Kasa.objects.filter(id = gonderen).update(bakiye = bakiye_dusme)
                        Kasa.objects.filter(id = alici).update(bakiye = bakiye_yukseltme)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kasa_virman_olusturma_izni:
                        gonderen = request.POST.get("gonderen")
                        alici = request.POST.get("alici")
                        islemtarihi = request.POST.get("islemtarihi")
                        tutar = float(str(request.POST.get("tutar")).replace(",","."))
                        aciklama = request.POST.get("aciklama")
                        virman.objects.create(virman_ait_oldugu = request.user.kullanicilar_db,
                            virman_tarihi = islemtarihi,gonderen_kasa = get_object_or_404(Kasa,id = gonderen)
                            ,alici_kasa = get_object_or_404(Kasa,id = alici),tutar = tutar,
                            aciklama = aciklama
                            )
                        bakiye_dusme = get_object_or_404(Kasa,id = gonderen).bakiye
                        bakiye_yukseltme = get_object_or_404(Kasa,id = alici).bakiye
                        bakiye_dusme = bakiye_dusme - float(tutar)
                        bakiye_yukseltme = bakiye_yukseltme + float(tutar)
                        Kasa.objects.filter(id = gonderen).update(bakiye = bakiye_dusme)
                        Kasa.objects.filter(id = alici).update(bakiye = bakiye_yukseltme)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                gonderen = request.POST.get("gonderen")
                alici = request.POST.get("alici")
                islemtarihi = request.POST.get("islemtarihi")
                tutar = float(str(request.POST.get("tutar")).replace(",","."))
                aciklama = request.POST.get("aciklama")
                virman.objects.create(virman_ait_oldugu = request.user,
                    virman_tarihi = islemtarihi,gonderen_kasa = get_object_or_404(Kasa,id = gonderen)
                    ,alici_kasa = get_object_or_404(Kasa,id = alici),tutar = tutar,
                    aciklama = aciklama
                    )
                bakiye_dusme = get_object_or_404(Kasa,id = gonderen).bakiye
                bakiye_yukseltme = get_object_or_404(Kasa,id = alici).bakiye
                bakiye_dusme = bakiye_dusme - float(tutar)
                bakiye_yukseltme = bakiye_yukseltme + float(tutar)
                Kasa.objects.filter(id = gonderen).update(bakiye = bakiye_dusme)
                Kasa.objects.filter(id = alici).update(bakiye = bakiye_yukseltme)

    return redirect("accounting:a_kasa_viev",hash)



def super_admin_virman(request,id):
    content = sozluk_yapisi()
    content["proje_tipleri"] = proje_tipi.objects.filter(proje_ait_bilgisi =  get_object_or_404(CustomUser,id = id))
    yetki(request)
    profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = get_object_or_404(CustomUser,id = id))
    content["santiyeler"] = profile
    if request.POST:
            gonderen = request.POST.get("gonderen")
            alici = request.POST.get("alici")
            islemtarihi = request.POST.get("islemtarihi")
            tutar = float(str(request.POST.get("tutar")).replace(",","."))
            aciklama = request.POST.get("aciklama")
            virman.objects.create(virman_ait_oldugu = get_object_or_404(CustomUser,id = id),
                virman_tarihi = islemtarihi,gonderen_kasa = get_object_or_404(Kasa,id = gonderen)
                ,alici_kasa = get_object_or_404(Kasa,id = alici),tutar = tutar,
                aciklama = aciklama
                )
            bakiye_dusme = get_object_or_404(Kasa,id = gonderen).bakiye
            bakiye_yukseltme = get_object_or_404(Kasa,id = alici).bakiye
            bakiye_dusme = bakiye_dusme - float(tutar)
            bakiye_yukseltme = bakiye_yukseltme + float(tutar)
            Kasa.objects.filter(id = gonderen).update(bakiye = bakiye_dusme)
            Kasa.objects.filter(id = alici).update(bakiye = bakiye_yukseltme)
            return redirect("accounting:kasa")
    return render(request,"muhasebe_page/super_admin_virman.html",content)


def virman_gondermeler(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =virman.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        content["kasalar"]  = Kasa.objects.filter(silinme_bilgisi = False)
    else:
        profile = virman.objects.filter(silinme_bilgisi = False,virman_ait_oldugu = request.user)
        content["kasalar"]  = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    if request.GET:
        gonderen_kasa = request.GET.get("gonderen_kasa")
        alici_kasa = request.GET.get("alici_kasa")
        tarih = request.GET.get("tarih")
        if super_admin_kontrolu(request):
            profile =virman.objects.filter()
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = virman.objects.filter(Q(virman_ait_oldugu = request.user) & Q(silinme_bilgisi = False))
        if gonderen_kasa:
            profile = profile.filter(gonderen_kasa= get_object_or_none(Kasa,id =gonderen_kasa ) )
        if alici_kasa:
            profile = profile.filter(alici_kasa=  get_object_or_none(Kasa,id =alici_kasa ))
        if tarih :
            profile = profile.filter(virman_tarihi = tarih)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/virman_raporu.html",content)

def virman_gondermeler_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        profile = virman.objects.filter(silinme_bilgisi = False,virman_ait_oldugu = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        content["kasalar"]  = Kasa.objects.filter(silinme_bilgisi = False)
    else:
        profile = virman.objects.filter(silinme_bilgisi = False,virman_ait_oldugu = request.user)
        content["kasalar"]  = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    if request.GET:
        gonderen_kasa = request.GET.get("gonderen_kasa")
        alici_kasa = request.GET.get("alici_kasa")
        tarih = request.GET.get("tarih")
        if super_admin_kontrolu(request):
            d = decode_id(hash)
            content["hashler"] = hash
            users = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = users
            profile = virman.objects.filter(silinme_bilgisi = False,virman_ait_oldugu = users)
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = virman.objects.filter(Q(virman_ait_oldugu = request.user) & Q(silinme_bilgisi = False))
        if gonderen_kasa:
            profile = profile.filter(gonderen_kasa= get_object_or_none(Kasa,id =gonderen_kasa ) )
        if alici_kasa:
            profile = profile.filter(alici_kasa=  get_object_or_none(Kasa,id =alici_kasa ))
        if tarih :
            profile = profile.filter(virman_tarihi = tarih)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/virman_raporu.html",content)
#virman olayları
#virman olayları
#ürünler olayları
def urun_viev(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =urunler.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.urun_gorme:
                    profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user.kullanicilar_db)
                    kategori = urun_kategorileri.objects.filter(kategrori_ait_oldugu = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user)
            kategori = urun_kategorileri.objects.filter(kategrori_ait_oldugu = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =urunler.objects.filter(Q(urun_ait_oldugu__first_name__icontains = search)|Q(urun_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.urun_gorme:
                        profile = urunler.objects.filter(Q(urun_ait_oldugu = request.user.kullanicilar_db) & Q(urun_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = urunler.objects.filter(Q(urun_ait_oldugu = request.user) & Q(urun_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["urun_kategorisi"] = kategori
    return render(request,"muhasebe_page/urunler.html",content)
def urun_viev_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =urunler.objects.filter(Q(urun_ait_oldugu__first_name__icontains = search)|Q(urun_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = urunler.objects.filter(Q(urun_ait_oldugu = request.user) & Q(urun_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"muhasebe_page/urunler.html",content)
#Ürün ekleme

def urun_ekle(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            kasa_Adi   = request.POST.get("kasaadi")
            bakiye = request.POST.get("bakiye")
            kategori = request.POST.get("kategori")
            urun_turu = request.POST.get("urun_turu")
            stok = request.POST.get("stok")
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
            if stok == "1":
                stok = True
            else:
                stok = False
            urunler.objects.create(urun_ait_oldugu = get_object_or_404(CustomUser,id = kullanici_bilgisi )
                                ,urun_adi = kasa_Adi,urun_fiyati = bakiye,urun_kategorisi = get_object_or_404(urun_kategorileri,id =kategori)
                                ,urun_turu_secim =  urun_turu,stok_mu = stok,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan
                                )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.urun_olusturma:
                        kasa_Adi   = request.POST.get("kasaadi")
                        bakiye = request.POST.get("bakiye")
                        kategori = request.POST.get("kategori")
                        urun_turu = request.POST.get("urun_turu")
                        stok = request.POST.get("stok")
                        maaslarda_kullan = request.POST.get("maaslarda_kullan")
                        avanslarda_kullan = request.POST.get("avanslarda_kullan")
                        if maaslarda_kullan == "1":
                            maaslarda_kullan = False
                        else:
                            maaslarda_kullan = True
                        if avanslarda_kullan == "1":
                            avanslarda_kullan = False
                        else:
                            avanslarda_kullan = True
                        if stok == "1":
                            stok = True
                        else:
                            stok = False
                        urunler.objects.create(urun_ait_oldugu = request.user.kullanicilar_db
                            ,urun_adi = kasa_Adi,urun_fiyati = bakiye,urun_kategorisi = get_object_or_404(urun_kategorileri,id =kategori)
                            ,urun_turu_secim =  urun_turu,stok_mu = stok ,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan )
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                kasa_Adi   = request.POST.get("kasaadi")
                bakiye = request.POST.get("bakiye")
                kategori = request.POST.get("kategori")
                urun_turu = request.POST.get("urun_turu")
                stok = request.POST.get("stok")
                maaslarda_kullan = request.POST.get("maaslarda_kullan")
                avanslarda_kullan = request.POST.get("avanslarda_kullan")
                if maaslarda_kullan == "1":
                    maaslarda_kullan = False
                else:
                    maaslarda_kullan = True
                if avanslarda_kullan == "1":
                    avanslarda_kullan = False
                else:
                    avanslarda_kullan = True
                if stok == "1":
                    stok = True
                else:
                    stok = False
                urunler.objects.create(urun_ait_oldugu = request.user
                    ,urun_adi = kasa_Adi,urun_fiyati = bakiye,urun_kategorisi = get_object_or_404(urun_kategorileri,id =kategori)
                    ,urun_turu_secim =  urun_turu,stok_mu = stok,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)

    return redirect("accounting:urun_viev")

#ürün ekle
#ürün_sil
def urun_sil(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        urunler.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.urun_silme:
                    urunler.objects.filter(urun_ait_oldugu = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            urunler.objects.filter(urun_ait_oldugu = request.user,id = id).update(silinme_bilgisi = True)
    return redirect("accounting:urun_viev")


#ürün Sil

#ürün Düzenle
#kasa düzenle
def urun_duzenle(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("kasaadi")
        silinmedurumu = request.POST.get("silinmedurumu")
        bakiye = request.POST.get("bakiye")
        kategori = request.POST.get("kategori")
        urun_turu = request.POST.get("urun_turu")
        stok = request.POST.get("stok")
        maaslarda_kullan = request.POST.get("maaslarda_kullan")
        avanslarda_kullan = request.POST.get("avanslarda_kullan")
        if maaslarda_kullan == "1":
            maaslarda_kullan = False
        else:
            maaslarda_kullan = True
        if avanslarda_kullan == "1":
            avanslarda_kullan = False
        else:
            avanslarda_kullan = True
        if stok == "1":
            stok = True
        else:
            stok = False
        if silinmedurumu == "1":
            silinmedurumu = False
            urunler.objects.filter(id = id).update(urun_ait_oldugu = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,urun_adi = proje_tip_adi,urun_fiyati = bakiye,silinme_bilgisi = silinmedurumu,urun_kategorisi = get_object_or_404(urun_kategorileri,id =kategori)
                                                   ,urun_turu_secim =  urun_turu,stok_mu = stok,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        elif silinmedurumu == "2":
            silinmedurumu = True
            urunler.objects.filter(id = id).update(urun_ait_oldugu = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,urun_adi = proje_tip_adi,urun_fiyati = bakiye,silinme_bilgisi = silinmedurumu,urun_kategorisi = get_object_or_404(urun_kategorileri,id =kategori)
                                                   ,urun_turu_secim =  urun_turu,stok_mu = stok,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
        else:
            urunler.objects.filter(id = id).update(urun_ait_oldugu = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,urun_adi = proje_tip_adi,urun_fiyati = bakiye,urun_kategorisi = get_object_or_404(urun_kategorileri,id =kategori)
                                                   ,urun_turu_secim =  urun_turu,stok_mu = stok,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.urun_guncelleme:
                    proje_tip_adi   = request.POST.get("kasaadi")
                    bakiye = request.POST.get("bakiye")
                    kategori = request.POST.get("kategori")
                    urun_turu = request.POST.get("urun_turu")
                    stok = request.POST.get("stok")
                    maaslarda_kullan = request.POST.get("maaslarda_kullan")
                    avanslarda_kullan = request.POST.get("avanslarda_kullan")
                    if maaslarda_kullan == "1":
                        maaslarda_kullan = False
                    else:
                        maaslarda_kullan = True
                    if avanslarda_kullan == "1":
                        avanslarda_kullan = False
                    else:
                        avanslarda_kullan = True
                    if stok == "1":
                        stok = True
                    else:
                        stok = False
                    urunler.objects.filter(urun_ait_oldugu = request.user.kullanicilar_db,id = id).update(urun_adi = proje_tip_adi
                            ,urun_fiyati = bakiye,urun_kategorisi = get_object_or_404(urun_kategorileri,id =kategori),urun_turu_secim =  urun_turu,stok_mu = stok,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("kasaadi")
            bakiye = request.POST.get("bakiye")
            kategori = request.POST.get("kategori")
            urun_turu = request.POST.get("urun_turu")
            stok = request.POST.get("stok")
            maaslarda_kullan = request.POST.get("maaslarda_kullan")
            avanslarda_kullan = request.POST.get("avanslarda_kullan")
            if maaslarda_kullan == "1":
                maaslarda_kullan = False
            else:
                maaslarda_kullan = True
            if avanslarda_kullan == "1":
                avanslarda_kullan = False
            else:
                avanslarda_kullan = True
            if stok == "1":
                stok = True
            else:
                stok = False
            urunler.objects.filter(urun_ait_oldugu = request.user,id = id).update(urun_adi = proje_tip_adi
                    ,urun_fiyati = bakiye,urun_kategorisi = get_object_or_404(urun_kategorileri,id =kategori),urun_turu_secim =  urun_turu,stok_mu = stok,avans_icin_kullan =avanslarda_kullan,
                                maas_icin_kullan = maaslarda_kullan)
    return redirect("accounting:urun_viev")

#ürün Düzenle
#ürünler olayları

#Gelirler Sayfası
def gelirler_sayfasi(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Gelir_Bilgisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_faturasi_gorme_izni:
                    profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db,silinme_bilgisi = False).order_by("-fatura_tarihi")
                    content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user,silinme_bilgisi = False).order_by("-fatura_tarihi")
            content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)

    content["santiyeler_i"] = profile
    return render(request,"muhasebe_page/deneme_gelir_duzeltme.html",content)
def gelirler_sayfasi_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users).order_by("-fatura_tarihi")
        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-fatura_tarihi")
        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    if request.GET:
        search = request.GET.get("search")
        tarih = request.GET.get("tarih")
        if search:
            if super_admin_kontrolu(request):
                profile =Gelir_Bilgisi.objects.filter(Q(fatura_no__icontains = search)|Q(cari_bilgisi__cari_adi__icontains = search)| Q(gelir_kime_ait_oldugu__first_name__icontains = search)| Q(aciklama__icontains = search) | Q(gelir_kategorisi__gelir_kategori_adi__icontains = search))
                kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
                content["kullanicilar"] =kullanicilar
            else:
                profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).filter(Q(fatura_no__icontains = search)|Q(cari_bilgisi__cari_adi__icontains = search)| Q(aciklama__icontains = search) | Q(gelir_kategorisi__gelir_kategori_adi__icontains = search))
                content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        if tarih :
            profile = profile.filter(Q(fatura_tarihi__lte  = tarih) & Q(vade_tarihi__gte  = tarih) )


    content["santiyeler_i"] = profile
    content["santiyeler"] = profile[:1]
    content["giderler_bilgisi"] = profile
    return render(request,"muhasebe_page/deneme_gelir_duzeltme.html",content)
#
def gelir_ekle(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        urunler_bilgisi = ""
        cari_bilgileri = ""
        kategori_bilgisi = ""
        etiketler =  ""
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_faturasi_kesme_izni:
                    profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                    urunler_bilgisi = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user.kullanicilar_db)
                    cari_bilgileri = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user.kullanicilar_db)
                    kategori_bilgisi = gelir_kategorisi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                    etiketler = gelir_etiketi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
            urunler_bilgisi = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user)
            cari_bilgileri = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user)
            kategori_bilgisi = gelir_kategorisi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user)
            etiketler = gelir_etiketi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user)
    content["gelir_kategoerisi"] = kategori_bilgisi
    content["gelir_etiketi"] = etiketler
    content["kasa"] = profile
    content["urunler"]  = urunler_bilgisi
    content["cari_bilgileri"] = cari_bilgileri
    return render(request,"muhasebe_page/gelir_faturasi.html",content)
#
def gelir_ekle_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = users)
        urunler_bilgisi = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = users)
        cari_bilgileri = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = users)
        kategori_bilgisi = gelir_kategorisi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = users)
        etiketler = gelir_etiketi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = users)
    else:
        profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        urunler_bilgisi = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user)
        cari_bilgileri = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user)
        kategori_bilgisi = gelir_kategorisi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user)
        etiketler = gelir_etiketi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user)
    content["gelir_kategoerisi"] = kategori_bilgisi
    content["gelir_etiketi"] = etiketler
    content["kasa"] = profile
    content["urunler"]  = urunler_bilgisi
    content["cari_bilgileri"] = cari_bilgileri
    return render(request,"muhasebe_page/gelir_faturasi.html",content)

def gelir_duzenle(request ,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        urunler_bilgisi =  ""
        cari_bilgileri = ""
        kategori_bilgisi = ""
        etiketler = ""
        gelir_bilgisi_ver =  get_object_or_none(Gelir_Bilgisi,id = id)
        urunleri = ""
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_faturasi_duzenleme_izni:
                    profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                    urunler_bilgisi = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user.kullanicilar_db)
                    cari_bilgileri = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user.kullanicilar_db)
                    kategori_bilgisi = gelir_kategorisi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                    etiketler = gelir_etiketi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                    gelir_bilgisi_ver =  get_object_or_none(Gelir_Bilgisi,id = id)
                    urunleri = gelir_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
            urunler_bilgisi = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user)
            cari_bilgileri = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = request.user)
            kategori_bilgisi = gelir_kategorisi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user)
            etiketler = gelir_etiketi.objects.filter(silinme_bilgisi = False,gelir_kategoris_ait_bilgisi = request.user)
            gelir_bilgisi_ver =  get_object_or_none(Gelir_Bilgisi,id = id)
            urunleri = gelir_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
    content["gelir_kategoerisi"] = kategori_bilgisi
    content["gelir_etiketi"] = etiketler
    content["kasa"] = profile
    content["urunler"]  = urunler_bilgisi
    content["cari_bilgileri"] = cari_bilgileri
    content["bilgi"] = gelir_bilgisi_ver
    content["urunler"] = urunleri
    return render(request,"muhasebe_page/gelir_faturasi_duzeltme.html",content)
def denme(request):
    return render(request,"a.html")
from django.http import JsonResponse
def search(request):
    term = request.GET.get('term', '')
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_faturasi_kesme_izni or a.izinler.gider_faturasi_kesme_izni:
                    user = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
    else:
        user = request.user
    results = urunler.objects.filter(urun_adi__icontains=term, urun_ait_oldugu=user)
    suggestions = [{'label': result.urun_adi, 'value': result.urun_fiyati} for result in results]
    print("oldu mu yav")
    return JsonResponse(suggestions, safe=False)
def cariler_bilgisi(request):
    term = request.GET.get('term', '')
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_faturasi_kesme_izni or a.izinler.gider_faturasi_kesme_izni:
                    user = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
    else:
        user = request.user
    results = cari.objects.filter(cari_adi__icontains=term, cari_kart_ait_bilgisi=user)
    suggestions = [{'label': result.cari_adi, 'value':result.aciklama} for result in results]
    return JsonResponse(suggestions, safe=False)
def search_2(request,hash):
    term = request.GET.get('term', '')
    d = decode_id(hash)
    user = get_object_or_404(CustomUser,id = d)
    #user = request.user
    results = urunler.objects.filter(urun_adi__icontains=term, urun_ait_oldugu=user)
    suggestions = [{'label': result.urun_adi, 'value': result.urun_fiyati} for result in results]
    print("oldu mu yav")
    return JsonResponse(suggestions, safe=False)
def cariler_bilgisi_2(request,hash):
    term = request.GET.get('term', '')
    d = decode_id(hash)
    user = get_object_or_404(CustomUser,id = d)
    results = cari.objects.filter(cari_adi__icontains=term, cari_kart_ait_bilgisi=user)
    suggestions = [{'label': result.cari_adi, 'value':result.aciklama} for result in results]
    return JsonResponse(suggestions, safe=False)

def gelir_faturasi_kaydet(request):
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gelir_faturasi_kesme_izni:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
    else:
        kullanici = request.user
    if request.POST:
        musteri_bilgisi  = request.POST.get("musteri_bilgisi")
        daterange = request.POST.get("daterange")       
        cari_aciklma = request.POST.get("cari_aciklma")
        etiketler = request.POST.getlist("etiketler")
        faturano = request.POST.get("faturano")
        urunadi = request.POST.getlist("urunadi")
        miktari = request.POST.getlist("miktari")
        bfiyatInput = request.POST.getlist("bfiyatInput")
        indirim = request.POST.getlist("indirim")
        aciklama = request.POST.getlist("aciklama")
        doviz_kuru = request.POST.get("doviz_kuru")
        profile = request.FILES.get("fatura_belgesi")
        gelir_kate = request.POST.get("gelir_kategorisi_gonder")
        kate = get_object_or_none(gelir_kategorisi,id = gelir_kate)
        print(kate,"kategorisi gönder",gelir_kate)
        cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici)
        if cari_bilgisi:
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gelir_Bilgisi.objects.create(gelir_kime_ait_oldugu = kullanici,
            cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisii =kate,doviz = doviz_kuru,aciklama = cari_aciklma
                                         )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gelir_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i])
                    if urun:
                        if indirim[i] == "":
                            a = 0
                        else:
                            a = indirim[i]
                        gelir_urun_bilgisi_bi = gelir_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(a),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=kullanici,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gelir_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )

        else:
            cari_bilgisi = cari.objects.create(cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici,aciklama = cari_aciklma)
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gelir_Bilgisi.objects.create(gelir_kime_ait_oldugu = kullanici,
            cari_bilgisi = get_object_or_none(cari,id = cari_bilgisi.id),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisii =kate,doviz = doviz_kuru,aciklama = cari_aciklma
            )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gelir_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i])
                    if urun:
                        gelir_urun_bilgisi_bi = gelir_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=kullanici,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gelir_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
        toplam_tutar = 0
        gelir_urun_bilgisi_al =gelir_urun_bilgisi.objects.filter(gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id)) 
        for i in gelir_urun_bilgisi_al:
            toplam_tutar += (i.urun_fiyati)*(i.urun_adeti)-i.urun_indirimi
        Gelir_Bilgisi.objects.filter(id =new_project.id ).update(toplam_tutar =toplam_tutar,kalan_tutar =toplam_tutar )
        gelir_qr.objects.create(gelir_kime_ait_oldugu = get_object_or_none(Gelir_Bilgisi,id = new_project.id))
        if profile:
            u = Gelir_Bilgisi.objects.get(id = new_project.id )
            u.fatura_gorseli = profile
            u.save()
    return redirect("accounting:gelirler_sayfasi")

def gelir_faturasi_kaydet_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        musteri_bilgisi  = request.POST.get("musteri_bilgisi")
        daterange = request.POST.get("daterange")
        gelir_kategorisii = request.POST.get("gelir_kategorisi")
        cari_aciklma = request.POST.get("cari_aciklma")
        etiketler = request.POST.getlist("etiketler")
        faturano = request.POST.get("faturano")
        urunadi = request.POST.getlist("urunadi")
        miktari = request.POST.getlist("miktari")
        bfiyatInput = request.POST.getlist("bfiyatInput")
        indirim = request.POST.getlist("indirim")
        aciklama = request.POST.getlist("aciklama")
        doviz_kuru = request.POST.get("doviz_kuru")
        profile = request.FILES.get("fatura_belgesi")
        cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = users)
        if cari_bilgisi:
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gelir_Bilgisi.objects.create(gelir_kime_ait_oldugu = users,
            cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = users),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisi = get_object_or_none( gelir_kategorisi,id =gelir_kategorisii),doviz = doviz_kuru,aciklama = cari_aciklma
                                         )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gelir_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=users,urun_adi = urunadi[i])
                    if urun:
                        if indirim[i] == "":
                            a = 0
                        else:
                            a = indirim[i]
                        gelir_urun_bilgisi_bi = gelir_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  users,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=request.user,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(a),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=users,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gelir_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  users,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )

        else:
            cari_bilgisi = cari.objects.create(cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = users,aciklama = cari_aciklma)
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gelir_Bilgisi.objects.create(gelir_kime_ait_oldugu = users,
            cari_bilgisi = get_object_or_none(cari,id = cari_bilgisi.id),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisi = get_object_or_none( gelir_kategorisi,id =gelir_kategorisii),doviz = doviz_kuru,aciklama = cari_aciklma
            )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gelir_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=users,urun_adi = urunadi[i])
                    if urun:
                        gelir_urun_bilgisi_bi = gelir_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  users,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=request.user,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=users,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gelir_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  users,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
        toplam_tutar = 0
        gelir_urun_bilgisi_al =gelir_urun_bilgisi.objects.filter(gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id)) 
        for i in gelir_urun_bilgisi_al:
            toplam_tutar += (i.urun_fiyati)*(i.urun_adeti)-i.urun_indirimi
        Gelir_Bilgisi.objects.filter(id =new_project.id ).update(toplam_tutar =toplam_tutar,kalan_tutar =toplam_tutar )
        gelir_qr.objects.create(gelir_kime_ait_oldugu = get_object_or_none(Gelir_Bilgisi,id = new_project.id))
        if profile:
            u = Gelir_Bilgisi.objects.get(id = new_project.id )
            u.fatura_gorseli = profile
            u.save()
    return redirect("accounting:gelirler_sayfasi_2",hash)


def gelir_odemesi_ekle(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.gelir_faturasi_makbuz_kesme_izni:
                pass
            else:
                return redirect("main:yetkisiz")
        else:
            return redirect("main:yetkisiz")
    else:
        pass
    if request.POST:
        faturabilgisi = request.POST.get("faturabilgisi")
        odemeturu = request.POST.get("odemeturu")
        kasabilgisi = request.POST.get("kasabilgisi")
        islemtarihi = request.POST.get("islemtarihi")
        islemtutari = request.POST.get("islemtutari")
        makbuznumarasi = request.POST.get("makbuznumarasi")
        aciklama_bilgisi = request.POST.get("aciklama_bilgisi")
        dosya = request.FILES.get("dosya")
        Gelir_odemesi.objects.create(
            gelir_kime_ait_oldugu =get_object_or_none(Gelir_Bilgisi, id = faturabilgisi),
            gelir_turu = odemeturu,kasa_bilgisi = get_object_or_none(Kasa,id = kasabilgisi),
            tutar  =islemtutari,tarihi = islemtarihi,gelir_makbuzu = dosya,
            makbuz_no = makbuznumarasi,aciklama = aciklama_bilgisi
        )
        toplam_tutar = 0
        a = get_object_or_none(Gelir_Bilgisi, id = faturabilgisi)
        toplam_tutar = a.kalan_tutar-float(islemtutari)
        b = Gelir_Bilgisi.objects.filter(id =faturabilgisi ).update(kalan_tutar =toplam_tutar )
    return redirect("accounting:gelirler_sayfasi")
def gider_duzenle(request ,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_duzenleme_izni:
                    profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                    urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = request.user.kullanicilar_db)
                    cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = request.user.kullanicilar_db)
                    kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                    etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                    gelir_bilgisi_ver =  get_object_or_none(Gider_Bilgisi,id = id)
                    urunleri = gider_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
            urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = request.user)
            cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = request.user)
            kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = request.user)
            etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = request.user)
            gelir_bilgisi_ver =  get_object_or_none(Gider_Bilgisi,id = id)
            urunleri = gider_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
        content["gelir_kategoerisi"] = kategori_bilgisi
        content["gelir_etiketi"] = etiketler
        content["kasa"] = profile
        content["urunler"]  = urunler_bilgisi
        content["cari_bilgileri"] = cari_bilgileri
        content["bilgi"] = gelir_bilgisi_ver
        content["urunler"] = urunleri
    return render(request,"muhasebe_page/gider_faturasi_duzeltme.html",content)
#Gelirler Sayfası
def makbuz_sil(request):
    if request.POST:
        gelir_gider = request.POST.get("makbuz")
        makbuz_bilgisi = request.POST.get("makbuzidsi")
        if gelir_gider == "0":
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gelir_faturasi_makbuz_silme_izni:
                        pass
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                pass
            gider = get_object_or_none(Gelir_odemesi,id = makbuz_bilgisi ).gelir_kime_ait_oldugu
            Gelir_odemesi.objects.filter(id = makbuz_bilgisi).delete()
            gider_odemelrei = Gelir_odemesi.objects.filter(gelir_kime_ait_oldugu =gider)
            toplam_deger = gider.toplam_tutar
            for i in gider_odemelrei:
                toplam_deger -= i.tutar
            Gelir_Bilgisi.objects.filter(id = gider.id).update(kalan_tutar = toplam_deger)
            return redirect("accounting:gelirler_sayfasi")
        elif gelir_gider == "1":
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gider_faturasi_makbuz_silme_izni:
                        pass
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                pass
            gider = get_object_or_none(Gider_odemesi,id = makbuz_bilgisi ).gelir_kime_ait_oldugu
            Gider_odemesi.objects.filter(id = makbuz_bilgisi).delete()
            gider_odemelrei = Gider_odemesi.objects.filter(gelir_kime_ait_oldugu =gider)
            toplam_deger = gider.toplam_tutar
            for i in gider_odemelrei:
                toplam_deger -= i.tutar
            Gider_Bilgisi.objects.filter(id = gider.id).update(kalan_tutar = toplam_deger)
            return redirect("accounting:giderler_sayfasi")
#Gider Sayfası
def giderler_sayfasi(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Gider_Bilgisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_gorme_izni:
                    profile = Gider_Bilgisi.objects.filter(silinme_bilgisi = False,gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-fatura_tarihi")
                    content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Gider_Bilgisi.objects.filter(silinme_bilgisi = False,gelir_kime_ait_oldugu = request.user).order_by("-fatura_tarihi")
            content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    content["santiyeler_i"] = profile
    content["santiyeler"] = profile[:1]
    content["giderler_bilgisi"] = profile
    return render(request,"muhasebe_page/deneme_gider.html",content)
def giderler_sayfasi_borc(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Gider_Bilgisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_gorme_izni:
                    bilgi_ver = Gider_Bilgisi.objects.filter(silinme_bilgisi = False,gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-fatura_tarihi")
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            bilgi_ver = Gider_Bilgisi.objects.filter(silinme_bilgisi = False,gelir_kime_ait_oldugu = request.user).order_by("-fatura_tarihi")
        sonuc = []
        for i in bilgi_ver:
            y =  gider_urun_bilgisi.objects.filter(gider_bilgis = i)
            urun_tutari = 0
            for j in y:
                urun_tutari = urun_tutari + (float(j.urun_adeti)*float(j.urun_fiyati))
            y =  Gider_odemesi.objects.filter(gelir_kime_ait_oldugu = i)
            odeme_tutari = 0
            for j in y:
                odeme_tutari = odeme_tutari + float(j.tutar)
            if urun_tutari > odeme_tutari:
                sonuc.append(i)
        profile = sonuc
        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    if request.GET:
        search = request.GET.get("search")
        tarih = request.GET.get("tarih")
        if search:
            if super_admin_kontrolu(request):
                profile =Gider_Bilgisi.objects.filter(Q(fatura_no__icontains = search)|Q(cari_bilgisi__cari_adi__icontains = search)| Q(gelir_kime_ait_oldugu__first_name__icontains = search)| Q(aciklama__icontains = search) | Q(gelir_kategorisi__gelir_kategori_adi__icontains = search))
                kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
                content["kullanicilar"] =kullanicilar
            else:
                profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).filter(Q(fatura_no__icontains = search)|Q(cari_bilgisi__cari_adi__icontains = search)|  Q(aciklama__icontains = search) | Q(gelir_kategorisi__gider_kategori_adi__icontains = search))
                content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        if tarih :
            profile = profile.filter(Q(fatura_tarihi__lte  = tarih) & Q(vade_tarihi__gte  = tarih) )
   

    content["santiyeler_i"] = profile
    content["santiyeler"] = profile
    content["top"]  = profile
    return render(request,"muhasebe_page/deneme_gider.html",content)
#
def giderler_sayfasi_borc_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        bilgi_ver = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users).order_by("-fatura_tarihi")
        sonuc = []
        for i in bilgi_ver:
            y =  gider_urun_bilgisi.objects.filter(gider_bilgis = i)
            urun_tutari = 0
            for j in y:
                urun_tutari = urun_tutari + (float(j.urun_adeti)*float(j.urun_fiyati))
            y =  Gider_odemesi.objects.filter(gelir_kime_ait_oldugu = i)
            odeme_tutari = 0
            for j in y:
                odeme_tutari = odeme_tutari + float(j.tutar)
            if urun_tutari > odeme_tutari:
                sonuc.append(i)
        profile = sonuc
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        bilgi_ver = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-fatura_tarihi")
        sonuc = []
        for i in bilgi_ver:
            y =  gider_urun_bilgisi.objects.filter(gider_bilgis = i)
            urun_tutari = 0
            for j in y:
                urun_tutari = urun_tutari + (float(j.urun_adeti)*float(j.urun_fiyati))
            y =  Gider_odemesi.objects.filter(gelir_kime_ait_oldugu = i)
            odeme_tutari = 0
            for j in y:
                odeme_tutari = odeme_tutari + float(j.tutar)
            if urun_tutari > odeme_tutari:
                sonuc.append(i)
        profile = sonuc
        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    if request.GET:
        search = request.GET.get("search")
        tarih = request.GET.get("tarih")
        if search:
            if super_admin_kontrolu(request):
                profile =Gider_Bilgisi.objects.filter(Q(fatura_no__icontains = search)|Q(cari_bilgisi__cari_adi__icontains = search)| Q(gelir_kime_ait_oldugu__first_name__icontains = search)| Q(aciklama__icontains = search) | Q(gelir_kategorisi__gelir_kategori_adi__icontains = search))
                kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
                content["kullanicilar"] =kullanicilar
            else:
                profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).filter(Q(fatura_no__icontains = search)|Q(cari_bilgisi__cari_adi__icontains = search)|  Q(aciklama__icontains = search) | Q(gelir_kategorisi__gider_kategori_adi__icontains = search))
                content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        if tarih :
            profile = profile.filter(Q(fatura_tarihi__lte  = tarih) & Q(vade_tarihi__gte  = tarih) )


    content["santiyeler_i"] = profile
    content["santiyeler"] = profile
    content["top"]  = profile
    return render(request,"muhasebe_page/deneme_gider.html",content)
#
def giderler_sayfasi_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users).order_by("-fatura_tarihi")
        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-fatura_tarihi")
        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    if request.GET:
        search = request.GET.get("search")
        tarih = request.GET.get("tarih")
        if search:
            if super_admin_kontrolu(request):
                profile =Gider_Bilgisi.objects.filter(Q(fatura_no__icontains = search)|Q(cari_bilgisi__cari_adi__icontains = search)| Q(gelir_kime_ait_oldugu__first_name__icontains = search)| Q(aciklama__icontains = search) | Q(gelir_kategorisi__gelir_kategori_adi__icontains = search))
                kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
                content["kullanicilar"] =kullanicilar
            else:
                profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).filter(Q(fatura_no__icontains = search)|Q(cari_bilgisi__cari_adi__icontains = search)|  Q(aciklama__icontains = search) | Q(gelir_kategorisi__gider_kategori_adi__icontains = search))
                content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        if tarih :
            profile = profile.filter(Q(fatura_tarihi__lte  = tarih) & Q(vade_tarihi__gte  = tarih) )

    content["santiyeler_i"] = profile
    return render(request,"muhasebe_page/deneme_gider.html",content)

def gider_ekle(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        urunler_bilgisi = ""
        cari_bilgileri = ""
        kategori_bilgisi =""
        etiketler =  ""

    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_kesme_izni:
                    profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                    urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = request.user.kullanicilar_db)
                    cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = request.user.kullanicilar_db)
                    kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                    etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
            urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = request.user)
            cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = request.user)
            kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = request.user)
            etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = request.user)
    content["gelir_kategoerisi"] = kategori_bilgisi
    content["gelir_etiketi"] = etiketler
    content["kasa"] = profile
    content["urunler"]  = urunler_bilgisi
    content["cari_bilgileri"] = cari_bilgileri
    return render(request,"muhasebe_page/gider_faturasi.html",content)
#

def gider_ekle_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = users)
        urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = users)
        cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = users)
        kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = users)
        etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        

    else:
        profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = request.user)
        cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = request.user)
        kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = request.user)
        etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = request.user)
    content["gelir_kategoerisi"] = kategori_bilgisi
    content["gelir_etiketi"] = etiketler
    content["kasa"] = profile
    content["urunler"]  = urunler_bilgisi
    content["cari_bilgileri"] = cari_bilgileri
    return render(request,"muhasebe_page/gider_faturasi.html",content)
#
def gider_faturasi_kaydet(request):
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_kesme_izni:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
    else:
        kullanici = request.user
    if request.POST:
        musteri_bilgisi  = request.POST.get("musteri_bilgisi")
        daterange = request.POST.get("daterange")
        gelir_kategorisii = request.POST.get("gelir_kategorisi_gonder")
        cari_aciklma = request.POST.get("cari_aciklma")
        etiketler = request.POST.getlist("etiketler")
        faturano = request.POST.get("faturano")
        urunadi = request.POST.getlist("urunadi")
        miktari = request.POST.getlist("miktari")
        bfiyatInput = request.POST.getlist("bfiyatInput")
        indirim = request.POST.getlist("indirim")
        aciklama = request.POST.getlist("aciklama")
        aciklama_id = request.POST.getlist('aciklama_id')
        doviz_kuru = request.POST.get("doviz_kuru")
        profile = request.FILES.get("fatura_belgesi")
        print(gelir_kategorisii,"veri verme")
        cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici)
        if cari_bilgisi:
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gider_Bilgisi.objects.create(gelir_kime_ait_oldugu = kullanici,
            cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisii_id = gelir_kategorisii,doviz = doviz_kuru,aciklama = cari_aciklma
                                         )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gider_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i])
                    if urun:
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=kullanici,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )

        else:
            cari_bilgisi = cari.objects.create(cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici,aciklama = cari_aciklma)
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gider_Bilgisi.objects.create(gelir_kime_ait_oldugu = kullanici,
            cari_bilgisi = get_object_or_none(cari,id = cari_bilgisi.id),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisii_id =get_object_or_none(gelir_kategorisi,id = gelir_kategorisii),doviz = doviz_kuru,aciklama = cari_aciklma
                                         )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gider_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i])
                    if urun:
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=kullanici,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
        toplam_tutar = 0
        gelir_urun_bilgisi_al =gider_urun_bilgisi.objects.filter(gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id)) 
        for i in gelir_urun_bilgisi_al:
            toplam_tutar += (i.urun_fiyati)*(i.urun_adeti)-i.urun_indirimi
        Gider_Bilgisi.objects.filter(id =new_project.id ).update(toplam_tutar =toplam_tutar,kalan_tutar =toplam_tutar )
        
        gider_qr.objects.create(gelir_kime_ait_oldugu = get_object_or_none(Gider_Bilgisi,id = new_project.id))
        print(aciklama_id,"gelen id")
        if profile:
            u = Gider_Bilgisi.objects.get(id = new_project.id )
            u.fatura_gorseli = profile
            u.save()
    return redirect("accounting:giderler_sayfasi")

def gider_faturasi_kaydet_2(request,hash):
    print("kayıt")
    content = {}
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        musteri_bilgisi  = request.POST.get("musteri_bilgisi")
        daterange = request.POST.get("daterange")
        gelir_kategorisii = request.POST.get("gelir_kategorisi")
        cari_aciklma = request.POST.get("cari_aciklma")
        etiketler = request.POST.getlist("etiketler")
        faturano = request.POST.get("faturano")
        urunadi = request.POST.getlist("urunadi")
        miktari = request.POST.getlist("miktari")
        bfiyatInput = request.POST.getlist("bfiyatInput")
        indirim = request.POST.getlist("indirim")
        aciklama = request.POST.getlist("aciklama")
        aciklama_id = request.POST.getlist('aciklama_id')
        doviz_kuru = request.POST.get("doviz_kuru")
        profile = request.FILES.get("fatura_belgesi")
        cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = users)
        if cari_bilgisi:
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gider_Bilgisi.objects.create(gelir_kime_ait_oldugu = users,
            cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = users),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisi = get_object_or_none( gelir_kategorisi,id =gelir_kategorisii),doviz = doviz_kuru,aciklama = cari_aciklma
                                         )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gider_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=users,urun_adi = urunadi[i])
                    if urun:
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  users,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=users,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=users,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  users,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )

        else:
            cari_bilgisi = cari.objects.create(cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = users,aciklama = cari_aciklma)
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gider_Bilgisi.objects.create(gelir_kime_ait_oldugu = users,
            cari_bilgisi = get_object_or_none(cari,id = cari_bilgisi.id),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisi = get_object_or_none( gelir_kategorisi,id =gelir_kategorisii),doviz = doviz_kuru,aciklama = cari_aciklma
                                         )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gider_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=users,urun_adi = urunadi[i])
                    if urun:
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  users,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=users,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=users,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  users,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
        toplam_tutar = 0
        gelir_urun_bilgisi_al =gider_urun_bilgisi.objects.filter(gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id)) 
        for i in gelir_urun_bilgisi_al:
            toplam_tutar += (i.urun_fiyati)*(i.urun_adeti)-i.urun_indirimi
        Gider_Bilgisi.objects.filter(id =new_project.id ).update(toplam_tutar =toplam_tutar,kalan_tutar =toplam_tutar )
        
        gider_qr.objects.create(gelir_kime_ait_oldugu = get_object_or_none(Gider_Bilgisi,id = new_project.id))
        print(aciklama_id,"gelen id")
        if profile:
            u = Gider_Bilgisi.objects.get(id = new_project.id )
            u.fatura_gorseli = profile
            u.save()
    return redirect("accounting:giderler_sayfasi")
#
def gider_odemesi_ekle(request):
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_makbuz_kesme_izni:
                    pass
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
    else:   
        pass        
    if request.POST:
        faturabilgisi = request.POST.get("faturabilgisi")
        odemeturu = request.POST.get("odemeturu")
        kasabilgisi = request.POST.get("kasabilgisi")
        islemtarihi = request.POST.get("islemtarihi")
        islemtutari = request.POST.get("islemtutari")
        makbuznumarasi = request.POST.get("makbuznumarasi")
        aciklama_bilgisi = request.POST.get("aciklama_bilgisi")
        dosya = request.FILES.get("dosya")
        Gider_odemesi.objects.create(
            gelir_kime_ait_oldugu =get_object_or_none(Gider_Bilgisi, id = faturabilgisi),
            gelir_turu = odemeturu,kasa_bilgisi = get_object_or_none(Kasa,id = kasabilgisi),
            tutar  =islemtutari,tarihi = islemtarihi,gelir_makbuzu = dosya,
            makbuz_no = makbuznumarasi,aciklama = aciklama_bilgisi
        )
        toplam_tutar = 0
        a = get_object_or_none(Gider_Bilgisi, id = faturabilgisi)
        toplam_tutar = a.kalan_tutar-float(islemtutari)
        b = Gider_Bilgisi.objects.filter(id =faturabilgisi ).update(kalan_tutar =toplam_tutar )
    return redirect("accounting:giderler_sayfasi")


def fatura_sil(request):
    if request.POST:
        print("Fatura sil Çalıştı")
        gelir_gider = request.POST.get("gelir_gelir")
        id_bilgisi  = request.POST.get("idbilgisicek")
        if gelir_gider == "0":
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gelir_faturasi_silme_izni:
                        pass
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                pass
            Gelir_Bilgisi.objects.filter(id = id_bilgisi).update(silinme_bilgisi = True)
            Gelir_odemesi.objects.filter(gelir_kime_ait_oldugu =get_object_or_404(Gelir_Bilgisi,id =id_bilgisi )).update(silinme_bilgisi = True)
            return redirect("accounting:gelirler_sayfasi")

        elif gelir_gider == "1":
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gider_faturasi_silme_izni:
                        pass
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                pass
            Gider_Bilgisi.objects.filter(id = id_bilgisi).update(silinme_bilgisi = True)
            Gider_odemesi.objects.filter(gelir_kime_ait_oldugu=get_object_or_404(Gider_Bilgisi,id =id_bilgisi )).update(silinme_bilgisi = True)
            return redirect("accounting:giderler_sayfasi")
#Gider Sayfası
from django.shortcuts import get_object_or_404

def gelir_gider_duzelt(request):
    if request.POST:
        bilgi = request.POST.get("bilgi")
        degisen = request.POST.get("degisen")
        musteri_bilgisi = request.POST.get("musteri_bilgisi")
        daterange = request.POST.get("daterange")
        gelir_kategorisii = request.POST.get("gelir_kategorisi_gonder")
        cari_aciklma = request.POST.get("cari_aciklma")
        etiketler = request.POST.getlist("etiketler")
        faturano = request.POST.get("faturano")
        urunadi = request.POST.getlist("urunadi")
        miktari = request.POST.getlist("miktari")
        bfiyatInput = request.POST.getlist("bfiyatInput")
        indirim = request.POST.getlist("indirim")
        aciklama = request.POST.getlist("aciklama")
        doviz_kuru = request.POST.get("doviz_kuru")
        profile = request.FILES.get("fatura_belgesi")
        if bilgi == "0":
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gelir_faturasi_duzenleme_izni:
                        kullanici = request.user.kullanicilar_db
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                kullanici = request.user
            gelir_bilgisi = get_object_or_404(Gelir_Bilgisi, id=degisen)
            gelir_bilgisi.gelir_kime_ait_oldugu = kullanici
            cari_bilgisi = get_object_or_none(cari, cari_adi=musteri_bilgisi, cari_kart_ait_bilgisi=kullanici)
            if not cari_bilgisi:
                cari_bilgisi = cari.objects.create(cari_adi=musteri_bilgisi, cari_kart_ait_bilgisi=kullanici, aciklama=cari_aciklma)
            date_range_parts = daterange.split(' - ')
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')
            if profile:
                gelir_bilgisi.fatura_gorseli = profile
            
            gelir_bilgisi.fatura_tarihi = fatura_tarihi
            gelir_bilgisi.vade_tarihi = vade_tarihi
            #gelir_bilgisi.fatura_no = faturano
            gelir_bilgisi.gelir_kategorisii = get_object_or_none(gelir_kategorisi, id=gelir_kategorisii)
            gelir_bilgisi.save()

            gelir_etiketi_sec = gelir_etiketi.objects.filter(id__in=etiketler)
            gelir_bilgisi.gelir_etiketi_sec.set(gelir_etiketi_sec)
            gelir_urun_bilgisi.objects.filter(gider_bilgis = get_object_or_404(Gelir_Bilgisi, id=degisen)).delete()
            for i in range(len(urunadi)):
                if urunadi[i] and miktari[i] and bfiyatInput[i]:
                    urun = get_object_or_none(urunler, urun_ait_oldugu=kullanici, urun_adi=urunadi[i])
                    if not urun:
                        urun = urunler.objects.create(urun_ait_oldugu=kullanici, urun_adi=urunadi[i], urun_fiyati=float(bfiyatInput[i]))

                    gelir_urun_bilgisi.objects.create(
                        urun_ait_oldugu=kullanici,
                        urun_bilgisi=urun,
                        urun_fiyati=bfiyatInput[i],
                        urun_indirimi=float(indirim[i]),
                        urun_adeti=int(miktari[i]),
                        gider_bilgis=gelir_bilgisi,
                        aciklama=aciklama[i]
                    )
            toplam_tutar = 0
            gelir_urun_bilgisi_al =gelir_urun_bilgisi.objects.filter(gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = degisen)) 
            for i in gelir_urun_bilgisi_al:
                toplam_tutar += (i.urun_fiyati)*(i.urun_adeti)-i.urun_indirimi
            kalan_bilgisi = toplam_tutar
            gelir_odemeleri = Gelir_odemesi.objects.filter(gelir_kime_ait_oldugu =  get_object_or_none(Gelir_Bilgisi,id = degisen))
            for i in gelir_odemeleri:
                kalan_bilgisi -= i.tutar
            Gelir_Bilgisi.objects.filter(id =degisen ).update(toplam_tutar =toplam_tutar,kalan_tutar =kalan_bilgisi )
        
            return redirect("accounting:gelirler_sayfasi")

        elif bilgi == "1":
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.gider_faturasi_duzenleme_izni:
                        kullanici = request.user.kullanicilar_db
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                kullanici = request.user
            gider_bilgisi = get_object_or_404(Gider_Bilgisi, id=degisen)
            gider_bilgisi.gelir_kime_ait_oldugu = kullanici
            cari_bilgisi = get_object_or_none(cari, cari_adi=musteri_bilgisi, cari_kart_ait_bilgisi=kullanici)
            if not cari_bilgisi:
                cari_bilgisi = cari.objects.create(cari_adi=musteri_bilgisi, cari_kart_ait_bilgisi=kullanici, aciklama=cari_aciklma)
            date_range_parts = daterange.split(' - ')
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            gider_bilgisi.fatura_tarihi = fatura_tarihi
            gider_bilgisi.vade_tarihi = vade_tarihi
            #gider_bilgisi.fatura_no = faturano
            if profile:
                gider_bilgisi.fatura_gorseli = profile
            gider_bilgisi.gelir_kategorisii = get_object_or_none(gider_kategorisi, id=gelir_kategorisii)
            gider_bilgisi.doviz = doviz_kuru
            gider_bilgisi.save()

            gelir_etiketi_sec = gider_etiketi.objects.filter(id__in=etiketler)
            gider_bilgisi.gelir_etiketi_sec.set(gelir_etiketi_sec)
            toplam_tutar = 0
            gider_urun_bilgisi.objects.filter(gider_bilgis = get_object_or_404(Gider_Bilgisi, id=degisen)).delete()
            for i in range(len(urunadi)):
                if urunadi[i] and miktari[i] and bfiyatInput[i]:
                    urun = get_object_or_none(urunler, urun_ait_oldugu=kullanici, urun_adi=urunadi[i])
                    if not urun:
                        urun = urunler.objects.create(urun_ait_oldugu=kullanici, urun_adi=urunadi[i], urun_fiyati=float(bfiyatInput[i]))

                    gider_urun_bilgisi.objects.create(
                        urun_ait_oldugu=kullanici,
                        urun_bilgisi=urun,
                        urun_fiyati=bfiyatInput[i],
                        urun_indirimi=float(indirim[i]),
                        urun_adeti=int(miktari[i]),
                        gider_bilgis=gider_bilgisi,
                        aciklama=aciklama[i]
                    )
            gelir_urun_bilgisi_al =gider_urun_bilgisi.objects.filter(gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = degisen)) 
            for i in gelir_urun_bilgisi_al:
                toplam_tutar += (i.urun_fiyati)*(i.urun_adeti)-i.urun_indirimi
            kalan_bilgisi = toplam_tutar
            gelir_odemeleri = Gider_odemesi.objects.filter(gelir_kime_ait_oldugu =  get_object_or_none(Gider_Bilgisi,id = degisen))
            for i in gelir_odemeleri:
                kalan_bilgisi -= i.tutar
            Gider_Bilgisi.objects.filter(id =degisen ).update(toplam_tutar =toplam_tutar,kalan_tutar =kalan_bilgisi )
        
            return redirect("accounting:giderler_sayfasi")

    return redirect("accounting:giderler_sayfasi")

def gider_gelir_ekleme(request):
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_kesme_izni or a.izinler.gelir_faturasi_kesme_izni:
                    user = request.user
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
    else:
        user = request.user
    tur = request.POST.get("tur")
    adi = request.POST.get("adi")
    aciklama= request.POST.get("aciklama")
    renk = request.POST.get("renk")
    print("selam")
    if tur == "0":
        gelir_kategorisi.objects.create(
            gelir_kategoris_ait_bilgisi = user,
            gelir_kategori_adi = adi,
            gelir_kategorisi_renk = renk,
            aciklama = aciklama
        )
        return redirect("accounting:gelir_ekle")
    elif tur == "1":
        gider_kategorisi.objects.create(
            gider_kategoris_ait_bilgisi = user,
            gider_kategori_adi = adi,
            gider_kategorisi_renk = renk,
            aciklama = aciklama
        )
        return redirect("accounting:gider_ekle")
def gider_gelir_ekleme_2(request,hash):
    d = decode_id(hash)
    user = get_object_or_404(CustomUser,id = d)
    tur = request.POST.get("tur")
    adi = request.POST.get("adi")
    aciklama= request.POST.get("aciklama")
    renk = request.POST.get("renk")
    print("selam")
    if tur == "0":
        gelir_kategorisi.objects.create(
            gelir_kategoris_ait_bilgisi = user,
            gelir_kategori_adi = adi,
            gelir_kategorisi_renk = renk,
            aciklama = aciklama
        )
        return redirect("accounting:gelir_ekle_2",hash)
    elif tur == "1":
        gider_kategorisi.objects.create(
            gider_kategoris_ait_bilgisi = user,
            gider_kategori_adi = adi,
            gider_kategorisi_renk = renk,
            aciklama = aciklama
        )
        return redirect("accounting:gider_ekle_2",hash)

from django.utils import timezone
def gider_gelir_etiketekleme(request):
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_kesme_izni or a.izinler.gelir_faturasi_kesme_izni:
                    user = request.user
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
    else:
        user = request.user
    tur = request.POST.get("tur2")
    adi = request.POST.get("adi2")

    if tur == "0":
        gelir_etiketi.objects.create(
            gelir_kategoris_ait_bilgisi = user,
            gelir_etiketi_adi = adi,kayit_tarihi = timezone.now()
        )
        return redirect("accounting:gelir_ekle")
    elif tur == "1":
        gider_etiketi.objects.create(
            gider_kategoris_ait_bilgisi = user,
            gider_etiketi_adi = adi,kayit_tarihi = timezone.now()
        )
        return redirect("accounting:gider_ekle")

def gider_gelir_etiketekleme_2(request,hash):
    d = decode_id(hash)
    user = get_object_or_404(CustomUser,id = d)
    user = request.user
    tur = request.POST.get("tur2")
    adi = request.POST.get("adi2")

    if tur == "0":
        gelir_etiketi.objects.create(
            gelir_kategoris_ait_bilgisi = user,
            gelir_etiketi_adi = adi,kayit_tarihi = timezone.now()
        )
        return redirect("accounting:gelir_ekle")
    elif tur == "1":
        gider_etiketi.objects.create(
            gider_kategoris_ait_bilgisi = user,
            gider_etiketi_adi = adi,kayit_tarihi = timezone.now()
        )
        return redirect("accounting:gider_ekle")


#Gelirler özeti
def gelirler_ozeti(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Gelir_Bilgisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_ozeti_gorme:
                    profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-fatura_tarihi")
                    content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-fatura_tarihi")
            content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)

    content["santiyeler_i"] = profile
    return render(request,"muhasebe_page/gelir_ozeti.html",content)
#Gelirler  hash
def gelirler_ozeti_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =Gelir_Bilgisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users).order_by("-fatura_tarihi")
        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = users)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_ozeti_gorme:
                    profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-fatura_tarihi")
                    content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-fatura_tarihi")
            content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)

    content["santiyeler_i"] = profile
    return render(request,"muhasebe_page/gelir_ozeti.html",content)
#Gider Sayfası
#Gider Sayfası
def giderler_ozeti(request):#gider_ozeti_gorme
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Gider_Bilgisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_ozeti_gorme:
                    profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-fatura_tarihi")
                    content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-fatura_tarihi")
            content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    content["santiyeler_i"] = profile
    content["santiyeler"] = profile[:1]
    content["giderler_bilgisi"] = profile
    return render(request,"muhasebe_page/gider_ozeti.html",content)

def giderler_ozeti_2(request,hash):#gider_ozeti_gorme
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =Gider_Bilgisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users).order_by("-fatura_tarihi")
        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = users)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_ozeti_gorme:
                    profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-fatura_tarihi")
                    content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-fatura_tarihi")
            content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
    content["santiyeler_i"] = profile
    content["santiyeler"] = profile[:1]
    content["giderler_bilgisi"] = profile
    return render(request,"muhasebe_page/gider_ozeti.html",content)


#Hesapğ eksta
def hesap_ekstra_durumu(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Gider_Bilgisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.hesap_ekstra_gorme:
                    profile = list(Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db))+list(Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user))
                    profile.sort(key=lambda x: x.kayit_tarihi)
                    
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = list(Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user))+list(Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user))
            profile.sort(key=lambda x: x.kayit_tarihi)
            

    content["santiyeler_i"] = profile
    return render(request,"muhasebe_page/hesap_eksta.html",content)
def hesap_ekstra_durumu_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =Gider_Bilgisi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = list(Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users))+list(Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users))
        profile.sort(key=lambda x: x.kayit_tarihi)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.hesap_ekstra_gorme:
                    profile = list(Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db))+list(Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db))
                    profile.sort(key=lambda x: x.kayit_tarihi)
                    
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = list(Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user))+list(Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user))
            profile.sort(key=lambda x: x.kayit_tarihi)
            

    content["santiyeler_i"] = profile
    return render(request,"muhasebe_page/hesap_eksta.html",content)
def muhasebe_ayarlari(request):
    content = sozluk_yapisi()
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.muhasabe_ayarlari_gorme:
                    b = faturalardaki_gelir_gider_etiketi_ozel.objects.filter(kullanici = request.user.kullanicilar_db).last()
                    bilgi = faturalar_icin_bilgiler.objects.filter(gelir_kime_ait_oldugu  = request.user.kullanicilar_db).last()
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
    else:
        bilgi = faturalar_icin_bilgiler.objects.filter(gelir_kime_ait_oldugu  = request.user).last()
        b = faturalardaki_gelir_gider_etiketi_ozel.objects.filter(kullanici = request.user).last()
    if b : 
        content["fatura_icerigi"] = b
    else:
        content["fatura_icerigi"] = "0"
    if bilgi : 
        content["faturalar_bilgisi"] = bilgi
    else:
        content["faturalar_bilgisi"] = "0"
    if request.POST:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.muhasabe_ayarlari_guncelleme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            kullanici = request.user
        dark_logo = request.FILES.get("dark_logo")
        gideretiketi = request.POST.get("gideretiketi")
        gelir_etiketi = request.POST.get("gelir_etiketi")
        adres_bilgisi = request.POST.get("adres")
        emailadresi = request.POST.get("emailadresi")
        telefonadresi = request.POST.get("telefonadresi")
        faturalar_icin_bilgiler.objects.filter(gelir_kime_ait_oldugu  = kullanici).delete()
        faturalar_icin_bilgiler.objects.create(gelir_kime_ait_oldugu  = kullanici,
        adress = adres_bilgisi,email =emailadresi,telefon = telefonadresi )
        if dark_logo:
            faturalar_icin_logo.objects.create(gelir_makbuzu = dark_logo
            ,gelir_kime_ait_oldugu  = kullanici
            )
        fatura_etiketi  = faturalardaki_gelir_gider_etiketi_ozel.objects.filter(kullanici  = kullanici).last()
        if fatura_etiketi:
            if fatura_etiketi.gider_etiketi == gideretiketi and fatura_etiketi.gelir_etiketi == gelir_etiketi :
                print("2 tane if pas geçti")
            else:
                faturalardaki_gelir_gider_etiketi_ozel.objects.create(
                kullanici  = kullanici,
                gelir_etiketi = gelir_etiketi,
                gider_etiketi = gideretiketi 
                )
                gelirlerim = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = kullanici)
                for i in gelirlerim:
                    fatura_no_bilgi = i.fatura_no
                    
                    # Etiket çıkarma işlemi
                    tag = ""
                    for j in fatura_no_bilgi:
                        if j.isdigit():  # İlk rakamdan itibaren etiketin bittiğini varsayıyoruz
                            break
                        tag += j
                    
                    # Eski etiketi yeni etiketle değiştirme
                    yeni_fatura_no = fatura_no_bilgi.replace(tag, gelir_etiketi, 1)
                    
                    # Güncelleme
                    Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu=kullanici, id=i.id).update(fatura_no=yeni_fatura_no)
                gelirlerim = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = kullanici)
                for i in gelirlerim:
                    fatura_no_bilgi = i.fatura_no
                    
                    # Etiket çıkarma işlemi
                    tag = ""
                    for j in fatura_no_bilgi:
                        if j.isdigit():  # İlk rakamdan itibaren etiketin bittiğini varsayıyoruz
                            break
                        tag += j
                    
                    # Eski etiketi yeni etiketle değiştirme
                    yeni_fatura_no = fatura_no_bilgi.replace(tag, gideretiketi, 1)
                    
                    # Güncelleme
                    Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu=kullanici, id=i.id).update(fatura_no=yeni_fatura_no)
        else:
            faturalardaki_gelir_gider_etiketi_ozel.objects.create(
            kullanici  = kullanici,
            gelir_etiketi = gelir_etiketi,
            gider_etiketi = gideretiketi 
            )
            gelirlerim = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = kullanici)
            for i in gelirlerim:
                    fatura_no_bilgi = i.fatura_no
                    
                    # Etiket çıkarma işlemi
                    tag = ""
                    for j in fatura_no_bilgi:
                        if j.isdigit():  # İlk rakamdan itibaren etiketin bittiğini varsayıyoruz
                            break
                        tag += j
                    
                    # Eski etiketi yeni etiketle değiştirme
                    yeni_fatura_no = fatura_no_bilgi.replace(tag, gelir_etiketi, 1)
                    
                    # Güncelleme
                    Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu=kullanici, id=i.id).update(fatura_no=yeni_fatura_no)
            gelirlerim = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = kullanici)
            for i in gelirlerim:
                    fatura_no_bilgi = i.fatura_no
                    
                    # Etiket çıkarma işlemi
                    tag = ""
                    for j in fatura_no_bilgi:
                        if j.isdigit():  # İlk rakamdan itibaren etiketin bittiğini varsayıyoruz
                            break
                        tag += j
                    
                    # Eski etiketi yeni etiketle değiştirme
                    yeni_fatura_no = fatura_no_bilgi.replace(tag, gideretiketi, 1)
                    
                    # Güncelleme
                    Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu=kullanici, id=i.id).update(fatura_no=yeni_fatura_no)
    return render(request,"muhasebe_page/muhasebe_ayarlari.html",content)
def muhasebe_ayarlari_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    b = faturalardaki_gelir_gider_etiketi_ozel.objects.filter(kullanici = users).last()
    bilgi = faturalar_icin_bilgiler.objects.filter(gelir_kime_ait_oldugu  = users).last()
    if b : 
        content["fatura_icerigi"] = b
    else:
        content["fatura_icerigi"] = "0"
    if bilgi : 
        content["faturalar_bilgisi"] = bilgi
    else:
        content["faturalar_bilgisi"] = "0"
    if request.POST:
        dark_logo = request.FILES.get("dark_logo")
        gideretiketi = request.POST.get("gideretiketi")
        gelir_etiketi = request.POST.get("gelir_etiketi")
        adres_bilgisi = request.POST.get("adres")
        emailadresi = request.POST.get("emailadresi")
        telefonadresi = request.POST.get("telefonadresi")
        faturalar_icin_bilgiler.objects.filter(gelir_kime_ait_oldugu  = users).delete()
        faturalar_icin_bilgiler.objects.create(gelir_kime_ait_oldugu  = users,
        adress = adres_bilgisi,email =emailadresi,telefon = telefonadresi )
        if dark_logo:
            faturalar_icin_logo.objects.create(gelir_makbuzu = dark_logo
            ,gelir_kime_ait_oldugu  = users
            )
        fatura_etiketi  = faturalardaki_gelir_gider_etiketi_ozel.objects.filter(kullanici  = users).last()
        if fatura_etiketi:
            if fatura_etiketi.gider_etiketi == gideretiketi and fatura_etiketi.gelir_etiketi == gelir_etiketi :
                print("2 tane if pas geçti")
            else:
                faturalardaki_gelir_gider_etiketi_ozel.objects.create(
                kullanici  = users,
                gelir_etiketi = gelir_etiketi,
                gider_etiketi = gideretiketi 
                )
                gelirlerim = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users)
                for i in gelirlerim:
                    fatura_no_bilgi = i.fatura_no
                    
                    # Etiket çıkarma işlemi
                    tag = ""
                    for j in fatura_no_bilgi:
                        if j.isdigit():  # İlk rakamdan itibaren etiketin bittiğini varsayıyoruz
                            break
                        tag += j
                    
                    # Eski etiketi yeni etiketle değiştirme
                    yeni_fatura_no = fatura_no_bilgi.replace(tag, gelir_etiketi, 1)
                    
                    # Güncelleme
                    Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu=users, id=i.id).update(fatura_no=yeni_fatura_no)
                gelirlerim = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users)
                for i in gelirlerim:
                    fatura_no_bilgi = i.fatura_no
                    
                    # Etiket çıkarma işlemi
                    tag = ""
                    for j in fatura_no_bilgi:
                        if j.isdigit():  # İlk rakamdan itibaren etiketin bittiğini varsayıyoruz
                            break
                        tag += j
                    
                    # Eski etiketi yeni etiketle değiştirme
                    yeni_fatura_no = fatura_no_bilgi.replace(tag, gideretiketi, 1)
                    
                    # Güncelleme
                    Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu=users, id=i.id).update(fatura_no=yeni_fatura_no)
        else:
            faturalardaki_gelir_gider_etiketi_ozel.objects.create(
            kullanici  = users,
            gelir_etiketi = gelir_etiketi,
            gider_etiketi = gideretiketi 
            )
            gelirlerim = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users)
            for i in gelirlerim:
                    fatura_no_bilgi = i.fatura_no
                    
                    # Etiket çıkarma işlemi
                    tag = ""
                    for j in fatura_no_bilgi:
                        if j.isdigit():  # İlk rakamdan itibaren etiketin bittiğini varsayıyoruz
                            break
                        tag += j
                    
                    # Eski etiketi yeni etiketle değiştirme
                    yeni_fatura_no = fatura_no_bilgi.replace(tag, gelir_etiketi, 1)
                    
                    # Güncelleme
                    Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu=users, id=i.id).update(fatura_no=yeni_fatura_no)
            gelirlerim = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users)
            for i in gelirlerim:
                    fatura_no_bilgi = i.fatura_no
                    
                    # Etiket çıkarma işlemi
                    tag = ""
                    for j in fatura_no_bilgi:
                        if j.isdigit():  # İlk rakamdan itibaren etiketin bittiğini varsayıyoruz
                            break
                        tag += j
                    
                    # Eski etiketi yeni etiketle değiştirme
                    yeni_fatura_no = fatura_no_bilgi.replace(tag, gideretiketi, 1)
                    
                    # Güncelleme
                    Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu=users, id=i.id).update(fatura_no=yeni_fatura_no)
    return render(request,"muhasebe_page/muhasebe_ayarlari.html",content)

def fatura_goster(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        gelir_bilgisi_ver =  get_object_or_none(Gelir_Bilgisi,id = id)
        urunleri = gelir_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
    else:
        profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = request.user)
        cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = request.user)
        kategori_bilgisi = gelir_kategorisi.objects.filter(gelir_kategoris_ait_bilgisi = request.user)
        etiketler = gelir_etiketi.objects.filter(gelir_kategoris_ait_bilgisi = request.user)
        gelir_bilgisi_ver =  get_object_or_none(Gelir_Bilgisi,id = id)
        urunleri = gelir_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
    content["fatura_bilgi"] = faturalar_icin_bilgiler.objects.filter(gelir_kime_ait_oldugu =gelir_bilgisi_ver.gelir_kime_ait_oldugu).last()
    content["logosu"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu =gelir_bilgisi_ver.gelir_kime_ait_oldugu).last()
    content["bilgi"] = gelir_bilgisi_ver
    content["urunler"] = urunleri
    return render(request,"muhasebe_page/gelir_faturasi_goster.html",content)
def fatura_goster2(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        gelir_bilgisi_ver =  get_object_or_none(Gider_Bilgisi,id = id)
        urunleri = gider_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
    else:
        profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = request.user)
        cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = request.user)
        kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = request.user)
        etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = request.user)
        gelir_bilgisi_ver =  get_object_or_none(Gider_Bilgisi,id = id)
        urunleri = gider_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
    content["fatura_bilgi"] = faturalar_icin_bilgiler.objects.filter(gelir_kime_ait_oldugu =gelir_bilgisi_ver.gelir_kime_ait_oldugu).last()
    content["logosu"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu =gelir_bilgisi_ver.gelir_kime_ait_oldugu).last()
    content["bilgi"] = gelir_bilgisi_ver
    content["urunler"] = urunleri
    return render(request,"muhasebe_page/gider_faturasi_goster.html",content)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, GradientFill, Alignment
from io import BytesIO
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

def toplam_odenme_tutar(id):
    a = Gelir_odemesi.objects.filter(gelir_kime_ait_oldugu=id)
    topla = 0
    for i in a:
        topla += i.tutar
    return topla

def toplam_tutar_cikarma(id):
    a = gelir_urun_bilgisi.objects.filter(gider_bilgis=id)
    topla = 0
    for i in a:
        topla += (i.urun_fiyati * i.urun_adeti) - i.urun_indirimi
    return topla
def toplam_tutar_cikarmai(id):
    a = gider_urun_bilgisi.objects.filter(gider_bilgis = id)
    topla = 0
    for i in a:
        topla = topla + ((i.urun_fiyati*i.urun_adeti)-i.urun_indirimi)
    return topla
def toplam_odenme_tutari(id):
    a = Gider_odemesi.objects.filter(gelir_kime_ait_oldugu = id)
    topla = 0
    for i in a:
        topla = topla + i.tutar
    return topla
def ekstra(id,k):
    bilgi =  faturalardaki_gelir_gider_etiketi.objects.last()
    if bilgi.gelir_etiketi in k:
        a = gelir_urun_bilgisi.objects.filter(gider_bilgis = id)
        topla = 0
        for i in a:
            topla = topla + ((i.urun_fiyati*i.urun_adeti)-i.urun_indirimi)
        return topla
    else:
        a = gider_urun_bilgisi.objects.filter(gider_bilgis = id)
        topla = 0
        for i in a:
            topla = topla + ((i.urun_fiyati*i.urun_adeti)-i.urun_indirimi)
        return topla

def ekstra_odeme(id,k):
    bilgi =  faturalardaki_gelir_gider_etiketi.objects.last()
    if bilgi.gelir_etiketi in k:
        a = Gelir_odemesi.objects.filter(gelir_kime_ait_oldugu = id)
        topla = 0
        for i in a:
            topla = topla +  i.tutar
        return topla
    else:
        a = Gider_odemesi.objects.filter(gelir_kime_ait_oldugu = id)
        topla = 0
        for i in a:
            topla = topla +  i.tutar
        return topla
def download_excel(request):
    tablo = []
    if request.POST:
        a = request.POST.get("sonuc_cek")
        if a == "0":
            islem = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu=request.user)
            for i in islem:
                tablo.append([
                    str(i.fatura_no),
                    str(i.cari_bilgisi.cari_adi),
                    str(i.aciklama),
                    str(i.fatura_tarihi.strftime("%d.%m.%Y")),
                    str(round(float(toplam_tutar_cikarma(i.id)),2)),
                    str(round(float(toplam_odenme_tutar(i.id)),2))
                ])
        elif a == "1":
            islem = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu=request.user)
            for i in islem:
                tablo.append([
                    str(i.fatura_no),
                    str(i.cari_bilgisi.cari_adi),
                    str(i.aciklama),
                    str(i.fatura_tarihi.strftime("%d.%m.%Y")),
                    str(round(float(toplam_tutar_cikarmai(i.id)),2)),
                    str(round(float(toplam_odenme_tutari(i.id)),2))
                ])
        elif a == "2":
            profile = list(Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user))+list(Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user))
            profile.sort(key=lambda x: x.kayit_tarihi)
            for i in profile:
                tablo.append([
                    str(i.fatura_no),
                    str(i.cari_bilgisi.cari_adi),
                    str(i.aciklama),
                    str(i.fatura_tarihi.strftime("%d.%m.%Y")),
                    str(round(float(ekstra(i,i.fatura_no)),2)),
                    str(round(float(ekstra_odeme(i,i.fatura_no)),2))
                ])

    wb = Workbook()
    ws = wb.active

    # Başlık satırını ekle
    ws.append(['Fatura No', 'Müşteri', 'Açıklama', 'Tarih', 'Tutar', 'Ödeme Tutarı'])

    # İçerik satırlarını ekle
    for row in tablo:
        ws.append(row)

    # Stil düzenlemeleri (opsiyonel)
    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Excel dosyasını bir BytesIO nesnesine yaz
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # HttpResponse ile dosyayı indirme bağlantısı olarak sun
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if a == "0":
        response['Content-Disposition'] = 'attachment; filename=incomesummary.xlsx'
    if a == "1":
        response['Content-Disposition'] = 'attachment; filename=expensesummary.xlsx'
    if a == "2":
        response['Content-Disposition'] = 'attachment; filename=accountsummary.xlsx'
    response.write(excel_data.getvalue())
    return response



def download_pdf(request):
    a = request.POST.get("sonuc_cek")
    # PDF dosyasını oluştur
    response = HttpResponse(content_type='application/pdf')
    if a == "0":
        response['Content-Disposition'] = 'attachment; filename=incomesummary.pdf'
    if a == "1":
        response['Content-Disposition'] = 'attachment; filename=expensesummary.pdf'
    if a == "2":
        response['Content-Disposition'] = 'attachment; filename=accountsummary.pdf'


    # PDF içeriği oluştur
    pdf = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    data = [['Fatura No', 'Müşteri', 'Açıklama', 'Tarih', 'Tutar', 'Ödeme Tutarı']]

    if request.POST:
        a = request.POST.get("sonuc_cek")
        if a == "0":
            islem = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu=request.user)
            for i in islem:
                data.append([
                    str(i.fatura_no),
                    str(i.cari_bilgisi.cari_adi),
                    str(i.aciklama),
                    str(i.fatura_tarihi.strftime("%d.%m.%Y")),
                    str(round(float(toplam_tutar_cikarma(i.id)),2)),
                    str(round(float(toplam_odenme_tutar(i.id)),2))
                ])
        elif a == "1":
            islem = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu=request.user)
            for i in islem:
                data.append([
                    str(i.fatura_no),
                    str(i.cari_bilgisi.cari_adi),
                    str(i.aciklama),
                    str(i.fatura_tarihi.strftime("%d.%m.%Y")),
                    str(round(float(toplam_tutar_cikarmai(i.id)),2)),
                    str(round(float(toplam_odenme_tutari(i.id)),2))
                ])
        elif a == "2":
            profile = list(Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user))+list(Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user))
            profile.sort(key=lambda x: x.kayit_tarihi)
            for i in profile:
                data.append([
                    str(i.fatura_no),
                    str(i.cari_bilgisi.cari_adi),
                    str(i.aciklama),
                    str(i.fatura_tarihi.strftime("%d.%m.%Y")),
                    str(round(float(ekstra(i,i.fatura_no)),2)),
                    str(round(float(ekstra_odeme(i,i.fatura_no)),2))
                ])

    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    # PDF'e tabloyu ekle
    elements.append(table)
    pdf.build(elements)
    return response
    #return redirect("main:ana_sayfa")
def fatura_gosterqr(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        gelir_bilgisi_ver =  get_object_or_none(Gelir_Bilgisi,id = id)
        urunleri = gelir_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
    else:
        gelir_bilgisi_ver =  get_object_or_none(Gelir_Bilgisi,id = id)
        urunleri = gelir_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
    content["fatura_bilgi"] = faturalar_icin_bilgiler.objects.filter(gelir_kime_ait_oldugu =gelir_bilgisi_ver.gelir_kime_ait_oldugu).last()
    content["logosu"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu =gelir_bilgisi_ver.gelir_kime_ait_oldugu).last()
    content["bilgi"] = gelir_bilgisi_ver
    content["urunler"] = urunleri
    return render(request,"muhasebe_page/faturalari_goster_gelir.html",content)
def fatura_goster2qr(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        gelir_bilgisi_ver =  get_object_or_none(Gider_Bilgisi,id = id)
        urunleri = gider_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
    else:

        gelir_bilgisi_ver =  get_object_or_none(Gider_Bilgisi,id = id)
        urunleri = gider_urun_bilgisi.objects.filter(gider_bilgis = gelir_bilgisi_ver)
    content["fatura_bilgi"] = faturalar_icin_bilgiler.objects.filter(gelir_kime_ait_oldugu =gelir_bilgisi_ver.gelir_kime_ait_oldugu).last()
    content["logosu"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu =gelir_bilgisi_ver.gelir_kime_ait_oldugu).last()
    content["bilgi"] = gelir_bilgisi_ver
    content["urunler"] = urunleri
    return render(request,"muhasebe_page/faturalari_goster_gider.html",content)

def urunler_kategorisi(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =urun_kategorileri.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.urun_gorme:
                    profile = urun_kategorileri.objects.filter(silinme_bilgisi = False,kategrori_ait_oldugu = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urun_kategorileri.objects.filter(silinme_bilgisi = False,kategrori_ait_oldugu = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =urun_kategorileri.objects.filter(Q(kategrori_ait_oldugu__first_name__icontains = search)|Q(kategori_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.urun_gorme:
                        profile = urun_kategorileri.objects.filter(Q(kategrori_ait_oldugu = request.user.kullanicilar_db) & Q(kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = urun_kategorileri.objects.filter(Q(kategrori_ait_oldugu = request.user) & Q(kategori_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"stok/urun_kategorileri.html",content)

def urunler_kategorisi_ekle(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            kasa_Adi   = request.POST.get("kasaadi")
            urun_kategorileri.objects.create(kategrori_ait_oldugu = get_object_or_404(CustomUser,id = kullanici_bilgisi )
                                ,kategori_adi = kasa_Adi

                                )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.urun_olusturma:
                        kasa_Adi   = request.POST.get("kasaadi")
                        urun_kategorileri.objects.create(kategrori_ait_oldugu = request.user.kullanicilar_db
                            ,kategori_adi = kasa_Adi)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                kasa_Adi   = request.POST.get("kasaadi")
                urun_kategorileri.objects.create(kategrori_ait_oldugu = request.user
                    ,kategori_adi = kasa_Adi)

    return redirect("accounting:urunler_kategorisi")
def urunler_kategorisi_sil(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        urunler.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.urun_silme:
                    urun_kategorileri.objects.filter(kategrori_ait_oldugu = request.user.kullanicilar_db,id = id).delete()
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            urun_kategorileri.objects.filter(kategrori_ait_oldugu = request.user,id = id).delete()
    return redirect("accounting:urunler_kategorisi")

def urunler_kategorisi_duzenle(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("kasaadi")
        urun_kategorileri.objects.filter(id = id).update(kategrori_ait_oldugu = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,kategori_adi = proje_tip_adi)
            
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.urun_guncelleme:
                    proje_tip_adi   = request.POST.get("kasaadi")
                    
                    urun_kategorileri.objects.filter(kategrori_ait_oldugu = request.user.kullanicilar_db,id = id).update(kategori_adi = proje_tip_adi
                            )
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("kasaadi")
            urun_kategorileri.objects.filter(kategrori_ait_oldugu = request.user,id = id).update(kategori_adi = proje_tip_adi
                            )
    return redirect("accounting:urunler_kategorisi")

def satin_alma_talabi(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =urun_talepleri.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.satin_alma_talebi_gorme:
                    profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user.kullanicilar_db)
                    urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user.kullanicilar_db,silinme_bilgisi = False,urun_turu_secim = "2" )
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user)
            urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user,silinme_bilgisi = False,urun_turu_secim = "2")
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =urun_talepleri.objects.filter(Q(talebin_ait_oldugu__first_name__icontains = search)|Q(urun__urun_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_gorme:
                        profile = urun_talepleri.objects.filter(Q(talebin_ait_oldugu = request.user.kullanicilar_db) & Q(urun__urun_adi__icontains = search)& Q(silinme_bilgisi = False))
                        
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = urun_talepleri.objects.filter(Q(talebin_ait_oldugu = request.user) & Q(urun__urun_adi__icontains = search)& Q(silinme_bilgisi = False))
                
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["urunlerimiz"] = urunler_gonder
    return render(request,"stok/satin_alama_talebi.html",content)

def satin_alma_talabi_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =urun_talepleri.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = users)
        urunler_gonder = urunler.objects.filter(urun_ait_oldugu =users,silinme_bilgisi = False,urun_turu_secim = "2")
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.satin_alma_talebi_gorme:
                    profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user.kullanicilar_db)
                    urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user.kullanicilar_db,silinme_bilgisi = False,urun_turu_secim = "2" )
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user)
            urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user,silinme_bilgisi = False,urun_turu_secim = "2")
  
    content["santiyeler"] = profile

    content["urunlerimiz"] = urunler_gonder
    return render(request,"stok/satin_alama_talebi.html",content)

def satin_alma_talebi_ekle(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            
        else:
            urun_bilgisi = request.POST.get("urun")
            miktar = request.POST.get("miktar")
            fiyat = request.POST.get("fiyat")
            tedarikci = request.POST.get("tedarikci")
            aciklama = request.POST.get("aciklama")
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_olusturma:
                        urun_talepleri.objects.create(talebin_ait_oldugu = request.user.kullanicilar_db,
                        talebi_olusturan =request.user,urun = get_object_or_none(urunler,id =urun_bilgisi),
                        miktar = float(miktar),fiyati = float(fiyat),
                        tedarikci =tedarikci,aciklama = aciklama,talep_Olusturma_tarihi = datetime.now() )
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                urun_talepleri.objects.create(talebin_ait_oldugu = request.user,
                        talebi_olusturan =request.user,urun = get_object_or_none(urunler,id =urun_bilgisi),
                        miktar = float(miktar),fiyati = float(fiyat),
                        tedarikci =tedarikci,aciklama = aciklama,talep_Olusturma_tarihi = datetime.now()  )

    return redirect("accounting:satin_alma_talabi")

def satin_alma_talebi_sil(request,id):
    if True:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_silme:
                        urun_talepleri.objects.filter(id = id,talebi_olusturan = request.user).update(silinme_bilgisi = True)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                urun_talepleri.objects.filter(id = id,talebin_ait_oldugu = request.user).update(silinme_bilgisi = True)

    return redirect("accounting:satin_alma_talabi")


def satin_alma_talabi_onaylama(request):
    return redirect("accounting:satin_alma_talabi")
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =urun_talepleri.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.satin_alma_talebi_onaylama_gorme:
                    profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user.kullanicilar_db).filter(Q(talep_durumu = "1" ) | Q(talep_durumu ="3"))
                    urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user.kullanicilar_db,silinme_bilgisi = False,urun_turu_secim = "2" )
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user).filter(Q(talep_durumu = "1" ) | Q(talep_durumu ="3"))
            urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user,silinme_bilgisi = False,urun_turu_secim = "2")
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =urun_talepleri.objects.filter(Q(talebin_ait_oldugu__first_name__icontains = search)|Q(urun__urun_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_onaylama_gorme:
                        profile = urun_talepleri.objects.filter(Q(talebin_ait_oldugu = request.user.kullanicilar_db) & Q(urun__urun_adi__icontains = search)& Q(silinme_bilgisi = False)).filter(Q(talep_durumu = "1" ) | Q(talep_durumu ="3"))
                        
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = urun_talepleri.objects.filter(Q(talebin_ait_oldugu = request.user) & Q(urun__urun_adi__icontains = search)& Q(silinme_bilgisi = False)).filter(Q(talep_durumu = "1" ) | Q(talep_durumu ="3"))
                
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["urunlerimiz"] = urunler_gonder
    return render(request,"stok/satin_alama_talebi_onaylama.html",content)


def satin_alma_talebi_onayla(request,id):
    if True:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_silme:
                        urun_talepleri.objects.filter(id = id,talebi_olusturan = request.user).update(talep_durumu = "2",talebi_onaylayan = request.user,talep_durum_tarihi=datetime.now())
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                urun_talepleri.objects.filter(id = id,talebin_ait_oldugu = request.user).update(talep_durumu = "2",talebi_onaylayan = request.user,talep_durum_tarihi=datetime.now())

    return redirect("accounting:satin_alma_talabi_onaylama")

def satin_alma_talebi_red(request,id):
    if True:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_silme:
                        urun_talepleri.objects.filter(id = id,talebi_olusturan = request.user).update(talep_durumu = "3",talebi_onaylayan = request.user,talep_durum_tarihi=datetime.now())
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                urun_talepleri.objects.filter(id = id,talebin_ait_oldugu = request.user).update(talep_durumu = "3",talebi_onaylayan = request.user,talep_durum_tarihi=datetime.now())

    return redirect("accounting:satin_alma_talabi_onaylama")

def satin_alma_(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =urun_talepleri.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.satin_alma_talebi_onaylama_gorme:
                    profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user.kullanicilar_db,talep_durumu = "2" )
                    urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user.kullanicilar_db,silinme_bilgisi = False,urun_turu_secim = "2")
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user,talep_durumu = "2")
            urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user,silinme_bilgisi = False,urun_turu_secim = "2")
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =urun_talepleri.objects.filter(Q(talebin_ait_oldugu__first_name__icontains = search)|Q(urun__urun_adi__icontains = search),satin_alinma_durumu = False)
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_onaylama_gorme:
                        profile = urun_talepleri.objects.filter(Q(talebin_ait_oldugu = request.user.kullanicilar_db) & Q(urun__urun_adi__icontains = search)& Q(silinme_bilgisi = False),talep_durumu = "2" ,satin_alinma_durumu = False)
                        
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = urun_talepleri.objects.filter(Q(talebin_ait_oldugu = request.user) & Q(urun__urun_adi__icontains = search)& Q(silinme_bilgisi = False),talep_durumu = "2" ,satin_alinma_durumu = False)
                
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["urunlerimiz"] = urunler_gonder
    return render(request,"stok/satin_alma.html",content)

def satin_alma_onayla(request,id):
    if True:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_silme:
                        urun_talepleri.objects.filter(id = id,talebi_olusturan = request.user).update(satin_alinma_durumu = True,satin_almayi_onaylayan = request.user,satin_alinma_tarihi=datetime.now())
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                urun_talepleri.objects.filter(id = id,talebin_ait_oldugu = request.user).update(satin_alinma_durumu = True,satin_almayi_onaylayan = request.user,satin_alinma_tarihi=datetime.now())

    return redirect("accounting:satin_alma_talabi_onaylama")

def satin_alma_kabuller(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =urun_talepleri.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.satin_alma_talebi_onaylama_gorme:
                    profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user.kullanicilar_db,talep_durumu = "2" ,satin_alinma_durumu = True)
                    urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user.kullanicilar_db,silinme_bilgisi = False,urun_turu_secim = "2")
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user,talep_durumu = "2",satin_alinma_durumu = True )
            urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user,silinme_bilgisi = False,urun_turu_secim = "2")
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =urun_talepleri.objects.filter(Q(talebin_ait_oldugu__first_name__icontains = search)|Q(urun__urun_adi__icontains = search),satin_alinma_durumu = True)
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_onaylama_gorme:
                        profile = urun_talepleri.objects.filter(Q(talebin_ait_oldugu = request.user.kullanicilar_db) & Q(urun__urun_adi__icontains = search)& Q(silinme_bilgisi = False),talep_durumu = "2" ,satin_alinma_durumu = True)
                        
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = urun_talepleri.objects.filter(Q(talebin_ait_oldugu = request.user) & Q(urun__urun_adi__icontains = search)& Q(silinme_bilgisi = False),talep_durumu = "2" ,satin_alinma_durumu = True)
                
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["urunlerimiz"] = urunler_gonder
    return render(request,"stok/satin_alma_kabul.html",content)


def stok(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =urunler.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.stok_talebi_onaylama_gorme:
                    profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user.kullanicilar_db,urun_turu_secim = "2",stok_mu = True )
                    kategori = urun_kategorileri.objects.filter(kategrori_ait_oldugu = request.user.kullanicilar_db)
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user,urun_turu_secim = "2",stok_mu = True )
            kategori = urun_kategorileri.objects.filter(kategrori_ait_oldugu = request.user)
    
    content["santiyeler"] = profile
    content["urun_kategorisi"] = kategori
    return render(request,"stok/stok.html",content)

def stok_girisi_yap(request):
    if request.POST:
        if super_admin_kontrolu(request):
            pass
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.stok_talebi_onaylama_gorme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            kullanici = request.user
        urun = request.POST.get("urun")
        transactionType = request.POST.get("transactionType")
        transactionAmount = request.POST.get("transactionAmount")
        stok_giris_cikis.objects.create(
            stok_kime_ait = kullanici,stok_giren = request.user,
            stok_giren_urun= get_object_or_none(urunler,id = urun),
            stok_durumu = transactionType,stok_adeti  = transactionAmount
        )
        print("kaydedildi")
    return redirect("accounting:stok")
def stok_sayisi(id):
    toplam = 0
    a = gider_urun_bilgisi.objects.filter(urun_bilgisi = id,gider_bilgis__silinme_bilgisi = False)
    b = gelir_urun_bilgisi.objects.filter(urun_bilgisi = id,gider_bilgis__silinme_bilgisi = False)
    stok_giris_cikisi = stok_giris_cikis.objects.filter(stok_giren_urun = id,stok_durumu = "0")
    for i in a:
        toplam = toplam+float(i.urun_adeti)
    for i in b:
        toplam = toplam-float(i.urun_adeti)
    for i in stok_giris_cikisi:
        toplam = toplam+float(i.stok_adeti)
    stok_giris_cikisi = stok_giris_cikis.objects.filter(stok_giren_urun = id,stok_durumu = "1")
    for i in stok_giris_cikisi:
        toplam = toplam-float(i.stok_adeti)
    zimmet = zimmet_olayi.objects.filter(zimmet_verilen_urun = id).filter(Q(zimmet_durumu = "0")|Q(zimmet_durumu = "2"))
    for i in zimmet:
        toplam = toplam - float(i.zimmet_miktari)
    return toplam
def urun_bilgisi(request,id):
    if True:
        urun = get_object_or_none(urunler , id = id)
        zimmetler = zimmet_olayi.objects.filter(zimmet_verilen_urun = urun)
        fatura_data = {
            'isim':urun.urun_adi,
        'kategori': urun.urun_kategorisi.kategori_adi,
        'stok': stok_sayisi(urun),
        "kalemler": [
            {
                "Personel": str(kalem.zimmet_alan_personel.isim) +" "+str(kalem.zimmet_alan_personel.soyisim),
                "adet": kalem.zimmet_miktari,
                "alis": kalem.zimmet_verilis_tarihi.strftime("%d.%m.%Y"),
    "veris":kalem.zimmet_teslim_edilme_tarihi.strftime("%d.%m.%Y") if kalem.zimmet_teslim_edilme_tarihi else "" ,
            } for kalem in zimmetler
        ]
        }
        print(fatura_data)
        return JsonResponse(fatura_data)

def zimmetler(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =urunler.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.zimmet_gorme:
                    profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user.kullanicilar_db,urun_turu_secim = "2",stok_mu = True )
                    kategori = zimmet_olayi.objects.filter(zimmet_kime_ait = request.user.kullanicilar_db)
                    personeller = calisanlar.objects.filter(calisan_kime_ait =request.user.kullanicilar_db )
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user,urun_turu_secim = "2",stok_mu = True )
            kategori = zimmet_olayi.objects.filter(zimmet_kime_ait = request.user)
            personeller = calisanlar.objects.filter(calisan_kime_ait =request.user )   
    content["santiyeler"] = profile
    content["urun_kategorisi"] = kategori
    content["personeller"] = personeller
    return render(request,"stok/zimmetler.html",content)

def zimmet_ekle(request):
    if request.POST:
        if super_admin_kontrolu(request):
            pass
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.zimmet_olusturma:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            kullanici = request.user
        personel = request.POST.get("personel")
        urun = request.POST.get("malzeme")
        miktar = request.POST.get("miktar")
        teslimTarihi = request.POST.get("teslimTarihi")
        zimmet_olayi.objects.create(
            zimmet_kime_ait = kullanici,zimmeti_veren = request.user,
            zimmet_alan_personel = get_object_or_none(calisanlar,id = personel),
            zimmet_verilen_urun = get_object_or_none(urunler,id = urun),
            zimmet_durumu = "0",zimmet_miktari = miktar,
            zimmet_verilis_tarihi = teslimTarihi
        )
    return redirect("accounting:zimmetler")
def zimmeti_teslim_Al(request,id,iz):
    if True:
        if super_admin_kontrolu(request):
            pass
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.zimmet_olusturma:
                    pass
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            pass
        
        if iz == 1:
            zimmet_olayi.objects.filter(id = id).update(zimmet_durumu = "1",zimmet_teslim_edilme_tarihi = datetime.now())
        elif iz == 2:
            zimmet_olayi.objects.filter(id = id).update(zimmet_durumu = "2",zimmet_teslim_edilme_tarihi = datetime.now())
    return redirect("accounting:zimmetler")

def zimmet(request,id):
    if True:
        kalem = get_object_or_none(zimmet_olayi , id = id)
        if kalem.zimmet_alan_personel.profile:
            fatura_data = {
                "id" : id,
                "resim" : kalem.zimmet_alan_personel.profile.url,
                    "urun":kalem.zimmet_verilen_urun.urun_adi,
                    "personel": str(kalem.zimmet_alan_personel.isim) +" "+str(kalem.zimmet_alan_personel.soyisim),
                    "adet": str(kalem.zimmet_miktari),
                    "alis": kalem.zimmet_verilis_tarihi.strftime("%d.%m.%Y"),
                    "veris":kalem.zimmet_teslim_edilme_tarihi.strftime("%d.%m.%Y") if kalem.zimmet_teslim_edilme_tarihi else "" ,
                    "durum":kalem.zimmet_durumu,
                    
            }
        else:
            fatura_data = {
                "id" : id,
                "resim" : "0",
                    "urun":kalem.zimmet_verilen_urun.urun_adi,
                    "personel": str(kalem.zimmet_alan_personel.isim) +" "+str(kalem.zimmet_alan_personel.soyisim),
                    "adet": str(kalem.zimmet_miktari),
                    "alis": kalem.zimmet_verilis_tarihi.strftime("%d.%m.%Y"),
                    "veris":kalem.zimmet_teslim_edilme_tarihi.strftime("%d.%m.%Y") if kalem.zimmet_teslim_edilme_tarihi else "" ,
                    "durum":kalem.zimmet_durumu,
                    
            }
        return JsonResponse(fatura_data)

def avans_maas(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.personeller_odeme_gorme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            kullanici = request.user
        maas_ve_avanslar_iptal_edilen_faturalar = calisanlar_calismalari_odemeleri.objects.filter(
            calisan__calisan_kime_ait=kullanici,
            fatura__isnull=False
            
            ).exclude(
                fatura__silinme_bilgisi=True
            )
        maas_ve_avanslar_faturasi_olmayan = calisanlar_calismalari_odemeleri.objects.filter(
            calisan__calisan_kime_ait=kullanici,
            fatura__isnull=True)
        maas_ve_avanslar__faturalar = calisanlar_calismalari_odemeleri.objects.filter(
            calisan__calisan_kime_ait=kullanici,
            fatura__isnull=False
            
            ).exclude(
                fatura__silinme_bilgisi=False
            )     
    content["silinen_faturali"] = maas_ve_avanslar_iptal_edilen_faturalar
    content["faturasi_olmayan"] = maas_ve_avanslar_faturasi_olmayan
    content["faturasi_olan"] = maas_ve_avanslar__faturalar
    return render(request,"personel/maas_avans_sayfasi.html",content)

def avans_maas_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        kullanici = users
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.personeller_odeme_gorme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            kullanici = request.user
    maas_ve_avanslar_iptal_edilen_faturalar = calisanlar_calismalari_odemeleri.objects.filter(
            calisan__calisan_kime_ait=kullanici,
            fatura__isnull=False
            
            ).exclude(
                fatura__silinme_bilgisi=True
            )
    maas_ve_avanslar_faturasi_olmayan = calisanlar_calismalari_odemeleri.objects.filter(
            calisan__calisan_kime_ait=kullanici,
            fatura__isnull=True)
    maas_ve_avanslar__faturalar = calisanlar_calismalari_odemeleri.objects.filter(
            calisan__calisan_kime_ait=kullanici,
            fatura__isnull=False
            
            ).exclude(
                fatura__silinme_bilgisi=False
            )     
    content["silinen_faturali"] = maas_ve_avanslar_iptal_edilen_faturalar
    content["faturasi_olmayan"] = maas_ve_avanslar_faturasi_olmayan
    content["faturasi_olan"] = maas_ve_avanslar__faturalar
    return render(request,"personel/maas_avans_sayfasi.html",content)
# Personel Gider Faturası Kesme
def personel_gider_faturasi_kesme_2(request,id,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    avans_mi_maas_mi = get_object_or_404(calisanlar_calismalari_odemeleri,id = id)
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        kullanci = users
        if avans_mi_maas_mi.odeme_turu:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = kullanci,avans_icin_kullan = True)
            urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = kullanci,avans_icin_kullan = True).last()
            cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = kullanci)
            kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,avans_icin_kullan = True)
            etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,avans_icin_kullan = True)
        else:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = kullanci,maas_icin_kullan = True)
            urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = kullanci,maas_icin_kullan = True).last()
            cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = kullanci)
            kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,maas_icin_kullan = True)
            etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,maas_icin_kullan = True)

    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_kesme_izni:
                    kullanci = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            kullanci = request.user
        if avans_mi_maas_mi.odeme_turu:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = kullanci,avans_icin_kullan = True)
            urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = kullanci,avans_icin_kullan = True).last()
            cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = kullanci)
            kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,avans_icin_kullan = True)
            etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,avans_icin_kullan = True)
        else:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = kullanci,maas_icin_kullan = True)
            urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = kullanci,maas_icin_kullan = True).last()
            cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = kullanci)
            kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,maas_icin_kullan = True)
            etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,maas_icin_kullan = True)
    content["veri"] = avans_mi_maas_mi
    content["gelir_kategoerisi"] = kategori_bilgisi
    content["gelir_etiketi"] = etiketler
    content["kasa"] = profile
    content["urunler"]  = urunler_bilgisi
    content["cari_bilgileri"] = cari_bilgileri
    return render(request,"muhasebe_page/gider_faturasi_personel_kesimi.html",content)
def personel_gider_faturasi_kesme(request,id):
    content = sozluk_yapisi()
    avans_mi_maas_mi = get_object_or_404(calisanlar_calismalari_odemeleri,id = id)
    if super_admin_kontrolu(request):
        profile =Kasa.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        urunler_bilgisi = ""
        cari_bilgileri = ""
        kategori_bilgisi =""
        etiketler =  ""

    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_kesme_izni:
                    kullanci = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            kullanci = request.user
        if avans_mi_maas_mi.odeme_turu:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = kullanci,avans_icin_kullan = True)
            urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = kullanci,avans_icin_kullan = True).last()
            cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = kullanci)
            kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,avans_icin_kullan = True)
            etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,avans_icin_kullan = True)
        else:
            profile = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = kullanci,maas_icin_kullan = True)
            urunler_bilgisi = urunler.objects.filter(urun_ait_oldugu = kullanci,maas_icin_kullan = True).last()
            cari_bilgileri = cari.objects.filter(cari_kart_ait_bilgisi = kullanci)
            kategori_bilgisi = gider_kategorisi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,maas_icin_kullan = True)
            etiketler = gider_etiketi.objects.filter(gider_kategoris_ait_bilgisi = kullanci,maas_icin_kullan = True)
    content["veri"] = avans_mi_maas_mi
    content["gelir_kategoerisi"] = kategori_bilgisi
    content["gelir_etiketi"] = etiketler
    content["kasa"] = profile
    content["urunler"]  = urunler_bilgisi
    content["cari_bilgileri"] = cari_bilgileri
    return render(request,"muhasebe_page/gider_faturasi_personel_kesimi.html",content)
# 
def gider_faturasi_kaydet_personel(request):
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_kesme_izni:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
    else:
        kullanici = request.user
    if request.POST:
        faturaya_bagla = request.POST.get("faturaya_bagla")
        musteri_bilgisi  = request.POST.get("musteri_bilgisi")
        daterange = request.POST.get("daterange")
        gelir_kategorisii = request.POST.get("gelir_kategorisi")
        cari_aciklma = request.POST.get("cari_aciklma")
        etiketler = request.POST.getlist("etiketler")
        faturano = request.POST.get("faturano")
        urunadi = request.POST.getlist("urunadi")
        miktari = request.POST.getlist("miktari")
        bfiyatInput = request.POST.getlist("bfiyatInput")
        indirim = request.POST.getlist("indirim")
        aciklama = request.POST.getlist("aciklama")
        aciklama_id = request.POST.getlist('aciklama_id')
        doviz_kuru = request.POST.get("doviz_kuru")
        profile = request.FILES.get("fatura_belgesi")
        print(gelir_kategorisii,"veri verme")
        cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici)
        if cari_bilgisi:
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gider_Bilgisi.objects.create(gelir_kime_ait_oldugu = kullanici,
            cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisii_id = gelir_kategorisii,doviz = doviz_kuru,aciklama = cari_aciklma
                                         )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gider_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i])
                    if urun:
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=kullanici,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )

        else:
            cari_bilgisi = cari.objects.create(cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici,aciklama = cari_aciklma)
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gider_Bilgisi.objects.create(gelir_kime_ait_oldugu = kullanici,
            cari_bilgisi = get_object_or_none(cari,id = cari_bilgisi.id),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisii_id =gelir_kategorisii,doviz = doviz_kuru,aciklama = cari_aciklma
                                         )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gider_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i])
                    if urun:
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=kullanici,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
        toplam_tutar = 0
        gelir_urun_bilgisi_al =gider_urun_bilgisi.objects.filter(gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id)) 
        for i in gelir_urun_bilgisi_al:
            toplam_tutar += (i.urun_fiyati)*(i.urun_adeti)-i.urun_indirimi
        Gider_Bilgisi.objects.filter(id =new_project.id ).update(toplam_tutar =toplam_tutar,kalan_tutar =0 )
        
        gider_qr.objects.create(gelir_kime_ait_oldugu = get_object_or_none(Gider_Bilgisi,id = new_project.id))
        print(aciklama_id,"gelen id")
        fatura_gorsel = get_object_or_none(calisanlar_calismalari_odemeleri,id = faturaya_bagla )
        if fatura_gorsel.dosya:
            u = Gider_Bilgisi.objects.get(id = new_project.id )
            u.fatura_gorseli = fatura_gorsel.dosya
            u.save()
        if profile:
            u = Gider_Bilgisi.objects.get(id = new_project.id )
            u.fatura_gorseli = profile
            u.save()
        calisanlar_calismalari_odemeleri.objects.filter(id = faturaya_bagla).update(fatura = get_object_or_none(Gider_Bilgisi,id = new_project.id))
        if fatura_gorsel.odeme_turu:
            makbuzu = Gider_odemesi.objects.create(
                gelir_kime_ait_oldugu = get_object_or_none(Gider_Bilgisi,id = new_project.id),
                kasa_bilgisi = Kasa.objects.filter(silinme_bilgisi = False,avans_icin_kullan = True,kasa_kart_ait_bilgisi = kullanici).last(),
                tutar = get_object_or_none(Gider_Bilgisi,id = new_project.id).toplam_tutar,tarihi = get_object_or_none(Gider_Bilgisi,id = new_project.id).fatura_tarihi,
                makbuz_no = get_object_or_none(Gider_Bilgisi,id = new_project.id).fatura_no,aciklama=get_object_or_none(Gider_Bilgisi,id = new_project.id).aciklama,
                islemi_yapan = request.user
            )
            if fatura_gorsel.dosya:
                u = Gider_odemesi.objects.get(id = makbuzu.id )
                u.gelir_makbuzu = fatura_gorsel.dosya
                u.save()
        else:
            makbuzu = Gider_odemesi.objects.create(
                gelir_kime_ait_oldugu = get_object_or_none(Gider_Bilgisi,id = new_project.id),
                kasa_bilgisi = Kasa.objects.filter(silinme_bilgisi = False,maas_icin_kullan = True,kasa_kart_ait_bilgisi = kullanici).last(),
                tutar = get_object_or_none(Gider_Bilgisi,id = new_project.id).toplam_tutar,tarihi = get_object_or_none(Gider_Bilgisi,id = new_project.id).fatura_tarihi,
                makbuz_no = get_object_or_none(Gider_Bilgisi,id = new_project.id).fatura_no,aciklama=get_object_or_none(Gider_Bilgisi,id = new_project.id).aciklama,
                islemi_yapan = request.user
            )
            if fatura_gorsel.dosya:
                u = Gider_odemesi.objects.get(id = makbuzu.id )
                u.gelir_makbuzu = fatura_gorsel.dosya
                u.save()
    return redirect("accounting:giderler_sayfasi")



def gider_faturasi_kaydet_personel_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gider_faturasi_kesme_izni:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
    else:
        kullanici = users
    if request.POST:
        faturaya_bagla = request.POST.get("faturaya_bagla")
        musteri_bilgisi  = request.POST.get("musteri_bilgisi")
        daterange = request.POST.get("daterange")
        gelir_kategorisii = request.POST.get("gelir_kategorisi")
        cari_aciklma = request.POST.get("cari_aciklma")
        etiketler = request.POST.getlist("etiketler")
        faturano = request.POST.get("faturano")
        urunadi = request.POST.getlist("urunadi")
        miktari = request.POST.getlist("miktari")
        bfiyatInput = request.POST.getlist("bfiyatInput")
        indirim = request.POST.getlist("indirim")
        aciklama = request.POST.getlist("aciklama")
        aciklama_id = request.POST.getlist('aciklama_id')
        doviz_kuru = request.POST.get("doviz_kuru")
        profile = request.FILES.get("fatura_belgesi")
        print(gelir_kategorisii,"veri verme")
        cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici)
        if cari_bilgisi:
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gider_Bilgisi.objects.create(gelir_kime_ait_oldugu = kullanici,
            cari_bilgisi = get_object_or_none(cari,cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisii_id = gelir_kategorisii,doviz = doviz_kuru,aciklama = cari_aciklma
                                         )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gider_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i])
                    if urun:
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=kullanici,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )

        else:
            cari_bilgisi = cari.objects.create(cari_adi = musteri_bilgisi,cari_kart_ait_bilgisi = kullanici,aciklama = cari_aciklma)
            date_range_parts = daterange.split(' - ')

            # Tarihleri ayrı ayrı alma ve uygun formata dönüştürme
            fatura_tarihi_str, vade_tarihi_str = date_range_parts
            fatura_tarihi = datetime.strptime(fatura_tarihi_str, '%m/%d/%Y')
            vade_tarihi = datetime.strptime(vade_tarihi_str, '%m/%d/%Y')

            new_project =Gider_Bilgisi.objects.create(gelir_kime_ait_oldugu = kullanici,
            cari_bilgisi = get_object_or_none(cari,id = cari_bilgisi.id),
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = faturano,
            gelir_kategorisii_id =gelir_kategorisii,doviz = doviz_kuru,aciklama = cari_aciklma
                                         )
            new_project.save()
            gelir_etiketi_sec = []
            for i in etiketler:
                gelir_etiketi_sec.append(gider_etiketi.objects.get(id=int(i)))
            new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
            for i in range(0,len(urunadi)):
                if urunadi[i] != "" and miktari[i] != "" and bfiyatInput[i] != "":
                    urun = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i])
                    if urun:
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=kullanici,urun_adi = urunadi[i]),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
                    else:
                        urun = urunler.objects.create(urun_ait_oldugu=kullanici,urun_adi = urunadi[i],
                                                      urun_fiyati = float(bfiyatInput[i]))
                        gelir_urun_bilgisi_bi = gider_urun_bilgisi.objects.create(
                            urun_ait_oldugu =  kullanici,urun_bilgisi = get_object_or_none(urunler,id = urun.id),
                            urun_fiyati = bfiyatInput[i],urun_indirimi = float(indirim[i]),urun_adeti = int(miktari[i]),
                            gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id),
                            aciklama = aciklama[i]
                        )
        toplam_tutar = 0
        gelir_urun_bilgisi_al =gider_urun_bilgisi.objects.filter(gider_bilgis =  get_object_or_none(Gider_Bilgisi,id = new_project.id)) 
        for i in gelir_urun_bilgisi_al:
            toplam_tutar += (i.urun_fiyati)*(i.urun_adeti)-i.urun_indirimi
        Gider_Bilgisi.objects.filter(id =new_project.id ).update(toplam_tutar =toplam_tutar,kalan_tutar =0 )
        
        gider_qr.objects.create(gelir_kime_ait_oldugu = get_object_or_none(Gider_Bilgisi,id = new_project.id))
        print(aciklama_id,"gelen id")
        fatura_gorsel = get_object_or_none(calisanlar_calismalari_odemeleri,id = faturaya_bagla )
        if fatura_gorsel.dosya:
            u = Gider_Bilgisi.objects.get(id = new_project.id )
            u.fatura_gorseli = fatura_gorsel.dosya
            u.save()
        if profile:
            u = Gider_Bilgisi.objects.get(id = new_project.id )
            u.fatura_gorseli = profile
            u.save()
        calisanlar_calismalari_odemeleri.objects.filter(id = faturaya_bagla).update(fatura = get_object_or_none(Gider_Bilgisi,id = new_project.id))
        if fatura_gorsel.odeme_turu:
            makbuzu = Gider_odemesi.objects.create(
                gelir_kime_ait_oldugu = get_object_or_none(Gider_Bilgisi,id = new_project.id),
                kasa_bilgisi = Kasa.objects.filter(silinme_bilgisi = False,avans_icin_kullan = True,kasa_kart_ait_bilgisi = kullanici).last(),
                tutar = get_object_or_none(Gider_Bilgisi,id = new_project.id).toplam_tutar,tarihi = get_object_or_none(Gider_Bilgisi,id = new_project.id).fatura_tarihi,
                makbuz_no = get_object_or_none(Gider_Bilgisi,id = new_project.id).fatura_no,aciklama=get_object_or_none(Gider_Bilgisi,id = new_project.id).aciklama,
                islemi_yapan = request.user
            )
            if fatura_gorsel.dosya:
                u = Gider_odemesi.objects.get(id = makbuzu.id )
                u.gelir_makbuzu = fatura_gorsel.dosya
                u.save()
        else:
            makbuzu = Gider_odemesi.objects.create(
                gelir_kime_ait_oldugu = get_object_or_none(Gider_Bilgisi,id = new_project.id),
                kasa_bilgisi = Kasa.objects.filter(silinme_bilgisi = False,maas_icin_kullan = True,kasa_kart_ait_bilgisi = kullanici).last(),
                tutar = get_object_or_none(Gider_Bilgisi,id = new_project.id).toplam_tutar,tarihi = get_object_or_none(Gider_Bilgisi,id = new_project.id).fatura_tarihi,
                makbuz_no = get_object_or_none(Gider_Bilgisi,id = new_project.id).fatura_no,aciklama=get_object_or_none(Gider_Bilgisi,id = new_project.id).aciklama,
                islemi_yapan = request.user
            )
            if fatura_gorsel.dosya:
                u = Gider_odemesi.objects.get(id = makbuzu.id )
                u.gelir_makbuzu = fatura_gorsel.dosya
                u.save()
    return redirect("accounting:giderler_sayfasi_2",hash)

##########################################################3

def satin_alma_talebi_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = users
            urun_bilgisi = request.POST.get("urun")
            miktar = request.POST.get("miktar")
            fiyat = request.POST.get("fiyat")
            tedarikci = request.POST.get("tedarikci")
            aciklama = request.POST.get("aciklama")
            urun_talepleri.objects.create(talebin_ait_oldugu = kullanici_bilgisi,
                        talebi_olusturan =request.user,urun = get_object_or_none(urunler,id =urun_bilgisi),
                        miktar = float(miktar),fiyati = float(fiyat),
                        tedarikci =tedarikci,aciklama = aciklama,talep_Olusturma_tarihi = datetime.now()  )
       

    return redirect("accounting:satin_alma_talabi_2",hash)

def satin_alma_talebi_sil_2(request,id,hash):
    if True:
        #yetkili_adi
        if super_admin_kontrolu(request):
            content = sozluk_yapisi()
            d = decode_id(hash)
            content["hashler"] = hash
            users = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = users
            kullanici_bilgisi  = request.POST.get("kullanici")
            urun_talepleri.objects.filter(id = id,talebin_ait_oldugu = users).update(silinme_bilgisi = True)

        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_silme:
                        urun_talepleri.objects.filter(id = id,talebi_olusturan = request.user).update(silinme_bilgisi = True)
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                urun_talepleri.objects.filter(id = id,talebin_ait_oldugu = request.user).update(silinme_bilgisi = True)

    return redirect("accounting:satin_alma_talabi_2",hash)


def satin_alma_talabi_onaylama_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    return redirect("accounting:satin_alma_talabi_2",hash)
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =urun_talepleri.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.satin_alma_talebi_onaylama_gorme:
                    profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user.kullanicilar_db).filter(Q(talep_durumu = "1" ) | Q(talep_durumu ="3"))
                    urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user.kullanicilar_db,silinme_bilgisi = False,urun_turu_secim = "2" )
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = request.user).filter(Q(talep_durumu = "1" ) | Q(talep_durumu ="3"))
            urunler_gonder = urunler.objects.filter(urun_ait_oldugu =request.user,silinme_bilgisi = False,urun_turu_secim = "2")
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =urun_talepleri.objects.filter(Q(talebin_ait_oldugu__first_name__icontains = search)|Q(urun__urun_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.satin_alma_talebi_onaylama_gorme:
                        profile = urun_talepleri.objects.filter(Q(talebin_ait_oldugu = request.user.kullanicilar_db) & Q(urun__urun_adi__icontains = search)& Q(silinme_bilgisi = False)).filter(Q(talep_durumu = "1" ) | Q(talep_durumu ="3"))
                        
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                profile = urun_talepleri.objects.filter(Q(talebin_ait_oldugu = request.user) & Q(urun__urun_adi__icontains = search)& Q(silinme_bilgisi = False)).filter(Q(talep_durumu = "1" ) | Q(talep_durumu ="3"))
                
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["urunlerimiz"] = urunler_gonder
    return render(request,"stok/satin_alama_talebi_onaylama.html",content)


def satin_alma_talebi_onayla_2(request,id,hash):
    if True:
        #yetkili_adi
        if super_admin_kontrolu(request):
            content = sozluk_yapisi()
            d = decode_id(hash)
            content["hashler"] = hash
            users = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = users
            kullanici_bilgisi  = request.POST.get("kullanici")
            urun_talepleri.objects.filter(id = id,talebi_olusturan = users).update(talep_durumu = "2",talebi_onaylayan = request.user,talep_durum_tarihi=datetime.now())
       

    return redirect("accounting:satin_alma_talabi_onaylama_2",hash)

def satin_alma_talebi_red_2(request,id,hash):
    if True:
        #yetkili_adi
        if super_admin_kontrolu(request):
            content = sozluk_yapisi()
            d = decode_id(hash)
            content["hashler"] = hash
            users = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = users
            urun_talepleri.objects.filter(id = id,talebi_olusturan =users).update(talep_durumu = "3",talebi_onaylayan = request.user,talep_durum_tarihi=datetime.now())
        

    return redirect("accounting:satin_alma_talabi_onaylama_2",hash)

def satin_alma_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =urun_talepleri.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = users,talep_durumu = "2")
        urunler_gonder = urunler.objects.filter(urun_ait_oldugu =users,silinme_bilgisi = False,urun_turu_secim = "2")
   
    content["santiyeler"] = profile
    content["urunlerimiz"] = urunler_gonder
    return render(request,"stok/satin_alma.html",content)

def satin_alma_onayla_2(request,id,hash):
    if True:
        #yetkili_adi
        if super_admin_kontrolu(request):
            content = sozluk_yapisi()
            d = decode_id(hash)
            content["hashler"] = hash
            users = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = users
            urun_talepleri.objects.filter(id = id,talebin_ait_oldugu = request.user).update(satin_alinma_durumu = True,satin_almayi_onaylayan = request.user,satin_alinma_tarihi=datetime.now())
        
    return redirect("accounting:satin_alma_talabi_onaylama_2",hash)

def satin_alma_kabuller_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =urun_talepleri.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = urun_talepleri.objects.filter(silinme_bilgisi = False,talebin_ait_oldugu = users,talep_durumu = "2",satin_alinma_durumu = True )
        urunler_gonder = urunler.objects.filter(urun_ait_oldugu =users,silinme_bilgisi = False,urun_turu_secim = "2")
    
    content["santiyeler"] = profile
    content["urunlerimiz"] = urunler_gonder
    return render(request,"stok/satin_alma_kabul.html",content)


def stok_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =urunler.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = users,urun_turu_secim = "2",stok_mu = True )
        kategori = urun_kategorileri.objects.filter(kategrori_ait_oldugu = users)
    content["santiyeler"] = profile
    content["urun_kategorisi"] = kategori
    return render(request,"stok/stok.html",content)

def stok_girisi_yap_2(request,hash):
    if request.POST:
        if super_admin_kontrolu(request):
            content = sozluk_yapisi()
            d = decode_id(hash)
            content["hashler"] = hash
            users = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = users
            kullanici = users
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.stok_talebi_onaylama_gorme:
                        kullanici = request.user.kullanicilar_db
                    else:
                        return redirect("main:yetkisiz")
                else:
                    return redirect("main:yetkisiz")
            else:
                kullanici = request.user
        urun = request.POST.get("urun")
        transactionType = request.POST.get("transactionType")
        transactionAmount = request.POST.get("transactionAmount")
        stok_giris_cikis.objects.create(
            stok_kime_ait = kullanici,stok_giren = request.user,
            stok_giren_urun= get_object_or_none(urunler,id = urun),
            stok_durumu = transactionType,stok_adeti  = transactionAmount
        )
        print("kaydedildi")
    return redirect("accounting:stok_2",hash)


def zimmetler_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =urunler.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = users,urun_turu_secim = "2",stok_mu = True )
        kategori = zimmet_olayi.objects.filter(zimmet_kime_ait = users)
        personeller = calisanlar.objects.filter(calisan_kime_ait =users )
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.zimmet_gorme:
                    profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user.kullanicilar_db,urun_turu_secim = "2",stok_mu = True )
                    kategori = zimmet_olayi.objects.filter(zimmet_kime_ait = request.user.kullanicilar_db)
                    personeller = calisanlar.objects.filter(calisan_kime_ait =request.user.kullanicilar_db )
                else:
                    return redirect("main:yetkisiz")
            else:
                return redirect("main:yetkisiz")
        else:
            profile = urunler.objects.filter(silinme_bilgisi = False,urun_ait_oldugu = request.user,urun_turu_secim = "2",stok_mu = True )
            kategori = zimmet_olayi.objects.filter(zimmet_kime_ait = request.user)
            personeller = calisanlar.objects.filter(calisan_kime_ait =request.user )   
    content["santiyeler"] = profile
    content["urun_kategorisi"] = kategori
    content["personeller"] = personeller
    return render(request,"stok/zimmetler.html",content)

def zimmet_ekle_2(request,hash):
    if request.POST:
        if super_admin_kontrolu(request):
            content = sozluk_yapisi()
            d = decode_id(hash)
            content["hashler"] = hash
            users = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = users
            kullanici = users
        
        personel = request.POST.get("personel")
        urun = request.POST.get("malzeme")
        miktar = request.POST.get("miktar")
        teslimTarihi = request.POST.get("teslimTarihi")
        zimmet_olayi.objects.create(
            zimmet_kime_ait = kullanici,zimmeti_veren = request.user,
            zimmet_alan_personel = get_object_or_none(calisanlar,id = personel),
            zimmet_verilen_urun = get_object_or_none(urunler,id = urun),
            zimmet_durumu = "0",zimmet_miktari = miktar,
            zimmet_verilis_tarihi = teslimTarihi
        )
    return redirect("accounting:zimmetler_2",hash)
def zimmeti_teslim_Al_2(request,id,iz,hash):
    if True:
        if super_admin_kontrolu(request):
            content = sozluk_yapisi()
            d = decode_id(hash)
            content["hashler"] = hash
            users = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = users
        if iz == 1:
            zimmet_olayi.objects.filter(id = id).update(zimmet_durumu = "1",zimmet_teslim_edilme_tarihi = datetime.now())
        elif iz == 2:
            zimmet_olayi.objects.filter(id = id).update(zimmet_durumu = "2",zimmet_teslim_edilme_tarihi = datetime.now())
    return redirect("accounting:zimmetler")

################################################

def urunler_kategorisi_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =urun_kategorileri.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = urun_kategorileri.objects.filter(silinme_bilgisi = False,kategrori_ait_oldugu = users)
    
    content["santiyeler"] = profile

    return render(request,"stok/urun_kategorileri.html",content)

def urunler_kategorisi_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kasa_Adi   = request.POST.get("kasaadi")
            urun_kategorileri.objects.create(kategrori_ait_oldugu = users
                    ,kategori_adi = kasa_Adi)

    return redirect("accounting:urunler_kategorisi_2",hash)
def urunler_kategorisi_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        urun_kategorileri.objects.filter(kategrori_ait_oldugu = users,id = id).delete()
    return redirect("accounting:urunler_kategorisi_2",hash)

def urunler_kategorisi_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        proje_tip_adi   = request.POST.get("kasaadi")
        urun_kategorileri.objects.filter(kategrori_ait_oldugu = users,id = id).update(kategori_adi = proje_tip_adi
                            )
    return redirect("accounting:urunler_kategorisi_2",hash)
