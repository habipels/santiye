from django.shortcuts import render,redirect,get_object_or_404
from django.http import HttpResponse
from users.models import *
from django.contrib.auth.models import User
from django.core.paginator import Paginator , EmptyPage, PageNotAnInteger
from django.db.models.query_utils import Q
from site_settings.models import *
from django.utils.translation  import gettext as _
from django.utils.translation import get_language, activate, gettext
from site_info.models import *
from muhasebe.models import *
from django.contrib.admin.models import LogEntry
from hashids import Hashids
from django.middleware.csrf import get_token
from django.http import JsonResponse
import requests
def get_client_ip(request):
    """Kullanıcının IP adresini alır"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip
import geoip2.database
from django.http import JsonResponse
import pytz

def get_country(request):
    # Kullanıcının IP adresini al (varsayılan olarak, Django'dan gelen IP'yi alabilirsiniz)
    ip_address = request.META.get('REMOTE_ADDR', None)

    # Eğer IP adresi localhost (127.0.0.1) ise, bunu manuel olarak değiştirebiliriz
    if ip_address == '127.0.0.1':
        ip_address = '78.173.5.100'  # Google DNS IP'si gibi bir genel IP'yi kullanabilirsiniz

    # Eğer IP adresi mevcutsa, GeoIP2 veritabanından ülkeyi sorgulayalım
    if ip_address:
        try:
            # GeoIP2 veritabanını okuma işlemi
            with geoip2.database.Reader('geoip/GeoLite2-Country.mmdb') as reader:
                # IP adresine göre ülkeyi alıyoruz
                response = reader.country(ip_address)
                country = response.country.iso_code  # Ülke kodunu alıyoruz (örn. 'TR' için Türkiye)
                return JsonResponse({'country': country})
        
        except geoip2.errors.AddressNotFoundError:
            # Eğer ülke bulunamazsa 'Unknown' dönebiliriz
            return JsonResponse({'country': 'Unknown'})
        except Exception as e:
            # Diğer hataları yakalamak için genel bir hata yönetimi ekleyebiliriz
            return JsonResponse({'error': str(e)})

    # Eğer IP adresi yoksa, hata mesajı döneriz
    return JsonResponse({'error': 'IP address not found'})
from django.urls import reverse
def redirect_with_language(view_name, *args, **kwargs):
    lang = get_language()
    url = reverse(view_name, args=args, kwargs=kwargs)
    return redirect(f'/{lang}{url}')

def get_time_zone_from_country(country):
    # Ülke ISO kodlarına göre zaman dilimlerini döndüren harita
    time_zones = {
        'AF': 'Asia/Kabul',
        'AL': 'Europe/Tirane',
        'DZ': 'Africa/Algiers',
        'AS': 'Pacific/Pago_Pago',
        'AD': 'Europe/Andorra',
        'AO': 'Africa/Luanda',
        'AR': 'America/Argentina/Buenos_Aires',
        'AM': 'Asia/Yerevan',
        'AU': 'Australia/Sydney',
        'AT': 'Europe/Vienna',
        'AZ': 'Asia/Baku',
        'BS': 'America/Nassau',
        'BH': 'Asia/Bahrain',
        'BD': 'Asia/Dhaka',
        'BB': 'America/Barbados',
        'BY': 'Europe/Minsk',
        'BE': 'Europe/Brussels',
        'BZ': 'America/Belize',
        'BJ': 'Africa/Porto-Novo',
        'BT': 'Asia/Thimphu',
        'BO': 'America/La_Paz',
        'BA': 'Europe/Sarajevo',
        'BW': 'Africa/Gaborone',
        'BR': 'America/Sao_Paulo',
        'BN': 'Asia/Brunei',
        'BG': 'Europe/Sofia',
        'BF': 'Africa/Ouagadougou',
        'BI': 'Africa/Bujumbura',
        'KH': 'Asia/Phnom_Penh',
        'CM': 'Africa/Douala',
        'CA': 'America/Toronto',
        'CV': 'Atlantic/Cape_Verde',
        'KY': 'America/Cayman',
        'CF': 'Africa/Bangui',
        'TD': 'Africa/Ndjamena',
        'CL': 'America/Santiago',
        'CN': 'Asia/Shanghai',
        'CO': 'America/Bogota',
        'KM': 'Indian/Comoro',
        'CG': 'Africa/Brazzaville',
        'CD': 'Africa/Kinshasa',
        'CK': 'Pacific/Rarotonga',
        'CR': 'America/Costa_Rica',
        'CI': 'Africa/Abidjan',
        'HR': 'Europe/Zagreb',
        'CU': 'America/Havana',
        'CY': 'Asia/Nicosia',
        'CZ': 'Europe/Prague',
        'DK': 'Europe/Copenhagen',
        'DJ': 'Africa/Djibouti',
        'DM': 'America/Dominica',
        'DO': 'America/Santo_Domingo',
        'EC': 'America/Guayaquil',
        'EG': 'Africa/Cairo',
        'SV': 'America/El_Salvador',
        'GQ': 'Africa/Malabo',
        'ER': 'Africa/Asmara',
        'EE': 'Europe/Tallinn',
        'ET': 'Africa/Addis_Ababa',
        'FJ': 'Pacific/Fiji',
        'FI': 'Europe/Helsinki',
        'FR': 'Europe/Paris',
        'GA': 'Africa/Libreville',
        'GB': 'Europe/London',
        'GE': 'Asia/Tbilisi',
        'GH': 'Africa/Accra',
        'GR': 'Europe/Athens',
        'GD': 'America/Grenada',
        'GT': 'America/Guatemala',
        'GN': 'Africa/Conakry',
        'GW': 'Africa/Bissau',
        'GY': 'America/Guyana',
        'HT': 'America/Port-au-Prince',
        'HN': 'America/Tegucigalpa',
        'HK': 'Asia/Hong_Kong',
        'HU': 'Europe/Budapest',
        'IS': 'Atlantic/Reykjavik',
        'IN': 'Asia/Kolkata',
        'ID': 'Asia/Jakarta',
        'IR': 'Asia/Tehran',
        'IQ': 'Asia/Baghdad',
        'IE': 'Europe/Dublin',
        'IL': 'Asia/Jerusalem',
        'IT': 'Europe/Rome',
        'JM': 'America/Jamaica',
        'JP': 'Asia/Tokyo',
        'JO': 'Asia/Amman',
        'KZ': 'Asia/Almaty',
        'KE': 'Africa/Nairobi',
        'KI': 'Pacific/Tarawa',
        'KW': 'Asia/Kuwait',
        'KG': 'Asia/Bishkek',
        'LA': 'Asia/Vientiane',
        'LV': 'Europe/Riga',
        'LB': 'Asia/Beirut',
        'LS': 'Africa/Maseru',
        'LR': 'Africa/Monrovia',
        'LY': 'Africa/Tripoli',
        'LT': 'Europe/Vilnius',
        'LU': 'Europe/Luxembourg',
        'MO': 'Asia/Macau',
        'MK': 'Europe/Skopje',
        'MG': 'Indian/Antananarivo',
        'MW': 'Africa/Blantyre',
        'MY': 'Asia/Kuala_Lumpur',
        'MV': 'Indian/Maldives',
        'ML': 'Africa/Bamako',
        'MT': 'Europe/Malta',
        'MH': 'Pacific/Majuro',
        'MQ': 'America/Martinique',
        'MR': 'Africa/Nouakchott',
        'MU': 'Indian/Mauritius',
        'YT': 'Indian/Mayotte',
        'MX': 'America/Mexico_City',
        'FM': 'Pacific/Guam',
        'MD': 'Europe/Chisinau',
        'MC': 'Europe/Monaco',
        'MN': 'Asia/Ulaanbaatar',
        'ME': 'Europe/Belgrade',
        'MS': 'America/Port_of_Spain',
        'MA': 'Africa/Casablanca',
        'MZ': 'Africa/Maputo',
        'MM': 'Asia/Yangon',
        'MW': 'Africa/Blantyre',
        'NA': 'Africa/Windhoek',
        'NP': 'Asia/Kathmandu',
        'NI': 'America/Managua',
        'NE': 'Africa/Niamey',
        'NG': 'Africa/Lagos',
        'NO': 'Europe/Oslo',
        'NP': 'Asia/Kathmandu',
        'PK': 'Asia/Karachi',
        'PA': 'America/Panama',
        'PG': 'Pacific/Port_Moresby',
        'PY': 'America/Asuncion',
        'PE': 'America/Lima',
        'PH': 'Asia/Manila',
        'PL': 'Europe/Warsaw',
        'PT': 'Europe/Lisbon',
        'PR': 'America/Puerto_Rico',
        'QA': 'Asia/Qatar',
        'RO': 'Europe/Bucharest',
        'RU': 'Europe/Moscow',
        'RW': 'Africa/Kigali',
        'SA': 'Asia/Riyadh',
        'RS': 'Europe/Belgrade',
        'SC': 'Indian/Mahe',
        'SL': 'Africa/Freetown',
        'SG': 'Asia/Singapore',
        'SK': 'Europe/Bratislava',
        'SI': 'Europe/Ljubljana',
        'SE': 'Europe/Stockholm',
        'CH': 'Europe/Zurich',
        'SY': 'Asia/Damascus',
        'TJ': 'Asia/Dushanbe',
        'TH': 'Asia/Bangkok',
        'TG': 'Africa/Lome',
        'TO': 'Pacific/Tongatapu',
        'TT': 'America/Port_of_Spain',
        'TN': 'Africa/Tunis',
        'TR': 'Europe/Istanbul',
        'TM': 'Asia/Ashgabat',
        'UG': 'Africa/Kampala',
        'UA': 'Europe/Kiev',
        'AE': 'Asia/Dubai',
        'GB': 'Europe/London',
        'US': 'America/New_York',
        'UY': 'America/Montevideo',
        'UZ': 'Asia/Tashkent',
        'VU': 'Pacific/Efate',
        'VE': 'America/Caracas',
        'VN': 'Asia/Ho_Chi_Minh',
        'YE': 'Asia/Aden',
        'ZM': 'Africa/Lusaka',
        'ZW': 'Africa/Harare',
        # Diğer ülkeler için zaman dilimlerini burada ekleyebilirsiniz
    }
    return time_zones.get(country, 'UTC')  # Varsayılan olarak UTC döner
from django.utils.timezone import make_naive

from django.utils.timezone import make_aware
import pytz

import pytz
from django.utils import timezone

def get_kayit_tarihi_from_request(request):
    ip_address = request.META.get('REMOTE_ADDR', None)
    if ip_address == '127.0.0.1':
        ip_address = '78.173.5.100'  # Lokal IP için genel bir IP kullan

    try:
        with geoip2.database.Reader('geoip/GeoLite2-Country.mmdb') as reader:
            response = reader.country(ip_address)
            country = response.country.iso_code
            country_time_zone = get_time_zone_from_country(country)
            tz = pytz.timezone(country_time_zone)

            utc_time = timezone.now()  # BU ZATEN UTC
            local_time = utc_time.astimezone(tz)

            # saniye ve mikrosaniyeyi sıfırlıyoruz
            local_time = local_time.replace(second=0, microsecond=0)

            # ŞİMDİ: tekrar UTC'ye dönüyoruz
            final_utc_time = local_time.astimezone(pytz.UTC)

            print("Kayıt zamanı UTC olarak:", final_utc_time)
            return final_utc_time
    except Exception:
        utc_time = timezone.now().replace(second=0, microsecond=0)
        return utc_time


def get_csrf_token(request):
    token = get_token(request)
    #print(token)
    return JsonResponse({'csrfToken': token})
# Salt değeri ve minimum hash uzunluğu belirleyin
HASHIDS_SALT = "habip_elis_12345"
HASHIDS_MIN_LENGTH = 32

hashids = Hashids(salt=HASHIDS_SALT, min_length=HASHIDS_MIN_LENGTH)

def encode_id(id):
    return hashids.encode(id)

def decode_id(hash_id):
    ids = hashids.decode(hash_id)
    return ids[0] if ids else None
def page_not_found_view(request, exception):
    return render(request, '404.html')

def custom_error_500(request):
    return redirect("/")   
"""
trans = translate(language='tr')
    z = BlogPost.objects.all()
    content = {"trans":trans,"z":z,"dil":dil_bilgisi}
"""
def yetkisiz(request):
    return render(request,"yetkisiz.html",sozluk_yapisi())
def super_admin_kontrolu(request):
    if request.user.is_superuser:
            return 1
    else:
        return 0
def dil_bilgisi():
    return get_language()
def translate(language):
    cur_language = get_language()
    try:
        activate(language)
        text = gettext('Hello')

    finally:
        activate(cur_language)
    return text

def sozluk_yapisi():
    trans = translate(language='tr')
    dil = dil_bilgisi()
    sozluk = {"trans":trans,"dil":dil_bilgisi()}
    sozluk ["diller"] = dil_ayarla.objects.all()
    sozluk ["sayfalogusu"] = sayfa_logosu.objects.last()
    sozluk ["sayfa_iconu"] = sayfa_iconu.objects.last()
    sozluk ["site_adi"] = site_adi.objects.last()
    sozluk ["layout"] = layout.objects.last()
    sozluk ["color_sheme"] = color_sheme.objects.last()
    sozluk["sidebar_boyutu"] = sidebar_boyutu.objects.last()
    sozluk["side_bar_gorunum"] = side_bar_gorunum.objects.last()
    sozluk["layout_pozisyonu"] = layout_pozisyonu.objects.last()
    sozluk["topbar_color"] = topbar_color.objects.last()
    sozluk["sidebar_rengi"] = sidebar_rengi.objects.last()
    sozluk["layout_uzunlugu"] = layout_uzunlugu.objects.last()
    sozluk["layout_sitili"] = layout_sitili.objects.last()
    sozluk["etiketler"] = faturalardaki_gelir_gider_etiketi.objects.last()
    sozluk["latest_actions"] =LogEntry.objects.all().order_by('-action_time')[:10]
    sozluk["binalar"]=bina_goruntuleri.objects.all()
    sozluk["trans"]=trans
    sozluk["dil"]=dil
    return sozluk
def loglar(request):
    content = sozluk_yapisi()
    if request.user.is_authenticated:
        if request.user.is_superuser :
            content["latest_actions"] =LogEntry.objects.filter(object_repr = get_object_or_none(calisan_maas_durumlari,id = 4)  ).order_by('-action_time')
        if request.GET.get("search"):
            search = request.GET.get("search")
            if search:
                profile = LogEntry.objects.all().order_by('-action_time')
            else:
                profile = LogEntry.objects.all().order_by('-action_time')
        else:
            profile = LogEntry.objects.all().order_by('-action_time')
        page_num = request.GET.get('page', 1)
        paginator = Paginator(profile, 10) # 6 employees per page
        try:
            page_obj = paginator.page(page_num)
        except PageNotAnInteger:
            # if page is not an integer, deliver the first page
            page_obj = paginator.page(1)
        except EmptyPage:
            # if the page is out of range, deliver the last page
            page_obj = paginator.page(paginator.num_pages)
        content["santiyeler"] = profile
        content["top"]  = profile
        content["medya"] = page_obj

    else:
        return redirect_with_language("/users/login/")

    return render(request,"santiye_yonetimi/loglar.html",content)
# views.py
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json

import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

import requests

service={
  "type": "service_account",
  "project_id": "biadago-cloud",
  "private_key_id": "70e24c8bcac3a97b664fb3a14c79322736d59576",
  "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQDWvSHVgKNhVZmr\n1tS5q+YmEgguBgrgOMXbvy/yFUArRYkO1i5mIS8tGtuQEsWX3Xytomk7/bKFsrkk\nk42eGAdtNX9PaR70WZBgo2ScEsp9cfnALLR9Qobmgef3SDmGdJpYgTZSc6isCtC/\nOrARC7pQlZjHbafRFzFONRJPfVHHpi67J1LGJaOND9/+CK2+r8qhEWujGBmgKR3i\nQ9pfKTXWmJsabx1d0aaSTBoLUPfWaUa75EP321JFhVRRVWg485L2BF7RceU4J1cK\n3Po7knfOClfzW8lfSWSO7p6OaAgHfz9kn5vxXMgh7DncfJXFw5dJ8++V9kqsS4jw\ntrY13mHJAgMBAAECggEANey51d0fah5uYCYrNlMSEQZfMnuG+KaZHSVGO4MVoagt\nEmI7tZ7os2l2sJfeMdRHbm0GAzAiyAtJDgPVwNkk6EpSZZUu1kq1hGcTqVPYsKbS\nAm3Xh0sRCEqf/0uOUpLufYI5K3xq44U1xYfN1gH9cCYY/x+s0EeENLEEH50HT9yY\nK7v9RfNx69anfflIlrRt3VHm5Ocj+bGKeBzoMQ+YiIC91fNAa85jr+heOdvnObmP\n2ipEQSmhmyze1Z+Fdxz+GBXA74f0c78tLjBU3WiGHBf2EXMSUCtbTSsiuF9DZpyN\nf1/U+cKJ1EDmajUvwpSEj1tDwihXPE1hFuruYwioZwKBgQD0lTQUU16KGbBeLxUT\nKH6GZlQJbl7v6jHAKzHn/8QaZfsfoF2VKw3f4yRaAESS48NdvIOC8rPsEp9yjzNs\nV38jRFyAI5WgFVIhZR3gyIiFHjhyBLKMWQAh024zWBw9GvFwCG8Uc2aNAyPGgjLH\nn/ZWY1RdxLpf8D8eF4+RIvk6swKBgQDgw0no40cHC2CjMBrtLWOOMstxNEeIsYxf\n7+csRwXyWWkIW4JHOyTpKGBQjC+SNic8KvQ7+DywrCS7igz3qcKRIvBhUq9qtt6v\nDtpX8k48MO9VeumcDDHLy/27vchC40ikW92VWa3X5ZE+FQjux0K3cP1do3gUEgE+\nHVbBLcIfkwKBgQDwVQR9zIYjUabahY1B7BKX4klFkyy6tvf4CvnZLJv4DKm8pAoR\nH+NcUohP39+CL0iz/R+FNxPRL2N6YHh5R2josK3sRAss6IZxxjibvrFXjSCN+Uux\nWWsl0eqBjV0CNk10dvUftV3ZxnILB7j6K5cVwDkQgtVYnGyJF0G9rg4UvQKBgAQY\nSr5tdZvRPz953uO3UfsDPeWgGDWLVo1g54tM9/TEYD+Au0zk7PU6gRa2lx9I0Uot\nVinJigGGAV1RVI8mjp7qTgrX4M5G6qOx15SGm5pJIfMivCLVrgqSetryyDU/wtEL\nw2u3KI2oZw8EfxcqljKVYmhUVBm5gkBJdI0scj71AoGBAKkpu+NVy55oql/fYMRv\ngXd7KSAGE5UdR0MYRPe8SJ4o5G1rrDahjW7N9LWwJQwWxek5xHCkFJOndE1g4jP9\nVbjKqD4iqyiyuSi/6wWWdEbOMZrp7t+MTGRPELXns3Fjl/aDmDNd3uyW2QhIgqjl\nqbOmlytGXqUhIS+DW4WU8ALI\n-----END PRIVATE KEY-----\n",
  "client_email": "firebase-adminsdk-x8ood@biadago-cloud.iam.gserviceaccount.com",
  "client_id": "118028014141792543512",
  "auth_uri": "https://accounts.google.com/o/oauth2/auth",
  "token_uri": "https://oauth2.googleapis.com/token",
  "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
  "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-x8ood%40biadago-cloud.iam.gserviceaccount.com",
  "universe_domain": "googleapis.com"
}
import firebase_admin
from firebase_admin import credentials, messaging

# Uygulama zaten başlatılmadıysa başlat
if not firebase_admin._apps:
    cred = credentials.Certificate(service)  # Dosya yolunu doğru ver
    firebase_admin.initialize_app(cred)

def send_fcm_message(token, title, body):
    message = messaging.Message(
        notification=messaging.Notification(
            title=title,
            body=body,
        ),
        token=token,
    )
    print("Mesaj gönderiliyor:", message)
    response = messaging.send(message)
    return response  # Mesaj ID'si döner

@csrf_exempt
def save_device_token(request):
    print("Token kaydetme isteği alındı")
    if request.method == "POST":
        data = json.loads(request.body)
        token = data.get("token")
        print("Token geldi:", token)

        # Token veritabanına kaydet (örneğin kullanıcı modeli güncelle)
        if request.user.is_authenticated:
            CustomUser.objects.filter(id=request.user.id).update(token=token, platform="web")
        else:
            # İstersen anonim kullanıcılar için farklı işlem yap
            pass

        # Test amaçlı bildirim gönder
        resp_json = send_fcm_message(
            token,
            "Test Bildirimi",
            "Token başarıyla kaydedildi ve test bildirimi gönderildi."
        )
        print("FCM gönderim durumu:", resp_json)
        
        return JsonResponse({"status": "ok"})
    return JsonResponse({"error": "Invalid method"}, status=405)


#superadmin Kontrol
def yetki(request):
    if request.user.is_superuser:

        pass
    else:

        return redirect_with_language("main:yetkisiz")
# Create your views here,
# Anasayfa

def homepage(request):
    content = sozluk_yapisi()
    if request.user.is_authenticated:
        pass
    else:
        #print("login")
        return redirect_with_language("users:login")
    if request.user.is_authenticated:
        content = sozluk_yapisi()
        CustomUser.objects.filter(id = request.user.id).update(kullanici_tercih_dili =content["dil"])
        if super_admin_kontrolu(request):
            profile =Gelir_Bilgisi.objects.all()
            kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dashboard_gorme:
                        profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-id")
                        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user,silinme_bilgisi = False).order_by("-id")
                content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        page_num = request.GET.get('page', 1)
        paginator = Paginator(profile, 5) # 6 employees per page

        try:
            page_obj = paginator.page(page_num)
        except PageNotAnInteger:
                # if page is not an integer, deliver the first page
            page_obj = paginator.page(1)
        except EmptyPage:
                # if the page is out of range, deliver the last page
            page_obj = paginator.page(paginator.num_pages)

        content["santiyeler"] = profile
        if super_admin_kontrolu(request):
            profile =Gider_Bilgisi.objects.all()
            kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dashboard_gorme:
                        profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db,silinme_bilgisi = False).order_by("-id")
                        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user,silinme_bilgisi = False).order_by("-id")
                content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        page_num = request.GET.get('page', 1)
        paginator = Paginator(profile, 5) # 6 employees per page

        try:
            page_obj = paginator.page(page_num)
        except PageNotAnInteger:
                # if page is not an integer, deliver the first page
            page_obj = paginator.page(1)
        except EmptyPage:
                # if the page is out of range, deliver the last page
            page_obj = paginator.page(paginator.num_pages)
        if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dashboard_gorme:
                        bilgi_ver = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db,silinme_bilgisi = False).order_by("-fatura_tarihi")
                        sonuc = []
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
        else:
            bilgi_ver = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user,silinme_bilgisi = False).order_by("-fatura_tarihi")
            sonuc = []
        for i in bilgi_ver:
            y =  gider_urun_bilgisi.objects.filter(gider_bilgis = i)
            urun_tutari = 0
            for j in y:
                urun_tutari = urun_tutari + (float(j.urun_adeti)*float(j.urun_fiyati))
            y =  Gider_odemesi.objects.filter(gelir_kime_ait_oldugu = i)
            odeme_tutari = 0
            for j in y:
                odeme_tutari = odeme_tutari + float(j.tutar)
            if urun_tutari > odeme_tutari:
                sonuc.append(i)
            if len(sonuc) >= 5 :
                break
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dashboard_gorme:
                    content["gider"] = sonuc
                    content["bilgi"] = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db,silinme_bilgisi = False).order_by("-id")[:5]
                    content["son_gorevler"] = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =request.user.kullanicilar_db,silinme_bilgisi = False ).order_by("-id")[:5]
                    content["son_gorevler_bina"] = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =request.user.kullanicilar_db ,silinme_bilgisi = False).exclude(blok = None).last()
                    konum = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =request.user.kullanicilar_db,silinme_bilgisi = False ).exclude(blok = None).last()
                    kul = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")

            else:
                return redirect_with_language("main:yetkisiz")
        else:
            content["gider"] = sonuc
            content["bilgi"] = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user,silinme_bilgisi = False).order_by("-id")[:5]
            content["son_gorevler"] = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =request.user ,silinme_bilgisi = False).order_by("-id")[:5]
            content["son_gorevler_bina"] = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =request.user ,silinme_bilgisi = False).exclude(blok = None).last()
            konum = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =request.user,silinme_bilgisi = False ).exclude(blok = None).last()
            kul = request.user
    try: 
        weather_data = None
        ip_info = None
        
        # Kullanıcının IP adresini alıyoruz
        ip = get_client_ip(request) #
        
        # ipinfo.io API'sini kullanarak IP'ye göre konum alıyoruz
        ipinfo_api_url = f"http://ipinfo.io/{ip}/json"
        ip_response = requests.get(ipinfo_api_url)
        if ip_response.status_code == 200:
            ip_info = ip_response.json()
            loc = ip_info.get('loc')
            
            if loc:  # Eğer 'loc' None değilse
                #print(loc)
                location = loc.split(',')
                lat, lon = location[0], location[1]
                # OpenWeatherMap API'yi kullanarak hava durumu alıyoruz
                api_key = 'dee0661903df4f2c76ccfd8afab8be69'
                weather_api_url = f'http://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&units=metric&appid={api_key}'
                
                weather_response = requests.get(weather_api_url)
                #print(weather_response)
                if weather_response.status_code == 200:
                    weather_data = weather_response.json()
                a = weather_data["weather"][0]
                icon = a["icon"]
                content['weather_data'] = weather_data
                content['ip_info'] = ip_info
                content['icon'] = icon
                content['sehir'] = weather_data["name"]
        try:
            if konum.blok.proje_santiye_Ait.lat and konum.blok.proje_santiye_Ait.lon:
                lat, lon = konum.blok.proje_santiye_Ait.lat, konum.blok.proje_santiye_Ait.lon
                    # OpenWeatherMap API'yi kullanarak hava durumu alıyoruz
                api_key = 'dee0661903df4f2c76ccfd8afab8be69'
                weather_api_url = f'http://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&units=metric&appid={api_key}'
                
                weather_response = requests.get(weather_api_url)
                if weather_response.status_code == 200:
                    weather_data = weather_response.json()
                a = weather_data["weather"][0]
                icon = a["icon"]
                content['weather_data'] = weather_data
                content['ip_info'] = ip_info
                content['icon'] = icon
                content['sehir'] = weather_data["name"]
        except:
            pass     
    except:
        pass
    return render(request,"index.html",content)
def ana_sayfa(request):
    content = sozluk_yapisi()
    if request.user.is_authenticated:
        pass
    else:
        return redirect_with_language("/users/login/")
    if request.user.is_authenticated:
        content = sozluk_yapisi()
        if super_admin_kontrolu(request):
            profile =Gelir_Bilgisi.objects.all()
            kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dashboard_gorme:
                        profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-id")
                        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-id")
                content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        page_num = request.GET.get('page', 1)
        paginator = Paginator(profile, 5) # 6 employees per page

        try:
            page_obj = paginator.page(page_num)
        except PageNotAnInteger:
                # if page is not an integer, deliver the first page
            page_obj = paginator.page(1)
        except EmptyPage:
                # if the page is out of range, deliver the last page
            page_obj = paginator.page(paginator.num_pages)

        content["santiyeler"] = profile
        if super_admin_kontrolu(request):
            profile =Gider_Bilgisi.objects.all()
            kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dashboard_gorme:
                        profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-id")
                        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user.kullanicilar_db)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-id")
                content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = request.user)
        page_num = request.GET.get('page', 1)
        paginator = Paginator(profile, 5) # 6 employees per page

        try:
            page_obj = paginator.page(page_num)
        except PageNotAnInteger:
                # if page is not an integer, deliver the first page
            page_obj = paginator.page(1)
        except EmptyPage:
                # if the page is out of range, deliver the last page
            page_obj = paginator.page(paginator.num_pages)
        if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dashboard_gorme:
                        bilgi_ver = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-fatura_tarihi")
                        sonuc = []
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
        else:
            bilgi_ver = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-fatura_tarihi")
            sonuc = []
        for i in bilgi_ver:
            y =  gider_urun_bilgisi.objects.filter(gider_bilgis = i)
            urun_tutari = 0
            for j in y:
                urun_tutari = urun_tutari + (float(j.urun_adeti)*float(j.urun_fiyati))
            y =  Gider_odemesi.objects.filter(gelir_kime_ait_oldugu = i)
            odeme_tutari = 0
            for j in y:
                odeme_tutari = odeme_tutari + float(j.tutar)
            if urun_tutari > odeme_tutari:
                sonuc.append(i)
            if len(sonuc) >= 5 :
                break
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dashboard_gorme:
                    content["gider"] = sonuc
                    content["bilgi"] = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user.kullanicilar_db).order_by("-id")[:5]
                    content["son_gorevler"] = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =request.user.kullanicilar_db ).order_by("-id")[:5]
                else:
                    return redirect_with_language("main:yetkisiz")

            else:
                return redirect_with_language("main:yetkisiz")
        else:
            content["gider"] = sonuc
            content["bilgi"] = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = request.user).order_by("-id")[:5]
            content["son_gorevler"] = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =request.user ).order_by("-id")[:5]
    weather_data = None
    ip_info = None
    
    # Kullanıcının IP adresini alıyoruz
    ip = get_client_ip(request) #
    
    # ipinfo.io API'sini kullanarak IP'ye göre konum alıyoruz
    ipinfo_api_url = f"http://ipinfo.io/{ip}/json"
    ip_response = requests.get(ipinfo_api_url)
    #print(ip_response.json())
    if ip_response.status_code == 200:
        ip_info = ip_response.json()
        loc = ip_info.get('loc')
        
        if loc:  # Eğer 'loc' None değilse
            #print(loc)
            location = loc.split(',')
            lat, lon = location[0], location[1]
            
            # OpenWeatherMap API'yi kullanarak hava durumu alıyoruz
            api_key = 'dee0661903df4f2c76ccfd8afab8be69'
            weather_api_url = f'http://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&units=metric&appid={api_key}'
            
            weather_response = requests.get(weather_api_url)
            #print(weather_response)
            if weather_response.status_code == 200:
                weather_data = weather_response.json()
            a = weather_data["weather"][0]
            icon = a["icon"]
            content['weather_data'] = weather_data
            content['ip_info'] = ip_info
            content['icon'] = icon
            content['sehir'] = weather_data["name"]
            
    else:
        return redirect_with_language("/users/login/")

    return render(request,"index.html",content)
def homepage_2(request,hash):
    content = sozluk_yapisi()
    if request.user.is_authenticated:
        content = sozluk_yapisi()
        if super_admin_kontrolu(request):
            d = decode_id(hash)
            content["hashler"] = hash
            users = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = users
            kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
            profile = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users).order_by("-id")
            content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = users)
            bilgi_ver = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users).order_by("-fatura_tarihi")
            sonuc = []
            for i in bilgi_ver:
                y =  gider_urun_bilgisi.objects.filter(gider_bilgis = i)
                urun_tutari = 0
                for j in y:
                    urun_tutari = urun_tutari + (float(j.urun_adeti)*float(j.urun_fiyati))
                y =  Gider_odemesi.objects.filter(gelir_kime_ait_oldugu = i)
                odeme_tutari = 0
                for j in y:
                    odeme_tutari = odeme_tutari + float(j.tutar)
                if urun_tutari > odeme_tutari:
                    sonuc.append(i)
                if len(sonuc) >= 5 :
                    break
            content["gider"] = sonuc
            content["bilgi"] = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = users).order_by("-id")[:5]
            content["son_gorevler"] = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =users ).order_by("-id")[:5]
            content["son_gorevler_bina"] = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =users).exclude(blok = None).last()
            konum = IsplaniPlanlari.objects.filter(proje_ait_bilgisi =users ).exclude(blok = None).last()
            kul = users
            weather_data = None
            ip_info = None
            
            # Kullanıcının IP adresini alıyoruz
            ip = get_client_ip(request) #
            
            # ipinfo.io API'sini kullanarak IP'ye göre konum alıyoruz
            ipinfo_api_url = f"http://ipinfo.io/{ip}/json"
            ip_response = requests.get(ipinfo_api_url)
            if ip_response.status_code == 200:
                ip_info = ip_response.json()
                loc = ip_info.get('loc')
                
                if loc:  # Eğer 'loc' None değilse
                    #print(loc)
                    location = loc.split(',')
                    lat, lon = location[0], location[1]
                    # OpenWeatherMap API'yi kullanarak hava durumu alıyoruz
                    api_key = 'dee0661903df4f2c76ccfd8afab8be69'
                    weather_api_url = f'http://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&units=metric&appid={api_key}'
                    
                    weather_response = requests.get(weather_api_url)
                    #print(weather_response)
                    if weather_response.status_code == 200:
                        weather_data = weather_response.json()
                    a = weather_data["weather"][0]
                    icon = a["icon"]
                    content['weather_data'] = weather_data
                    content['ip_info'] = ip_info
                    content['icon'] = icon
                    content['sehir'] = weather_data["name"]
            try:
                if konum.blok.proje_santiye_Ait.lat and konum.blok.proje_santiye_Ait.lon:
                    lat, lon = konum.blok.proje_santiye_Ait.lat, konum.blok.proje_santiye_Ait.lon
                        # OpenWeatherMap API'yi kullanarak hava durumu alıyoruz
                    api_key = 'dee0661903df4f2c76ccfd8afab8be69'
                    weather_api_url = f'http://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&units=metric&appid={api_key}'
                    
                    weather_response = requests.get(weather_api_url)
                    if weather_response.status_code == 200:
                        weather_data = weather_response.json()
                    a = weather_data["weather"][0]
                    icon = a["icon"]
                    content['weather_data'] = weather_data
                    content['ip_info'] = ip_info
                    content['icon'] = icon
                    content['sehir'] = weather_data["name"]
            except:
                pass    
        else:
            pass
        content["santiyeler"] = profile

    return render(request,"index.html",content)

#site ayarı
def site_ayarlari(request):
    content = sozluk_yapisi()
    if request.user.is_authenticated:
        if request.user.is_superuser :

            pass
        else:

            return redirect_with_language("main:yetkisiz")
    else:
        return redirect_with_language("/users/login/")

    return render(request,"santiye_yonetimi/site_ayarlari.html",content)

def site_ayari_kaydet(request):
    if request.POST:
        data_layout = request.POST.get("data-layout")
        if data_layout:
            layout.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),isim =data_layout,data_layout =data_layout )
        data_bs_theme = request.POST.get("data-bs-theme")
        if data_bs_theme:
            color_sheme.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),isim =data_bs_theme,data_bs_theme =data_bs_theme )
        data_sidebar_visibility = request.POST.get("data-sidebar-visibility")
        if data_sidebar_visibility :
            side_bar_gorunum.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),isim =data_sidebar_visibility,data_sidebar_visibility =data_sidebar_visibility )
        data_layout_width = request.POST.get("data-layout-width")
        if data_layout_width:
            layout_uzunlugu.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),isim =data_layout_width,data_layout_width =data_layout_width )
        data_layout_position = request.POST.get("data-layout-position")
        if data_layout_position:
            layout_pozisyonu.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),isim =data_layout_position,data_layout_position =data_layout_position )
        data_topbar = request.POST.get("data-topbar")
        if data_topbar:
            topbar_color.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),isim =data_topbar,data_topbar =data_topbar )
        data_sidebar_size = request.POST.get("data-sidebar-size")
        if data_sidebar_size:
            sidebar_boyutu.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),isim =data_sidebar_size,data_sidebar_size =data_sidebar_size )
        data_layout_style = request.POST.get("data-layout-style")
        if data_layout_style:
            layout_sitili.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),isim =data_layout_style,data_layout_style =data_layout_style )
        data_sidebar = request.POST.get("data-sidebar")
        if data_sidebar:
            sidebar_rengi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),isim =data_sidebar,data_sidebar =data_sidebar )
        dark_logo = request.FILES.get("dark_logo")

        site_adi_bilgisi = request.POST.get("site_adi")
        footeryazisi = request.POST.get("footeryazisi")
        if site_adi_bilgisi:
            site_adi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),site_adi_sekme_tr = site_adi_bilgisi)
        if footeryazisi:
            site_adi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),footer = footeryazisi)
        gideretiketi = request.POST.get("gideretiketi")
        geliretiketi = request.POST.get("gelir_etiketi")
        if gideretiketi and geliretiketi :
            faturalardaki_gelir_gider_etiketi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gider_etiketi = gideretiketi,gelir_etiketi = geliretiketi )


        if dark_logo:
            u = sayfa_logosu.objects.last()
            u.image = dark_logo
            u.save()
        light_logo = request.FILES.get("light_logo")
        if light_logo:
            u = sayfa_logosu.objects.last()
            u.dark_image = light_logo
            u.save()
        icon = request.FILES.get("icon")
        if icon:
            u = sayfa_iconu.objects.last()
            u.sayfa_icon = icon
            u.save()
    return redirect_with_language("main:site_ayarlari")
#site ayarı

# Create your views here.
#şantiye Ekleme
def santiye_ekle(request):
    if request.user.is_authenticated:
        if request.user.is_superuser :

            pass
        else:

            return redirect_with_language("main:yetkisiz")
        if request.POST:
            yetkiliAdSoyad = request.POST.get("yetkiliAdSoyad")
            email = request.POST.get("email")
            santiyeAdi = request.POST.get("santiyeAdi")
            sfire = request.POST.get("sfire")
            #print(yetkiliAdSoyad,email,santiyeAdi,sfire)
            newUser = CustomUser(username =email,email=email,first_name = santiyeAdi,last_name =yetkiliAdSoyad )
            newUser.set_password(sfire)
            newUser.save()
            return redirect_with_language("main:santiye_listele")

    else:
        return redirect_with_language("/users/login/")

    return render(request,"santiye_yonetimi/santiye_ekleme.html",sozluk_yapisi())
#şantiye Listleme Ve Ayarları

def santiye_listele(request):
    content = sozluk_yapisi()
    if request.user.is_authenticated:
        if request.user.is_superuser :

            pass
        else:

            return redirect_with_language("main:yetkisiz")
        if request.GET.get("search"):
            search = request.GET.get("search")
            if search:
                profile = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).filter(Q(last_name__icontains = search)|Q(first_name__icontains = search)|Q(email__icontains = search) ).order_by("-id")
            else:
                profile = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        else:
            profile = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        page_num = request.GET.get('page', 1)
        paginator = Paginator(profile, 10) # 6 employees per page
        try:
            page_obj = paginator.page(page_num)
        except PageNotAnInteger:
            # if page is not an integer, deliver the first page
            page_obj = paginator.page(1)
        except EmptyPage:
            # if the page is out of range, deliver the last page
            page_obj = paginator.page(paginator.num_pages)
        content["santiyeler"] = profile
        content["top"]  = profile
        content["medya"] = page_obj

    else:
        return redirect_with_language("/users/login/")

    return render(request,"santiye_yonetimi/santiye_ayarlari.html",content)

#şantiye Silme
def santiye_sil(request):
    content = {}
    if request.user.is_authenticated:
        yetki(request)
        content["santiyeler"] = CustomUser.objects.all()
        if request.POST:
            sil = request.POST.get("buttonId")
            CustomUser.objects.filter(id = sil).update(kullanici_silme_bilgisi = True)
        return redirect_with_language("main:santiye_listele")
    else:
        return redirect_with_language("/users/login/")

    return redirect_with_language("main:santiye_listele")

#şantiye düzeltme
def santiye_duzelt(request):
    content = {}
    if request.user.is_authenticated:
        if request.user.is_superuser :

            pass
        else:

            return redirect_with_language("main:yetkisiz")
        content["santiyeler"] = CustomUser.objects.all()
        if request.POST:
            sil = request.POST.get("buttonId")
            yetkili_adi = request.POST.get("yetkili_adi")
            email = request.POST.get("email")
            santiyeadi = request.POST.get("santiyeadi")
            CustomUser.objects.filter(id = sil).update(email=email,first_name = santiyeadi,last_name =yetkili_adi)
        return redirect_with_language("main:santiye_listele")
    else:
        return redirect_with_language("/users/login/")

    return redirect_with_language("main:santiye_listele")

#dil_ayarlari
def dil_ayari_listele(request):
    content = sozluk_yapisi()
    if request.user.is_authenticated:
        if request.user.is_superuser :

            pass
        else:

            return redirect_with_language("main:yetkisiz")
        if request.GET.get("search"):
            search = request.GET.get("search")
            if search:
                profile = dil_ayarla.objects.filter(Q(dil_adi__icontains = search)|Q(dil_kisaltması__icontains = search)).order_by("-id")
            else:
                profile = dil_ayarla.objects.all().order_by("-id")
        else:
            profile = dil_ayarla.objects.all().order_by("-id")
        page_num = request.GET.get('page', 1)
        paginator = Paginator(profile, 10) # 6 employees per page
        try:
            page_obj = paginator.page(page_num)
        except PageNotAnInteger:
            # if page is not an integer, deliver the first page
            page_obj = paginator.page(1)
        except EmptyPage:
            # if the page is out of range, deliver the last page
            page_obj = paginator.page(paginator.num_pages)
        content["santiyeler"] = profile
        content["top"]  = profile
        content["medya"] = page_obj

    else:
        return redirect_with_language("/users/login/")

    return render(request,"santiye_yonetimi/dil_ayarlari.html",content)

#dil Ekleme
def dil_ekle(request):
    if request.user.is_authenticated:
        if request.user.is_superuser :

            pass
        else:

            return redirect_with_language("main:yetkisiz")
        if request.POST:
            yetkili_adi = request.POST.get("yetkili_adi")
            dilkisitlamasi = request.POST.get("dilkisitlamasi")
            santiyeadi = request.FILES.get("santiyeadi")
            dil_aktiflik_durumu = request.POST.get("dil_aktiflik_durumu")
            if dil_aktiflik_durumu == "1":
                dil_aktiflik_durumu = True
            else:
                dil_aktiflik_durumu = False

            dil =dil_ayarla(dil_adi =yetkili_adi,dil_kisaltması =  dilkisitlamasi,dil_bayragi_icon =santiyeadi,
                                      dil_aktiflik_durumu = dil_aktiflik_durumu )

            dil.save()

            return redirect_with_language("main:dil_ayari_listele")

    else:
        return redirect_with_language("/users/login/")

#dil düzeltme
def dil_duzelt(request):
    if request.user.is_authenticated:
        if request.user.is_superuser :

            pass
        else:

            return redirect_with_language("main:yetkisiz")
        if request.POST:
            yetkili_adi = request.POST.get("yetkili_adi")
            dilkisitlamasi = request.POST.get("dilkisitlamasi")
            santiyeadi = request.FILES.get("santiyeadi")
            dil_aktiflik_durumu = request.POST.get("dil_aktiflik_durumu")
            buttonId = request.POST.get("buttonId")
            if dil_aktiflik_durumu == "1":
                dil_aktiflik_durumu = True
            else:
                dil_aktiflik_durumu = False
            dil_ayarla.objects.filter(id =buttonId ).delete()
            if santiyeadi:
                dil =dil_ayarla(dil_adi =yetkili_adi,dil_kisaltması =  dilkisitlamasi,dil_bayragi_icon =santiyeadi,
                                        dil_aktiflik_durumu = dil_aktiflik_durumu )

                dil.save()
            else:
                dil =dil_ayarla(dil_adi =yetkili_adi,dil_kisaltması =  dilkisitlamasi,
                                        dil_aktiflik_durumu = dil_aktiflik_durumu )

                dil.save()

            return redirect_with_language("main:dil_ayari_listele")

    else:
        return redirect_with_language("/users/login/")

#dil sil
def dil_sil(request):
    if request.user.is_authenticated:
        if request.user.is_superuser :

            pass
        else:

            return redirect_with_language("main:yetkisiz")
        if request.POST:
            sil = request.POST.get("buttonId")
            dil_ayarla.objects.filter(id = sil).delete()
        return redirect_with_language("main:dil_ayari_listele")
    else:
        return redirect_with_language("main:dil_ayari_listele")
#Proje Tipi


def proje_tipi_(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =proje_tipi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.projeler_gorme:
                    profile = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =proje_tipi.objects.filter(Q(proje_ait_bilgisi__last_name__icontains = search)|Q(Proje_tipi_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.projeler_gorme:
                        profile = proje_tipi.objects.filter(Q(proje_ait_bilgisi = request.user.kullanicilar_db) & Q(Proje_tipi_adi__icontains = search)& Q(silinme_bilgisi = False))

                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = proje_tipi.objects.filter(Q(proje_ait_bilgisi = request.user) & Q(Proje_tipi_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/proje_tipi.html",content)
#Proje Tipi
def proje_tipi_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        users = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = users
        profile = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            d = decode_id(hash)
            users = get_object_or_404(CustomUser,id = d)
            content["hashler"] = hash
            content["hash_bilgi"] = users
            profile = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = users)
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = proje_tipi.objects.filter(Q(proje_ait_bilgisi = request.user) & Q(Proje_tipi_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/proje_tipi.html",content)
#Proje Ekleme
def proje_ekleme(request):
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            proje_tipi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,Proje_tipi_adi = proje_tip_adi)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.projeler_olusturma:
                        proje_tip_adi   = request.POST.get("yetkili_adi")
                        proje_tipi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user.kullanicilar_db,Proje_tipi_adi = proje_tip_adi)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                proje_tip_adi   = request.POST.get("yetkili_adi")
                proje_tipi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user,Proje_tipi_adi = proje_tip_adi)
    return redirect_with_language("main:proje_tipi_")
#Proje Adı Silme
def proje_Adi_sil(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        proje_tipi.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.projeler_silme:
                    proje_tipi.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            proje_tipi.objects.filter(proje_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect_with_language("main:proje_tipi_")
#Proje Düzenlme
import folium
def proje_duzenle(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        silinmedurumu = request.POST.get("silinmedurumu")
        if silinmedurumu == "1":
            silinmedurumu = False
            proje_tipi.objects.filter(id = id).update(proje_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,Proje_tipi_adi = proje_tip_adi,silinme_bilgisi = silinmedurumu)
        elif silinmedurumu == "2":
            silinmedurumu = True
            proje_tipi.objects.filter(id = id).update(proje_ait_bilgisi = get_object_or_404(CustomUser,id = kullanici_bilgisi ) ,Proje_tipi_adi = proje_tip_adi,silinme_bilgisi = silinmedurumu)

    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.projeler_duzenleme:
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    proje_tipi.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,id = id).update(Proje_tipi_adi = proje_tip_adi)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("yetkili_adi")
            proje_tipi.objects.filter(proje_ait_bilgisi = request.user,id = id).update(Proje_tipi_adi = proje_tip_adi)
    return redirect_with_language("main:proje_tipi_")
#Şantiye Projesi Ekleme
def santiye_projesi_ekle_(request):
    content = sozluk_yapisi()
    m = folium.Map(location=[20, 0], zoom_start=2)
    # Haritayı HTML'ye dönüştürme
    map_html = m._repr_html_()
    content['map'] = map_html
    content["birim_bilgisi"] = birimler.objects.filter(silinme_bilgisi = False)
    content["proje_tipleri"] = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi =  request.user)
    if super_admin_kontrolu(request):
        profile =santiye.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_gorme:
                    profile = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db)
                    content["proje_tipleri"] = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi =  request.user.kullanicilar_db)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
            content["proje_tipleri"] = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi =  request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =santiye.objects.filter(Q(proje_ait_bilgisi__last_name__icontains = search)|Q(Proje_tipi_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.santiye_gorme:
                        profile = santiye.objects.filter(Q(proje_ait_bilgisi = request.user.kullanicilar_db) & Q(Proje_tipi_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = santiye.objects.filter(Q(proje_ait_bilgisi = request.user) & Q(Proje_tipi_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/santiye_projesi.html",content)
def santiye_projesi_ekle_2(request,hash):
    content = sozluk_yapisi()
   
    if super_admin_kontrolu(request):
        content["birim_bilgisi"] = birimler.objects.filter(silinme_bilgisi = False)
        d = decode_id(hash)
        users = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = users
        profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        content["proje_tipleri"] = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi =  users)
    else:
        profile = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            d = decode_id(hash)
            users = get_object_or_404(CustomUser,id = d)
            content["hashler"] = hash
            content["hash_bilgi"] = users
            profile =santiye.objects.filter(Q(proje_ait_bilgisi = users) & Q(proje_ait_bilgisi__last_name__icontains = search)|Q(Proje_tipi_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = santiye.objects.filter(Q(proje_ait_bilgisi = request.user) & Q(Proje_tipi_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/santiye_projesi.html",content)
#asdas
def santiye_projesi_bloklar_ekle_(request,id):
    content = sozluk_yapisi()
    content["id_bilgisi"] = id
    content["proje_tipleri"] = proje_tipi.objects.filter(proje_ait_bilgisi =  request.user)
    if super_admin_kontrolu(request):
        profile =bloglar.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.blog_gorme:
                    profile = bloglar.objects.filter(proje_santiye_Ait__id = id)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = bloglar.objects.filter(proje_santiye_Ait__id = id)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =bloglar.objects.filter(Q(proje_santiye_Ait__proje_ait_bilgisi__last_name__icontains = search)|Q(proje_santiye_Ait__Proje_tipi_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = bloglar.objects.filter(Q(proje_ait_bilgisi = request.user) & Q(blog_adi__icontains = search))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/santiye_projesi_blok_ekle.html",content)

def blog_ekle(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.blog_olusturma:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    
    if request.POST:
        santiye_bilgisi = request.POST.get("santiye_bilgisi")
        blok_adi = request.POST.get("blok_adi")
        kat_sayisi = request.POST.get("kat_sayisi")
        baslangictarihi = request.POST.get("baslangictarihi")
        bitistarihi =request.POST.get("bitistarihi")
        bloglar.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
            proje_ait_bilgisi = get_object_or_404(santiye,id = santiye_bilgisi).proje_ait_bilgisi,
            proje_santiye_Ait = get_object_or_404(santiye,id = santiye_bilgisi),
            blog_adi = blok_adi,kat_sayisi = kat_sayisi,
            baslangic_tarihi = baslangictarihi,bitis_tarihi = bitistarihi
        )
        y = "/siteblog/"+santiye_bilgisi+"/"
    return redirect_with_language("main:santiye_projesi_ekle_")

def blog_duzenle(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.blog_duzenleme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    if request.POST:
        santiye_bilgisi = request.POST.get("santiye_bilgisi")
        blog = request.POST.get("blog")
        blok_adi = request.POST.get("blok_adi")
        kat_sayisi = request.POST.get("kat_sayisi")
        baslangictarihi = request.POST.get("baslangictarihi")
        bitistarihi =request.POST.get("bitistarihi")
        bloglar.objects.filter(id = blog).update(
            proje_ait_bilgisi = get_object_or_404(santiye,id = santiye_bilgisi).proje_ait_bilgisi,
            proje_santiye_Ait = get_object_or_404(santiye,id = santiye_bilgisi),
            blog_adi = blok_adi,kat_sayisi = kat_sayisi,
            baslangic_tarihi = baslangictarihi,bitis_tarihi = bitistarihi
        )
        y = "/siteblog/"+santiye_bilgisi+"/"
    return redirect_with_language("main:santiye_projesi_ekle_")
def blog_sil(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.blog_silme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    if request.POST:
        buttonId = request.POST.get("buttonId")
        geri = request.POST.get("geri")
        blog_bilgisi = get_object_or_404(bloglar,id = buttonId)
        bloglar.objects.filter(id = buttonId).delete()
        y = "/siteblog/"+geri+"/"
    return redirect_with_language("main:santiye_projesi_ekle_")
def santiye_ekleme_sahibi(request):
    if request.POST:
        if super_admin_kontrolu(request):
            kullanici = request.POST.get("kullanici")
            link = "/addsitesuperadmin/"+kullanici
            return redirect_with_language(link)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.santiye_olusturma:
                            projetipi = request.POST.get("projetipi")
                            proje_adi = request.POST.get("yetkili_adi")
                            resim_secme = request.POST.get("options")
                            a = santiye.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user.kullanicilar_db,proje_tipi = get_object_or_404(proje_tipi,id = projetipi),
                                                proje_adi = proje_adi,bina_goruntuleri_aitlik = get_object_or_none(bina_goruntuleri,id = resim_secme)
                                    )
                                                
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                projetipi = request.POST.get("projetipi")
                proje_adi = request.POST.get("yetkili_adi")
                resim_secme = request.POST.get("options")
                a = santiye.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user,proje_tipi = get_object_or_404(proje_tipi,id = projetipi),
                                    proje_adi = proje_adi,bina_goruntuleri_aitlik = get_object_or_none(bina_goruntuleri,id = resim_secme)
                                    )
    return redirect_with_language("main:santiye_projesi_ekle_")
def santiye_ekleme_sahibi_2(request,hash):
    if request.POST:
        content = sozluk_yapisi()
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        if super_admin_kontrolu(request):
            projetipi = request.POST.get("projetipi")
            proje_adi = request.POST.get("yetkili_adi")

            a = santiye.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = users,proje_tipi = get_object_or_404(proje_tipi,id = projetipi),
                                    proje_adi = proje_adi
                                    )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.santiye_olusturma:
                            projetipi = request.POST.get("projetipi")
                            proje_adi = request.POST.get("yetkili_adi")

                            a = santiye.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user.kullanicilar_db,proje_tipi = get_object_or_404(proje_tipi,id = projetipi),
                                                proje_adi = proje_adi
                                                )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                projetipi = request.POST.get("projetipi")
                proje_adi = request.POST.get("yetkili_adi")

                a = santiye.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user,proje_tipi = get_object_or_404(proje_tipi,id = projetipi),
                                    proje_adi = proje_adi
                                    )
    return redirect_with_language("main:santiye_projesi_ekle_2",hash)
def santiye_ekleme_super_admin(request,id):
    content = sozluk_yapisi()
    content["proje_tipleri"] = proje_tipi.objects.filter(proje_ait_bilgisi =  get_object_or_404(CustomUser,id = id))
    if request.user.is_superuser :

        pass
    else:

        return redirect_with_language("main:yetkisiz")
    if request.POST:
        projetipi = request.POST.get("projetipi")
        proje_adi = request.POST.get("yetkili_adi")
        baslangic_tarihi = request.POST.get("baslangic_tarihi")
        bitis_tarihi = request.POST.get("bitis_tarihi")
        latitude = request.POST.get("latitude")
        longitude = request.POST.get("longitude")
        a = santiye.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = get_object_or_404(CustomUser,id = id),proje_tipi = get_object_or_404(proje_tipi,id = projetipi),
                               proje_adi = proje_adi,baslangic_tarihi = baslangic_tarihi,
                               tahmini_bitis_tarihi = bitis_tarihi,lat = latitude,lon = longitude
                               )
        return redirect_with_language("main:santiye_projesi_ekle_")
    return render(request,"santiye_yonetimi/super_admin_santiye_ekleme.html",content)


def santiye_projesi_sil(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        santiye.objects.filter(id = id).update(silinme_bilgisi = True)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_silme:
                    santiye.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,id = id).update(silinme_bilgisi = True)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            santiye.objects.filter(proje_ait_bilgisi = request.user,id = id).update(silinme_bilgisi = True)
    return redirect_with_language("main:santiye_projesi_ekle_")


def santiye_projesi_duzenle(request):
    content = {}
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        silinmedurumu = request.POST.get("silinmedurumu")
        resim_secme = request.POST.get("options")
        if silinmedurumu == "1":
            silinmedurumu = False
            santiye.objects.filter(id = id).update(proje_adi = proje_tip_adi,silinme_bilgisi = silinmedurumu)
        elif silinmedurumu == "2":
            silinmedurumu = True
            santiye.objects.filter(id = id).update(proje_adi = proje_tip_adi,silinme_bilgisi = silinmedurumu)

    else:
        resim_secme = request.POST.get("options")
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_duzenleme:
                    proje_tip_adi   = request.POST.get("yetkili_adi")
                    santiye.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,id = id).update(proje_adi = proje_tip_adi,bina_goruntuleri_aitlik = get_object_or_none(bina_goruntuleri,id = resim_secme))
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            proje_tip_adi   = request.POST.get("yetkili_adi")
            santiye.objects.filter(proje_ait_bilgisi = request.user,id = id).update(proje_adi = proje_tip_adi,bina_goruntuleri_aitlik = get_object_or_none(bina_goruntuleri,id = resim_secme))
    return redirect_with_language("main:santiye_projesi_ekle_")

#şantiye Kalemleri
def santtiye_kalemleri(request,id):
    content = sozluk_yapisi()
    content["birim_bilgisi"] = birimler.objects.filter(silinme_bilgisi = False)
    if request.user.is_authenticated:
        if request.user.is_superuser:
            kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
            profile = santiye_kalemleri.objects.filter(proje_santiye_Ait = get_object_or_404(santiye,id = id))
            content["santiyeler_bilgileri"] = bloglar.objects.all()
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kalemleri_gorme:
                            content["santiyeler_bilgileri"] = bloglar.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,proje_santiye_Ait_id = id )
                            profile = santiye_kalemleri.objects.filter(proje_santiye_Ait = get_object_or_404(santiye,id = id) ,silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                content["santiyeler_bilgileri"] = bloglar.objects.filter(proje_ait_bilgisi = request.user,proje_santiye_Ait_id = id )
                profile = santiye_kalemleri.objects.filter(proje_santiye_Ait = get_object_or_404(santiye,id = id) ,silinme_bilgisi = False,proje_ait_bilgisi = request.user)

        if request.GET.get("search"):
            search = request.GET.get("search")
            if request.user.is_superuser:
                kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
                content["kullanicilar"] =kullanicilar
                profile = santiye_kalemleri.objects.filter(proje_santiye_Ait = get_object_or_404(santiye,id = id)).filter(Q(proje_ait_bilgisi__last_name__icontains =search)| Q(kalem_adi__icontains =search)|  Q(proje_santiye_Ait__proje_adi__icontains =search))
                content["santiyeler_bilgileri"] = santiye.objects.all()
            else:
                if request.user.kullanicilar_db:
                    a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                    if a:
                        if a.izinler.kalemleri_gorme:
                                content["santiyeler_bilgileri"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db)
                                profile = santiye_kalemleri.objects.filter(proje_santiye_Ait = get_object_or_404(santiye,id = id) ,silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db).filter(Q(kalem_adi__icontains =search) | Q(proje_santiye_Ait__proje_adi__icontains =search))
                        else:
                            return redirect_with_language("main:yetkisiz")
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    content["santiyeler_bilgileri"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
                    profile = santiye_kalemleri.objects.filter(proje_santiye_Ait = get_object_or_404(santiye,id = id) ,silinme_bilgisi = False,proje_ait_bilgisi = request.user).filter(Q(kalem_adi__icontains =search) | Q(proje_santiye_Ait__proje_adi__icontains =search))

        page_num = request.GET.get('page', 1)
        paginator = Paginator(profile, 10) # 6 employees per page
        try:
            page_obj = paginator.page(page_num)
        except PageNotAnInteger:
                # if page is not an integer, deliver the first page
            page_obj = paginator.page(1)
        except EmptyPage:
                # if the page is out of range, deliver the last page
            page_obj = paginator.page(paginator.num_pages)
        content["santiyeler"] = profile
        content["top"]  = profile
        content["medya"] = page_obj
    else:
        return redirect_with_language("/users/login/")
    return render(request,"santiye_yonetimi/santiye_kalemleri.html",content)

def santiyeye_kalem_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            return redirect_with_language("main:santiye_kalem_ekle_admin",kullanici)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kalemleri_olusturma:
                        projetipi = request.POST.getlist("projetipi")
                        yetkili_adi = request.POST.get("yetkili_adi")
                        santiye_agirligi = request.POST.get("katsayisi")
                        finansal_agirlik = request.POST.get("blogsayisi")
                        metraj = request.POST.get("metraj")
                        tutar = request.POST.get("tutar")
                        birim_bilgisi = request.POST.get("birim_bilgisi")
                        kata_veya_binaya_daihil = request.POST.get("kata_veya_binaya_daihil")
                        id = bloglar.objects.filter(id__in = projetipi).first()
                        kalem = santiye_kalemleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            proje_ait_bilgisi = request.user.kullanicilar_db,
                            proje_santiye_Ait =id.proje_santiye_Ait,
                            kalem_adi = yetkili_adi,santiye_agirligi = santiye_agirligi,
                            santiye_finansal_agirligi = finansal_agirlik,
                            birimi = get_object_or_404(birimler,id =birim_bilgisi ),metraj = metraj,
                            tutari = tutar
                        )
                        if kata_veya_binaya_daihil == "0":
                            blog_lar = bloglar.objects.filter(id__in = projetipi)
                            for i in blog_lar:
                                for j in range(0,int(i.kat_sayisi)):
                                    santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                        proje_ait_bilgisi = request.user.kullanicilar_db,
                                        proje_santiye_Ait = id.proje_santiye_Ait,
                                        kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =kalem.id ),
                                        kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                                    )
                        elif kata_veya_binaya_daihil == "1":
                            blog_lar = bloglar.objects.filter(id__in = projetipi)
                            for i in blog_lar:
                                for j in range(0,int(i.kat_sayisi)):
                                    santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                        proje_ait_bilgisi = request.user.kullanicilar_db,
                                        proje_santiye_Ait = id.proje_santiye_Ait,
                                        kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =kalem.id ),
                                        kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                                    )
                                    break
                        elif kata_veya_binaya_daihil == "2":
                            blog_lar = bloglar.objects.filter(id__in = projetipi)
                            for i in blog_lar:
                                for j in range(0,4):
                                    santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                        proje_ait_bilgisi = request.user.kullanicilar_db,
                                        proje_santiye_Ait = id.proje_santiye_Ait,
                                        kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =kalem.id ),
                                        kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                                    )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                projetipi = request.POST.getlist("projetipi")
                yetkili_adi = request.POST.get("yetkili_adi")
                santiye_agirligi = request.POST.get("katsayisi")
                finansal_agirlik = request.POST.get("blogsayisi")
                metraj = request.POST.get("metraj")
                tutar = request.POST.get("tutar")
                birim_bilgisi = request.POST.get("birim_bilgisi")
                kata_veya_binaya_daihil = request.POST.get("kata_veya_binaya_daihil")
                id = bloglar.objects.filter(id__in = projetipi).first()
                kalem = santiye_kalemleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    proje_ait_bilgisi = request.user,
                    proje_santiye_Ait =id.proje_santiye_Ait,
                    kalem_adi = yetkili_adi,santiye_agirligi = santiye_agirligi,
                    santiye_finansal_agirligi = finansal_agirlik,
                    birimi = get_object_or_404(birimler,id =birim_bilgisi ),metraj = metraj,
                    tutari = tutar
                )
                if kata_veya_binaya_daihil == "0":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,int(i.kat_sayisi)):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = request.user,
                                proje_santiye_Ait = id.proje_santiye_Ait,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =kalem.id ),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
                elif kata_veya_binaya_daihil == "1":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,int(i.kat_sayisi)):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = request.user,
                                proje_santiye_Ait = id.proje_santiye_Ait,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =kalem.id ),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
                            break
                elif kata_veya_binaya_daihil == "2":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,4):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = request.user,
                                proje_santiye_Ait = id.proje_santiye_Ait,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =kalem.id ),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
                        
    return redirect_with_language("main:santiye_projesi_ekle_")

def kalem_sil(request):
    if request.POST:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.kalemleri_silme:
                    buttonId = request.POST.get("buttonId")
                    geri_don = request.POST.get("geri_don")
                    santiye_kalemleri.objects.filter(id = buttonId).update(
                        silinme_bilgisi = True
                    )
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            buttonId = request.POST.get("buttonId")
            geri_don = request.POST.get("geri_don")
            santiye_kalemleri.objects.filter(id = buttonId).update(
                silinme_bilgisi = True
            )
    return redirect_with_language("main:santiye_projesi_ekle_")
def santiye_kalemleri_duzenle(request):
    if request.POST:
        buttonId = request.POST.get("buttonId")
        santiye_kalemlerin_dagilisi.objects.filter(kalem_bilgisi__id =buttonId).delete()
        yetkili_adi = request.POST.get("yetkili_adi")
        santiye_agirligi = request.POST.get("katsayisi")
        finansal_agirlik = request.POST.get("blogsayisi")
        geri_don = request.POST.get("geri_don")
        metraj = request.POST.get("metraj")
        tutar = request.POST.get("tutar")
        birim_bilgisi = request.POST.get("birim_bilgisi")
        kata_veya_binaya_daihil = request.POST.get("kata_veya_binaya_daihil")
        projetipi = request.POST.getlist("projetipi")
        if request.user.is_superuser:
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "3":
                santiye_kalemleri.objects.filter(id  = buttonId).update(
                        kalem_adi = yetkili_adi,santiye_agirligi = santiye_agirligi,
                        santiye_finansal_agirligi = finansal_agirlik,
                        birimi = get_object_or_404(birimler,id =birim_bilgisi ),metraj = metraj,
                tutari = tutar)
            elif silinmedurumu == "2":
                santiye_kalemleri.objects.filter(id  = buttonId).update(
                        kalem_adi = yetkili_adi,santiye_agirligi = santiye_agirligi,
                        santiye_finansal_agirligi = finansal_agirlik,
                        silinme_bilgisi = True,birimi = get_object_or_404(birimler,id =birim_bilgisi ),metraj = metraj,
                tutari = tutar)

            elif silinmedurumu == "1":
                santiye_kalemleri.objects.filter(id  = buttonId).update(
                        kalem_adi = yetkili_adi,santiye_agirligi = santiye_agirligi,
                        santiye_finansal_agirligi = finansal_agirlik,
                        silinme_bilgisi = False,birimi = get_object_or_404(birimler,id =birim_bilgisi ),metraj = metraj,
                tutari = tutar
                    )
        else:
            if request.user.kullanicilar_db:
                    a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                    if a:
                        if a.izinler.kalemleri_duzenleme:
                            santiye_kalemleri.objects.filter(id  = buttonId).update(
                            proje_ait_bilgisi = request.user.kullanicilar_db,
                            kalem_adi = yetkili_adi,santiye_agirligi = santiye_agirligi,
                            santiye_finansal_agirligi = finansal_agirlik,birimi = get_object_or_404(birimler,id =birim_bilgisi ),metraj = metraj,
                            tutari = tutar
                            )
                            if kata_veya_binaya_daihil == "0":
                                blog_lar = bloglar.objects.filter(id__in = projetipi)
                                for i in blog_lar:
                                    for j in range(0,int(i.kat_sayisi)):
                                        santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                            proje_ait_bilgisi = request.user.kullanicilar_db,
                                            proje_santiye_Ait_id = geri_don,
                                            kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =buttonId ),
                                            kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                                        )
                            elif kata_veya_binaya_daihil == "1":
                                blog_lar = bloglar.objects.filter(id__in = projetipi)
                                for i in blog_lar:
                                    for j in range(0,int(i.kat_sayisi)):
                                        santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                            proje_ait_bilgisi = request.user.kullanicilar_db,
                                            proje_santiye_Ait_id = geri_don,
                                            kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =buttonId ),
                                            kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                                        )
                                        break
                            elif kata_veya_binaya_daihil == "2":
                                blog_lar = bloglar.objects.filter(id__in = projetipi)
                                for i in blog_lar:
                                    for j in range(0,4):
                                        santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                            proje_ait_bilgisi = request.user.kullanicilar_db,
                                            proje_santiye_Ait_id = geri_don,
                                            kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =buttonId ),
                                            kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                                        )
                        else:
                            return redirect_with_language("main:yetkisiz")
                    else:
                        return redirect_with_language("main:yetkisiz")
            else:
                santiye_kalemleri.objects.filter(id  = buttonId).update(
                            proje_ait_bilgisi = request.user,
                            kalem_adi = yetkili_adi,santiye_agirligi = santiye_agirligi,
                            santiye_finansal_agirligi = finansal_agirlik,birimi = get_object_or_404(birimler,id =birim_bilgisi ),metraj = metraj,
                        tutari = tutar
                        )
                if kata_veya_binaya_daihil == "0":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,int(i.kat_sayisi)):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = request.user,
                                proje_santiye_Ait_id = geri_don,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =buttonId ),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
                elif kata_veya_binaya_daihil == "1":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,int(i.kat_sayisi)):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = request.user,
                                proje_santiye_Ait_id = geri_don,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =buttonId),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
                            break
                elif kata_veya_binaya_daihil == "2":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,4):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = request.user,
                                proje_santiye_Ait_id = geri_don,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =buttonId ),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
        return redirect_with_language("main:santiye_projesi_ekle_")

def kalem_blog_dagilis_sil(request,id,ik):
    a = santiye_kalemlerin_dagilisi.objects.filter(blog_bilgisi__id = id).first()
    a = a.proje_santiye_Ait.id
    santiye_kalemlerin_dagilisi.objects.filter(kalem_bilgisi__id= ik,blog_bilgisi__id = id).delete()
    return redirect_with_language("main:santtiye_kalemleri",a)

def santiye_kalem_ve_blog(request):
    content = sozluk_yapisi()
    content["proje_tipleri"] = proje_tipi.objects.filter(proje_ait_bilgisi =  request.user)
    if super_admin_kontrolu(request):
        profile =bloglar.objects.all()
        kullanicilar = CustomUser.objects.filter(proje_santiye_Ait__silinme_bilgisi = False ,kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.ilerleme_takibi_gorme:
                        profile = bloglar.objects.filter(proje_santiye_Ait__silinme_bilgisi = False ,proje_ait_bilgisi = request.user.kullanicilar_db)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
        else:
            profile = bloglar.objects.filter(proje_santiye_Ait__silinme_bilgisi = False ,proje_ait_bilgisi = request.user)
    
    content["santiyeler"] = profile
    return render(request,"santiye_yonetimi/santiye_blog_kalem.html",content)

def santiye_kalem_ve_blog_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    
    content["proje_tipleri"] = proje_tipi.objects.filter(proje_ait_bilgisi =  request.user)
    if super_admin_kontrolu(request):
        profile = bloglar.objects.filter(proje_santiye_Ait__silinme_bilgisi = False ,proje_ait_bilgisi = users)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = bloglar.objects.filter(proje_ait_bilgisi = request.user)
    
    content["santiyeler"] = profile

    return render(request,"santiye_yonetimi/santiye_blog_kalem.html",content)

def blogtan_kaleme_ilerleme_takibi(request,id,slug):
    content = sozluk_yapisi()
    content["id"] = get_object_or_404(bloglar,id = id)
    content["blog_id"] = id
    if request.user.is_authenticated:
        if request.user.is_superuser:
            content["santiyeler_bilgileri"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = get_object_or_404(bloglar,id = id).proje_ait_bilgisi)
            kalemler = santiye_kalemlerin_dagilisi.objects.filter(blog_bilgisi__id = id)
            kalem_id = []
            for i in kalemler:
                if i.kalem_bilgisi.id in kalem_id:
                    pass
                else:
                    kalem_id.append(i.kalem_bilgisi.id)

            profile =  santiye_kalemleri.objects.filter(id__in = kalem_id,silinme_bilgisi = False)
            page_num = request.GET.get('page', 1)
            paginator = Paginator(profile, 10) # 6 employees per page
            try:
                page_obj = paginator.page(page_num)
            except PageNotAnInteger:
                # if page is not an integer, deliver the first page
                page_obj = paginator.page(1)
            except EmptyPage:
                # if the page is out of range, deliver the last page
                page_obj = paginator.page(paginator.num_pages)
            content["santiyeler"] = profile
            content["top"]  = profile
            content["medya"] = page_obj

        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kalemleri_gorme:
                        content["santiyeler_bilgileri"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db)
                        kalemler = santiye_kalemlerin_dagilisi.objects.filter(blog_bilgisi__id = id)
                        kalem_id = []
                        for i in kalemler:
                            if i.kalem_bilgisi.id in kalem_id:
                                pass
                            else:
                                kalem_id.append(i.kalem_bilgisi.id)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                content["santiyeler_bilgileri"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
                kalemler = santiye_kalemlerin_dagilisi.objects.filter(blog_bilgisi__id = id)
                kalem_id = []
                for i in kalemler:
                    if i.kalem_bilgisi.id in kalem_id:
                        pass
                    else:
                        kalem_id.append(i.kalem_bilgisi.id)

            profile =  santiye_kalemleri.objects.filter(id__in = kalem_id,silinme_bilgisi = False)
            page_num = request.GET.get('page', 1)
            paginator = Paginator(profile, 10) # 6 employees per page
            try:
                page_obj = paginator.page(page_num)
            except PageNotAnInteger:
                # if page is not an integer, deliver the first page
                page_obj = paginator.page(1)
            except EmptyPage:
                # if the page is out of range, deliver the last page
                page_obj = paginator.page(paginator.num_pages)
            content["santiyeler"] = profile
            content["top"]  = profile
            content["medya"] = page_obj
    else:
        return redirect_with_language("/users/login/")
    return render(request,"santiye_yonetimi/ilerleme_takibi.html",content)
def blogtan_kaleme_ilerleme_takibi_hash(request,id,slug,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    content["id"] = get_object_or_404(bloglar,id = id)
    content["blog_id"] = id
    if request.user.is_authenticated:
        if request.user.is_superuser:
            content["santiyeler_bilgileri"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = get_object_or_404(bloglar,id = id).proje_ait_bilgisi)
            kalemler = santiye_kalemlerin_dagilisi.objects.filter(blog_bilgisi__id = id)
            kalem_id = []
            for i in kalemler:
                if i.kalem_bilgisi.id in kalem_id:
                    pass
                else:
                    kalem_id.append(i.kalem_bilgisi.id)

            profile =  santiye_kalemleri.objects.filter(id__in = kalem_id,silinme_bilgisi = False)
            page_num = request.GET.get('page', 1)
            paginator = Paginator(profile, 10) # 6 employees per page
            try:
                page_obj = paginator.page(page_num)
            except PageNotAnInteger:
                # if page is not an integer, deliver the first page
                page_obj = paginator.page(1)
            except EmptyPage:
                # if the page is out of range, deliver the last page
                page_obj = paginator.page(paginator.num_pages)
            content["santiyeler"] = profile
            content["top"]  = profile
            content["medya"] = page_obj

        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.kalemleri_gorme:
                        content["santiyeler_bilgileri"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db)
                        kalemler = santiye_kalemlerin_dagilisi.objects.filter(blog_bilgisi__id = id)
                        kalem_id = []
                        for i in kalemler:
                            if i.kalem_bilgisi.id in kalem_id:
                                pass
                            else:
                                kalem_id.append(i.kalem_bilgisi.id)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                content["santiyeler_bilgileri"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
                kalemler = santiye_kalemlerin_dagilisi.objects.filter(blog_bilgisi__id = id)
                kalem_id = []
                for i in kalemler:
                    if i.kalem_bilgisi.id in kalem_id:
                        pass
                    else:
                        kalem_id.append(i.kalem_bilgisi.id)

            profile =  santiye_kalemleri.objects.filter(id__in = kalem_id,silinme_bilgisi = False)
            page_num = request.GET.get('page', 1)
            paginator = Paginator(profile, 10) # 6 employees per page
            try:
                page_obj = paginator.page(page_num)
            except PageNotAnInteger:
                # if page is not an integer, deliver the first page
                page_obj = paginator.page(1)
            except EmptyPage:
                # if the page is out of range, deliver the last page
                page_obj = paginator.page(paginator.num_pages)
            content["santiyeler"] = profile
            content["top"]  = profile
            content["medya"] = page_obj
    else:
        return redirect_with_language("/users/login/")
    return render(request,"santiye_yonetimi/ilerleme_takibi.html",content)

def ilerleme_kaydet(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.ilerleme_takibi_duzenleme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        geri_don = request.POST.get("geri_don")
        veri_cek = request.POST.get("veri_cek")
        kalem = request.POST.getlist("kalem")
        tumbilgi = request.POST.getlist("tumbilgi")
        a = []
        for i in tumbilgi:
            k = i.split(",")
            for j in k :
                a.append(j)
        
        for i in kalem:
            a.remove(i)
            santiye_kalemlerin_dagilisi.objects.filter(id = int(i),tamamlanma_bilgisi = False).update(tamamlanma_bilgisi = True, degistirme_tarihi=timezone.now() )
        for i in a:
            if i != ""  :
                if i in kalem:
                    pass
                else:
                    santiye_kalemlerin_dagilisi.objects.filter(id = int(i),tamamlanma_bilgisi = True).update(tamamlanma_bilgisi = False, degistirme_tarihi=timezone.now() )
    return redirect_with_language("main:blogtan_kaleme_ilerleme_takibi",geri_don,veri_cek)

def ilerleme_kaydet_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.ilerleme_takibi_duzenleme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        geri_don = request.POST.get("geri_don")
        veri_cek = request.POST.get("veri_cek")
        kalem = request.POST.getlist("kalem")
        tumbilgi = request.POST.getlist("tumbilgi")
        a = []
        for i in tumbilgi:
            k = i.split(",")
            for j in k :
                a.append(j)
        
        for i in kalem:
            a.remove(i)
            santiye_kalemlerin_dagilisi.objects.filter(id = int(i),tamamlanma_bilgisi = False).update(tamamlanma_bilgisi = True, degistirme_tarihi=timezone.now() )
        for i in a:
            if i != ""  :
                if i in kalem:
                    pass
                else:
                    santiye_kalemlerin_dagilisi.objects.filter(id = int(i),tamamlanma_bilgisi = True).update(tamamlanma_bilgisi = False, degistirme_tarihi=timezone.now() )
    return redirect_with_language("main:blogtan_kaleme_ilerleme_takibi_hash",geri_don,veri_cek,hash)


def santiye_kalem_ekle_admin(request,id):
    content = sozluk_yapisi()
    content["santiyeler_bilgileri"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = id)
    if request.POST:
        projetipi = request.POST.get("projetipi")
        yetkili_adi = request.POST.get("yetkili_adi")
        santiye_agirligi = request.POST.get("katsayisi")
        finansal_agirlik = request.POST.get("blogsayisi")
        metraj = request.POST.get("metraj")
        tutar = request.POST.get("tutar")
        birim_bilgisi = request.POST.get("birim_bilgisi")
        
        kalem = santiye_kalemleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                proje_ait_bilgisi = get_object_or_404(CustomUser,id = id),
                proje_santiye_Ait = get_object_or_404(santiye,id =projetipi ),
                kalem_adi = yetkili_adi,santiye_agirligi = santiye_agirligi,
                santiye_finansal_agirligi = finansal_agirlik,
                birimi = get_object_or_404(birimler,id =birim_bilgisi ),metraj = metraj,
                tutari = tutar
            )
        blog_lar = bloglar.objects.filter(proje_santiye_Ait = get_object_or_404(santiye,id =projetipi ))
        kat_sayisi = int(get_object_or_404(santiye,id =projetipi ).kat_sayisi)
        for i in blog_lar:
            for j in range(0,kat_sayisi):
                santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    proje_ait_bilgisi = get_object_or_404(CustomUser,id = id) ,
                    proje_santiye_Ait = get_object_or_404(santiye,id =projetipi ),
                    kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =kalem.id ),
                    kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                    )
        return redirect_with_language("main:santtiye_kalemleri",get_object_or_404(santiye,id =projetipi ).id)
    return render(request,"santiye_yonetimi/santiyeyekalem_ekle_admin.html",content)
#şantiye Kalemleri
#Proje Bilgisi
def projeler_sayfasi(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile =projeler.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = projeler.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)

    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =projeler.objects.filter(Q(proje_ait_bilgisi__last_name__icontains = search)|Q(Proje_tipi_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = projeler.objects.filter(Q(proje_ait_bilgisi = request.user) & Q(silinme_bilgisi = False)).filter(Q(aciklama__icontains = search)| Q(proje_Adi__icontains = search))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["blog_bilgisi"]  =bloglar.objects.filter(proje_ait_bilgisi = request.user,proje_santiye_Ait__silinme_bilgisi = False)
    return render(request,"santiye_yonetimi/proje_sayfai.html",content)

#proje Ekleme
def proje_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            return redirect_with_language("main:proje_ekle_admin",id = kullanici)
        else:
            yetkili_adi = request.POST.get("yetkili_adi")
            tarih_bilgisi = request.POST.get("tarih_bilgisi")
            aciklama = request.POST.get("aciklama")
            durumu  = request.POST.get("durumu")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            blogbilgisi = request.POST.getlist("blogbilgisi")
            new_project = projeler(
                proje_ait_bilgisi = request.user,
                proje_Adi = yetkili_adi,
                tarih = tarih_bilgisi,
                aciklama = aciklama,
                durum = durumu,silinme_bilgisi = False
            )
            new_project.save()
            bloglar_bilgisi = []
            for i in blogbilgisi:
                bloglar_bilgisi.append(bloglar.objects.get(id=int(i)))
            new_project.blog_bilgisi.add(*bloglar_bilgisi)
            images = request.FILES.getlist('file')
            for images in images:
                proje_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),dosya=images,proje_ait_bilgisi = get_object_or_404(projeler,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet

    return redirect_with_language("main:projeler_sayfasi")

def proje_ekle_admin(request,id):
    content = sozluk_yapisi()
    content["blog_bilgisi"]  =bloglar.objects.filter(proje_ait_bilgisi = get_object_or_404(CustomUser,id = id),proje_santiye_Ait__silinme_bilgisi = False)
    if request.POST:
        yetkili_adi = request.POST.get("yetkili_adi")
        tarih_bilgisi = request.POST.get("tarih_bilgisi")
        aciklama = request.POST.get("aciklama")
        durumu  = request.POST.get("durumu")
        if durumu == "1":
            durumu = True
        else:
            durumu = False
        blogbilgisi = request.POST.getlist("blogbilgisi")
        new_project = projeler(
                proje_ait_bilgisi = get_object_or_404(CustomUser,id = id ),
                proje_Adi = yetkili_adi,
                tarih = tarih_bilgisi,
                aciklama = aciklama,
                durum = durumu,silinme_bilgisi = False
            )
        new_project.save()
        bloglar_bilgisi = []
        for i in blogbilgisi:
            bloglar_bilgisi.append(bloglar.objects.get(id=int(i)))
        new_project.blog_bilgisi.add(*bloglar_bilgisi)

        images = request.FILES.getlist('file')
        for images in images:
            proje_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),dosya=images,proje_ait_bilgisi = get_object_or_404(projeler,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
        return redirect_with_language("main:projeler_sayfasi")

    return render(request,"santiye_yonetimi/proje_ekle_admin.html",content)
#Proje Ekeleme
#proje silme
def proje_silme(request):
    if request.POST:
        buttonId = request.POST.get("buttonId")
        projeler.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:projeler_sayfasi")

#proje silme
#proje düzenleme
def proje_duzenle_bilgi(request):
    c = request.POST
    if c:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            yetkili_adi = request.POST.get("yetkili_adi")
            tarih_bilgisi = request.POST.get("tarih_bilgisi")
            aciklama = request.POST.get("aciklama")
            durumu  = request.POST.get("durumu")
            buttonIdInput = request.POST.get("buttonId")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            blogbilgisi = request.POST.getlist("blogbilgisi")
            projeler.objects.filter(id = buttonIdInput).update(
                proje_ait_bilgisi = get_object_or_404(CustomUser,id =kullanici ),
                proje_Adi = yetkili_adi,
                tarih = tarih_bilgisi,
                aciklama = aciklama,
                durum = durumu,silinme_bilgisi = False
            )
            z = get_object_or_404(projeler,id = buttonIdInput)
            bloglar_bilgisi = []
            for i in blogbilgisi:
                bloglar_bilgisi.append(bloglar.objects.get(id=int(i)))
            z.blog_bilgisi.add(*bloglar_bilgisi)
            images = request.FILES.getlist('file')
            for images in images:
                proje_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),dosya=images,proje_ait_bilgisi = get_object_or_404(projeler,id = buttonIdInput))  # Urun_resimleri modeline resimleri kaydet
        else:
            yetkili_adi = request.POST.get("yetkili_adi")
            tarih_bilgisi = request.POST.get("tarih_bilgisi")
            aciklama = request.POST.get("aciklama")
            durumu  = request.POST.get("durumu")
            buttonIdInput = request.POST.get("buttonId")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            blogbilgisi = request.POST.getlist("blogbilgisi")
            projeler.objects.filter(id = buttonIdInput).update(
                proje_Adi = yetkili_adi,
                tarih = tarih_bilgisi,
                aciklama = aciklama,
                durum = durumu,silinme_bilgisi = False
            )
            z = get_object_or_404(projeler,id = buttonIdInput)
            bloglar_bilgisi = []
            for i in blogbilgisi:
                bloglar_bilgisi.append(bloglar.objects.get(id=int(i)))
            z.blog_bilgisi.add(*bloglar_bilgisi)
            images = request.FILES.getlist('file')
            for images in images:
                proje_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),dosya=images,proje_ait_bilgisi = get_object_or_404(projeler,id = buttonIdInput))  # Urun_resimleri modeline resimleri kaydet
    return redirect_with_language("main:projeler_sayfasi")
#proje düzenleme

#taseron olaylari
def taseron_sayfasi(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile =taseronlar.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.taseronlar_gorme:
                    profile = taseronlar.objects.filter(silinme_bilgisi = False,taseron_ait_bilgisi = request.user.kullanicilar_db)
                    if a.izinler.santiye_gorme:
                        content["blog_bilgisi"] =santiye.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = taseronlar.objects.filter(silinme_bilgisi = False,taseron_ait_bilgisi = request.user)
            content["blog_bilgisi"]  =santiye.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =taseronlar.objects.filter(Q(taseron_ait_bilgisi__last_name__icontains = search)|Q(taseron_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.taseronlar_gorme:
                        profile = taseronlar.objects.filter(Q(taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = taseronlar.objects.filter(Q(taseron_ait_bilgisi = request.user) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))

            
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    
    return render(request,"santiye_yonetimi/taseronlar.html",content)
#taseron olaylari
#taseron olaylari
def taseron_sayfasi_2(request,hash):
    content = sozluk_yapisi()

    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =taseronlar.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = taseronlar.objects.filter(silinme_bilgisi = False,taseron_ait_bilgisi = users)
        content["blog_bilgisi"]  =santiye.objects.filter(proje_ait_bilgisi = users,silinme_bilgisi = False)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.taseronlar_gorme:
                    profile = taseronlar.objects.filter(silinme_bilgisi = False,taseron_ait_bilgisi = request.user.kullanicilar_db)
                    if a.izinler.santiye_gorme:
                        content["blog_bilgisi"] =santiye.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = taseronlar.objects.filter(silinme_bilgisi = False,taseron_ait_bilgisi = request.user)
            content["blog_bilgisi"]  =santiye.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =taseronlar.objects.filter(Q(taseron_ait_bilgisi__last_name__icontains = search)|Q(taseron_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.taseronlar_gorme:
                        profile = taseronlar.objects.filter(Q(taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = taseronlar.objects.filter(Q(taseron_ait_bilgisi = request.user) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))

            
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    
    return render(request,"santiye_yonetimi/taseronlar.html",content)
#taseron olaylari

def taseron_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            if True:
                taseron_adi = request.POST.get("taseron_adi")
                telefonnumarasi = request.POST.get("telefonnumarasi")
                email_adresi = request.POST.get("email_adresi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                new_project = taseronlar(
                    taseron_ait_bilgisi = users,
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
                new_project.save()
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    liste = str(i).split(",")
                    proje = projeler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = users,
                            blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                            kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])
                            )
                    bloglar_bilgisi.append(projeler.objects.get(id=int(proje.id)))
                new_project.proje_bilgisi.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                isim = 1
                for images in images:
                    taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
                car = cari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    cari_kart_ait_bilgisi = users,
                    cari_adi = taseron_adi,
                    telefon_numarasi = telefonnumarasi,
                    aciklama = aciklama
                )
                cari_taseron_baglantisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    gelir_kime_ait_oldugu = get_object_or_404(taseronlar,id = new_project.id ),
                    cari_bilgisi = get_object_or_404(cari,id = car.id)
                )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.taseronlar_olusturma:
                        taseron_adi = request.POST.get("taseron_adi")
                        telefonnumarasi = request.POST.get("telefonnumarasi")
                        email_adresi = request.POST.get("email_adresi")
                        blogbilgisi = request.POST.getlist("blogbilgisi")
                        aciklama = request.POST.get("aciklama")
                        new_project = taseronlar(
                            taseron_ait_bilgisi = request.user.kullanicilar_db,
                            taseron_adi = taseron_adi,
                            email = email_adresi,
                            aciklama = aciklama,
                            telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                        )
                        new_project.save()
                        bloglar_bilgisi = []
                        for i in blogbilgisi:
                            liste = str(i).split(",")
                            proje = projeler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user.kullanicilar_db,
                                    blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                                    kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])
                                    )
                            
                            bloglar_bilgisi.append(projeler.objects.get(id=int(proje.id)))
                        new_project.proje_bilgisi.add(*bloglar_bilgisi)
                        images = request.FILES.getlist('file')
                        isim = 1
                        for images in images:
                            taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                        car = cari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            cari_kart_ait_bilgisi = request.user.kullanicilar_db,
                            cari_adi = taseron_adi,
                            telefon_numarasi = telefonnumarasi,
                            aciklama = aciklama
                        )
                        cari_taseron_baglantisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            gelir_kime_ait_oldugu = get_object_or_404(taseronlar,id = new_project.id ),
                            cari_bilgisi = get_object_or_404(cari,id = car.id)
                        )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                taseron_adi = request.POST.get("taseron_adi")
                telefonnumarasi = request.POST.get("telefonnumarasi")
                email_adresi = request.POST.get("email_adresi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                new_project = taseronlar(
                    taseron_ait_bilgisi = request.user,
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
                new_project.save()
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    liste = str(i).split(",")
                    proje = projeler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user,
                            blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                            kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])
                            )
                    bloglar_bilgisi.append(projeler.objects.get(id=int(proje.id)))
                new_project.proje_bilgisi.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                isim = 1
                for images in images:
                    taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
                car = cari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    cari_kart_ait_bilgisi = request.user,
                    cari_adi = taseron_adi,
                    telefon_numarasi = telefonnumarasi,
                    aciklama = aciklama
                )
                cari_taseron_baglantisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    gelir_kime_ait_oldugu = get_object_or_404(taseronlar,id = new_project.id ),
                    cari_bilgisi = get_object_or_404(cari,id = car.id)
                )
    return redirect_with_language("main:taseron_sayfasi_2",hash)

def taseron_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            return redirect_with_language("main:taseron_ekle_admin",kullanici)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.taseronlar_olusturma:
                        taseron_adi = request.POST.get("taseron_adi")
                        telefonnumarasi = request.POST.get("telefonnumarasi")
                        email_adresi = request.POST.get("email_adresi")
                        blogbilgisi = request.POST.getlist("blogbilgisi")
                        aciklama = request.POST.get("aciklama")
                        new_project = taseronlar(
                            taseron_ait_bilgisi = request.user.kullanicilar_db,
                            taseron_adi = taseron_adi,
                            email = email_adresi,
                            aciklama = aciklama,
                            telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                        )
                        new_project.save()
                        bloglar_bilgisi = []
                        for i in blogbilgisi:
                            liste = str(i).split(",")
                            proje = projeler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user.kullanicilar_db,
                                    blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                                    kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])
                                    )
                            
                            bloglar_bilgisi.append(projeler.objects.get(id=int(proje.id)))
                        new_project.proje_bilgisi.add(*bloglar_bilgisi)
                        images = request.FILES.getlist('file')
                        isim = 1
                        for images in images:
                            taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                        car = cari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            cari_kart_ait_bilgisi = request.user.kullanicilar_db,
                            cari_adi = taseron_adi,
                            telefon_numarasi = telefonnumarasi,
                            aciklama = aciklama
                        )
                        cari_taseron_baglantisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            gelir_kime_ait_oldugu = get_object_or_404(taseronlar,id = new_project.id ),
                            cari_bilgisi = get_object_or_404(cari,id = car.id)
                        )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                taseron_adi = request.POST.get("taseron_adi")
                telefonnumarasi = request.POST.get("telefonnumarasi")
                email_adresi = request.POST.get("email_adresi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                new_project = taseronlar(
                    taseron_ait_bilgisi = request.user,
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
                new_project.save()
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    liste = str(i).split(",")
                    proje = projeler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user,
                            blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                            kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])
                            )
                    bloglar_bilgisi.append(projeler.objects.get(id=int(proje.id)))
                new_project.proje_bilgisi.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                isim = 1
                for images in images:
                    taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
                car = cari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    cari_kart_ait_bilgisi = request.user,
                    cari_adi = taseron_adi,
                    telefon_numarasi = telefonnumarasi,
                    aciklama = aciklama
                )
                cari_taseron_baglantisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    gelir_kime_ait_oldugu = get_object_or_404(taseronlar,id = new_project.id ),
                    cari_bilgisi = get_object_or_404(cari,id = car.id)
                )
    return redirect_with_language("main:taseron_sayfasi")

def taseron_ekle_admin(request,id):
    content = sozluk_yapisi()
    content["blog_bilgisi"] =santiye.objects.filter(proje_ait_bilgisi__id = id,silinme_bilgisi = False)
    if request.POST:
        if True:
            taseron_adi = request.POST.get("taseron_adi")
            telefonnumarasi = request.POST.get("telefonnumarasi")
            email_adresi = request.POST.get("email_adresi")
            blogbilgisi = request.POST.getlist("blogbilgisi")
            aciklama = request.POST.get("aciklama")
            new_project = taseronlar(
                taseron_ait_bilgisi = get_object_or_404(CustomUser,id = id),
                taseron_adi = taseron_adi,
                email = email_adresi,
                aciklama = aciklama,
                telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
            )
            new_project.save()
            bloglar_bilgisi = []
            for i in blogbilgisi:
                liste = str(i).split(",")
                proje = projeler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi__id = id,
                        blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                        kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])
                        )
                bloglar_bilgisi.append(projeler.objects.get(id=int(proje.id)))
            new_project.proje_bilgisi.add(*bloglar_bilgisi)
            images = request.FILES.getlist('file')
            isim = 1
            for images in images:
                taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                isim = isim+1
            car = cari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                cari_kart_ait_bilgisi = get_object_or_404(CustomUser,id = id),
                cari_adi = taseron_adi,
                telefon_numarasi = telefonnumarasi,
                aciklama = aciklama
            )
            cari_taseron_baglantisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                gelir_kime_ait_oldugu = get_object_or_404(taseronlar,id = new_project.id ),
                cari_bilgisi = get_object_or_404(cari,id = car.id)
            )
        return redirect_with_language("main:taseron_sayfasi")
    return render(request,"santiye_yonetimi/admin_taseron_ekle.html",content)
#proje silme
def taseron_silme_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.taseronlar_silme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        buttonId = request.POST.get("buttonId")
        taseronlar.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:taseron_sayfasi_2",hash)
def taseron_silme(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.taseronlar_silme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        buttonId = request.POST.get("buttonId")
        taseronlar.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:taseron_sayfasi")
#taşeron Düzenleme
def taseron_duzelt_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            id_bilgisi = request.POST.get("id_bilgisi")
            taseron_adi = request.POST.get("taseron_adi")
            telefonnumarasi = request.POST.get("telefonnumarasi")
            email_adresi = request.POST.get("email_adresi")
            blogbilgisi = request.POST.getlist("blogbilgisi")
            aciklama = request.POST.get("aciklama")
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "1":
                taseronlar.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
            elif silinmedurumu == "2":
                taseronlar.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = True
                )
            else:
                taseronlar.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi
                )
            bloglar_bilgisi = []
            for i in blogbilgisi:
                bloglar_bilgisi.append(projeler.objects.get(id=int(i)))
            get_object_or_404(taseronlar,id =id_bilgisi).proje_bilgisi.add(*bloglar_bilgisi)
            images = request.FILES.getlist('file')
            isim = 1
            for images in images:
                taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                isim = isim+1
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.taseronlar_duzenleme:
                        id_bilgisi = request.POST.get("id_bilgisi")
                        taseron_adi = request.POST.get("taseron_adi")
                        telefonnumarasi = request.POST.get("telefonnumarasi")
                        email_adresi = request.POST.get("email_adresi")
                        blogbilgisi = request.POST.getlist("blogbilgisi")
                        aciklama = request.POST.get("aciklama")
                        new_project = taseronlar.objects.filter(id =id_bilgisi ).update(
                            taseron_adi = taseron_adi,
                            email = email_adresi,
                            aciklama = aciklama,
                            telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                        )
                        bloglar_bilgisi = []
                        for i in blogbilgisi:
                            liste = str(i).split(",")
                            j  = get_object_or_none(projeler,blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                                    kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0]),proje_ait_bilgisi = request.user.kullanicilar_db)
                            if j:
                                #print(j,"geldi")
                                bloglar_bilgisi.append(projeler.objects.get(id=int(j.id)))
                            else:
                                proje = projeler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user.kullanicilar_db,
                                        blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                                        kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])
                                        )
                                bloglar_bilgisi.append(projeler.objects.get(id=int(proje.id)))
                        get_object_or_404(taseronlar,id =id_bilgisi).proje_bilgisi.clear()
                        get_object_or_404(taseronlar,id =id_bilgisi).proje_bilgisi.add(*bloglar_bilgisi)
                        images = request.FILES.getlist('file')
                        isim = 1
                        for images in images:
                            taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
        
                id_bilgisi = request.POST.get("id_bilgisi")
                taseron_adi = request.POST.get("taseron_adi")
                telefonnumarasi = request.POST.get("telefonnumarasi")
                email_adresi = request.POST.get("email_adresi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                new_project = taseronlar.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    liste = str(i).split(",")
                    j  = projeler.objects.filter(blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                        kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])).last()
                    if j:
                        #print(j,"geldi")       
                        bloglar_bilgisi.append(projeler.objects.get(id=int(j.id)))
                    else:
                        proje = projeler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user,
                        blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                        kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])
                        )
                        bloglar_bilgisi.append(projeler.objects.get(id=int(proje.id)))
                get_object_or_404(taseronlar,id =id_bilgisi).proje_bilgisi.clear()
                get_object_or_404(taseronlar,id =id_bilgisi).proje_bilgisi.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                isim = 1
                for images in images:
                    taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
    return redirect_with_language("main:taseron_sayfasi_2",hash)

def taseron_duzelt(request):
    if request.POST:
        if request.user.is_superuser:
            id_bilgisi = request.POST.get("id_bilgisi")
            taseron_adi = request.POST.get("taseron_adi")
            telefonnumarasi = request.POST.get("telefonnumarasi")
            email_adresi = request.POST.get("email_adresi")
            blogbilgisi = request.POST.getlist("blogbilgisi")
            aciklama = request.POST.get("aciklama")
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "1":
                taseronlar.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
            elif silinmedurumu == "2":
                taseronlar.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = True
                )
            else:
                taseronlar.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi
                )
            bloglar_bilgisi = []
            for i in blogbilgisi:
                bloglar_bilgisi.append(projeler.objects.get(id=int(i)))
            get_object_or_404(taseronlar,id =id_bilgisi).proje_bilgisi.add(*bloglar_bilgisi)
            images = request.FILES.getlist('file')
            isim = 1
            for images in images:
                taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                isim = isim+1
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.taseronlar_duzenleme:
                        id_bilgisi = request.POST.get("id_bilgisi")
                        taseron_adi = request.POST.get("taseron_adi")
                        telefonnumarasi = request.POST.get("telefonnumarasi")
                        email_adresi = request.POST.get("email_adresi")
                        blogbilgisi = request.POST.getlist("blogbilgisi")
                        aciklama = request.POST.get("aciklama")
                        new_project = taseronlar.objects.filter(id =id_bilgisi ).update(
                            taseron_adi = taseron_adi,
                            email = email_adresi,
                            aciklama = aciklama,
                            telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                        )
                        bloglar_bilgisi = []
                        for i in blogbilgisi:
                            liste = str(i).split(",")
                            j  = get_object_or_none(projeler,blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                                    kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0]),proje_ait_bilgisi = request.user.kullanicilar_db)
                            if j:
                                #print(j,"geldi")
                                bloglar_bilgisi.append(projeler.objects.get(id=int(j.id)))
                            else:
                                proje = projeler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user.kullanicilar_db,
                                        blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                                        kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])
                                        )
                                bloglar_bilgisi.append(projeler.objects.get(id=int(proje.id)))
                        get_object_or_404(taseronlar,id =id_bilgisi).proje_bilgisi.clear()
                        get_object_or_404(taseronlar,id =id_bilgisi).proje_bilgisi.add(*bloglar_bilgisi)
                        images = request.FILES.getlist('file')
                        isim = 1
                        for images in images:
                            taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
        
                id_bilgisi = request.POST.get("id_bilgisi")
                taseron_adi = request.POST.get("taseron_adi")
                telefonnumarasi = request.POST.get("telefonnumarasi")
                email_adresi = request.POST.get("email_adresi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                new_project = taseronlar.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    liste = str(i).split(",")
                    j  = projeler.objects.filter(blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                        kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])).last()
                    if j:
                        #print(j,"geldi")       
                        bloglar_bilgisi.append(projeler.objects.get(id=int(j.id)))
                    else:
                        proje = projeler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = request.user,
                        blog_bilgisi = get_object_or_none(bloglar,id = liste[1]),
                        kalem_bilgisi = get_object_or_none(santiye_kalemleri,id = liste[0])
                        )
                        bloglar_bilgisi.append(projeler.objects.get(id=int(proje.id)))
                get_object_or_404(taseronlar,id =id_bilgisi).proje_bilgisi.clear()
                get_object_or_404(taseronlar,id =id_bilgisi).proje_bilgisi.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                isim = 1
                for images in images:
                    taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(taseronlar,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
    return redirect_with_language("main:taseron_sayfasi")

#proje silme
def ust_yuklenici_sayfasi_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users

    if super_admin_kontrolu(request):
        profile =ust_yuklenici.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = ust_yuklenici.objects.filter(silinme_bilgisi = False,taseron_ait_bilgisi = users)
        content["blog_bilgisi"]  =santiye.objects.filter(proje_ait_bilgisi = users,silinme_bilgisi = False)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.ust_yuklenici_gorme:
                    profile = ust_yuklenici.objects.filter(silinme_bilgisi = False,taseron_ait_bilgisi = request.user.kullanicilar_db)
                    
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = ust_yuklenici.objects.filter(silinme_bilgisi = False,taseron_ait_bilgisi = request.user)
            content["blog_bilgisi"]  =santiye.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =ust_yuklenici.objects.filter(Q(taseron_ait_bilgisi__last_name__icontains = search)|Q(taseron_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.ust_yuklenici_gorme:
                        profile = ust_yuklenici.objects.filter(Q(taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = ust_yuklenici.objects.filter(Q(taseron_ait_bilgisi = request.user) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))

            
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    
    return render(request,"santiye_yonetimi/ust_yuklenici.html",content)
#Üst Yüklenici olaylari
def ust_yuklenici_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            if True:
                taseron_adi = request.POST.get("taseron_adi")
                telefonnumarasi = request.POST.get("telefonnumarasi")
                email_adresi = request.POST.get("email_adresi")
                
                aciklama = request.POST.get("aciklama")
                new_project = ust_yuklenici(
                    taseron_ait_bilgisi = users,
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
                new_project.save()
                
                images = request.FILES.getlist('file')
                #print(images)
                isim = 1
                for images in images:
                    ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.taseronlar_olusturma:
                        taseron_adi = request.POST.get("taseron_adi")
                        telefonnumarasi = request.POST.get("telefonnumarasi")
                        email_adresi = request.POST.get("email_adresi")
                        aciklama = request.POST.get("aciklama")
                        new_project = ust_yuklenici(
                            taseron_ait_bilgisi = request.user.kullanicilar_db,
                            taseron_adi = taseron_adi,
                            email = email_adresi,
                            aciklama = aciklama,
                            telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                        )
                        new_project.save()
                        images = request.FILES.getlist('file')
                        #print(images)
                        isim = 1
                        for images in images:
                            ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                taseron_adi = request.POST.get("taseron_adi")
                telefonnumarasi = request.POST.get("telefonnumarasi")
                email_adresi = request.POST.get("email_adresi")
                
                aciklama = request.POST.get("aciklama")
                new_project = ust_yuklenici(
                    taseron_ait_bilgisi = request.user,
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
                new_project.save()
                
                images = request.FILES.getlist('file')
                #print(images)
                isim = 1
                for images in images:
                    ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1

    return redirect_with_language("main:ust_yuklenici_sayfasi_2",hash)
#proje silme
def ust_yuklenici_silme_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.ust_yuklenici_silme:
                    pass
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            pass
        #print("Üst Yüklenici Sil")
        buttonId = request.POST.get("buttonId")
        ust_yuklenici.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:ust_yuklenici_sayfasi_2",hash)
#Üst Yüklenivci Düzenleme
def ust_yuklenici_duzelt_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            id_bilgisi = request.POST.get("id_bilgisi")
            taseron_adi = request.POST.get("taseron_adi")
            telefonnumarasi = request.POST.get("telefonnumarasi")
            email_adresi = request.POST.get("email_adresi")
            aciklama = request.POST.get("aciklama")
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "1":
                ust_yuklenici.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
            elif silinmedurumu == "2":
                ust_yuklenici.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = True
                )
            else:
                ust_yuklenici.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi
                )
        
            images = request.FILES.getlist('file')
            isim = 1
            for images in images:
                ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                isim = isim+1
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.taseronlar_duzenleme:
                        id_bilgisi = request.POST.get("id_bilgisi")
                        taseron_adi = request.POST.get("taseron_adi")
                        telefonnumarasi = request.POST.get("telefonnumarasi")
                        email_adresi = request.POST.get("email_adresi")
                        
                        aciklama = request.POST.get("aciklama")
                        new_project = ust_yuklenici.objects.filter(id =id_bilgisi ).update(
                            taseron_adi = taseron_adi,
                            email = email_adresi,
                            aciklama = aciklama,
                            telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                        )
                        images = request.FILES.getlist('file')
                        isim = 1
                        for images in images:
                            ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
        
                id_bilgisi = request.POST.get("id_bilgisi")
                taseron_adi = request.POST.get("taseron_adi")
                telefonnumarasi = request.POST.get("telefonnumarasi")
                email_adresi = request.POST.get("email_adresi")
                
                aciklama = request.POST.get("aciklama")
                new_project = ust_yuklenici.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
                
                images = request.FILES.getlist('file')
                isim = 1
                for images in images:
                    ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
    return redirect_with_language("main:ust_yuklenici_sayfasi_2",hash)

#Üst Yükleneci olaylari
def ust_yuklenici_sayfasi(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile =ust_yuklenici.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.ust_yuklenici_gorme:
                    profile = ust_yuklenici.objects.filter(silinme_bilgisi = False,taseron_ait_bilgisi = request.user.kullanicilar_db)
                    
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = ust_yuklenici.objects.filter(silinme_bilgisi = False,taseron_ait_bilgisi = request.user)
            content["blog_bilgisi"]  =santiye.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =ust_yuklenici.objects.filter(Q(taseron_ait_bilgisi__last_name__icontains = search)|Q(taseron_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.ust_yuklenici_gorme:
                        profile = ust_yuklenici.objects.filter(Q(taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = ust_yuklenici.objects.filter(Q(taseron_ait_bilgisi = request.user) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))

            
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    
    return render(request,"santiye_yonetimi/ust_yuklenici.html",content)
#Üst Yüklenici olaylari
def ust_yuklenici_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            return redirect_with_language("main:taseron_ekle_admin",kullanici)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.taseronlar_olusturma:
                        taseron_adi = request.POST.get("taseron_adi")
                        telefonnumarasi = request.POST.get("telefonnumarasi")
                        email_adresi = request.POST.get("email_adresi")
                        aciklama = request.POST.get("aciklama")
                        new_project = ust_yuklenici(
                            taseron_ait_bilgisi = request.user.kullanicilar_db,
                            taseron_adi = taseron_adi,
                            email = email_adresi,
                            aciklama = aciklama,
                            telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                        )
                        new_project.save()
                        images = request.FILES.getlist('file')
                        #print(images)
                        isim = 1
                        for images in images:
                            ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                taseron_adi = request.POST.get("taseron_adi")
                telefonnumarasi = request.POST.get("telefonnumarasi")
                email_adresi = request.POST.get("email_adresi")
                
                aciklama = request.POST.get("aciklama")
                new_project = ust_yuklenici(
                    taseron_ait_bilgisi = request.user,
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
                new_project.save()
                
                images = request.FILES.getlist('file')
                #print(images)
                isim = 1
                for images in images:
                    ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = new_project.id))  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1

    return redirect_with_language("main:ust_yuklenici_sayfasi")
#proje silme
def ust_yuklenici_silme(request):
    
    if request.POST:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.ust_yuklenici_silme:
                    pass
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            pass
        #print("Üst Yüklenici Sil")
        buttonId = request.POST.get("buttonId")
        ust_yuklenici.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:ust_yuklenici_sayfasi")
#Üst Yüklenivci Düzenleme
def ust_yuklenici_duzelt(request):
    if request.POST:
        if request.user.is_superuser:
            id_bilgisi = request.POST.get("id_bilgisi")
            taseron_adi = request.POST.get("taseron_adi")
            telefonnumarasi = request.POST.get("telefonnumarasi")
            email_adresi = request.POST.get("email_adresi")
            aciklama = request.POST.get("aciklama")
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "1":
                ust_yuklenici.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
            elif silinmedurumu == "2":
                ust_yuklenici.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = True
                )
            else:
                ust_yuklenici.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi
                )
        
            images = request.FILES.getlist('file')
            isim = 1
            for images in images:
                ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                isim = isim+1
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.taseronlar_duzenleme:
                        id_bilgisi = request.POST.get("id_bilgisi")
                        taseron_adi = request.POST.get("taseron_adi")
                        telefonnumarasi = request.POST.get("telefonnumarasi")
                        email_adresi = request.POST.get("email_adresi")
                        
                        aciklama = request.POST.get("aciklama")
                        new_project = ust_yuklenici.objects.filter(id =id_bilgisi ).update(
                            taseron_adi = taseron_adi,
                            email = email_adresi,
                            aciklama = aciklama,
                            telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                        )
                        images = request.FILES.getlist('file')
                        isim = 1
                        for images in images:
                            ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
        
                id_bilgisi = request.POST.get("id_bilgisi")
                taseron_adi = request.POST.get("taseron_adi")
                telefonnumarasi = request.POST.get("telefonnumarasi")
                email_adresi = request.POST.get("email_adresi")
                
                aciklama = request.POST.get("aciklama")
                new_project = ust_yuklenici.objects.filter(id =id_bilgisi ).update(
                    taseron_adi = taseron_adi,
                    email = email_adresi,
                    aciklama = aciklama,
                    telefon_numarasi = telefonnumarasi,silinme_bilgisi = False
                )
                
                images = request.FILES.getlist('file')
                isim = 1
                for images in images:
                    ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),aciklama="",dosya_adi = isim,dosya=images,proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = id_bilgisi))  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
    return redirect_with_language("main:ust_yuklenici_sayfasi")


#sözleşmeler
#sözleşme olaylari
def sozlesmler_sayfasi_2(request,hash):
    content = sozluk_yapisi()
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile = taseron_sozlesme_dosyalari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = taseron_sozlesme_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = users)
        #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
        content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi=users,silinme_bilgisi = False)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.sozlesmeler_gorme:
                    profile = taseron_sozlesme_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db)
                    #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,silinme_bilgisi = False)
                    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = taseron_sozlesme_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user)
            #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
            content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi__proje_ait_bilgisi__last_name__icontains = search)|Q(taseron_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.sozlesmeler_gorme:
                        profile = taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                
           
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    
    return render(request,"santiye_yonetimi/xxsozlesmler.html",content)
#sözleşme olaylari
#sözleşme olaylari
def ana_yuklenici_sozlesmler_sayfasi_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users

    if super_admin_kontrolu(request):
        profile = taseron_sozlesme_dosyalari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = ust_yuklenici_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = users)
        #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
        content["taseronlar"] = ust_yuklenici.objects.filter(taseron_ait_bilgisi= users,silinme_bilgisi = False)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.ust_yuklenici_gorme:
                    profile = ust_yuklenici_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db)
                    #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,silinme_bilgisi = False)
                    content["taseronlar"] = ust_yuklenici.objects.filter(taseron_ait_bilgisi= request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = ust_yuklenici_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user)
            #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
            content["taseronlar"] = ust_yuklenici.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =ust_yuklenici_dosyalari.objects.filter(Q(taseron_ait_bilgisi__proje_ait_bilgisi__last_name__icontains = search)|Q(taseron_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.sozlesmeler_gorme:
                        profile = ust_yuklenici_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = ust_yuklenici_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                
           
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    
    return render(request,"santiye_yonetimi/ana_yuklenici_sozlesme.html",content)
#sözleşme olaylari
def ust_yuklenici_sozlesme_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("dosyaadi")
            tarih = request.POST.get("tarih")
            aciklama = request.POST.get("aciklama")
            durumu = request.POST.get("durumu")
            file = request.FILES.get("file")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                dosya = file,dosya_adi = dosyaadi,
                tarih = tarih,aciklama = aciklama,
                durum = durumu
            )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.ust_yuklenici_olusturma:
                        pass
                    else:
                        return redirect_with_language("main:yetkisiz")
            else:
                pass
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("dosyaadi")
            tarih = request.POST.get("tarih")
            aciklama = request.POST.get("aciklama")
            durumu = request.POST.get("durumu")
            file = request.FILES.get("file")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                dosya = file,dosya_adi = dosyaadi,
                tarih = tarih,aciklama = aciklama,
                durum = durumu
            )
    return redirect_with_language("main:ana_yuklenici_sozlesmler_sayfasi_2",hash)

def ust_yuklenici_sozlesme_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.ust_yuklenici_duzenleme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        id_bilgisi = request.POST.get("id_bilgisi")
        taseron = request.POST.get("taseron")
        dosyaadi = request.POST.get("dosyaadi")
        tarih = request.POST.get("tarih")
        aciklama = request.POST.get("aciklama")
        durumu = request.POST.get("durumu")
        file = request.FILES.get("file")
        if durumu == "1":
            durumu = True
        else:
            durumu = False
        if request.user.is_superuser:
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "3":
                ust_yuklenici_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu
                )
            elif silinmedurumu == "2":
                ust_yuklenici_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu,silinme_bilgisi = True
                )
            elif silinmedurumu == "1":
                ust_yuklenici_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu,silinme_bilgisi = False
                )
        else:

            ust_yuklenici_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu
                )
    return redirect_with_language("main:ust_yuklenici_sayfasi_2",hash)
#sözleşmeler sil
def ust_yuklenici_silme_sozlesme_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.ust_yuklenici_silme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        buttonId = request.POST.get("buttonId")
        ust_yuklenici_dosyalari.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:ust_yuklenici_sayfasi_2",hash)

#sözleşmeler sil
#sözleşme olaylari
def sozlesme_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("dosyaadi")
            tarih = request.POST.get("tarih")
            aciklama = request.POST.get("aciklama")
            durumu = request.POST.get("durumu")
            file = request.FILES.get("file")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                dosya = file,dosya_adi = dosyaadi,
                tarih = tarih,aciklama = aciklama,
                durum = durumu
            )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.sozlesmeler_olusturma:
                        pass
                    else:
                        return redirect_with_language("main:yetkisiz")
            else:
                pass
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("dosyaadi")
            tarih = request.POST.get("tarih")
            aciklama = request.POST.get("aciklama")
            durumu = request.POST.get("durumu")
            file = request.FILES.get("file")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                dosya = file,dosya_adi = dosyaadi,
                tarih = tarih,aciklama = aciklama,
                durum = durumu
            )
    return redirect_with_language("main:sozlesmler_sayfasi_2",hash)

def sozlesme_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.sozlesmeler_duzenleme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        id_bilgisi = request.POST.get("id_bilgisi")
        taseron = request.POST.get("taseron")
        dosyaadi = request.POST.get("dosyaadi")
        tarih = request.POST.get("tarih")
        aciklama = request.POST.get("aciklama")
        durumu = request.POST.get("durumu")
        file = request.FILES.get("file")
        if durumu == "1":
            durumu = True
        else:
            durumu = False
        if request.user.is_superuser:
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "3":
                taseron_sozlesme_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu
                )
            elif silinmedurumu == "2":
                taseron_sozlesme_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu,silinme_bilgisi = True
                )
            elif silinmedurumu == "1":
                taseron_sozlesme_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu,silinme_bilgisi = False
                )
        else:

            taseron_sozlesme_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu
                )
    return redirect_with_language("main:sozlesmler_sayfasi_2",hash)

#sözleşmeler sil
def sozlesme_silme_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.sozlesmeler_silme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        buttonId = request.POST.get("buttonId")
        taseron_sozlesme_dosyalari.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:sozlesmler_sayfasi_2",hash)

#sözleşmeler
#sözleşme olaylari
def sozlesmler_sayfasi(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile = taseron_sozlesme_dosyalari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.sozlesmeler_gorme:
                    profile = taseron_sozlesme_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db)
                    #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,silinme_bilgisi = False)
                    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = taseron_sozlesme_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user)
            #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
            content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi__proje_ait_bilgisi__last_name__icontains = search)|Q(taseron_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.sozlesmeler_gorme:
                        profile = taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                
           
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    
    return render(request,"santiye_yonetimi/xxsozlesmler.html",content)
#sözleşme olaylari
#sözleşme olaylari
def ana_yuklenici_sozlesmler_sayfasi(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile = taseron_sozlesme_dosyalari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.ust_yuklenici_gorme:
                    profile = ust_yuklenici_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db)
                    #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,silinme_bilgisi = False)
                    content["taseronlar"] = ust_yuklenici.objects.filter(taseron_ait_bilgisi= request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = ust_yuklenici_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user)
            #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
            content["taseronlar"] = ust_yuklenici.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =ust_yuklenici_dosyalari.objects.filter(Q(taseron_ait_bilgisi__proje_ait_bilgisi__last_name__icontains = search)|Q(taseron_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.sozlesmeler_gorme:
                        profile = ust_yuklenici_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = ust_yuklenici_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                
           
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    
    return render(request,"santiye_yonetimi/ana_yuklenici_sozlesme.html",content)
#sözleşme olaylari
def ust_yuklenici_sozlesme_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            return redirect_with_language("main:sozlesme_ekle_admin",kullanici)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.ust_yuklenici_olusturma:
                        pass
                    else:
                        return redirect_with_language("main:yetkisiz")
            else:
                pass
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("dosyaadi")
            tarih = request.POST.get("tarih")
            aciklama = request.POST.get("aciklama")
            durumu = request.POST.get("durumu")
            file = request.FILES.get("file")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            ust_yuklenici_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                dosya = file,dosya_adi = dosyaadi,
                tarih = tarih,aciklama = aciklama,
                durum = durumu
            )
    return redirect_with_language("main:ana_yuklenici_sozlesmler_sayfasi")

def ust_yuklenici_sozlesme_duzenle(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.ust_yuklenici_duzenleme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        id_bilgisi = request.POST.get("id_bilgisi")
        taseron = request.POST.get("taseron")
        dosyaadi = request.POST.get("dosyaadi")
        tarih = request.POST.get("tarih")
        aciklama = request.POST.get("aciklama")
        durumu = request.POST.get("durumu")
        file = request.FILES.get("file")
        if durumu == "1":
            durumu = True
        else:
            durumu = False
        if request.user.is_superuser:
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "3":
                ust_yuklenici_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu
                )
            elif silinmedurumu == "2":
                ust_yuklenici_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu,silinme_bilgisi = True
                )
            elif silinmedurumu == "1":
                ust_yuklenici_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu,silinme_bilgisi = False
                )
        else:

            ust_yuklenici_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(ust_yuklenici,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu
                )
    return redirect_with_language("main:ust_yuklenici_sayfasi")
#sözleşmeler sil
def ust_yuklenici_silme_sozlesme(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.ust_yuklenici_silme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        buttonId = request.POST.get("buttonId")
        ust_yuklenici_dosyalari.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:ust_yuklenici_sayfasi")

#sözleşmeler sil
#sözleşme olaylari
def sozlesme_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            return redirect_with_language("main:sozlesme_ekle_admin",kullanici)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.sozlesmeler_olusturma:
                        pass
                    else:
                        return redirect_with_language("main:yetkisiz")
            else:
                pass
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("dosyaadi")
            tarih = request.POST.get("tarih")
            aciklama = request.POST.get("aciklama")
            durumu = request.POST.get("durumu")
            file = request.FILES.get("file")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                dosya = file,dosya_adi = dosyaadi,
                tarih = tarih,aciklama = aciklama,
                durum = durumu
            )
    return redirect_with_language("main:sozlesmler_sayfasi")

def sozlesme_ekle_admin(request,id):
    content = sozluk_yapisi()
    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi__id= id,silinme_bilgisi = False)
    if request.POST:
        if True:
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("dosyaadi")
            tarih = request.POST.get("tarih")
            aciklama = request.POST.get("aciklama")
            durumu = request.POST.get("durumu")
            file = request.FILES.get("file")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            taseron_sozlesme_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                dosya = file,dosya_adi = dosyaadi,
                tarih = tarih,aciklama = aciklama,
                durum = durumu
            )
        return redirect_with_language("main:sozlesmler_sayfasi")
    return render(request,"santiye_yonetimi/admin_sozlesme_ekle.html",content)
#sözleşme düzenleme

def sozlesme_duzenle(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.sozlesmeler_duzenleme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        id_bilgisi = request.POST.get("id_bilgisi")
        taseron = request.POST.get("taseron")
        dosyaadi = request.POST.get("dosyaadi")
        tarih = request.POST.get("tarih")
        aciklama = request.POST.get("aciklama")
        durumu = request.POST.get("durumu")
        file = request.FILES.get("file")
        if durumu == "1":
            durumu = True
        else:
            durumu = False
        if request.user.is_superuser:
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "3":
                taseron_sozlesme_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu
                )
            elif silinmedurumu == "2":
                taseron_sozlesme_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu,silinme_bilgisi = True
                )
            elif silinmedurumu == "1":
                taseron_sozlesme_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu,silinme_bilgisi = False
                )
        else:

            taseron_sozlesme_dosyalari.objects.filter(id = id_bilgisi).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    durum = durumu
                )
    return redirect_with_language("main:sozlesmler_sayfasi")

#sözleşmeler sil
def sozlesme_silme(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.sozlesmeler_silme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        buttonId = request.POST.get("buttonId")
        taseron_sozlesme_dosyalari.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:sozlesmler_sayfasi")

#sözleşmeler sil
#sözleşmeler
#hakedişler

def hakedis_sayfasi_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users

    if super_admin_kontrolu(request):
        profile = taseron_hakedisles.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = taseron_hakedisles.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = users)
        content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi=users,silinme_bilgisi = False)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.hakedisler_gorme:
                    profile = taseron_hakedisles.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db)
                    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = taseron_hakedisles.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user)
            content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi__last_name__icontains = search)|Q(dosya_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.hakedisler_gorme:
                        profile = taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(proje_ait_bilgisi__taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi = request.user) & Q(proje_ait_bilgisi__taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    
    return render(request,"santiye_yonetimi/hakedis.html",content)

def hakedis_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            if True:
                taseron = request.POST.get("taseron")
                dosyaadi = request.POST.get("yetkili_adi")
                tarih = request.POST.get("tarih_bilgisi")
                aciklama = request.POST.get("aciklama")
                file = request.FILES.get("file")
                fatura_no = request.POST.get("fatura_no")
                
                taseron_hakedisles.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,
                    dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    fatura_numarasi = fatura_no
                )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.hakedisler_olusturma:
                        taseron = request.POST.get("taseron")
                        dosyaadi = request.POST.get("yetkili_adi")
                        tarih = request.POST.get("tarih_bilgisi")
                        aciklama = request.POST.get("aciklama")
                        file = request.FILES.get("file")
                        fatura_no = request.POST.get("fatura_no")
                       
                        taseron_hakedisles.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                            dosya = file,
                            dosya_adi = dosyaadi,
                            tarih = tarih,aciklama = aciklama,
                            fatura_numarasi = fatura_no
                        )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                taseron = request.POST.get("taseron")
                dosyaadi = request.POST.get("yetkili_adi")
                tarih = request.POST.get("tarih_bilgisi")
                aciklama = request.POST.get("aciklama")
                file = request.FILES.get("file")
                fatura_no = request.POST.get("fatura_no")
                
                taseron_hakedisles.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,
                    dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    fatura_numarasi = fatura_no
                )
    return redirect_with_language("main:hakedis_sayfasi_2",hash)
#hakedisekle
def hakedis_silme_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.hakedisler_silme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        buttonId = request.POST.get("buttonId")
        taseron_hakedisles.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:hakedis_sayfasi_2",hash)

#hakediş düzenle
def hakedis_duzenle_2(request,hash):
    if request.POST:
        if request.user.is_superuser:
            buttonId = request.POST.get("buttonId")
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("yetkili_adi")
            tarih = request.POST.get("tarih_bilgisi")
            aciklama = request.POST.get("aciklama")
            file = request.FILES.get("file")
            fatura_no = request.POST.get("fatura_no")
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "3":
                if file:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                        proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                        dosya = file,
                        dosya_adi = dosyaadi,
                        tarih = tarih,aciklama = aciklama,
                        fatura_numarasi = fatura_no
                    )
                else:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                        proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                        
                        dosya_adi = dosyaadi,
                        tarih = tarih,aciklama = aciklama,
                        fatura_numarasi = fatura_no
                    )
            elif silinmedurumu == "2":
                if file:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                        proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                        dosya = file,
                        dosya_adi = dosyaadi,
                        tarih = tarih,aciklama = aciklama,
                        
                        fatura_numarasi = fatura_no,silinme_bilgisi = True
                    )
                else:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                        proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                        
                        dosya_adi = dosyaadi,
                        tarih = tarih,aciklama = aciklama,
                        
                        fatura_numarasi = fatura_no,silinme_bilgisi = True
                    )
            elif silinmedurumu == "1":
                if file:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                        proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                        dosya = file,
                        dosya_adi = dosyaadi,
                        tarih = tarih,aciklama = aciklama,
                        fatura_numarasi = fatura_no,silinme_bilgisi = False
                    )
                else:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    
                    dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    fatura_numarasi = fatura_no,silinme_bilgisi = False
                )

        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.hakedisler_duzenleme:
                        pass
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")    
            else:
                pass
            buttonId = request.POST.get("buttonId")
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("yetkili_adi")
            tarih = request.POST.get("tarih_bilgisi")
            aciklama = request.POST.get("aciklama")
            file = request.FILES.get("file")
            fatura_no = request.POST.get("fatura_no")
            if file:
                taseron_hakedisles.objects.filter(id=buttonId).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,
                    dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                
                    fatura_numarasi = fatura_no
                )
            else:
                taseron_hakedisles.objects.filter(id=buttonId).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),

                    dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                
                    fatura_numarasi = fatura_no
                )
    return redirect_with_language("main:hakedis_sayfasi_2",hash)


def hakedis_sayfasi(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile = taseron_hakedisles.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.hakedisler_gorme:
                    profile = taseron_hakedisles.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db)
                    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = taseron_hakedisles.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user)
            content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi__last_name__icontains = search)|Q(dosya_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.hakedisler_gorme:
                        profile = taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(proje_ait_bilgisi__taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi = request.user) & Q(proje_ait_bilgisi__taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    
    return render(request,"santiye_yonetimi/hakedis.html",content)

def hakedis_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            return redirect_with_language("main:hakedis_ekle_admin",kullanici)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.hakedisler_olusturma:
                        taseron = request.POST.get("taseron")
                        dosyaadi = request.POST.get("yetkili_adi")
                        tarih = request.POST.get("tarih_bilgisi")
                        aciklama = request.POST.get("aciklama")
                        file = request.FILES.get("file")
                        fatura_no = request.POST.get("fatura_no")
                       
                        taseron_hakedisles.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                            dosya = file,
                            dosya_adi = dosyaadi,
                            tarih = tarih,aciklama = aciklama,
                            fatura_numarasi = fatura_no
                        )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                taseron = request.POST.get("taseron")
                dosyaadi = request.POST.get("yetkili_adi")
                tarih = request.POST.get("tarih_bilgisi")
                aciklama = request.POST.get("aciklama")
                file = request.FILES.get("file")
                fatura_no = request.POST.get("fatura_no")
                
                taseron_hakedisles.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,
                    dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    fatura_numarasi = fatura_no
                )
    return redirect_with_language("main:hakedis_sayfasi")
#hakedisekle
def hakedis_silme(request):
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.hakedisler_silme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        buttonId = request.POST.get("buttonId")
        taseron_hakedisles.objects.filter(id = buttonId).update(silinme_bilgisi = True)
    return redirect_with_language("main:hakedis_sayfasi")

def hakedis_ekle_admin(request,id):
    content = sozluk_yapisi()
    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi__id= id,silinme_bilgisi = False)
    if request.POST:
        if request.user.is_superuser:
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("yetkili_adi")
            tarih = request.POST.get("tarih_bilgisi")
            aciklama = request.POST.get("aciklama")
            durumu = request.POST.get("durumu")
            file = request.FILES.get("file")
            tutar = request.POST.get("tutar")
            fatura_no = request.POST.get("fatura_no")
            if durumu == "1":
                durumu = True
            else:
                durumu = False
            taseron_hakedisles.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                dosya = file,
                dosya_adi = dosyaadi,
                tarih = tarih,aciklama = aciklama,
                durum = durumu,
                tutar = tutar,
                fatura_numarasi = fatura_no
            )
        return redirect_with_language("main:hakedis_sayfasi")
    return render(request,"santiye_yonetimi/hakedis_admin_ekle.html",content)

#hakediş düzenle
def hakedis_duzenle(request):
    if request.POST:
        if request.user.is_superuser:
            buttonId = request.POST.get("buttonId")
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("yetkili_adi")
            tarih = request.POST.get("tarih_bilgisi")
            aciklama = request.POST.get("aciklama")
            file = request.FILES.get("file")
            fatura_no = request.POST.get("fatura_no")
            silinmedurumu = request.POST.get("silinmedurumu")
            if silinmedurumu == "3":
                if file:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                        proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                        dosya = file,
                        dosya_adi = dosyaadi,
                        tarih = tarih,aciklama = aciklama,
                        fatura_numarasi = fatura_no
                    )
                else:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                        proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                        
                        dosya_adi = dosyaadi,
                        tarih = tarih,aciklama = aciklama,
                        fatura_numarasi = fatura_no
                    )
            elif silinmedurumu == "2":
                if file:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                        proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                        dosya = file,
                        dosya_adi = dosyaadi,
                        tarih = tarih,aciklama = aciklama,
                        
                        fatura_numarasi = fatura_no,silinme_bilgisi = True
                    )
                else:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                        proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                        
                        dosya_adi = dosyaadi,
                        tarih = tarih,aciklama = aciklama,
                        
                        fatura_numarasi = fatura_no,silinme_bilgisi = True
                    )
            elif silinmedurumu == "1":
                if file:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                        proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                        dosya = file,
                        dosya_adi = dosyaadi,
                        tarih = tarih,aciklama = aciklama,
                        fatura_numarasi = fatura_no,silinme_bilgisi = False
                    )
                else:
                    taseron_hakedisles.objects.filter(id=buttonId).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    
                    dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                    fatura_numarasi = fatura_no,silinme_bilgisi = False
                )

        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.hakedisler_duzenleme:
                        pass
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")    
            else:
                pass
            buttonId = request.POST.get("buttonId")
            taseron = request.POST.get("taseron")
            dosyaadi = request.POST.get("yetkili_adi")
            tarih = request.POST.get("tarih_bilgisi")
            aciklama = request.POST.get("aciklama")
            file = request.FILES.get("file")
            fatura_no = request.POST.get("fatura_no")
            if file:
                taseron_hakedisles.objects.filter(id=buttonId).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),
                    dosya = file,
                    dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                
                    fatura_numarasi = fatura_no
                )
            else:
                taseron_hakedisles.objects.filter(id=buttonId).update(
                    proje_ait_bilgisi = get_object_or_404(taseronlar,id = taseron),

                    dosya_adi = dosyaadi,
                    tarih = tarih,aciklama = aciklama,
                
                    fatura_numarasi = fatura_no
                )
    return redirect_with_language("main:hakedis_sayfasi")


def depolama_sistemim(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile = klasorler.objects.all()

    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasorler.objects.filter(klasor_adi_db = None).filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasorler.objects.filter(klasor_adi_db = None).filter(silinme_bilgisi = False,dosya_sahibi = request.user)

    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(klasor_adi_db = None).filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasorler.objects.filter(klasor_adi_db = None).filter(Q(dosya_sahibi = request.user.kullanicilar_db) & Q(klasor_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasorler.objects.filter(klasor_adi_db = None).filter(Q(dosya_sahibi = request.user) & Q(klasor_adi__icontains = search)& Q(silinme_bilgisi = False))

            
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/depolama_sistemim.html",content)


def depolama_sistemim_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        #profile = klasorler.objects.all()
        profile = klasorler.objects.filter(klasor_adi_db = None).filter(silinme_bilgisi = False,dosya_sahibi = users)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasorler.objects.filter(klasor_adi_db = None).filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasorler.objects.filter(klasor_adi_db = None).filter(silinme_bilgisi = False,dosya_sahibi = request.user)

    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(klasor_adi_db = None).filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasorler.objects.filter(klasor_adi_db = None).filter(Q(dosya_sahibi = request.user.kullanicilar_db) & Q(klasor_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasorler.objects.filter(klasor_adi_db = None).filter(Q(dosya_sahibi = request.user) & Q(klasor_adi__icontains = search)& Q(silinme_bilgisi = False))

            
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/depolama_sistemim.html",content)

def klasor_olustur_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            if True:
                ust_klasor = request.POST.get("ust_klasor")
                if ust_klasor:
                    klasor = request.POST.get("klasor")

                    klasorler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                        dosya_sahibi = users,
                        klasor_adi = klasor,
                        klasor_adi_db = get_object_or_404(klasorler,id =ust_klasor )
                    )
                else:
                    klasor = request.POST.get("klasor")

                    klasorler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                        dosya_sahibi = users,
                        klasor_adi = klasor
                    )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_olusturma:
                        ust_klasor = request.POST.get("ust_klasor")
                        if ust_klasor:
                            klasor = request.POST.get("klasor")

                            klasorler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                dosya_sahibi = request.user.kullanicilar_db,
                                klasor_adi = klasor,
                                klasor_adi_db = get_object_or_404(klasorler,id =ust_klasor )
                            )
                        else:
                            klasor = request.POST.get("klasor")

                            klasorler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                dosya_sahibi = request.user.kullanicilar_db,
                                klasor_adi = klasor
                            )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                ust_klasor = request.POST.get("ust_klasor")
                if ust_klasor:
                    klasor = request.POST.get("klasor")

                    klasorler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                        dosya_sahibi = request.user,
                        klasor_adi = klasor,
                        klasor_adi_db = get_object_or_404(klasorler,id =ust_klasor )
                    )
                else:
                    klasor = request.POST.get("klasor")

                    klasorler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                        dosya_sahibi = request.user,
                        klasor_adi = klasor
                    )
    return redirect_with_language("main:depolama_sistemim_2",hash)

def klasor__yeniden_adlandir_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            klasor = request.POST.get("klasor")
            idbilgisi = request.POST.get("idbilgisi")
            klasorler.objects.filter(id = idbilgisi).update(
                    dosya_sahibi = users,
                    klasor_adi = klasor
                )
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_duzenleme:
                        klasor = request.POST.get("klasor")
                        idbilgisi = request.POST.get("idbilgisi")
                        klasorler.objects.filter(id = idbilgisi).update(
                            dosya_sahibi = request.user.kullanicilar_db,
                            klasor_adi = klasor
                        )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                klasor = request.POST.get("klasor")
                idbilgisi = request.POST.get("idbilgisi")
                klasorler.objects.filter(id = idbilgisi).update(
                    dosya_sahibi = request.user,
                    klasor_adi = klasor
                )
    return redirect_with_language("main:depolama_sistemim_2",hash)

def klasor_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_silme:
                    pass
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        if request.user.is_superuser:
            idbilgisi = request.POST.get("idbilgisi")
            klasorler.objects.filter(id = idbilgisi).update(
                silinme_bilgisi = True
            )
        else:
            pass
            
    return redirect_with_language("main:depolama_sistemim_2",hash)


def klasor_olustur(request):
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_olusturma:
                        ust_klasor = request.POST.get("ust_klasor")
                        if ust_klasor:
                            klasor = request.POST.get("klasor")

                            klasorler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                dosya_sahibi = request.user.kullanicilar_db,
                                klasor_adi = klasor,
                                klasor_adi_db = get_object_or_404(klasorler,id =ust_klasor )
                            )
                        else:
                            klasor = request.POST.get("klasor")

                            klasorler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                dosya_sahibi = request.user.kullanicilar_db,
                                klasor_adi = klasor
                            )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                ust_klasor = request.POST.get("ust_klasor")
                if ust_klasor:
                    klasor = request.POST.get("klasor")

                    klasorler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                        dosya_sahibi = request.user,
                        klasor_adi = klasor,
                        klasor_adi_db = get_object_or_404(klasorler,id =ust_klasor )
                    )
                else:
                    klasor = request.POST.get("klasor")

                    klasorler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                        dosya_sahibi = request.user,
                        klasor_adi = klasor
                    )
    return redirect_with_language("main:depolama_sistemim")

def klasor__yeniden_adlandir(request):
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_duzenleme:
                        klasor = request.POST.get("klasor")
                        idbilgisi = request.POST.get("idbilgisi")
                        klasorler.objects.filter(id = idbilgisi).update(
                            dosya_sahibi = request.user.kullanicilar_db,
                            klasor_adi = klasor
                        )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                klasor = request.POST.get("klasor")
                idbilgisi = request.POST.get("idbilgisi")
                klasorler.objects.filter(id = idbilgisi).update(
                    dosya_sahibi = request.user,
                    klasor_adi = klasor
                )
    return redirect_with_language("main:depolama_sistemim")

def klasor_sil(request):
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_silme:
                    pass
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            
            idbilgisi = request.POST.get("idbilgisi")
            klasorler.objects.filter(id = idbilgisi).update(
                silinme_bilgisi = True
            )
    return redirect_with_language("main:depolama_sistemim")

#klasöre Gir
def klasore_gir(request,id,slug):
    content = sozluk_yapisi()
    content["id_bilgisi"] = id
    if super_admin_kontrolu(request):
        profile = klasorler.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasorler.objects.filter(klasor_adi_db__id =id,silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db)
                    dosyalarim = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db,proje_ait_bilgisi__id =id)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasorler.objects.filter(klasor_adi_db__id =id,silinme_bilgisi = False,dosya_sahibi = request.user)
            dosyalarim = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user,proje_ait_bilgisi__id =id)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasorler.objects.filter(klasor_adi_db__id =id).filter(Q(dosya_sahibi = request.user.kullanicilar_db) & Q(klasor_adi__icontains = search)& Q(silinme_bilgisi = False))
                        dosyalarim = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db,proje_ait_bilgisi__id =id).filter(__icontains = search)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasorler.objects.filter(klasor_adi_db__id =id).filter(Q(dosya_sahibi = request.user) & Q(klasor_adi__icontains = search)& Q(silinme_bilgisi = False))
                dosyalarim = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user,proje_ait_bilgisi__id =id).filter(__icontains = search)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["dosyalarim"] = dosyalarim
    return render(request,"santiye_yonetimi/klasorici.html",content)
#klasöre Gir
#klasöre Gir
def klasore_gir_2(request,id,slug,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    content["id_bilgisi"] = id
    if super_admin_kontrolu(request):
        profile = klasorler.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = klasorler.objects.filter(klasor_adi_db__id =id,silinme_bilgisi = False,dosya_sahibi = users)
        dosyalarim = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = users,proje_ait_bilgisi__id =id)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasorler.objects.filter(klasor_adi_db__id =id,silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db)
                    dosyalarim = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db,proje_ait_bilgisi__id =id)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasorler.objects.filter(klasor_adi_db__id =id,silinme_bilgisi = False,dosya_sahibi = request.user)
            dosyalarim = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user,proje_ait_bilgisi__id =id)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasorler.objects.filter(klasor_adi_db__id =id).filter(Q(dosya_sahibi = request.user.kullanicilar_db) & Q(klasor_adi__icontains = search)& Q(silinme_bilgisi = False))
                        dosyalarim = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db,proje_ait_bilgisi__id =id).filter(__icontains = search)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasorler.objects.filter(klasor_adi_db__id =id).filter(Q(dosya_sahibi = request.user) & Q(klasor_adi__icontains = search)& Q(silinme_bilgisi = False))
                dosyalarim = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user,proje_ait_bilgisi__id =id).filter(__icontains = search)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["dosyalarim"] = dosyalarim
    return render(request,"santiye_yonetimi/klasorici.html",content)
#klasöre Gir

#klasore Dosya Ekle
def dosya_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_olusturma:
                        ust_klasor = request.POST.get("ust_klasor")
                        dosya_Adi = request.POST.get("klasor")
                        tarih = request.POST.get("tarih")
                        aciklama = request.POST.get("aciklama")
                        dosya = request.FILES.get("file")

                        klasor_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            dosya_sahibi=request.user.kullanicilar_db,
                            proje_ait_bilgisi=get_object_or_404(klasorler, id=ust_klasor),
                            dosya=dosya,
                            dosya_adi=dosya_Adi,
                            tarih=tarih,
                            aciklama=aciklama
                        )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                ust_klasor = request.POST.get("ust_klasor")
                dosya_Adi = request.POST.get("klasor")
                tarih = request.POST.get("tarih")
                aciklama = request.POST.get("aciklama")
                dosya = request.FILES.get("file")

                klasor_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    dosya_sahibi=request.user,
                    proje_ait_bilgisi=get_object_or_404(klasorler, id=ust_klasor),
                    dosya=dosya,
                    dosya_adi=dosya_Adi,
                    tarih=tarih,
                    aciklama=aciklama
                )
                #storage/mydir/files/34/yeni/

    #z = "storage/mydir/files/"+str(ust_klasor)+"/"+str(get_object_or_404(klasorler,id = ust_klasor).klasor_adi)+"/"
    return redirect_with_language("main:klasore_gir",str(ust_klasor),str(get_object_or_404(klasorler,id = ust_klasor).klasor_adi))
#klasore Dosya Ekle

def dosya_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            if True:
                ust_klasor = request.POST.get("ust_klasor")
                dosya_Adi = request.POST.get("klasor")
                tarih = request.POST.get("tarih")
                aciklama = request.POST.get("aciklama")
                dosya = request.FILES.get("file")

                klasor_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    dosya_sahibi=request.user,
                    proje_ait_bilgisi=get_object_or_404(klasorler, id=ust_klasor),
                    dosya=dosya,
                    dosya_adi=dosya_Adi,
                    tarih=tarih,
                    aciklama=aciklama
                )
            z = "control/storage/mydir/"+str(ust_klasor)+"/"+str(get_object_or_404(klasorler,id = ust_klasor).klasor_adi)+"/"+hash
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_olusturma:
                        ust_klasor = request.POST.get("ust_klasor")
                        dosya_Adi = request.POST.get("klasor")
                        tarih = request.POST.get("tarih")
                        aciklama = request.POST.get("aciklama")
                        dosya = request.FILES.get("file")

                        klasor_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            dosya_sahibi=request.user.kullanicilar_db,
                            proje_ait_bilgisi=get_object_or_404(klasorler, id=ust_klasor),
                            dosya=dosya,
                            dosya_adi=dosya_Adi,
                            tarih=tarih,
                            aciklama=aciklama
                        )
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                ust_klasor = request.POST.get("ust_klasor")
                dosya_Adi = request.POST.get("klasor")
                tarih = request.POST.get("tarih")
                aciklama = request.POST.get("aciklama")
                dosya = request.FILES.get("file")

                klasor_dosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    dosya_sahibi=request.user,
                    proje_ait_bilgisi=get_object_or_404(klasorler, id=ust_klasor),
                    dosya=dosya,
                    dosya_adi=dosya_Adi,
                    tarih=tarih,
                    aciklama=aciklama
                )
        z = "/storage/mydir/"+str(ust_klasor)+"/"+str(get_object_or_404(klasorler,id = ust_klasor).klasor_adi)+"/"
    return redirect_with_language(z)

#klasore Dosya Ekle
#dosya_ sil
def dosya_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            if True:
                ust_klasor = request.POST.get("ust_klasor")
                dosya_Adi = request.POST.get("klasor")

                klasor_dosyalari.objects.filter(id = dosya_Adi).update(silinme_bilgisi = True)
            if ust_klasor:
                z = "control/storage/mydir/"+str(ust_klasor)+"/"+str(get_object_or_404(klasorler,id = ust_klasor).klasor_adi)+"/"+hash
                return redirect_with_language(z)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_silme:
                        pass
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                pass
            ust_klasor = request.POST.get("ust_klasor")
            dosya_Adi = request.POST.get("klasor")

            klasor_dosyalari.objects.filter(id = dosya_Adi).update(silinme_bilgisi = True)
    if ust_klasor:
        z = "/storage/mydir/"+str(ust_klasor)+"/"+str(get_object_or_404(klasorler,id = ust_klasor).klasor_adi)+"/"
        return redirect_with_language(z)
    else:
        return redirect_with_language("main:depolama_sistemim_2",hash)

#dosya_ sil
def dosya_sil(request):
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_silme:
                        pass
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                pass
            ust_klasor = request.POST.get("ust_klasor")
            dosya_Adi = request.POST.get("klasor")

            klasor_dosyalari.objects.filter(id = dosya_Adi).update(silinme_bilgisi = True)
    if ust_klasor:
        z = "/storage/mydir/"+str(ust_klasor)+"/"+str(get_object_or_404(klasorler,id = ust_klasor).klasor_adi)+"/"
        return redirect_with_language(z)
    else:
        return redirect_with_language("main:depolama_sistemim")

def dosya_geri_getir(request):
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_duzenleme:
                        pass
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                pass
            ust_klasor = request.POST.get("ust_klasor")
            dosya_Adi = request.POST.get("klasor")

            klasor_dosyalari.objects.filter(id = dosya_Adi).update(silinme_bilgisi = False)
    return redirect_with_language("main:silinen_dosyalari")
def dosya_geri_getir_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            ust_klasor = request.POST.get("ust_klasor")
            dosya_Adi = request.POST.get("klasor")

            klasor_dosyalari.objects.filter(id = dosya_Adi).update(silinme_bilgisi = False)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_duzenleme:
                        pass
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                pass
            ust_klasor = request.POST.get("ust_klasor")
            dosya_Adi = request.POST.get("klasor")

            klasor_dosyalari.objects.filter(id = dosya_Adi).update(silinme_bilgisi = False)
    return redirect_with_language("main:silinen_dosyalari_2",hash)
#dosya_sil
from functools import reduce
import operator
#dokumanlari_gosterme

def dokumanlar(request):
    dosya_turu = [".xlsx",".pdf",".xlx",".txt",".docx",".doc",".ppt",".pptx"]
    content = sozluk_yapisi()
    content["id_bilgisi"] = id
    if super_admin_kontrolu(request):
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).filter(reduce(operator.or_, (Q(dosya__icontains = x) for x in dosya_turu)))
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).filter(reduce(operator.or_, (Q(dosya__icontains = x) for x in dosya_turu)))
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).filter(__icontains = search)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).filter(__icontains = search)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/dokuman.html",content)
def dokumanlar_2(request,hash):
    dosya_turu = [".xlsx",".pdf",".xlx",".txt",".docx",".doc",".ppt",".pptx"]
    content = sozluk_yapisi()
    content["id_bilgisi"] = id
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = users).filter(reduce(operator.or_, (Q(dosya__icontains = x) for x in dosya_turu)))
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).filter(reduce(operator.or_, (Q(dosya__icontains = x) for x in dosya_turu)))
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).filter(reduce(operator.or_, (Q(dosya__icontains = x) for x in dosya_turu)))
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).filter(__icontains = search)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).filter(__icontains = search)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/dokuman.html",content)

#dokumanlari_gosterme
#dokumanlari_gosterme

#media
def media_dosyalari(request):
    dosya_turu = [".jpg",".jpeg",".png",".ico",".css",".JFIF",".GIF",".WEBP"]
    content = sozluk_yapisi()
    content["id_bilgisi"] = id
    if super_admin_kontrolu(request):
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).filter(reduce(operator.or_, (Q(dosya__icontains = x) for x in dosya_turu)))
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).filter(reduce(operator.or_, (Q(dosya__icontains = x) for x in dosya_turu)))
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).filter(__icontains = search)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).filter(__icontains = search)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/dokuman.html",content)

#media
def media_dosyalari_2(request,hash):
    dosya_turu = [".jpg",".jpeg",".png",".ico",".css",".JFIF",".GIF",".WEBP"]
    content = sozluk_yapisi()
    content["id_bilgisi"] = id
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = users).filter(reduce(operator.or_, (Q(dosya__icontains = x) for x in dosya_turu)))
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).filter(reduce(operator.or_, (Q(dosya__icontains = x) for x in dosya_turu)))
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).filter(reduce(operator.or_, (Q(dosya__icontains = x) for x in dosya_turu)))
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).filter(__icontains = search)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).filter(__icontains = search)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/dokuman.html",content)


#media
#zamana göre

def zamana_dosyalari(request):

    content = sozluk_yapisi()
    content["id_bilgisi"] = id
    if super_admin_kontrolu(request):
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).order_by("-id")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).order_by("-id")
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).filter(__icontains = search)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).filter(__icontains = search)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/dokuman.html",content)
def zamana_dosyalari_2(request,hash):

    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    content["id_bilgisi"] = id
    if super_admin_kontrolu(request):
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = users).order_by("-id")
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).order_by("-id")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).order_by("-id")
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user.kullanicilar_db).filter(__icontains = search)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasor_dosyalari.objects.filter(silinme_bilgisi = False,dosya_sahibi = request.user).filter(__icontains = search)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/dokuman.html",content)


#zamana göre

def silinen_dosyalari(request):
    content = sozluk_yapisi()
    content["id_bilgisi"] = id
    if super_admin_kontrolu(request):
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasor_dosyalari.objects.filter(silinme_bilgisi = True,dosya_sahibi = request.user.kullanicilar_db).order_by("-id")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasor_dosyalari.objects.filter(silinme_bilgisi = True,dosya_sahibi = request.user).order_by("-id")
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = True,dosya_sahibi = request.user.kullanicilar_db).filter(__icontains = search)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasor_dosyalari.objects.filter(silinme_bilgisi = True,dosya_sahibi = request.user).filter(__icontains = search)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/dokuman.html",content)
#zamana göre

def silinen_dosyalari_2(request,hash):
    content = sozluk_yapisi()
    content["id_bilgisi"] = id
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = True,dosya_sahibi = users).order_by("-id")
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.dosya_yoneticisi_gorme:
                    profile = klasor_dosyalari.objects.filter(silinme_bilgisi = True,dosya_sahibi = request.user.kullanicilar_db).order_by("-id")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = klasor_dosyalari.objects.filter(silinme_bilgisi = True,dosya_sahibi = request.user).order_by("-id")
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =klasorler.objects.filter(Q(dosya_sahibi__last_name__icontains = search)|Q(klasor_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.dosya_yoneticisi_gorme:
                        profile = klasor_dosyalari.objects.filter(silinme_bilgisi = True,dosya_sahibi = request.user.kullanicilar_db).filter(__icontains = search)
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = klasor_dosyalari.objects.filter(silinme_bilgisi = True,dosya_sahibi = request.user).filter(__icontains = search)
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 25) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    return render(request,"santiye_yonetimi/dokuman.html",content)

#sözleşme olaylari
def sozlesmler_depolamam_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile = taseron_sozlesme_dosyalari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = taseron_sozlesme_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = users)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.sozlesmeler_gorme:
                    profile = taseron_sozlesme_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = taseron_sozlesme_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user)

    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi__proje_ait_bilgisi__last_name__icontains = search)|Q(taseron_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.sozlesmeler_gorme:
                        profile = taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
    return render(request,"santiye_yonetimi/sozlesmeler_depo.html",content)
def hakedis_depolamam_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile = taseron_hakedisles.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = taseron_hakedisles.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = users)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.hakedisler_gorme:
                    profile = taseron_hakedisles.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db)
                    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
            profile = taseron_hakedisles.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user)

    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi__last_name__icontains = search)|Q(dosya_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.hakedisler_gorme:
                        profile = taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(proje_ait_bilgisi__taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi = request.user) & Q(proje_ait_bilgisi__taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
    return render(request,"santiye_yonetimi/hakedis_depo.html",content)

#admin
#sözleşme olaylari
def sozlesmler_depolamam(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile = taseron_sozlesme_dosyalari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.sozlesmeler_gorme:
                    profile = taseron_sozlesme_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = taseron_sozlesme_dosyalari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user)

    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi__proje_ait_bilgisi__last_name__icontains = search)|Q(taseron_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.sozlesmeler_gorme:
                        profile = taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = taseron_sozlesme_dosyalari.objects.filter(Q(taseron_ait_bilgisi = request.user) & Q(taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
    return render(request,"santiye_yonetimi/sozlesmeler_depo.html",content)
def hakedis_depolamam(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile = taseron_hakedisles.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.hakedisler_gorme:
                    profile = taseron_hakedisles.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db)
                    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
            profile = taseron_hakedisles.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi__taseron_ait_bilgisi = request.user)

    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi__last_name__icontains = search)|Q(dosya_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.hakedisler_gorme:
                        profile = taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi = request.user.kullanicilar_db) & Q(proje_ait_bilgisi__taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = taseron_hakedisles.objects.filter(Q(proje_ait_bilgisi__taseron_ait_bilgisi = request.user) & Q(proje_ait_bilgisi__taseron_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    #content["blog_bilgisi"]  =projeler.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    content["taseronlar"] = taseronlar.objects.filter(taseron_ait_bilgisi= request.user,silinme_bilgisi = False)
    return render(request,"santiye_yonetimi/hakedis_depo.html",content)

#sözleşme olaylari
#dokumanlari_gosterme
def yapilacaklar_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile = IsplaniPlanlari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        content["blog_bilgisi"]  =CustomUser.objects.filter(kullanicilar_db = users,kullanici_silme_bilgisi = False,is_active = True)
        profile = IsplaniPlanlari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = users)
        content["katmanlar"] = katman.objects.filter(proje_ait_bilgisi = users ,silinme_bilgisi = False)
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.yapilacaklar_gorme:
                    profile = IsplaniPlanlari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db)
                    if a.izinler.yapilacaklar_olusturma:
                        content["blog_bilgisi"]  =CustomUser.objects.filter(kullanicilar_db = request.user.kullanicilar_db,kullanici_silme_bilgisi = False,is_active = True)
                        content["katmanlar"] = katman.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,silinme_bilgisi = False)
    
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            content["blog_bilgisi"]  =CustomUser.objects.filter(kullanicilar_db = request.user,kullanici_silme_bilgisi = False,is_active = True)
            profile = IsplaniPlanlari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
            content["katmanlar"] = katman.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = profile
    
    return render(request,"santiye_yonetimi/yapilacaklar.html",content)
#yapilacakalr
def yapilacaklar(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile = IsplaniPlanlari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.yapilacaklar_gorme:
                    profile = IsplaniPlanlari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db)
                    if a.izinler.yapilacaklar_olusturma:
                        content["blog_bilgisi"]  =CustomUser.objects.filter(kullanicilar_db = request.user.kullanicilar_db,kullanici_silme_bilgisi = False,is_active = True)
                        content["katmanlar"] = katman.objects.filter(proje_ait_bilgisi = request.user.kullanicilar_db,silinme_bilgisi = False)
    
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            content["blog_bilgisi"]  =CustomUser.objects.filter(kullanicilar_db = request.user,kullanici_silme_bilgisi = False,is_active = True)
            profile = IsplaniPlanlari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
            content["katmanlar"] = katman.objects.filter(proje_ait_bilgisi = request.user,silinme_bilgisi = False)
    
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = profile
    
    return render(request,"santiye_yonetimi/yapilacaklar.html",content)
#yapilacakalr
def yapilacak_gonder_json(request,id):
    
    
    a = get_object_or_404(IsplaniPlanlari, id = id)
    sonuc = {}

    # title varsa ekle
    
    sonuc["id"] = id
    if hasattr(a, 'title'):
        sonuc["title"] = a.title

    # aciklama varsa ekle
    if hasattr(a, 'aciklama'):
        sonuc["aciklama"] = a.aciklama

    # teslim_tarihi varsa ekle
    if hasattr(a, 'teslim_tarihi'):
        sonuc["teslim_tarihi"] = a.teslim_tarihi

    # status varsa ekle
    if hasattr(a, 'status'):
        sonuc["status"] = a.status

    # oncelik_durumu varsa ekle
    if hasattr(a, 'oncelik_durumu'):
        sonuc["oncelik_durumu"] = a.oncelik_durumu

    # katman_id varsa ekle
    if hasattr(a, 'katman') and hasattr(a.katman, 'id'):
        sonuc["katman_id"] = a.katman.id
    
        
    
    # blok_id varsa ekle
    if hasattr(a, 'blok') and hasattr(a.blok, 'id'):
        sonuc["blok_id"] = a.blok.id
    if hasattr(a, 'locasyonx') :
        sonuc["locasyonx"] = a.locasyonx
    if hasattr(a, 'locasyony') :
        sonuc["locasyony"] = a.locasyony
    # katman varsa ekle
    if hasattr(a, 'katman') and hasattr(a.katman, 'katman_adi'):
        sonuc["katman"] = a.katman.katman_adi
    if hasattr(a, 'katman') and hasattr(a.katman, 'katman_dosyasi'):
        sonuc["katman_dosyasi"] = a.katman.katman_dosyasi.url

    # blok varsa ekle
    if hasattr(a, 'blok') and hasattr(a.blok, 'blog_adi'):
        sonuc["blok"] = a.blok.blog_adi

    # kat varsa ekle
    if hasattr(a, 'kat'):
        sonuc["kat"] = a.kat
    try:
        sonuc["isaretli"] = IsplaniDosyalari.objects.filter(proje_ait_bilgisi = a,pin = "pin").last().dosya.url
    except:
        sonuc["isaretli"] =""
    #print(sonuc)
    return JsonResponse(sonuc)
import base64
from django.core.files.base import ContentFile
def yapilacalar_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.yapilacaklar_olusturma:
                        baslik = request.POST.get("baslik")
                        durum = request.POST.get("durum")
                        aciliyet =request.POST.get("aciliyet")
                        teslim_tarihi = request.POST.get("teslim_tarihi")
                        blogbilgisi = request.POST.getlist("blogbilgisi")
                        aciklama = request.POST.get("aciklama")
                        katman_bilgisi = request.POST.get("katman")
                        yapi_gonder = request.POST.get("yapi_gonder")
                        kat = request.POST.get("kat")
                        base64_image = request.POST.get('base_64_format', '')
                        pin_lokasyunuxd = request.POST.get("pin_lokasyunux") 
                        pin_lokasyunuyd = request.POST.get("pin_lokasyunuy") 
                        if base64_image !="" .startswith('data:image/png;base64,'):
                            base64_image = base64_image.replace('data:image/png;base64,', '')
                            file_extension = 'png'
                        elif base64_image !="" .startswith('data:image/jpeg;base64,'):
                            base64_image = base64_image.replace('data:image/jpeg;base64,', '')
                            file_extension = 'jpeg'
                        
                        if base64_image !=""  :
                            image_data = base64.b64decode(base64_image)
                            image_file = ContentFile(image_data, name=f'image.{file_extension}')
                        if kat== None or kat  == ""  :
                            kat = 0

                        new_project = IsplaniPlanlari(
                            proje_ait_bilgisi = request.user.kullanicilar_db,
                            title = baslik,
                            status = durum,
                            aciklama = aciklama,
                            oncelik_durumu =aciliyet,
                            teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                            blok = get_object_or_none(bloglar,id = yapi_gonder),
                            katman = get_object_or_none(katman,id = katman_bilgisi),
                            kat = kat,locasyonx = pin_lokasyunuxd,locasyony = pin_lokasyunuyd
                        )
                        new_project.save()
                        bloglar_bilgisi = []
                        for i in blogbilgisi:
                            bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                        new_project.yapacaklar.add(*bloglar_bilgisi)
                        images = request.FILES.getlist('file')
                        #print(images)
                        isim = 1
                        for images in images:
                            IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user.kullanicilar_db,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                        if base64_image !="" :
                            IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user.kullanicilar_db,dosya=image_file,pin="pin")
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                baslik = request.POST.get("baslik")
                durum = request.POST.get("durum")
                aciliyet =request.POST.get("aciliyet")
                teslim_tarihi = request.POST.get("teslim_tarihi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                katman_bilgisi = request.POST.get("katman")
                yapi_gonder = request.POST.get("yapi_gonder")
                kat = request.POST.get("kat")
                base64_image = request.POST.get('base_64_format', '')
                pin_lokasyunuxd = request.POST.get("pin_lokasyunux") 
                pin_lokasyunuyd = request.POST.get("pin_lokasyunuy")
                if base64_image !="" .startswith('data:image/png;base64,'):
                    base64_image = base64_image.replace('data:image/png;base64,', '')
                    file_extension = 'png'
                elif base64_image !="" .startswith('data:image/jpeg;base64,'):
                    base64_image = base64_image.replace('data:image/jpeg;base64,', '')
                    file_extension = 'jpeg'
                if kat == None or kat  == ""  :
                    kat = 0
                #print(kat)
                new_project = IsplaniPlanlari(
                    proje_ait_bilgisi = request.user,
                    title = baslik,
                    status = durum,
                    aciklama = aciklama,
                    oncelik_durumu =aciliyet,
                    teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                    blok = get_object_or_none(bloglar,id = yapi_gonder),
                    katman = get_object_or_none(katman,id = katman_bilgisi),
                    kat = kat,locasyonx = pin_lokasyunuxd,locasyony = pin_lokasyunuyd
                )
                new_project.save()
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                new_project.yapacaklar.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                #print(images)
                isim = 1
                for images in images:
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
                if base64_image !=""  :
                    image_data = base64.b64decode(base64_image)
                    image_file = ContentFile(image_data, name=f'image.{file_extension}')
                if base64_image !="" :
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=image_file,pin="pin")
    return redirect_with_language("main:yapilacaklar")
#
def yapilacalar_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            if True:
                baslik = request.POST.get("baslik")
                durum = request.POST.get("durum")
                aciliyet =request.POST.get("aciliyet")
                teslim_tarihi = request.POST.get("teslim_tarihi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                katman_bilgisi = request.POST.get("katman")
                yapi_gonder = request.POST.get("yapi_gonder")
                kat = request.POST.get("kat")
                base64_image = request.POST.get('base_64_format', '')
                pin_lokasyunuxd = request.POST.get("pin_lokasyunux") 
                pin_lokasyunuyd = request.POST.get("pin_lokasyunuy")
                if base64_image !="" .startswith('data:image/png;base64,'):
                    base64_image = base64_image.replace('data:image/png;base64,', '')
                    file_extension = 'png'
                elif base64_image !="" .startswith('data:image/jpeg;base64,'):
                    base64_image = base64_image.replace('data:image/jpeg;base64,', '')
                    file_extension = 'jpeg'
                if kat == None or kat  == ""  :
                    kat = 0
                #print(kat)
                new_project = IsplaniPlanlari(
                    proje_ait_bilgisi = users,
                    title = baslik,
                    status = durum,
                    aciklama = aciklama,
                    oncelik_durumu =aciliyet,
                    teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                    blok = get_object_or_none(bloglar,id = yapi_gonder),
                    katman = get_object_or_none(katman,id = katman_bilgisi),
                    kat = kat,locasyonx = pin_lokasyunuxd,locasyony = pin_lokasyunuyd
                )
                new_project.save()
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                new_project.yapacaklar.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                #print(images)
                isim = 1
                for images in images:
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = users,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
                if base64_image !=""  :
                    image_data = base64.b64decode(base64_image)
                    image_file = ContentFile(image_data, name=f'image.{file_extension}')
                if base64_image !="" :
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = users,dosya=image_file,pin="pin")
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.yapilacaklar_olusturma:
                        baslik = request.POST.get("baslik")
                        durum = request.POST.get("durum")
                        aciliyet =request.POST.get("aciliyet")
                        teslim_tarihi = request.POST.get("teslim_tarihi")
                        blogbilgisi = request.POST.getlist("blogbilgisi")
                        aciklama = request.POST.get("aciklama")
                        katman_bilgisi = request.POST.get("katman")
                        yapi_gonder = request.POST.get("yapi_gonder")
                        kat = request.POST.get("kat")
                        base64_image = request.POST.get('base_64_format', '')
                        pin_lokasyunuxd = request.POST.get("pin_lokasyunux") 
                        pin_lokasyunuyd = request.POST.get("pin_lokasyunuy") 
                        if base64_image !="" .startswith('data:image/png;base64,'):
                            base64_image = base64_image.replace('data:image/png;base64,', '')
                            file_extension = 'png'
                        elif base64_image !="" .startswith('data:image/jpeg;base64,'):
                            base64_image = base64_image.replace('data:image/jpeg;base64,', '')
                            file_extension = 'jpeg'
                        
                        if base64_image !=""  :
                            image_data = base64.b64decode(base64_image)
                            image_file = ContentFile(image_data, name=f'image.{file_extension}')
                        if kat== None or kat  == ""  :
                            kat = 0

                        new_project = IsplaniPlanlari(
                            proje_ait_bilgisi = request.user.kullanicilar_db,
                            title = baslik,
                            status = durum,
                            aciklama = aciklama,
                            oncelik_durumu =aciliyet,
                            teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                            blok = get_object_or_none(bloglar,id = yapi_gonder),
                            katman = get_object_or_none(katman,id = katman_bilgisi),
                            kat = kat,locasyonx = pin_lokasyunuxd,locasyony = pin_lokasyunuyd
                        )
                        new_project.save()
                        bloglar_bilgisi = []
                        for i in blogbilgisi:
                            bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                        new_project.yapacaklar.add(*bloglar_bilgisi)
                        images = request.FILES.getlist('file')
                        #print(images)
                        isim = 1
                        for images in images:
                            IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user.kullanicilar_db,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                        if base64_image !="" :
                            IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user.kullanicilar_db,dosya=image_file,pin="pin")
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                baslik = request.POST.get("baslik")
                durum = request.POST.get("durum")
                aciliyet =request.POST.get("aciliyet")
                teslim_tarihi = request.POST.get("teslim_tarihi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                katman_bilgisi = request.POST.get("katman")
                yapi_gonder = request.POST.get("yapi_gonder")
                kat = request.POST.get("kat")
                base64_image = request.POST.get('base_64_format', '')
                pin_lokasyunuxd = request.POST.get("pin_lokasyunux") 
                pin_lokasyunuyd = request.POST.get("pin_lokasyunuy")
                if base64_image !="" .startswith('data:image/png;base64,'):
                    base64_image = base64_image.replace('data:image/png;base64,', '')
                    file_extension = 'png'
                elif base64_image !="" .startswith('data:image/jpeg;base64,'):
                    base64_image = base64_image.replace('data:image/jpeg;base64,', '')
                    file_extension = 'jpeg'
                if kat == None or kat  == ""  :
                    kat = 0
                #print(kat)
                new_project = IsplaniPlanlari(
                    proje_ait_bilgisi = request.user,
                    title = baslik,
                    status = durum,
                    aciklama = aciklama,
                    oncelik_durumu =aciliyet,
                    teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                    blok = get_object_or_none(bloglar,id = yapi_gonder),
                    katman = get_object_or_none(katman,id = katman_bilgisi),
                    kat = kat,locasyonx = pin_lokasyunuxd,locasyony = pin_lokasyunuyd
                )
                new_project.save()
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                new_project.yapacaklar.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                #print(images)
                isim = 1
                for images in images:
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
                if base64_image !=""  :
                    image_data = base64.b64decode(base64_image)
                    image_file = ContentFile(image_data, name=f'image.{file_extension}')
                if base64_image !="" :
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=image_file,pin="pin")
    return redirect_with_language("main:yapilacaklar_2",hash)
#
def yapilacalar_ekle_duzenleme(request):
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.yapilacaklar_olusturma:
                        id = request.POST.get("id")
                        baslik = request.POST.get("baslik")
                        durum = request.POST.get("durum")
                        aciliyet =request.POST.get("aciliyet")
                        teslim_tarihi = request.POST.get("teslim_tarihi")
                        blogbilgisi = request.POST.getlist("blogbilgisi")
                        aciklama = request.POST.get("aciklama")
                        katman_bilgisi = request.POST.get("katman")
                        yapi_gonder = request.POST.get("yapi_gonder")
                        kat = request.POST.get("kat")
                        base64_image = request.POST.get('base_64_format', '')
                        pin_lokasyunuxd = request.POST.get("pin_lokasyunuxd") 
                        pin_lokasyunuyd = request.POST.get("pin_lokasyunuyd")
                        if kat == None or kat  == ""  :
                            kat = 0
                        if base64_image !="" .startswith('data:image/png;base64,'):
                            base64_image = base64_image.replace('data:image/png;base64,', '')
                            file_extension = 'png'
                        elif base64_image !="" .startswith('data:image/jpeg;base64,'):
                            base64_image = base64_image.replace('data:image/jpeg;base64,', '')
                            file_extension = 'jpeg'
                        
                        image_data = base64.b64decode(base64_image)
                        image_file = ContentFile(image_data, name=f'image.{file_extension}')
                        IsplaniPlanlari.objects.filter(id = id).update(
                            proje_ait_bilgisi = request.user.kullanicilar_db,
                            title = baslik,
                            status = durum,
                            aciklama = aciklama,
                            oncelik_durumu =aciliyet,
                            teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                            blok = get_object_or_none(bloglar,id = yapi_gonder),
                            katman = get_object_or_none(katman,id = katman_bilgisi),
                            kat = kat,locasyonx = pin_lokasyunuxd,locasyony = pin_lokasyunuyd
                        )
                        new_project = get_object_or_none(IsplaniPlanlari,id = id)
                        new_project.save()
                        bloglar_bilgisi = []
                        for i in blogbilgisi:
                            bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                        new_project.yapacaklar.add(*bloglar_bilgisi)
                        images = request.FILES.getlist('file')
                        #print(images)
                        isim = 1
                        for images in images:
                            IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user.kullanicilar_db,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                        if base64_image !="" :
                            IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user.kullanicilar_db,dosya=image_file,pin="pin")
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                id = request.POST.get("id")
                baslik = request.POST.get("baslik")
                durum = request.POST.get("durum")
                aciliyet =request.POST.get("aciliyet")
                teslim_tarihi = request.POST.get("teslim_tarihi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                katman_bilgisi = request.POST.get("katman")
                yapi_gonder = request.POST.get("yapi_gonder")
                kat = request.POST.get("kat")
                base64_image = request.POST.get('base_64_format', '')
                pin_lokasyunuxd = request.POST.get("pin_lokasyunuxd") 
                pin_lokasyunuyd = request.POST.get("pin_lokasyunuyd") 
                if kat == None or kat  == ""  :
                    kat = 0
                if base64_image !="" .startswith('data:image/png;base64,'):
                    base64_image = base64_image.replace('data:image/png;base64,', '')
                    file_extension = 'png'
                elif base64_image !="" .startswith('data:image/jpeg;base64,'):
                    base64_image = base64_image.replace('data:image/jpeg;base64,', '')
                    file_extension = 'jpeg'
                        
                image_data = base64.b64decode(base64_image)
                image_file = ContentFile(image_data, name=f'image.{file_extension}')
                IsplaniPlanlari.objects.filter(id = id).update(
                    proje_ait_bilgisi = request.user,
                    title = baslik,
                    status = durum,
                    aciklama = aciklama,
                    oncelik_durumu =aciliyet,
                    teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                    blok = get_object_or_none(bloglar,id = yapi_gonder),
                    katman = get_object_or_none(katman,id = katman_bilgisi),
                    kat = kat,locasyonx = pin_lokasyunuxd,locasyony = pin_lokasyunuyd
                )
                new_project = get_object_or_none(IsplaniPlanlari,id = id)
                new_project.save()
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                new_project.yapacaklar.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                #print(images)
                isim = 1
                for images in images:
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
                if base64_image !="" :
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=image_file,pin="pin")
    return redirect_with_language("main:yapilacaklar")
#
def yapilacalar_ekle_duzenleme_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            if True:
                id = request.POST.get("id")
                baslik = request.POST.get("baslik")
                durum = request.POST.get("durum")
                aciliyet =request.POST.get("aciliyet")
                teslim_tarihi = request.POST.get("teslim_tarihi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                katman_bilgisi = request.POST.get("katman")
                yapi_gonder = request.POST.get("yapi_gonder")
                kat = request.POST.get("kat")
                base64_image = request.POST.get('base_64_format', '')
                pin_lokasyunuxd = request.POST.get("pin_lokasyunuxd") 
                pin_lokasyunuyd = request.POST.get("pin_lokasyunuyd") 
                if kat == None or kat  == ""  :
                    kat = 0
                if base64_image !="" .startswith('data:image/png;base64,'):
                    base64_image = base64_image.replace('data:image/png;base64,', '')
                    file_extension = 'png'
                elif base64_image !="" .startswith('data:image/jpeg;base64,'):
                    base64_image = base64_image.replace('data:image/jpeg;base64,', '')
                    file_extension = 'jpeg'
                        
                image_data = base64.b64decode(base64_image)
                image_file = ContentFile(image_data, name=f'image.{file_extension}')
                IsplaniPlanlari.objects.filter(id = id).update(
                    proje_ait_bilgisi = users,
                    title = baslik,
                    status = durum,
                    aciklama = aciklama,
                    oncelik_durumu =aciliyet,
                    teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                    blok = get_object_or_none(bloglar,id = yapi_gonder),
                    katman = get_object_or_none(katman,id = katman_bilgisi),
                    kat = kat,locasyonx = pin_lokasyunuxd,locasyony = pin_lokasyunuyd
                )
                new_project = get_object_or_none(IsplaniPlanlari,id = id)
                new_project.save()
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                new_project.yapacaklar.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                #print(images)
                isim = 1
                for images in images:
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = users,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
                if base64_image !="" :
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = users,dosya=image_file,pin="pin")
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.yapilacaklar_olusturma:
                        id = request.POST.get("id")
                        baslik = request.POST.get("baslik")
                        durum = request.POST.get("durum")
                        aciliyet =request.POST.get("aciliyet")
                        teslim_tarihi = request.POST.get("teslim_tarihi")
                        blogbilgisi = request.POST.getlist("blogbilgisi")
                        aciklama = request.POST.get("aciklama")
                        katman_bilgisi = request.POST.get("katman")
                        yapi_gonder = request.POST.get("yapi_gonder")
                        kat = request.POST.get("kat")
                        base64_image = request.POST.get('base_64_format', '')
                        pin_lokasyunuxd = request.POST.get("pin_lokasyunuxd") 
                        pin_lokasyunuyd = request.POST.get("pin_lokasyunuyd")
                        if kat == None or kat  == ""  :
                            kat = 0
                        if base64_image !="" .startswith('data:image/png;base64,'):
                            base64_image = base64_image.replace('data:image/png;base64,', '')
                            file_extension = 'png'
                        elif base64_image !="" .startswith('data:image/jpeg;base64,'):
                            base64_image = base64_image.replace('data:image/jpeg;base64,', '')
                            file_extension = 'jpeg'
                        
                        image_data = base64.b64decode(base64_image)
                        image_file = ContentFile(image_data, name=f'image.{file_extension}')
                        IsplaniPlanlari.objects.filter(id = id).update(
                            proje_ait_bilgisi = request.user.kullanicilar_db,
                            title = baslik,
                            status = durum,
                            aciklama = aciklama,
                            oncelik_durumu =aciliyet,
                            teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                            blok = get_object_or_none(bloglar,id = yapi_gonder),
                            katman = get_object_or_none(katman,id = katman_bilgisi),
                            kat = kat,locasyonx = pin_lokasyunuxd,locasyony = pin_lokasyunuyd
                        )
                        new_project = get_object_or_none(IsplaniPlanlari,id = id)
                        new_project.save()
                        bloglar_bilgisi = []
                        for i in blogbilgisi:
                            bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                        new_project.yapacaklar.add(*bloglar_bilgisi)
                        images = request.FILES.getlist('file')
                        #print(images)
                        isim = 1
                        for images in images:
                            IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user.kullanicilar_db,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
                        if base64_image !="" :
                            IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user.kullanicilar_db,dosya=image_file,pin="pin")
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                id = request.POST.get("id")
                baslik = request.POST.get("baslik")
                durum = request.POST.get("durum")
                aciliyet =request.POST.get("aciliyet")
                teslim_tarihi = request.POST.get("teslim_tarihi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.get("aciklama")
                katman_bilgisi = request.POST.get("katman")
                yapi_gonder = request.POST.get("yapi_gonder")
                kat = request.POST.get("kat")
                base64_image = request.POST.get('base_64_format', '')
                pin_lokasyunuxd = request.POST.get("pin_lokasyunuxd") 
                pin_lokasyunuyd = request.POST.get("pin_lokasyunuyd") 
                if kat == None or kat  == ""  :
                    kat = 0
                if base64_image !="" .startswith('data:image/png;base64,'):
                    base64_image = base64_image.replace('data:image/png;base64,', '')
                    file_extension = 'png'
                elif base64_image !="" .startswith('data:image/jpeg;base64,'):
                    base64_image = base64_image.replace('data:image/jpeg;base64,', '')
                    file_extension = 'jpeg'
                        
                image_data = base64.b64decode(base64_image)
                image_file = ContentFile(image_data, name=f'image.{file_extension}')
                IsplaniPlanlari.objects.filter(id = id).update(
                    proje_ait_bilgisi = request.user,
                    title = baslik,
                    status = durum,
                    aciklama = aciklama,
                    oncelik_durumu =aciliyet,
                    teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                    blok = get_object_or_none(bloglar,id = yapi_gonder),
                    katman = get_object_or_none(katman,id = katman_bilgisi),
                    kat = kat,locasyonx = pin_lokasyunuxd,locasyony = pin_lokasyunuyd
                )
                new_project = get_object_or_none(IsplaniPlanlari,id = id)
                new_project.save()
                bloglar_bilgisi = []
                for i in blogbilgisi:
                    bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                new_project.yapacaklar.add(*bloglar_bilgisi)
                images = request.FILES.getlist('file')
                #print(images)
                isim = 1
                for images in images:
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                    isim = isim+1
                if base64_image !="" :
                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=image_file,pin="pin")
    return redirect_with_language("main:yapilacaklar_2",hash)
#
def yapilacalar_ekle_toplu(request):
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.yapilacaklar_olusturma:
                        baslik = request.POST.get("baslik")
                        durum = request.POST.get("durum")
                        aciliyet =request.POST.get("aciliyet")
                        teslim_tarihi = request.POST.get("teslim_tarihi")
                        blogbilgisi = request.POST.getlist("blogbilgisi")
                        aciklama = request.POST.getlist("aciklama")
                        for i in range(0,len(aciklama)):
                            if aciklama[i]:
                                new_project = IsplaniPlanlari(
                                    proje_ait_bilgisi = request.user.kullanicilar_db,
                                    title = baslik,
                                    status = durum,
                                    aciklama = aciklama[i],
                                    oncelik_durumu =aciliyet,
                                    teslim_tarihi = teslim_tarihi,silinme_bilgisi = False
                                )
                                new_project.save()
                                bloglar_bilgisi = []
                                for i in blogbilgisi:
                                    bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                                new_project.yapacaklar.add(*bloglar_bilgisi)
                                images = request.FILES.getlist('file')
                                isim = 1
                                #print(images,"resim geldi")
                                for images in images:
                                    IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user.kullanicilar_db,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                                    isim = isim+1
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                baslik = request.POST.get("baslik")
                durum = request.POST.get("durum")
                aciliyet =request.POST.get("aciliyet")
                teslim_tarihi = request.POST.get("teslim_tarihi")
                blogbilgisi = request.POST.getlist("blogbilgisi")
                aciklama = request.POST.getlist("aciklama")
                for i in range(0,len(aciklama)):
                    if aciklama[i]:
                        new_project = IsplaniPlanlari(
                            proje_ait_bilgisi = request.user,
                            title = baslik,
                            status = durum,
                            aciklama = aciklama[i],
                            oncelik_durumu =aciliyet,
                            teslim_tarihi = teslim_tarihi,silinme_bilgisi = False
                        )
                        new_project.save()
                        bloglar_bilgisi = []
                        for i in blogbilgisi:
                            bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
                        new_project.yapacaklar.add(*bloglar_bilgisi)
                        images = request.FILES.getlist('file')
                        isim = 1
                        #print(images,"resim geldi")
                        for images in images:
                            IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                            isim = isim+1
    return redirect_with_language("main:yapilacaklar")

def yapilacalar_sil_2(request,hash):
    
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.yapilacaklar_silme:
                    pass
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
    else:
        pass
    if True:
        content = sozluk_yapisi()
        d = decode_id(hash)
        content["hashler"] = hash
        users = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = users
        if request.POST:
            id = request.POST.get("id_bilgisi")
            IsplaniPlanlari.objects.filter(id = id).update(silinme_bilgisi = True)
    return redirect_with_language("main:yapilacaklar_2",hash)

def yapilacalar_sil(request):
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.yapilacaklar_silme:
                    pass
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
    else:
        pass
    if request.POST:
        id = request.POST.get("id_bilgisi")
        IsplaniPlanlari.objects.filter(id = id).update(silinme_bilgisi = True)
    return redirect_with_language("main:yapilacaklar")

def yapilacak_durumu_yenileme(request):
    if request.POST:
        yenilenecekeklemeyapilacak = request.POST.get("yenilenecekeklemeyapilacak")
        aciklama = request.POST.get("aciklama")
        durum  =request.POST.get("durum")
        teslim_tarihi  =request.POST.get("teslim_tarihi")
        new_project = IsplaniPlanlariIlerleme(
                proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id =yenilenecekeklemeyapilacak ),
                status = durum,
                aciklama = aciklama,
                teslim_tarihi = teslim_tarihi,silinme_bilgisi = False,
                yapan_kisi = request.user
            )
        IsplaniPlanlari.objects.filter(id =yenilenecekeklemeyapilacak).update(status = durum)
        new_project.save()
        images = request.FILES.getlist('file')
        isim = 1
        for images in images:
            IsplaniIlerlemeDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlariIlerleme,id = new_project.id),yapan_kisi = request.user,dosya=images,dosya_sahibi = get_object_or_404(IsplaniPlanlari,id =yenilenecekeklemeyapilacak))  # Urun_resimleri modeline resimleri kaydet
            isim = isim+1
    return redirect_with_language("main:yapilacaklar")

def yapilacalar_duzenle(request):
    if request.POST:
        if request.user.superuser:
            pass
        else:
            id = request.POSt.get("id_bilgisi")
            baslik = request.POST.get("baslik")
            durum = request.POST.get("durum")
            aciliyet =request.POST.get("aciliyet")
            teslim_tarihi = request.POST.get("teslim_tarihi")
            blogbilgisi = request.POST.getlist("blogbilgisi")
            aciklama = request.POST.get("aciklama")
            new_project = IsplaniPlanlari.objects.filter(id = id).update(
                proje_ait_bilgisi = request.user,
                title = baslik,
                status = durum,
                aciklama = aciklama,
                oncelik_durumu =aciliyet,
                teslim_tarihi = teslim_tarihi,silinme_bilgisi = False
            )
            new_project.save()
            bloglar_bilgisi = []
            for i in blogbilgisi:
                bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
            new_project.yapacaklar.add(*bloglar_bilgisi)
            images = request.FILES.getlist('file')
            isim = 1
            for images in images:
                IsplaniDosyalari.objects.create(proje_ait_bilgisi = get_object_or_404(IsplaniPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                isim = isim+1
    return redirect_with_language("main:yapilacaklar")

#time_lline
def yapilacaklar_timeline(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile = YapilacakPlanlari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        profile = YapilacakPlanlari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user).order_by("teslim_tarihi")

    if request.GET:
        siralama = request.GET.get("siralama")
        status = request.GET.get("status")
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =YapilacakPlanlari.objects.filter(Q(proje_ait_bilgisi__last_name__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            profile = YapilacakPlanlari.objects.filter(Q(proje_ait_bilgisi = request.user) & Q(silinme_bilgisi = False))
        if status:
            profile = profile.filter(status = status)
        if search:
            profile = profile.filter(title__icontains=search )
        if siralama == "1":
            profile = profile.order_by("-id")
        elif siralama == "2":
            profile = profile.order_by("-title")
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = page_obj
    content["blog_bilgisi"]  =CustomUser.objects.filter(kullanicilar_db = request.user,kullanici_silme_bilgisi = False,is_active = True)
    return render(request,"santiye_yonetimi/time_line.html",content)

def yapilacalar_time_line_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            baslik = request.POST.get("baslik")
            durum = request.POST.get("durum")
            aciliyet =request.POST.get("aciliyet")
            teslim_tarihi = request.POST.get("teslim_tarihi")
            blogbilgisi = request.POST.getlist("blogbilgisi")
            aciklama = request.POST.get("aciklama")
            new_project = YapilacakPlanlari(
                proje_ait_bilgisi = request.user,
                title = baslik,
                status = durum,
                aciklama = aciklama,
                oncelik_durumu =aciliyet,
                teslim_tarihi = teslim_tarihi,silinme_bilgisi = False
            )
            new_project.save()
            bloglar_bilgisi = []
            for i in blogbilgisi:
                bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
            new_project.yapacaklar.add(*bloglar_bilgisi)
            images = request.FILES.getlist('file')
            isim = 1
            for images in images:
                YapilacakDosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = get_object_or_404(YapilacakPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                isim = isim+1
    return redirect_with_language("main:yapilacaklar_timeline")


def yapilacalar_time_line_sil(request):
    if request.POST:
        id = request.POST.get("buttonId")
        YapilacakPlanlari.objects.filter(id = id).update(silinme_bilgisi = True)
    return redirect_with_language("main:yapilacaklar_timeline")
#yapilacakalr
def yapilacalar_time_line_duzenle(request):
    if request.POST:
        if request.user.is_superuser:
            pass
        else:
            id = request.POST.get("id")
            baslik = request.POST.get("baslik")
            durum = request.POST.get("durum")
            aciliyet =request.POST.get("aciliyet")
            teslim_tarihi = request.POST.get("teslim_tarihi")
            blogbilgisi = request.POST.getlist("blogbilgisi")
            aciklama = request.POST.get("aciklama")
            new_project = YapilacakPlanlari.objects.filter(id = id).update(
                proje_ait_bilgisi = request.user,
                title = baslik,
                status = durum,
                aciklama = aciklama,
                oncelik_durumu =aciliyet,
                teslim_tarihi = teslim_tarihi,silinme_bilgisi = False
            )
            bloglar_bilgisi = []
            for i in blogbilgisi:
                bloglar_bilgisi.append(CustomUser.objects.get(id=int(i)))
            get_object_or_404(YapilacakPlanlari,id = id).yapacaklar.add(*bloglar_bilgisi)
            images = request.FILES.getlist('file')
            isim = 1
            for images in images:
                YapilacakDosyalari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = get_object_or_404(YapilacakPlanlari,id = new_project.id),dosya_sahibi = request.user,dosya=images)  # Urun_resimleri modeline resimleri kaydet
                isim = isim+1
    return redirect_with_language("main:yapilacaklar_timeline")

from .utils import *
from django.utils.safestring import mark_safe
def gant_list(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile = gantt_olayi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gant_gorme:
                    kullanici = request.user.kullanicilar_db
                    
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
        gant_sablonlari = gant_aitlikleri.objects.filter(gantt_sahibii = kullanici,silinme_bilgisi = False)
        blok_alamlar = bloglar.objects.filter(proje_ait_bilgisi = kullanici)
        content["bloklar"] = blok_alamlar
        content["gant_sablonlari"] = gant_sablonlari
    return render(request,"santiye_yonetimi/gant_list.html",content)
def gant_sablon_ekle(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile = gantt_olayi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gant_olusturma:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
        if request.POST:
            baslik = request.POST.get("kasaadi")   
            blok_bilgileri = request.POST.get("avanslarda_kullan")
            aciklama  = request.POST.get("aciklame")
            gant_aitlikleri.objects.create(
                gant_adi = baslik,
                gant_blok = get_object_or_404(bloglar,id = blok_bilgileri),
                gantt_sahibii = kullanici,
                ganti_degistiren_kisi = request.user,
                aciklama = aciklama)
            return redirect_with_language("main:gant_list")
def gant_sablon_duzenle(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile = gantt_olayi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gant_olusturma:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
        if request.POST:
            baslik = request.POST.get("kasaadi")   
            blok_bilgileri = request.POST.get("avanslarda_kullan")
            aciklama  = request.POST.get("aciklame")
            id = request.POST.get("id")
            gant_aitlikleri.objects.filter(id = id).update(
                gant_adi = baslik,
                gant_blok = get_object_or_404(bloglar,id = blok_bilgileri),
                gantt_sahibii = kullanici,
                ganti_degistiren_kisi = request.user,
                aciklama = aciklama)
            return redirect_with_language("main:gant_list")
def gant_sablon_silme(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile = gantt_olayi.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gant_olusturma:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
        if request.POST:
           
            id = request.POST.get("id")
            gant_aitlikleri.objects.filter(id = id).update(
                silinme_bilgisi = True)
            return redirect_with_language("main:gant_list")

def takvim_olaylari(request,id):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        profile = IsplaniPlanlari.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.gant_gorme:
                    kullanici = request.user.kullanicilar_db
                    #content["gant"]  =gantt_olayi.objects.filter(gantt_sahibii = request.user.kullanicilar_db).last()
    
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
        content["gant_aitlikleri"] = get_object_or_none(gant_aitlikleri,id = id, gantt_sahibii = kullanici) 
        content["gant"]  = gantt_olayi.objects.filter(gantt_sahibii = kullanici,gatn_aitlik = get_object_or_none(gant_aitlikleri,id = id, gantt_sahibii = kullanici) ).last()
            
    
    return render(request,"santiye_yonetimi/takvim.html",content)
#takvim
from django.http import JsonResponse
from django.shortcuts import redirect

def gant_kaydet(request):
    if request.method == 'POST':
        gant_verisi = request.POST.get("gant_verisi")
        gant_aitlik = request.POST.get("gant_aitlik")
        if not gant_verisi:
            return JsonResponse({'ok': False, 'message': 'Gantt verisi bulunamadı.'})

        if super_admin_kontrolu(request):
            # Super admin işlemleri
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar, kullanicilar=request.user)
                if a and a.izinler.gant_duzenleme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return JsonResponse({'ok': False, 'message': 'Yetkisiz erişim'}, status=403)
            else:
                kullanici = request.user

        gantt_olayi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
            gantt_sahibii=kullanici,
            ganti_degistiren_kisi=request.user,
            gantt_verisi=gant_verisi,
            gatn_aitlik = get_object_or_none(gant_aitlikleri, id=gant_aitlik, gantt_sahibii=kullanici),
        )

        return JsonResponse({'ok': True, 'message': 'Gantt kaydedildi'})
    
    return redirect_with_language("main:takvim_olaylari",gant_aitlik)
def gant_aktarma(request):
    if request.method == 'POST':
        gelen_gant = request.POST.get("gelen_gant")
        giden_gant = request.POST.get("giden_gant")
        if not gelen_gant:
            return JsonResponse({'ok': False, 'message': 'Gantt verisi bulunamadı.'})
        if super_admin_kontrolu(request):
            # Super admin işlemleri
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar, kullanicilar=request.user)
                if a and a.izinler.gant_duzenleme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                kullanici = request.user
        gant_gelen_cekme = gantt_olayi.objects.filter(gatn_aitlik__id = gelen_gant).last()
        gantt_olayi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
            gantt_sahibii=kullanici,
            ganti_degistiren_kisi=request.user,
            gantt_verisi=gant_gelen_cekme.gantt_verisi,
            gatn_aitlik = get_object_or_none(gant_aitlikleri, id=giden_gant, gantt_sahibii=kullanici),
        )
        return redirect_with_language("main:gant_list")
def santiye_raporu_2(request,id,hash):
    content = sozluk_yapisi()
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.santiye_raporu_gorme:
                profile =  get_object_or_404(bloglar,proje_ait_bilgisi = request.user.kullanicilar_db,id = id )
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        profile =  get_object_or_404(bloglar,proje_ait_bilgisi = users,id = id )
    content["santiye"] = profile
    return render(request,"santiye_yonetimi/santiye_raporu.html",content)

def santiye_raporu(request,id):
    content = sozluk_yapisi()
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.santiye_raporu_gorme:
                profile =  get_object_or_404(bloglar,proje_ait_bilgisi = request.user.kullanicilar_db,id = id )
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        profile =  get_object_or_404(bloglar,proje_ait_bilgisi = request.user,id = id )
    content["santiye"] = profile
    return render(request,"santiye_yonetimi/santiye_raporu.html",content)
def santiye_raporu_rapor_gonderme(request,id):
    content = sozluk_yapisi()
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.santiye_raporu_gorme:
                profile =  get_object_or_404(bloglar,proje_ait_bilgisi = request.user.kullanicilar_db,id = id )
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    else:
        profile =  get_object_or_404(bloglar,proje_ait_bilgisi = request.user,id = id )
    content["santiye"] = profile
    return render(request,"santiye_yonetimi/santiye_raporu_2.html",content)
def kullanici_yetkileri(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            return redirect_with_language("main:yetkisiz")
        profile = personel_izinleri.objects.filter(izinlerin_sahibi_kullanici = request.user)
        content["izinler"] = profile
    return render(request,"kullanici_yetkileri/yetkiler.html",content)
def kullanici_yetkileri_duzenle(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            return redirect_with_language("main:yetkisiz")
        profile = personel_izinleri.objects.filter(izinlerin_sahibi_kullanici = request.user)
        content["izinler"] = profile
        content["secili_grup"] = get_object_or_404(personel_izinleri,id = id)
    return render(request,"kullanici_yetkileri/yetkiler.html",content)
def kullanici_yetki_olustur(request):
    if request.user.kullanicilar_db:
        return redirect_with_language("main:yetkisiz")
    if request.POST:
        grup_adi = request.POST.get("grup_adi")
        personel_izinleri.objects.create(
            isim = grup_adi,
            izinlerin_sahibi_kullanici = request.user
        )
    return redirect_with_language("main:kullanici_yetkileri")
def kullanici_yetki_adi_duzenle(request):
    if request.user.kullanicilar_db:
        return redirect_with_language("main:yetkisiz")
    if request.POST:
        grup_adi = request.POST.get("grup_adi")
        id  = request.POST.get("id")
        personel_izinleri.objects.filter(id = id).update(
            isim = grup_adi,
            izinlerin_sahibi_kullanici = request.user
        )
    return redirect_with_language("main:kullanici_yetkileri")
def kullanici_yetki_sil(request):
    if request.user.kullanicilar_db:
        return redirect_with_language("main:yetkisiz")
    if request.POST:
        id  = request.POST.get("id")
        personel_izinleri.objects.filter(id = id,izinlerin_sahibi_kullanici = request.user).delete()
    return redirect_with_language("main:kullanici_yetkileri")
def kullanici_yetki_alma(request):
    if request.user.kullanicilar_db:
        return redirect_with_language("main:yetkisiz")
    if request.POST:
        guncellenen = request.POST.get("guncellenen")
        izinler = get_object_or_404(personel_izinleri,id = guncellenen) 
        id_bilgiis = izinler.id
        ##
        #izinleri Sıfırla
        izinler.dashboard_gorme = False
        izinler.dashboard_silme = False
        izinler.dashboard_duzenleme = False
        izinler.dashboard_olusturma = False
        #
        izinler.gelir_ozeti_gorme = False
        izinler.gelir_ozeti_olusturma = False
        izinler.gelir_ozeti_duzenleme =False
        izinler.gelir_ozeti_silme = False
        #
        izinler.gider_ozeti_gorme = False
        izinler.gider_ozeti_olusturma = False
        izinler.gider_ozeti_duzenleme = False
        izinler.gider_ozeti_silme = False
        #
        izinler.hesap_ekstra_gorme = False
        izinler.hesap_ekstra_olusturma = False
        izinler.hesap_ekstra_duzenleme = False
        izinler.hesap_ekstra_silme = False
        #
        izinler.ilerleme_takibi_gorme = False
        izinler.ilerleme_takibi_olusturma = False
        izinler.ilerleme_takibi_duzenleme = False
        izinler.ilerleme_takibi_silme = False
        #
        izinler.sozlesmeler_gorme = False
        izinler.sozlesmeler_olusturma = False
        izinler.sozlesmeler_duzenleme = False
        izinler.sozlesmeler_silme = False
        #
        izinler.yapilacaklar_gorme = False
        izinler.yapilacaklar_olusturma = False
        izinler.yapilacaklar_duzenleme = False
        izinler.yapilacaklar_silme = False
        #
        izinler.dosya_yoneticisi_gorme = False
        izinler.dosya_yoneticisi_olusturma = False
        izinler.dosya_yoneticisi_duzenleme = False
        izinler.dosya_yoneticisi_silme = False
        #
        izinler.projeler_gorme = False
        izinler.projeler_olusturma = False
        izinler.projeler_duzenleme = False
        izinler.projeler_silme =False
        #
        izinler.personeller_gorme = False
        izinler.personeller_olusturma = False
        izinler.personeller_duzenleme = False
        izinler.personeller_silme = False
        #
        izinler.gelir_faturasi_gorme_izni = False
        izinler.gelir_faturasi_kesme_izni = False
        izinler.gelir_faturasi_duzenleme_izni = False
        izinler.gelir_faturasi_silme_izni = False
        #
        izinler.gider_faturasi_gorme_izni = False
        izinler.gider_faturasi_kesme_izni = False
        izinler.gider_faturasi_duzenleme_izni = False
        izinler.gider_faturasi_silme_izni = False
        #
        izinler.kasa_virman_olusturma_izni = False
        izinler.kasa_virman_gorme_izni = False
        #
        izinler.kasa_gosterme_izni = False
        izinler.kasa_olusturma_izni = False
        izinler.kasa_guncelleme_izni = False
        izinler.Kasa_silme_izni = False
        #
        izinler.cari_gosterme_izni = False
        izinler.cari_olusturma = False
        izinler.cari_guncelleme_izni = False
        izinler.cari_silme_izni = False
        #
        izinler.personeller_gorme = False
        izinler.personeller_olusturma = False
        izinler.personeller_duzenleme = False
        izinler.personeller_silme = False
        #
        izinler.santiye_olusturma = False
        izinler.santiye_silme = False
        izinler.santiye_gorme = False
        izinler.santiye_duzenleme = False
        #
        izinler.blog_olusturma = False
        izinler.blog_silme = False
        izinler.blog_gorme = False
        izinler.blog_duzenleme = False
        #
        izinler.kalemleri_olusturma = False
        izinler.kalemleri_silme = False
        izinler.kalemleri_gorme = False
        izinler.kalemleri_duzenleme = False
        #
        izinler.santiye_raporu_olusturma = False
        izinler.santiye_raporu_silme = False
        izinler.santiye_raporu_gorme = False
        izinler.santiye_raporu_duzenleme = False
        #
        izinler.taseronlar_gorme = False
        izinler.taseronlar_olusturma = False
        izinler.taseronlar_duzenleme = False
        izinler.taseronlar_silme =False
        #
        izinler.hakedisler_gorme = False
        izinler.hakedisler_olusturma = False
        izinler.hakedisler_duzenleme = False
        izinler.hakedisler_silme = False
        #
        izinler.kasa_detay_izni = False
        izinler.cari_detay_izni = False
        ##
        izinler.gelir_kategorisi_gorme = False
        izinler.gelir_kategorisi_olusturma = False
        izinler.gelir_kategorisi_guncelleme = False
        izinler.gelir_kategorisi_silme = False
        #
        izinler.gider_kategorisi_gorme = False
        izinler.gider_kategorisi_olusturma = False
        izinler.gider_kategorisi_guncelleme = False
        izinler.gider_kategorisi_silme = False
        #
        izinler.gelir_etiketi_gorme = False
        izinler.gelir_etiketi_olusturma = False
        izinler.gelir_etiketi_guncelleme = False
        izinler.gelir_etiketi_silme = False
        #
        izinler.gider_etiketi_gorme = False
        izinler.gider_etiketi_olusturma = False
        izinler.gider_etiketi_guncelleme = False
        izinler.gider_etiketi_silme = False
        #
        izinler.urun_gorme = False
        izinler.urun_olusturma = False
        izinler.urun_guncelleme = False
        izinler.urun_silme = False
        #
        izinler.muhasabe_ayarlari_gorme = False
        izinler.muhasabe_ayarlari_guncelleme = False
        #
        izinler.gelir_faturasi_makbuz_gorme_izni = False
        izinler.gelir_faturasi_makbuz_kesme_izni = False
        izinler.gelir_faturasi_makbuz_duzenleme_izni = False
        izinler.gelir_faturasi_makbuz_silme_izni = False
        #
        izinler.gider_faturasi_makbuz_gorme_izni = False
        izinler.gider_faturasi_makbuz_kesme_izni = False
        izinler.gider_faturasi_makbuz_duzenleme_izni = False
        izinler.gider_faturasi_makbuz_silme_izni = False
        #Puantaj
        izinler.personeller_puantaj_olusturma = False
        izinler.personeller_puantaj_silme = False
        izinler.personeller_puantaj_gorme = False
        izinler.personeller_puantaj_duzenleme = False
        #
        # 
        izinler.personeller_odeme_olusturma = False
        izinler.personeller_odeme_silme = False
        izinler.personeller_odeme_gorme = False
        izinler.personeller_odeme_duzenleme = False
        #
        izinler.katman_olusturma = False
        izinler.katman_silme = False
        izinler.katman_gorme = False
        izinler.katman_duzenleme = False
        #
        izinler.gant_olusturma = False
        izinler.gant_gorme = False
        izinler.gant_duzenleme = False
        izinler.gant_silme = False
        #
        izinler.genel_rapor_olusturma = False
        izinler.genel_rapor_gorme = False
        izinler.genel_rapor_duzenleme = False
        izinler.genel_rapor_silme = False
        izinler.genel_rapor_onaylama = False
        izinler.satin_alma_talebi_olusturma = False
        izinler.satin_alma_talebi_silme = False
        izinler.satin_alma_talebi_gorme = False
        izinler.satin_alma_talebi_duzenleme = False

        izinler.satin_alma_talebi_onaylama_olusturma = False
        izinler.satin_alma_talebi_onaylama_silme = False
        izinler.satin_alma_talebi_onaylama_gorme = False
        izinler.satin_alma_talebi_onaylama_duzenleme = False

        izinler.stok_olusturma = False

        izinler.stok_talebi_onaylama_silme = False
        izinler.stok_talebi_onaylama_gorme = False
        izinler.stok_talebi_onaylama_duzenleme = False

        izinler.zimmet_olusturma = False
        izinler.zimmet_silme = False
        izinler.zimmet_gorme = False

        izinler.rapor_olusturucu_gorme = False
        izinler.rapor_olusturucu_olusturma = False

        izinler.musteri_olusturma = False
        izinler.musteri_gorme = False
        izinler.musteri_duzenleme = False
        izinler.musteri_silme = False
        izinler.crm_musteri_olusturma = False
        izinler.crm_musteri_silme = False
        izinler.crm_musteri_gorme = False
        izinler.crm_musteri_duzenleme = False
        izinler.crm_talep_olusturma = False
        izinler.crm_talep_silme = False
        izinler.crm_talep_gorme = False
        izinler.crm_talep_duzenleme = False
        izinler.crm_teklif_olusturma = False
        izinler.crm_teklif_silme = False
        izinler.crm_teklif_gorme = False
        izinler.crm_teklif_duzenleme = False
        izinler.crm_daire_olusturma = False
        izinler.crm_daire_silme = False
        izinler.crm_daire_gorme = False
        izinler.crm_daire_duzenleme = False
        izinler.crm_evrak_olusturma = False
        izinler.crm_evrak_silme = False
        izinler.crm_evrak_gorme = False
        izinler.crm_evrak_duzenleme = False

        izinler.ust_yuklenici_gorme = False
        izinler.ust_yuklenici_olusturma = False
        izinler.ust_yuklenici_duzenleme = False
        izinler.ust_yuklenici_silme = False


        izinler.save()
        ##
        personeller_puantaj_olusturma = request.POST.get("personeller_puantaj_olusturma")
        if personeller_puantaj_olusturma:
            izinler.personeller_puantaj_olusturma = True
        personeller_puantaj_silme = request.POST.get("personeller_puantaj_silme")
        if personeller_puantaj_silme:
            izinler.personeller_puantaj_silme = True
        personeller_puantaj_gorme = request.POST.get("personeller_puantaj_gorme")
        if personeller_puantaj_gorme:
            izinler.personeller_puantaj_gorme = True
        personeller_puantaj_duzenleme = request.POST.get("personeller_puantaj_duzenleme")
        if personeller_puantaj_duzenleme:
            izinler.personeller_puantaj_duzenleme = True
        #
        personeller_odeme_olusturma = request.POST.get("personeller_odeme_olusturma")
        if personeller_odeme_olusturma:
            izinler.personeller_odeme_olusturma = True
        personeller_odeme_silme = request.POST.get("personeller_odeme_silme")
        if personeller_odeme_silme:
            izinler.personeller_odeme_silme = True
        personeller_odeme_gorme = request.POST.get("personeller_odeme_gorme")
        if personeller_odeme_gorme:
            izinler.personeller_odeme_gorme = True
        personeller_odeme_duzenleme = request.POST.get("personeller_odeme_duzenleme")
        if personeller_odeme_duzenleme:
            izinler.personeller_odeme_duzenleme = True
        #
        katman_olusturma = request.POST.get("katman_olusturma")
        if katman_olusturma:
            izinler.katman_olusturma = True
        katman_silme = request.POST.get("katman_silme")
        if katman_silme:
            izinler.katman_silme = True
        katman_gorme = request.POST.get("katman_gorme")
        if katman_gorme:
            izinler.katman_gorme = True
        katman_duzenleme = request.POST.get("katman_duzenleme")
        if katman_duzenleme:
            izinler.katman_duzenleme = True
        ##
        dashboard_gorme = request.POST.get("dashboard_gorme")
        if dashboard_gorme:
            izinler.dashboard_gorme = True
        dashboard_silme = request.POST.get("dashboard_silme")
        if dashboard_silme:
            izinler.dashboard_silme = True
        dashboard_duzenleme = request.POST.get("dashboard_duzenleme")
        if dashboard_duzenleme:
            izinler.dashboard_duzenleme = True
        dashboard_olusturma = request.POST.get("dashboard_olusturma")
        if dashboard_olusturma:
            izinler.dashboard_olusturma = True
        #
        gelir_ozeti_gorme = request.POST.get("gelir_ozeti_gorme")
        if gelir_ozeti_gorme:
            izinler.gelir_ozeti_gorme = True
        gelir_ozeti_olusturma = request.POST.get("gelir_ozeti_olusturma")
        if gelir_ozeti_olusturma:
            izinler.gelir_ozeti_olusturma = True
        gelir_ozeti_duzenleme = request.POST.get("gelir_ozeti_duzenleme")
        if gelir_ozeti_duzenleme:
            izinler.gelir_ozeti_duzenleme = True
        gelir_ozeti_silme = request.POST.get("gelir_ozeti_silme")
        if gelir_ozeti_silme:
            izinler.gelir_ozeti_silme = True
        #
        gider_ozeti_gorme = request.POST.get("gider_ozeti_gorme")
        if gider_ozeti_gorme:
            izinler.gider_ozeti_gorme = True
        gider_ozeti_olusturma = request.POST.get("gider_ozeti_olusturma")
        if gider_ozeti_olusturma:
            izinler.gider_ozeti_olusturma = True
        gider_ozeti_duzenleme = request.POST.get("gider_ozeti_duzenleme")
        if gider_ozeti_duzenleme:
            izinler.gider_ozeti_duzenleme = True
        gider_ozeti_silme = request.POST.get("gider_ozeti_silme")
        if gider_ozeti_silme:
            izinler.gider_ozeti_silme = True
        #
        hesap_ekstra_gorme = request.POST.get("hesap_ekstra_gorme")
        if hesap_ekstra_gorme:
            izinler.hesap_ekstra_gorme = True
        hesap_ekstra_olusturma = request.POST.get("hesap_ekstra_olusturma")
        if hesap_ekstra_olusturma:
            izinler.hesap_ekstra_olusturma = True
        hesap_ekstra_duzenleme = request.POST.get("hesap_ekstra_duzenleme")
        if hesap_ekstra_duzenleme:
            izinler.hesap_ekstra_duzenleme = True
        hesap_ekstra_silme = request.POST.get("hesap_ekstra_silme")
        if hesap_ekstra_silme:
            izinler.hesap_ekstra_silme = True
        #
        kasa_virman_olusturma_izni = request.POST.get("kasa_virman_olusturma_izni")
        if kasa_virman_olusturma_izni:
            izinler.kasa_virman_olusturma_izni = True
        kasa_virman_gorme_izni = request.POST.get("kasa_virman_gorme_izni")
        if kasa_virman_gorme_izni:
            izinler.kasa_virman_gorme_izni = True
        #
        ilerleme_takibi_gorme = request.POST.get("ilerleme_takibi_gorme")
        if ilerleme_takibi_gorme:
            izinler.ilerleme_takibi_gorme = True
        ilerleme_takibi_olusturma = request.POST.get("ilerleme_takibi_olusturma")
        if ilerleme_takibi_olusturma:
            izinler.ilerleme_takibi_olusturma = True
        ilerleme_takibi_duzenleme = request.POST.get("ilerleme_takibi_duzenleme")
        if ilerleme_takibi_duzenleme:
            izinler.ilerleme_takibi_duzenleme = True
        ilerleme_takibi_silme = request.POST.get("ilerleme_takibi_silme")
        if ilerleme_takibi_silme:
            izinler.ilerleme_takibi_silme = True
        #
        sozlesmeler_gorme = request.POST.get("sozlesmeler_gorme")
        if sozlesmeler_gorme:
            izinler.sozlesmeler_gorme = True
        sozlesmeler_olusturma = request.POST.get("sozlesmeler_olusturma")
        if sozlesmeler_olusturma:
            izinler.sozlesmeler_olusturma = True
        sozlesmeler_duzenleme = request.POST.get("sozlesmeler_duzenleme")
        if sozlesmeler_duzenleme:
            izinler.sozlesmeler_duzenleme = True
        sozlesmeler_silme = request.POST.get("sozlesmeler_silme")
        if sozlesmeler_silme:
            izinler.sozlesmeler_silme = True
        #
        yapilacaklar_gorme = request.POST.get("yapilacaklar_gorme")
        if yapilacaklar_gorme:
            izinler.yapilacaklar_gorme = True
        yapilacaklar_olusturma = request.POST.get("yapilacaklar_olusturma")
        if yapilacaklar_olusturma:
            izinler.yapilacaklar_olusturma = True
        yapilacaklar_duzenleme = request.POST.get("yapilacaklar_duzenleme")
        if yapilacaklar_duzenleme:
            izinler.yapilacaklar_duzenleme = True
        yapilacaklar_silme = request.POST.get("yapilacaklar_silme")
        if yapilacaklar_silme:
            izinler.yapilacaklar_silme = True
        #
        dosya_yoneticisi_gorme = request.POST.get("dosya_yoneticisi_gorme")
        if dosya_yoneticisi_gorme:
            izinler.dosya_yoneticisi_gorme = True
        dosya_yoneticisi_olusturma = request.POST.get("dosya_yoneticisi_olusturma")
        if dosya_yoneticisi_olusturma:
            izinler.dosya_yoneticisi_olusturma = True
        dosya_yoneticisi_duzenleme = request.POST.get("dosya_yoneticisi_duzenleme")
        if dosya_yoneticisi_duzenleme:
            izinler.dosya_yoneticisi_duzenleme = True
        dosya_yoneticisi_silme = request.POST.get("dosya_yoneticisi_silme")
        if dosya_yoneticisi_silme:
            izinler.dosya_yoneticisi_silme = True
        #
        projeler_gorme = request.POST.get("projeler_gorme")
        if projeler_gorme:
            izinler.projeler_gorme = True
        projeler_olusturma = request.POST.get("projeler_olusturma")
        if projeler_olusturma:
            izinler.projeler_olusturma = True
        projeler_duzenleme = request.POST.get("projeler_duzenleme")
        if projeler_duzenleme:
            izinler.projeler_duzenleme = True
        projeler_silme = request.POST.get("projeler_silme")
        if projeler_silme:
            izinler.projeler_silme = True

        #
        personeller_gorme = request.POST.get("personeller_gorme")
        if personeller_gorme:
            izinler.personeller_gorme = True
        personeller_olusturma = request.POST.get("personeller_olusturma")
        if personeller_olusturma:
            izinler.personeller_olusturma = True
        personeller_duzenleme = request.POST.get("personeller_duzenleme")
        if personeller_duzenleme:
            izinler.personeller_duzenleme = True
        personeller_silme = request.POST.get("personeller_silme")
        if personeller_silme:
            izinler.personeller_silme = True
        #
        gelir_faturasi_gorme_izni = request.POST.get("gelir_faturasi_gorme_izni")
        if gelir_faturasi_gorme_izni:
            izinler.gelir_faturasi_gorme_izni = True
        gelir_faturasi_kesme_izni = request.POST.get("gelir_faturasi_kesme_izni")
        if gelir_faturasi_kesme_izni:
            izinler.gelir_faturasi_kesme_izni = True
        gelir_faturasi_duzenleme_izni = request.POST.get("gelir_faturasi_duzenleme_izni")
        if gelir_faturasi_duzenleme_izni:
            izinler.gelir_faturasi_duzenleme_izni = True
        gelir_faturasi_silme_izni = request.POST.get("gelir_faturasi_silme_izni")
        if gelir_faturasi_silme_izni:
            izinler.gelir_faturasi_silme_izni = True
        #
        gider_faturasi_gorme_izni = request.POST.get("gider_faturasi_gorme_izni")
        if gider_faturasi_gorme_izni:
            izinler.gider_faturasi_gorme_izni = True
        gider_faturasi_kesme_izni = request.POST.get("gider_faturasi_kesme_izni")
        if gider_faturasi_kesme_izni:
            izinler.gider_faturasi_kesme_izni = True
        gider_faturasi_duzenleme_izni = request.POST.get("gider_faturasi_duzenleme_izni")
        if gider_faturasi_duzenleme_izni:
            izinler.gider_faturasi_duzenleme_izni = True
        gider_faturasi_silme_izni = request.POST.get("gider_faturasi_silme_izni")
        if gider_faturasi_silme_izni:
            izinler.gider_faturasi_silme_izni = True
        #
        
        kasa_gosterme_izni = request.POST.get("kasa_gosterme_izni")
        if kasa_gosterme_izni:
            izinler.kasa_gosterme_izni = True
        kasa_olusturma_izni = request.POST.get("kasa_olusturma_izni")
        if kasa_olusturma_izni:
            izinler.kasa_olusturma_izni = True
        kasa_guncelleme_izni = request.POST.get("kasa_guncelleme_izni")
        if kasa_guncelleme_izni:
            izinler.kasa_guncelleme_izni = True
        Kasa_silme_izni = request.POST.get("Kasa_silme_izni")
        if Kasa_silme_izni:
            izinler.Kasa_silme_izni = True
        #
        cari_gosterme_izni = request.POST.get("cari_gosterme_izni")
        if cari_gosterme_izni:
            izinler.cari_gosterme_izni = True
        cari_olusturma = request.POST.get("cari_olusturma")
        if cari_olusturma:
            izinler.cari_olusturma = True
        cari_guncelleme_izni = request.POST.get("cari_guncelleme_izni")
        if cari_guncelleme_izni:
            izinler.cari_guncelleme_izni = True
        cari_silme_izni = request.POST.get("cari_silme_izni")
        if cari_silme_izni:
            izinler.cari_silme_izni = True
        #
        personeller_gorme = request.POST.get("personeller_gorme")
        if personeller_gorme:
            izinler.personeller_gorme = True
        personeller_olusturma = request.POST.get("personeller_olusturma")
        if personeller_olusturma:
            izinler.personeller_olusturma = True
        personeller_duzenleme = request.POST.get("personeller_duzenleme")
        if personeller_duzenleme:
            izinler.personeller_duzenleme = True
        personeller_silme = request.POST.get("personeller_silme")
        if personeller_silme:
            izinler.personeller_silme = True
        #
        santiye_gorme = request.POST.get("santiye_gorme")
        if santiye_gorme:
            izinler.santiye_gorme = True
        santiye_olusturma  = request.POST.get("santiye_olusturma")
        if santiye_olusturma:
            izinler.santiye_olusturma = True
        santiye_duzenleme = request.POST.get("santiye_duzenleme")
        if santiye_duzenleme:
            izinler.santiye_duzenleme = True
        santiye_silme = request.POST.get("santiye_silme")
        if santiye_silme:
            izinler.santiye_silme = True
        #
        blog_gorme = request.POST.get("blog_gorme")
        if blog_gorme:
            izinler.blog_gorme = True
        blog_olusturma  = request.POST.get("blog_olusturma")
        if blog_olusturma:
            izinler.blog_olusturma = True
        blog_duzenleme = request.POST.get("blog_duzenleme")
        if blog_duzenleme:
            izinler.blog_duzenleme = True
        blog_silme = request.POST.get("blog_silme")
        if blog_silme:
            izinler.blog_silme = True
        #
        kalemleri_gorme = request.POST.get("kalemleri_gorme")
        if kalemleri_gorme:
            izinler.kalemleri_gorme = True
        kalemleri_olusturma  = request.POST.get("kalemleri_olusturma")
        if kalemleri_olusturma:
            izinler.kalemleri_olusturma = True
        kalemleri_duzenleme = request.POST.get("kalemleri_duzenleme")
        if kalemleri_duzenleme:
            izinler.kalemleri_duzenleme = True
        kalemleri_silme = request.POST.get("kalemleri_silme")
        if kalemleri_silme:
            izinler.kalemleri_silme = True
        #
        santiye_raporu_gorme = request.POST.get("santiye_raporu_gorme")
        if santiye_raporu_gorme:
            izinler.santiye_raporu_gorme = True
        santiye_raporu_olusturma  = request.POST.get("santiye_raporu_olusturma")
        if santiye_raporu_olusturma:
            izinler.santiye_raporu_olusturma = True
        santiye_raporu_duzenleme = request.POST.get("santiye_raporu_duzenleme")
        if santiye_raporu_duzenleme:
            izinler.santiye_raporu_duzenleme = True
        santiye_raporu_silme = request.POST.get("santiye_raporu_silme")
        if santiye_raporu_silme:
            izinler.santiye_raporu_silme = True
        #
        taseronlar_gorme = request.POST.get("taseronlar_gorme")
        if taseronlar_gorme:
            izinler.taseronlar_gorme = True
        taseronlar_olusturma = request.POST.get("taseronlar_olusturma")
        if taseronlar_olusturma:
            izinler.taseronlar_olusturma = True
        taseronlar_duzenleme = request.POST.get("taseronlar_duzenleme")
        if taseronlar_duzenleme:
            izinler.taseronlar_duzenleme = True
        taseronlar_silme = request.POST.get("taseronlar_silme")
        if taseronlar_silme:
            izinler.taseronlar_silme = True
        #
        hakedisler_gorme = request.POST.get("hakedisler_gorme")
        if hakedisler_gorme:
            izinler.hakedisler_gorme = True
        hakedisler_olusturma = request.POST.get("hakedisler_olusturma")
        if hakedisler_olusturma:
            izinler.hakedisler_olusturma = True
        hakedisler_duzenleme = request.POST.get("hakedisler_duzenleme")
        if hakedisler_duzenleme:
            izinler.hakedisler_duzenleme = True
        hakedisler_silme = request.POST.get("hakedisler_silme")
        if hakedisler_silme:
            izinler.hakedisler_silme = True
        #
        kasa_detay_izni = request.POST.get("kasa_detay_izni")
        if kasa_detay_izni:
            izinler.kasa_detay_izni = True
        cari_detay_izni = request.POST.get("cari_detay_izni")
        if cari_detay_izni:
            izinler.cari_detay_izni = True
        
        #
        gelir_kategorisi_gorme = request.POST.get("gelir_kategorisi_gorme")
        if gelir_kategorisi_gorme:  
            izinler.gelir_kategorisi_gorme = True
        gelir_kategorisi_olusturma = request.POST.get("gelir_kategorisi_olusturma")
        if gelir_kategorisi_olusturma : 
            izinler.gelir_kategorisi_olusturma = True
        gelir_kategorisi_guncelleme = request.POST.get("gelir_kategorisi_guncelleme")
        if gelir_kategorisi_guncelleme : 
            izinler.gelir_kategorisi_guncelleme = True
        gelir_kategorisi_silme = request.POST.get("gelir_kategorisi_silme")
        if gelir_kategorisi_silme : 
            izinler.gelir_kategorisi_silme = True
        
        #
        gider_kategorisi_gorme = request.POST.get("gider_kategorisi_gorme")
        if gider_kategorisi_gorme:  
            izinler.gider_kategorisi_gorme = True
        gider_kategorisi_olusturma = request.POST.get("gider_kategorisi_olusturma")
        if gider_kategorisi_olusturma : 
            izinler.gider_kategorisi_olusturma = True
        gider_kategorisi_guncelleme = request.POST.get("gider_kategorisi_guncelleme")
        if gider_kategorisi_guncelleme : 
            izinler.gider_kategorisi_guncelleme = True
        gider_kategorisi_silme = request.POST.get("gider_kategorisi_silme")
        if gider_kategorisi_silme : 
            izinler.gider_kategorisi_silme = True
        #
        #
        gelir_etiketi_gorme = request.POST.get("gelir_etiketi_gorme")
        if gelir_etiketi_gorme:  
            izinler.gelir_etiketi_gorme = True
        gelir_etiketi_olusturma = request.POST.get("gelir_etiketi_olusturma")
        if gelir_etiketi_olusturma : 
            izinler.gelir_etiketi_olusturma = True
        gelir_etiketi_guncelleme = request.POST.get("gelir_etiketi_guncelleme")
        if gelir_etiketi_guncelleme : 
            izinler.gelir_etiketi_guncelleme = True
        gelir_etiketi_silme = request.POST.get("gelir_etiketi_silme")
        if gelir_etiketi_silme : 
            izinler.gelir_etiketi_silme = True

        #
        gider_etiketi_gorme = request.POST.get("gider_etiketi_gorme")
        if gider_etiketi_gorme:  
            izinler.gider_etiketi_gorme = True
        gider_etiketi_olusturma = request.POST.get("gider_etiketi_olusturma")
        if gider_etiketi_olusturma : 
            izinler.gider_etiketi_olusturma = True
        gider_etiketi_guncelleme = request.POST.get("gider_etiketi_guncelleme")
        if gider_etiketi_guncelleme : 
            izinler.gider_etiketi_guncelleme = True
        gider_etiketi_silme = request.POST.get("gider_etiketi_silme")
        if gider_etiketi_silme : 
            izinler.gider_etiketi_silme = True
        #
        #
        urun_gorme = request.POST.get("urun_gorme")
        if urun_gorme:  
            izinler.urun_gorme = True
        urun_olusturma = request.POST.get("urun_olusturma")
        if urun_olusturma : 
            izinler.urun_olusturma = True
        urun_guncelleme = request.POST.get("urun_guncelleme")
        if urun_guncelleme : 
            izinler.urun_guncelleme = True
        urun_silme = request.POST.get("urun_silme")
        if urun_silme : 
            izinler.urun_silme = True

        #
        muhasabe_ayarlari_gorme = request.POST.get("muhasabe_ayarlari_gorme")
        if muhasabe_ayarlari_gorme : 
            izinler.muhasabe_ayarlari_gorme = True
        muhasabe_ayarlari_guncelleme = request.POST.get("muhasabe_ayarlari_guncelleme")
        if muhasabe_ayarlari_guncelleme : 
            izinler.muhasabe_ayarlari_guncelleme = True

        #
        gelir_faturasi_makbuz_gorme_izni = request.POST.get("gelir_faturasi_makbuz_gorme_izni")
        if gelir_faturasi_makbuz_gorme_izni:  
            izinler.gelir_faturasi_makbuz_gorme_izni = True
        gelir_faturasi_makbuz_kesme_izni = request.POST.get("gelir_faturasi_makbuz_kesme_izni")
        if gelir_faturasi_makbuz_kesme_izni : 
            izinler.gelir_faturasi_makbuz_kesme_izni = True
        gelir_faturasi_makbuz_duzenleme_izni = request.POST.get("gelir_faturasi_makbuz_duzenleme_izni")
        if gelir_faturasi_makbuz_duzenleme_izni : 
            izinler.gelir_faturasi_makbuz_duzenleme_izni = True
        gelir_faturasi_makbuz_silme_izni = request.POST.get("gelir_faturasi_makbuz_silme_izni")
        if gelir_faturasi_makbuz_silme_izni : 
            izinler.gelir_faturasi_makbuz_silme_izni = True
        #
        gider_faturasi_makbuz_gorme_izni = request.POST.get("gider_faturasi_makbuz_gorme_izni")
        if gider_faturasi_makbuz_gorme_izni:  
            izinler.gider_faturasi_makbuz_gorme_izni = True
        gider_faturasi_makbuz_kesme_izni = request.POST.get("gider_faturasi_makbuz_kesme_izni")
        if gider_faturasi_makbuz_kesme_izni : 
            izinler.gider_faturasi_makbuz_kesme_izni = True
        gider_faturasi_makbuz_duzenleme_izni = request.POST.get("gider_faturasi_makbuz_duzenleme_izni")
        if gider_faturasi_makbuz_duzenleme_izni : 
            izinler.gider_faturasi_makbuz_duzenleme_izni = True
        gider_faturasi_makbuz_silme_izni = request.POST.get("gider_faturasi_makbuz_silme_izni")
        if gider_faturasi_makbuz_silme_izni : 
            izinler.gider_faturasi_makbuz_silme_izni = True
        
        #
        gant_olusturma = request.POST.get("gant_olusturma")
        if gant_olusturma:  
            izinler.gant_olusturma = True
        gant_gorme = request.POST.get("gant_gorme")
        if gant_gorme : 
            izinler.gant_gorme = True
        gant_duzenleme = request.POST.get("gant_duzenleme")
        if gant_duzenleme : 
            izinler.gant_duzenleme = True
        gant_silme = request.POST.get("gant_silme")
        if gant_silme :
            izinler.gant_silme = True

        #
        genel_rapor_olusturma = request.POST.get("genel_rapor_olusturma")
        if genel_rapor_olusturma:  
            izinler.genel_rapor_olusturma = True
        genel_rapor_gorme = request.POST.get("genel_rapor_gorme")
        if genel_rapor_gorme : 
            izinler.genel_rapor_gorme = True
        genel_rapor_duzenleme = request.POST.get("genel_rapor_duzenleme")
        if genel_rapor_duzenleme : 
            izinler.genel_rapor_duzenleme = True
        genel_rapor_silme = request.POST.get("genel_rapor_silme")
        if genel_rapor_silme : 
            izinler.genel_rapor_silme = True
        genel_rapor_onaylama = request.POST.get("genel_rapor_onaylama")
        if genel_rapor_onaylama : 
            izinler.genel_rapor_onaylama = True
        #stok satın alma 
        satin_alma_talebi_olusturma = request.POST.get("satin_alma_talebi_olusturma")
        if satin_alma_talebi_olusturma:
            izinler.satin_alma_talebi_olusturma = True

        satin_alma_talebi_silme = request.POST.get("satin_alma_talebi_silme")
        if satin_alma_talebi_silme:
            izinler.satin_alma_talebi_silme = True

        satin_alma_talebi_gorme = request.POST.get("satin_alma_talebi_gorme")
        if satin_alma_talebi_gorme:
            izinler.satin_alma_talebi_gorme = True

        satin_alma_talebi_duzenleme = request.POST.get("satin_alma_talebi_duzenleme")
        if satin_alma_talebi_duzenleme:
            izinler.satin_alma_talebi_duzenleme = True

        satin_alma_talebi_onaylama_olusturma = request.POST.get("satin_alma_talebi_onaylama_olusturma")
        if satin_alma_talebi_onaylama_olusturma:
            izinler.satin_alma_talebi_onaylama_olusturma = True

        satin_alma_talebi_onaylama_silme = request.POST.get("satin_alma_talebi_onaylama_silme")
        if satin_alma_talebi_onaylama_silme:
            izinler.satin_alma_talebi_onaylama_silme = True

        satin_alma_talebi_onaylama_gorme = request.POST.get("satin_alma_talebi_onaylama_gorme")
        if satin_alma_talebi_onaylama_gorme:
            izinler.satin_alma_talebi_onaylama_gorme = True

        satin_alma_talebi_onaylama_duzenleme = request.POST.get("satin_alma_talebi_onaylama_duzenleme")
        if satin_alma_talebi_onaylama_duzenleme:
            izinler.satin_alma_talebi_onaylama_duzenleme = True

        stok_olusturma = request.POST.get("stok_olusturma")
        if stok_olusturma:
            izinler.stok_olusturma = True

        stok_talebi_onaylama_silme = request.POST.get("stok_talebi_onaylama_silme")
        if stok_talebi_onaylama_silme:
            izinler.stok_talebi_onaylama_silme = True

        stok_talebi_onaylama_gorme = request.POST.get("stok_talebi_onaylama_gorme")
        if stok_talebi_onaylama_gorme:
            izinler.stok_talebi_onaylama_gorme = True

        stok_talebi_onaylama_duzenleme = request.POST.get("stok_talebi_onaylama_duzenleme")
        if stok_talebi_onaylama_duzenleme:
            izinler.stok_talebi_onaylama_duzenleme = True

        zimmet_olusturma = request.POST.get("zimmet_olusturma")
        if zimmet_olusturma:
            izinler.zimmet_olusturma = True

        zimmet_silme = request.POST.get("zimmet_silme")
        if zimmet_silme:
            izinler.zimmet_silme = True

        zimmet_gorme = request.POST.get("zimmet_gorme")
        if zimmet_gorme:
            izinler.zimmet_gorme = True

        rapor_olusturucu_gorme = request.POST.get("rapor_olusturucu_gorme")
        if rapor_olusturucu_gorme:
            izinler.rapor_olusturucu_gorme = True

        rapor_olusturucu_olusturma = request.POST.get("rapor_olusturucu_olusturma")
        if rapor_olusturucu_olusturma:
            izinler.rapor_olusturucu_olusturma = True

        musteri_olusturma = request.POST.get("musteri_olusturma")
        if musteri_olusturma:
            izinler.musteri_olusturma = True
        musteri_gorme = request.POST.get("musteri_gorme")
        if musteri_gorme:
            izinler.musteri_gorme = True
        musteri_duzenleme = request.POST.get("musteri_duzenleme")
        if musteri_duzenleme:
            izinler.musteri_duzenleme = True
        musteri_silme = request.POST.get("musteri_silme")
        if musteri_silme:
            izinler.musteri_silme = True
        crm_musteri_olusturma = request.POST.get("crm_musteri_olusturma")
        if crm_musteri_olusturma:
            izinler.crm_musteri_olusturma = True
        crm_musteri_silme = request.POST.get("crm_musteri_silme")
        if crm_musteri_silme:
            izinler.crm_musteri_silme = True
        crm_musteri_gorme = request.POST.get("crm_musteri_gorme")
        if crm_musteri_gorme:
            izinler.crm_musteri_gorme = True
        crm_musteri_duzenleme = request.POST.get("crm_musteri_duzenleme")
        if crm_musteri_duzenleme:
            izinler.crm_musteri_duzenleme = True
        crm_talep_olusturma = request.POST.get("crm_talep_olusturma")
        if crm_talep_olusturma:
            izinler.crm_talep_olusturma = True
        crm_talep_silme = request.POST.get("crm_talep_silme")
        if crm_talep_silme:
            izinler.crm_talep_silme = True
        crm_talep_gorme = request.POST.get("crm_talep_gorme")
        if crm_talep_gorme:
            izinler.crm_talep_gorme = True
        crm_talep_duzenleme = request.POST.get("crm_talep_duzenleme")
        if crm_talep_duzenleme:
            izinler.crm_talep_duzenleme = True
        crm_teklif_olusturma = request.POST.get("crm_teklif_olusturma")
        if crm_teklif_olusturma:
            izinler.crm_teklif_olusturma = True
        crm_teklif_silme = request.POST.get("crm_teklif_silme")
        if crm_teklif_silme:
            izinler.crm_teklif_silme = True
        crm_teklif_gorme = request.POST.get("crm_teklif_gorme")
        if crm_teklif_gorme:
            izinler.crm_teklif_gorme = True
        crm_teklif_duzenleme = request.POST.get("crm_teklif_duzenleme")
        if crm_teklif_duzenleme:
            izinler.crm_teklif_duzenleme = True
        crm_daire_olusturma = request.POST.get("crm_daire_olusturma")
        if crm_daire_olusturma:
            izinler.crm_daire_olusturma = True
        crm_daire_silme = request.POST.get("crm_daire_silme")
        if crm_daire_silme:
            izinler.crm_daire_silme = True
        crm_daire_gorme = request.POST.get("crm_daire_gorme")
        if crm_daire_gorme:
            izinler.crm_daire_gorme = True
        crm_daire_duzenleme = request.POST.get("crm_daire_duzenleme")
        if crm_daire_duzenleme:
            izinler.crm_daire_duzenleme = True
        crm_evrak_olusturma = request.POST.get("crm_evrak_olusturma")
        if crm_evrak_olusturma:
            izinler.crm_evrak_olusturma = True
        crm_evrak_silme = request.POST.get("crm_evrak_silme")
        if crm_evrak_silme:
            izinler.crm_evrak_silme = True
        crm_evrak_gorme = request.POST.get("crm_evrak_gorme")
        if crm_evrak_gorme:
            izinler.crm_evrak_gorme = True
        crm_evrak_duzenleme = request.POST.get("crm_evrak_duzenleme")
        if crm_evrak_duzenleme:
            izinler.crm_evrak_duzenleme = True

        ust_yuklenici_gorme = request.POST.get("ust_yuklenici_gorme")
        if ust_yuklenici_gorme:
            izinler.ust_yuklenici_gorme = True
        ust_yuklenici_olusturma = request.POST.get("ust_yuklenici_olusturma")
        if ust_yuklenici_olusturma:
            izinler.ust_yuklenici_olusturma = True
        ust_yuklenici_duzenleme = request.POST.get("ust_yuklenici_duzenleme")
        if ust_yuklenici_duzenleme:
            izinler.ust_yuklenici_duzenleme = True
        ust_yuklenici_silme = request.POST.get("ust_yuklenici_silme")
        if ust_yuklenici_silme:
            izinler.ust_yuklenici_silme = True
        # RFI İzinleri
        izinler.rfi_gorme = False
        izinler.rfi_olusturma = False
        izinler.rfi_duzenleme = False
        izinler.rfi_silme = False

        if request.POST.get("rfi_gorme"):
            izinler.rfi_gorme = True
        if request.POST.get("rfi_olusturma"):
            izinler.rfi_olusturma = True
        if request.POST.get("rfi_duzenleme"):
            izinler.rfi_duzenleme = True
        if request.POST.get("rfi_silme"):
            izinler.rfi_silme = True

        # RFI Listesi İzinleri
        izinler.rfi_listesi_gorme = False
        izinler.rfi_listesi_olusturma = False
        izinler.rfi_listesi_duzenleme = False
        izinler.rfi_listesi_silme = False

        if request.POST.get("rfi_listesi_gorme"):
            izinler.rfi_listesi_gorme = True
        if request.POST.get("rfi_listesi_olusturma"):
            izinler.rfi_listesi_olusturma = True
        if request.POST.get("rfi_listesi_duzenleme"):
            izinler.rfi_listesi_duzenleme = True
        if request.POST.get("rfi_listesi_silme"):
            izinler.rfi_listesi_silme = True

        # RFI Listesi Onaylama İzinleri
        izinler.rfi_listesi_onaylama_gorme = False
        izinler.rfi_listesi_onaylama_olustur = False
        izinler.rfi_listesi_onaylama_silme = False

        if request.POST.get("rfi_listesi_onaylama_gorme"):
            izinler.rfi_listesi_onaylama_gorme = True
        if request.POST.get("rfi_listesi_onaylama_olustur"):
            izinler.rfi_listesi_onaylama_olustur = True
        if request.POST.get("rfi_listesi_onaylama_silme"):
            izinler.rfi_listesi_onaylama_silme = True

        izinler.save()
        
        return redirect_with_language("main:kullanici_yetkileri_duzenle",id_bilgiis)

def cari_history_view(request, cari_id):
    cari_instance = cari.objects.get(id=cari_id)
    history = cari_instance.history.all()
    return render(request, 'cari_history.html', {'history': history})
def get_object_or_none(model, *args, **kwargs):
    try:
        return model.objects.get(*args, **kwargs)
    except :
        return None
    

def giderleri_excelden_ekle(request,id):
    import openpyxl
    Gider_excel_ekl = get_object_or_404(Gider_excel_ekleme,id = id)
    kullanici = Gider_excel_ekl.gelir_kime_ait_oldugu
    dataframe = openpyxl.load_workbook(Gider_excel_ekl.gelir_makbuzu)
    dataframe1 = dataframe.active
    data = []
    y = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu).count()
    for row in range(1, dataframe1.max_row):
        a = []
        a.append(row + 1)
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            a.append(col[row].value)
        data.append(a)
    for i in data:
        fatura_tarihi = i[1]
        vade_tarihi = i[2]
        cari_bilgisi = i[3]
        tutar = i[4]
        tutar = float(str(tutar).replace("$", ""))
        print(tutar)
        kur = i[5]
        kategori = i[6]
        urun = i[7]
        etiket1 = i[8]
        etiket2 = i[9]
        aciklama = i[10]
        makbuz_bilgisi = None
        odeme_durumu = i[11]
        print(odeme_durumu)
        print(type(odeme_durumu))
        if makbuz_bilgisi :
            makbuz_bilgisi = makbuz_bilgisi
        else:
            makbuz_bilgisi = None
        if get_object_or_none(cari,cari_adi=cari_bilgisi ,cari_kart_ait_bilgisi = kullanici ):
            cari_bilgisii = get_object_or_none(cari,cari_adi=cari_bilgisi ,cari_kart_ait_bilgisi = kullanici )
        else:

            cari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),cari_kart_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                ,cari_adi = cari_bilgisi,aciklama = "",telefon_numarasi = 0.0)
            cari_bilgisii = get_object_or_none(cari,cari_adi=cari_bilgisi ,cari_kart_ait_bilgisi = kullanici )
        if get_object_or_none(gider_kategorisi,gider_kategori_adi=kategori ,gider_kategoris_ait_bilgisi = kullanici ):
            kategorii = get_object_or_none(gider_kategorisi,gider_kategori_adi=kategori ,gider_kategoris_ait_bilgisi = kullanici )
        else:

            gider_kategorisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gider_kategoris_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                                            ,gider_kategori_adi = kategori,gider_kategorisi_renk = "#000000",aciklama = "")
            kategorii = get_object_or_none(gider_kategorisi,gider_kategori_adi=kategori ,gider_kategoris_ait_bilgisi = kullanici )
        if get_object_or_none(urunler,urun_adi=urun ,urun_ait_oldugu = kullanici ):
            uruni = get_object_or_none(urunler,urun_adi=urun ,urun_ait_oldugu = kullanici )
        else:

            urunler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),urun_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu
                                ,urun_adi = urun,urun_fiyati = tutar)
            uruni = get_object_or_none(urunler,urun_adi=urun ,urun_ait_oldugu = kullanici )
        if get_object_or_none(gider_etiketi,gider_etiketi_adi=etiket1 ,gider_kategoris_ait_bilgisi = kullanici ):
            etiket1i =  get_object_or_none(gider_etiketi,gider_etiketi_adi=etiket1 ,gider_kategoris_ait_bilgisi = kullanici )
        else:

            gider_etiketi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gider_kategoris_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                                         ,gider_etiketi_adi = etiket1)
            etiket1i =  get_object_or_none(gider_etiketi,gider_etiketi_adi=etiket1 ,gider_kategoris_ait_bilgisi = kullanici )
        if get_object_or_none(gider_etiketi,gider_etiketi_adi=etiket2 ,gider_kategoris_ait_bilgisi = kullanici ):
            etiket2i =  get_object_or_none(gider_etiketi,gider_etiketi_adi=etiket2 ,gider_kategoris_ait_bilgisi = kullanici )
        else:

            gider_etiketi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gider_kategoris_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                                         ,gider_etiketi_adi = etiket2)
            etiket2i =  get_object_or_none(gider_etiketi,gider_etiketi_adi=etiket2 ,gider_kategoris_ait_bilgisi = kullanici )
        y = y+1
        b = len(str(y))
        c = 8 - b
        m = faturalardaki_gelir_gider_etiketi.objects.last().gider_etiketi+(c*"0")+str(y)
        print(i)
        if odeme_durumu == 1:
            new_project =Gider_Bilgisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu,
                cari_bilgisi = cari_bilgisii,
                fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = m,
                gelir_kategorisii = kategorii,doviz = 0,aciklama = aciklama,toplam_tutar =tutar ,kalan_tutar = 0 )
        elif odeme_durumu == 0:
            new_project =Gider_Bilgisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu,
                cari_bilgisi = cari_bilgisii,
                fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = m,
                gelir_kategorisii = kategorii,doviz = 0,aciklama = aciklama,toplam_tutar =tutar ,kalan_tutar = tutar )
        new_project.save()
        gelir_etiketi_sec = []
        gelir_etiketi_sec.append(etiket1i)
        gelir_etiketi_sec.append(etiket2i)
        new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
        gider_urun_bilgisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            urun_ait_oldugu =  Gider_excel_ekl.gelir_kime_ait_oldugu,urun_bilgisi =uruni,
                            urun_fiyati = tutar,urun_indirimi = 0.0,urun_adeti = 1,
                            gider_bilgis =  get_object_or_404(Gider_Bilgisi,id = new_project.id),
                            aciklama = "")
        if odeme_durumu == 1:
            bir = Gider_odemesi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = get_object_or_404(Gider_Bilgisi,id = new_project.id ),kasa_bilgisi = Gider_excel_ekl.kasa,
                                        tutar =tutar,tarihi =fatura_tarihi,
                                        aciklama = "",makbuz_no =m ,gelir_makbuzu = makbuz_bilgisi )
            bir.set_gelir_makbuzu(makbuz_bilgisi)
        gider_qr.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = get_object_or_404(Gider_Bilgisi,id = new_project.id))
    return redirect_with_language("main:ana_sayfa")


def gelirleri_excelden_ekle(request,id):
    import openpyxl
    Gider_excel_ekl = get_object_or_404(Gelir_excel_ekleme,id = id)
    kullanici = Gider_excel_ekl.gelir_kime_ait_oldugu
    dataframe = openpyxl.load_workbook(Gider_excel_ekl.gelir_makbuzu)
    dataframe1 = dataframe.active
    data = []
    y = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu).count()
    for row in range(1, dataframe1.max_row):
        a = []
        a.append(row + 1)
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            a.append(col[row].value)
        data.append(a)
    for i in data:
        fatura_tarihi = i[1]
        vade_tarihi = i[2]
        cari_bilgisi = i[3]
        tutar = i[4]
        tutar = float(str(tutar).replace("$", ""))
        print(tutar)
        kur = i[5]
        kategori = i[6]
        urun = i[7]
        etiket1 = i[8]
        aciklama = i[9]
        odeme_durumu = i[11]
        makbuz_bilgisi = None
        print(i)
        if makbuz_bilgisi :
            makbuz_bilgisi = makbuz_bilgisi
        else:
            makbuz_bilgisi = None
        if get_object_or_none(cari,cari_adi=cari_bilgisi ,cari_kart_ait_bilgisi = kullanici ):
            cari_bilgisii = get_object_or_none(cari,cari_adi=cari_bilgisi ,cari_kart_ait_bilgisi = kullanici )
        else:

            cari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),cari_kart_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                ,cari_adi = cari_bilgisi,aciklama = "",telefon_numarasi = 0.0)
            cari_bilgisii = get_object_or_none(cari,cari_adi=cari_bilgisi ,cari_kart_ait_bilgisi = kullanici )
        if get_object_or_none(gelir_kategorisi,gelir_kategori_adi=kategori ,gelir_kategoris_ait_bilgisi = kullanici ):
            kategorii = get_object_or_none(gelir_kategorisi,gelir_kategori_adi=kategori ,gelir_kategoris_ait_bilgisi = kullanici )
        else:

            gelir_kategorisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kategoris_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                                            ,gelir_kategori_adi = kategori,gelir_kategorisi_renk = "#000000",aciklama = "")
            kategorii = get_object_or_none(gelir_kategorisi,gelir_kategori_adi=kategori ,gelir_kategoris_ait_bilgisi = kullanici )
        if get_object_or_none(urunler,urun_adi=urun ,urun_ait_oldugu = kullanici ):
            uruni = get_object_or_none(urunler,urun_adi=urun ,urun_ait_oldugu = kullanici )
        else:

            urunler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),urun_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu
                                ,urun_adi = urun,urun_fiyati = tutar)
            uruni = get_object_or_none(urunler,urun_adi=urun ,urun_ait_oldugu = kullanici )
        if get_object_or_none(gelir_etiketi,gelir_etiketi_adi=etiket1 ,gelir_kategoris_ait_bilgisi = kullanici ):
            etiket1i =  get_object_or_none(gelir_etiketi,gelir_etiketi_adi=etiket1 ,gelir_kategoris_ait_bilgisi = kullanici )
        else:

            gelir_etiketi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kategoris_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                                         ,gelir_etiketi_adi = etiket1)
            etiket1i =  get_object_or_none(gelir_etiketi,gelir_etiketi_adi=etiket1 ,gelir_kategoris_ait_bilgisi = kullanici )
        
        y = y+1
        b = len(str(y))
        c = 8 - b
        m = faturalardaki_gelir_gider_etiketi.objects.last().gelir_etiketi+(c*"0")+str(y)
        if odeme_durumu == 1:
            new_project =Gelir_Bilgisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu,
            cari_bilgisi = cari_bilgisii,
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = m,
            gelir_kategorisii = kategorii,doviz = 0,aciklama = aciklama,toplam_tutar =tutar ,kalan_tutar = 0 )
        else:
            new_project =Gelir_Bilgisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu,
            cari_bilgisi = cari_bilgisii,
            fatura_tarihi=fatura_tarihi,vade_tarihi=vade_tarihi,fatura_no = m,
            gelir_kategorisii = kategorii,doviz = 0,aciklama = aciklama,toplam_tutar =tutar ,kalan_tutar =tutar )
        new_project.save()
        gelir_etiketi_sec = []
        gelir_etiketi_sec.append(etiket1i)
        #gelir_etiketi_sec.append(etiket2i)
        new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
        gelir_urun_bilgisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            urun_ait_oldugu =  Gider_excel_ekl.gelir_kime_ait_oldugu,urun_bilgisi =uruni,
                            urun_fiyati = tutar,urun_indirimi = 0.0,urun_adeti = 1,
                            gider_bilgis =  get_object_or_404(Gelir_Bilgisi,id = new_project.id),
                            aciklama = "")
        if odeme_durumu == 1:
            bir = Gelir_odemesi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = get_object_or_404(Gelir_Bilgisi,id = new_project.id ),kasa_bilgisi = Gider_excel_ekl.kasa,
                                        tutar =tutar,tarihi =fatura_tarihi,
                                        aciklama = "",makbuz_no =m ,gelir_makbuzu = makbuz_bilgisi )
            bir.set_gelir_makbuzu(makbuz_bilgisi)
        gelir_qr.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = get_object_or_404(Gelir_Bilgisi,id = new_project.id))
    return redirect_with_language("main:ana_sayfa")
def giderleri_excelden_eklei(request,id):
    import openpyxl
    
    Gider_excel_ekl = get_object_or_404(Gider_excel_ekleme,id = id)
    dataframe = openpyxl.load_workbook(Gider_excel_ekl.gelir_makbuzu)
    dataframe1 = dataframe.active
    data = []
    sadece_cari = []
    sadece_etiket = []
    sadece_kategorisi = []
    sadece_urunler = []
    sadece_fiyat = []
    sozluk_cari ={}
    sozluk_etiket = {}
    sozluk_kategorisi = {}
    sozluk_urun = {}
    for row in range(1, dataframe1.max_row):
        a = []
        a.append(row + 1)
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            a.append(col[row].value)
        data.append(a)
        #print(a)
    #print("-" * 50)
    for i in data:
        if i[3] not in sadece_cari:
            sadece_cari.append(i[3])
        if i[4] not in sadece_fiyat:
            y = float((str(str(str(i[4]).replace("$",""))).replace(".","")).replace(",","."))
            sadece_fiyat.append(y)
        if i[5] not in sadece_urunler:
            sadece_urunler.append(i[5])
        if i[6] not in sadece_etiket:
            sadece_etiket.append(i[6])
        if i[7] not in sadece_etiket:
            sadece_etiket.append(i[7])
        if i[8] not in sadece_etiket:
            sadece_etiket.append(i[8])
        if i[9] not in sadece_kategorisi:
            sadece_kategorisi.append(i[9])
    for i in sadece_cari:
        l = cari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),cari_kart_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                ,cari_adi = i,aciklama = "",telefon_numarasi = 0.0)
        sozluk_cari[i] = l.id
    k = 0
    for i in sadece_urunler:
        l = urunler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),urun_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu
                                ,urun_adi = i,urun_fiyati = sadece_fiyat[k]

                                )
        #print(i)
        k = k+1
        sozluk_urun[i] = l.id
    for i in sadece_kategorisi:
        z = gider_kategorisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gider_kategoris_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                                            ,gider_kategori_adi = i,gider_kategorisi_renk = "#000000",aciklama = "")
        sozluk_kategorisi[i] = z.id
    for i in sadece_etiket:
        z = gider_etiketi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gider_kategoris_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                                         ,gider_etiketi_adi = i)
        sozluk_etiket[i] = z.id
    y = Gider_Bilgisi.objects.filter(gelir_kime_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu).count()
    for i in data:
        y = y+1
        b = len(str(y))
        c = 8 - b
        m = faturalardaki_gelir_gider_etiketi.objects.last().gider_etiketi+(c*"0")+str(y)
        new_project =Gider_Bilgisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu,
            cari_bilgisi = get_object_or_none(cari,id = sozluk_cari[i[3]]),
            fatura_tarihi=i[2],vade_tarihi=i[2],fatura_no = m,
            gelir_kategorisi = get_object_or_none( gider_kategorisi,id =sozluk_kategorisi[i[8]] ),doviz = i[11],aciklama = i[10]
                                         )
        #print(sozluk_cari[i[3]])
        new_project.save()
        gelir_etiketi_sec = []
        gelir_etiketi_sec.append(gider_etiketi.objects.get(id=int(sozluk_etiket[i[5]])))
        gelir_etiketi_sec.append(gider_etiketi.objects.get(id=int(sozluk_etiket[i[6]])))
        new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
        #bilgi_getirler = urunler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),urun_ait_oldugu =  Gider_excel_ekl.gelir_kime_ait_oldugu,urun_adi = sozluk_urun[i[5]],urun_fiyati = float((str(str(str(i[4]).replace("$",""))).replace(".","")).replace(",",".")))
        gider_urun_bilgisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            urun_ait_oldugu =  Gider_excel_ekl.gelir_kime_ait_oldugu,urun_bilgisi = get_object_or_404(urunler, 
                            id = sozluk_urun[i[5]]),
                            urun_fiyati = float((str(str(str(i[4]).replace("$",""))).replace(".","")).replace(",",".")),urun_indirimi = 0.0,urun_adeti = 1,
                            gider_bilgis =  get_object_or_404(Gider_Bilgisi,id = new_project.id),
                            aciklama = "")
        bir = Gider_odemesi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = get_object_or_404(Gider_Bilgisi,id = new_project.id ),kasa_bilgisi = Gider_excel_ekl.kasa,
                                     tutar =float((str(str(str(i[4]).replace("$",""))).replace(".","")).replace(",",".")),tarihi =i[2],
                                       aciklama = "",makbuz_no =str(i[13] ) ,gelir_makbuzu = str(i[12] ) )
        bir.set_gelir_makbuzu(str(i[13] ))
        gider_qr.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = get_object_or_404(Gider_Bilgisi,id = new_project.id))
    return redirect_with_language("main:ana_sayfa")

def gelirleri_excelden_eklei(request,id):
    import openpyxl
    Gider_excel_ekl = get_object_or_404(Gelir_excel_ekleme,id = id)
    dataframe = openpyxl.load_workbook(Gider_excel_ekl.gelir_makbuzu)
    dataframe1 = dataframe.active
    data = []
    sadece_cari = []
    sadece_etiket = []
    sadece_kategorisi = []
    sadece_urunler = []
    sadece_fiyat = []
    sozluk_cari ={}
    sozluk_etiket = {}
    sozluk_kategorisi = {}
    sozluk_urun = {}
    for row in range(1, dataframe1.max_row):
        a = []
        a.append(row + 1)
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            a.append(col[row].value)
        data.append(a)
        #print(a)
    #print("-" * 50)
    for i in data:
        if i[3] not in sadece_cari:
            sadece_cari.append(i[3])
        if i[4] not in sadece_fiyat:
            y = float(str(str(i[4]).replace("$","")).replace(",","."))
            sadece_fiyat.append(y)
        if i[5] not in sadece_urunler:
            sadece_urunler.append(i[5])
        if i[6] not in sadece_etiket:
            sadece_etiket.append(i[6])
        if i[7] not in sadece_etiket:
            sadece_etiket.append(i[7])
        if i[8] not in sadece_kategorisi:
            sadece_kategorisi.append(i[8])
    for i in sadece_cari:
        z = cari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),cari_kart_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                ,cari_adi = i,aciklama = "",telefon_numarasi = 0.0)
        sozluk_cari[i] = z.id
    k = 0
    for i in sadece_urunler:
        z = urunler.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),urun_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu
                                ,urun_adi = i,urun_fiyati = sadece_fiyat[k]

                                )
        k = k+1
        sozluk_urun[i] = z.id
    for i in sadece_kategorisi:
        z = gelir_kategorisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gider_kategoris_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                                            ,gider_kategori_adi = i,gider_kategorisi_renk = "#000000",aciklama = "")
        sozluk_kategorisi[i] = z.id
    for i in sadece_etiket:
        z = gelir_etiketi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gider_kategoris_ait_bilgisi = Gider_excel_ekl.gelir_kime_ait_oldugu
                                         ,gider_etiketi_adi = i)
        sozluk_etiket[i] = z.id
    y = Gelir_Bilgisi.objects.filter(gelir_kime_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu).count()
    for i in data:
        y = y+1
        b = len(str(y))
        c = 8 - b
        m = faturalardaki_gelir_gider_etiketi.objects.last().gelir_etiketi+(c*"0")+str(y)
        new_project =Gider_Bilgisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = Gider_excel_ekl.gelir_kime_ait_oldugu,
            cari_bilgisi = get_object_or_none(cari,cari_adi = sozluk_cari[i[3]],cari_kart_ait_bilgisi = request.user),
            fatura_tarihi=i[2],vade_tarihi=i[2],fatura_no = m,
            gelir_kategorisi = get_object_or_none( gelir_kategorisi,id =sozluk_kategorisi[i[8]] ),doviz = i[10],aciklama = i[9]
                                         )
        new_project.save()
        gelir_etiketi_sec = []
        gelir_etiketi_sec.append(gelir_etiketi.objects.get(id=int(sozluk_etiket[i[5]])))
        gelir_etiketi_sec.append(gelir_etiketi.objects.get(id=int(sozluk_etiket[i[6]])))
        new_project.gelir_etiketi_sec.add(*gelir_etiketi_sec)
        gelir_urun_bilgisi_bi = gelir_urun_bilgisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                            urun_ait_oldugu =  Gider_excel_ekl.gelir_kime_ait_oldugu,urun_bilgisi = get_object_or_none(urunler, urun_ait_oldugu=Gider_excel_ekl.gelir_kime_ait_oldugu,urun_adi = sozluk_urun[i[5]]),
                            urun_fiyati = float(str(str(i[4]).replace("$","")).replace(",",".")),urun_indirimi = 0.0,urun_adeti = 1,
                            gider_bilgis =  get_object_or_none(Gelir_Bilgisi,id = new_project.id),
                            aciklama = ""
                        )
        Gelir_odemesi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),gelir_kime_ait_oldugu = get_object_or_404(Gelir_odemesi,id = new_project.id ),kasa_bilgisi = Gider_excel_ekl.kasa,
                                     tutar =float(str(str(i[4]).replace("$","")).replace(",",".")),tarihi =i[2],makbuz_no = new_project.id,
                                       aciklama = "deneme"  )
    return redirect_with_language("main:ana_sayfa")

def sayfa_denemeleri(request):
    content = sozluk_yapisi()
    return render(request,"xxpuantaj.html",content)
def bildirimler(request):
    content = sozluk_yapisi()
    return render(request,"bildirimler.html",content)

def katman_sayfasi(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        profile =katman.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.katman_gorme:
                    profile = katman.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db)
                    content["insaatlar"] = santiye.objects.filter(proje_ait_bilgisi =  request.user.kullanicilar_db,silinme_bilgisi = False)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = katman.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
            content["insaatlar"] = santiye.objects.filter(proje_ait_bilgisi =  request.user,silinme_bilgisi = False)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = profile
    return render(request,"santiye_yonetimi/katman.html",content)

def katman_ekle(request):
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            return redirect_with_language("main:taseron_ekle_admin",kullanici)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a.izinler.katman_olusturma:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                kullanici = request.user
        katman_adi = request.POST.get("taseron_adi")
        santiye_al = request.POST.get("blogbilgisi")
        dosya = request.FILES.get("file")
        katman.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
            proje_ait_bilgisi = kullanici,
            proje_santiye_Ait = get_object_or_none(santiye,id = santiye_al),
            katman_adi  = katman_adi,
            katman_dosyasi = dosya
        )
    return redirect_with_language("main:katman_sayfasi")

def katman_sil(request):
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            return redirect_with_language("main:taseron_ekle_admin",kullanici)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a.izinler.katman_silme:
                    buttonIdInput = request.POST.get("buttonId")
                    katman.objects.filter(id = buttonIdInput).update(
                        silinme_bilgisi = True
                    )
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                buttonIdInput = request.POST.get("buttonId")
                katman.objects.filter(id = buttonIdInput).update(
                    silinme_bilgisi = True
                )
        
    return redirect_with_language("main:katman_sayfasi")

def katman_duzenle(request):
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            return redirect_with_language("main:taseron_ekle_admin",kullanici)
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a.izinler.katman_olusturma:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                kullanici = request.user
        katman_adi = request.POST.get("taseron_adi")
        santiye_al = request.POST.get("blogbilgisi")
        dosya = request.FILES.get("file")
        buttonId = request.POST.get("buttonId")
        if dosya:
            katman.objects.filter(id = buttonId).update(
                proje_ait_bilgisi = kullanici,
                proje_santiye_Ait = get_object_or_none(santiye,id = santiye_al),
                katman_adi  = katman_adi,
                katman_dosyasi = dosya
            )
        else:
            katman.objects.filter(id = buttonId).update(
            proje_ait_bilgisi = kullanici,
            proje_santiye_Ait = get_object_or_none(santiye,id = santiye_al),
            katman_adi  = katman_adi
        )
    return redirect_with_language("main:katman_sayfasi")


def get_yapi(request, santiye_id):
    yapilar = bloglar.objects.filter(proje_santiye_Ait__id=santiye_id).values('id', 'blog_adi',"kat_sayisi")
    yapilar_list = list(yapilar)
    return JsonResponse({'yapilar': yapilar_list})
######################################3


def kullanici_yetkileri_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile = personel_izinleri.objects.filter(izinlerin_sahibi_kullanici = users)
        content["izinler"] = profile
    else:
        if request.user.kullanicilar_db:
            return redirect_with_language("main:yetkisiz")
        profile = personel_izinleri.objects.filter(izinlerin_sahibi_kullanici = request.user)
        content["izinler"] = profile
    return render(request,"kullanici_yetkileri/yetkiler.html",content)
def kullanici_yetkileri_duzenle_2(request,id,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile = personel_izinleri.objects.filter(izinlerin_sahibi_kullanici = users)
        content["izinler"] = profile
        content["secili_grup"] = get_object_or_404(personel_izinleri,id = id)
    else:
        if request.user.kullanicilar_db:
            return redirect_with_language("main:yetkisiz")
        profile = personel_izinleri.objects.filter(izinlerin_sahibi_kullanici = request.user)
        content["izinler"] = profile
        content["secili_grup"] = get_object_or_404(personel_izinleri,id = id)
    return render(request,"kullanici_yetkileri/yetkiler.html",content)
def kullanici_yetki_olustur_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        return redirect_with_language("main:yetkisiz")
    if request.POST:
        grup_adi = request.POST.get("grup_adi")
        personel_izinleri.objects.create(
            isim = grup_adi,
            izinlerin_sahibi_kullanici = users
        )
    return redirect_with_language("main:kullanici_yetkileri_2",hash)
def kullanici_yetki_adi_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        return redirect_with_language("main:yetkisiz")
    if request.POST:
        grup_adi = request.POST.get("grup_adi")
        id  = request.POST.get("id")
        personel_izinleri.objects.filter(id = id).update(
            isim = grup_adi,
            izinlerin_sahibi_kullanici = users
        )
    return redirect_with_language("main:kullanici_yetkileri_2",hash)
def kullanici_yetki_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        return redirect_with_language("main:yetkisiz")
    if request.POST:
        id  = request.POST.get("id")
        personel_izinleri.objects.filter(id = id,izinlerin_sahibi_kullanici = users).delete()
    return redirect_with_language("main:kullanici_yetkileri_2",hash)
def kullanici_yetki_alma_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        return redirect_with_language("main:yetkisiz")
    if request.POST:
        guncellenen = request.POST.get("guncellenen")
        izinler = get_object_or_404(personel_izinleri,id = guncellenen) 
        ##
        #izinleri Sıfırla
        izinler.dashboard_gorme = False
        izinler.dashboard_silme = False
        izinler.dashboard_duzenleme = False
        izinler.dashboard_olusturma = False
        #
        izinler.gelir_ozeti_gorme = False
        izinler.gelir_ozeti_olusturma = False
        izinler.gelir_ozeti_duzenleme =False
        izinler.gelir_ozeti_silme = False
        #
        izinler.gider_ozeti_gorme = False
        izinler.gider_ozeti_olusturma = False
        izinler.gider_ozeti_duzenleme = False
        izinler.gider_ozeti_silme = False
        #
        izinler.hesap_ekstra_gorme = False
        izinler.hesap_ekstra_olusturma = False
        izinler.hesap_ekstra_duzenleme = False
        izinler.hesap_ekstra_silme = False
        #
        izinler.ilerleme_takibi_gorme = False
        izinler.ilerleme_takibi_olusturma = False
        izinler.ilerleme_takibi_duzenleme = False
        izinler.ilerleme_takibi_silme = False
        #
        izinler.sozlesmeler_gorme = False
        izinler.sozlesmeler_olusturma = False
        izinler.sozlesmeler_duzenleme = False
        izinler.sozlesmeler_silme = False
        #
        izinler.yapilacaklar_gorme = False
        izinler.yapilacaklar_olusturma = False
        izinler.yapilacaklar_duzenleme = False
        izinler.yapilacaklar_silme = False
        #
        izinler.dosya_yoneticisi_gorme = False
        izinler.dosya_yoneticisi_olusturma = False
        izinler.dosya_yoneticisi_duzenleme = False
        izinler.dosya_yoneticisi_silme = False
        #
        izinler.projeler_gorme = False
        izinler.projeler_olusturma = False
        izinler.projeler_duzenleme = False
        izinler.projeler_silme =False
        #
        izinler.personeller_gorme = False
        izinler.personeller_olusturma = False
        izinler.personeller_duzenleme = False
        izinler.personeller_silme = False
        #
        izinler.gelir_faturasi_gorme_izni = False
        izinler.gelir_faturasi_kesme_izni = False
        izinler.gelir_faturasi_duzenleme_izni = False
        izinler.gelir_faturasi_silme_izni = False
        #
        izinler.gider_faturasi_gorme_izni = False
        izinler.gider_faturasi_kesme_izni = False
        izinler.gider_faturasi_duzenleme_izni = False
        izinler.gider_faturasi_silme_izni = False
        #
        izinler.kasa_virman_olusturma_izni = False
        izinler.kasa_virman_gorme_izni = False
        #
        izinler.kasa_gosterme_izni = False
        izinler.kasa_olusturma_izni = False
        izinler.kasa_guncelleme_izni = False
        izinler.Kasa_silme_izni = False
        #
        izinler.cari_gosterme_izni = False
        izinler.cari_olusturma = False
        izinler.cari_guncelleme_izni = False
        izinler.cari_silme_izni = False
        #
        izinler.personeller_gorme = False
        izinler.personeller_olusturma = False
        izinler.personeller_duzenleme = False
        izinler.personeller_silme = False
        #
        izinler.santiye_olusturma = False
        izinler.santiye_silme = False
        izinler.santiye_gorme = False
        izinler.santiye_duzenleme = False
        #
        izinler.blog_olusturma = False
        izinler.blog_silme = False
        izinler.blog_gorme = False
        izinler.blog_duzenleme = False
        #
        izinler.kalemleri_olusturma = False
        izinler.kalemleri_silme = False
        izinler.kalemleri_gorme = False
        izinler.kalemleri_duzenleme = False
        #
        izinler.santiye_raporu_olusturma = False
        izinler.santiye_raporu_silme = False
        izinler.santiye_raporu_gorme = False
        izinler.santiye_raporu_duzenleme = False
        #
        izinler.taseronlar_gorme = False
        izinler.taseronlar_olusturma = False
        izinler.taseronlar_duzenleme = False
        izinler.taseronlar_silme =False
        #
        izinler.hakedisler_gorme = False
        izinler.hakedisler_olusturma = False
        izinler.hakedisler_duzenleme = False
        izinler.hakedisler_silme = False
        #
        izinler.kasa_detay_izni = False
        izinler.cari_detay_izni = False
        ##
        izinler.gelir_kategorisi_gorme = False
        izinler.gelir_kategorisi_olusturma = False
        izinler.gelir_kategorisi_guncelleme = False
        izinler.gelir_kategorisi_silme = False
        #
        izinler.gider_kategorisi_gorme = False
        izinler.gider_kategorisi_olusturma = False
        izinler.gider_kategorisi_guncelleme = False
        izinler.gider_kategorisi_silme = False
        #
        izinler.gelir_etiketi_gorme = False
        izinler.gelir_etiketi_olusturma = False
        izinler.gelir_etiketi_guncelleme = False
        izinler.gelir_etiketi_silme = False
        #
        izinler.gider_etiketi_gorme = False
        izinler.gider_etiketi_olusturma = False
        izinler.gider_etiketi_guncelleme = False
        izinler.gider_etiketi_silme = False
        #
        izinler.urun_gorme = False
        izinler.urun_olusturma = False
        izinler.urun_guncelleme = False
        izinler.urun_silme = False
        #
        izinler.muhasabe_ayarlari_gorme = False
        izinler.muhasabe_ayarlari_guncelleme = False
        #
        izinler.gelir_faturasi_makbuz_gorme_izni = False
        izinler.gelir_faturasi_makbuz_kesme_izni = False
        izinler.gelir_faturasi_makbuz_duzenleme_izni = False
        izinler.gelir_faturasi_makbuz_silme_izni = False
        #
        izinler.gider_faturasi_makbuz_gorme_izni = False
        izinler.gider_faturasi_makbuz_kesme_izni = False
        izinler.gider_faturasi_makbuz_duzenleme_izni = False
        izinler.gider_faturasi_makbuz_silme_izni = False
        #Puantaj
        izinler.personeller_puantaj_olusturma = False
        izinler.personeller_puantaj_silme = False
        izinler.personeller_puantaj_gorme = False
        izinler.personeller_puantaj_duzenleme = False
        #
        # 
        izinler.personeller_odeme_olusturma = False
        izinler.personeller_odeme_silme = False
        izinler.personeller_odeme_gorme = False
        izinler.personeller_odeme_duzenleme = False
        #
        izinler.katman_olusturma = False
        izinler.katman_silme = False
        izinler.katman_gorme = False
        izinler.katman_duzenleme = False
        izinler.save()
        ##
        personeller_puantaj_olusturma = request.POST.get("personeller_puantaj_olusturma")
        if personeller_puantaj_olusturma:
            izinler.personeller_puantaj_olusturma = True
        personeller_puantaj_silme = request.POST.get("personeller_puantaj_silme")
        if personeller_puantaj_silme:
            izinler.personeller_puantaj_silme = True
        personeller_puantaj_gorme = request.POST.get("personeller_puantaj_gorme")
        if personeller_puantaj_gorme:
            izinler.personeller_puantaj_gorme = True
        personeller_puantaj_duzenleme = request.POST.get("personeller_puantaj_duzenleme")
        if personeller_puantaj_duzenleme:
            izinler.personeller_puantaj_duzenleme = True
        #
        personeller_odeme_olusturma = request.POST.get("personeller_odeme_olusturma")
        if personeller_odeme_olusturma:
            izinler.personeller_odeme_olusturma = True
        personeller_odeme_silme = request.POST.get("personeller_odeme_silme")
        if personeller_odeme_silme:
            izinler.personeller_odeme_silme = True
        personeller_odeme_gorme = request.POST.get("personeller_odeme_gorme")
        if personeller_odeme_gorme:
            izinler.personeller_odeme_gorme = True
        personeller_odeme_duzenleme = request.POST.get("personeller_odeme_duzenleme")
        if personeller_odeme_duzenleme:
            izinler.personeller_odeme_duzenleme = True
        #
        katman_olusturma = request.POST.get("katman_olusturma")
        if katman_olusturma:
            izinler.katman_olusturma = True
        katman_silme = request.POST.get("katman_silme")
        if katman_silme:
            izinler.katman_silme = True
        katman_gorme = request.POST.get("katman_gorme")
        if katman_gorme:
            izinler.katman_gorme = True
        katman_duzenleme = request.POST.get("katman_duzenleme")
        if katman_duzenleme:
            izinler.katman_duzenleme = True
        ##
        dashboard_gorme = request.POST.get("dashboard_gorme")
        if dashboard_gorme:
            izinler.dashboard_gorme = True
        dashboard_silme = request.POST.get("dashboard_silme")
        if dashboard_silme:
            izinler.dashboard_silme = True
        dashboard_duzenleme = request.POST.get("dashboard_duzenleme")
        if dashboard_duzenleme:
            izinler.dashboard_duzenleme = True
        dashboard_olusturma = request.POST.get("dashboard_olusturma")
        if dashboard_olusturma:
            izinler.dashboard_olusturma = True
        #
        gelir_ozeti_gorme = request.POST.get("gelir_ozeti_gorme")
        if gelir_ozeti_gorme:
            izinler.gelir_ozeti_gorme = True
        gelir_ozeti_olusturma = request.POST.get("gelir_ozeti_olusturma")
        if gelir_ozeti_olusturma:
            izinler.gelir_ozeti_olusturma = True
        gelir_ozeti_duzenleme = request.POST.get("gelir_ozeti_duzenleme")
        if gelir_ozeti_duzenleme:
            izinler.gelir_ozeti_duzenleme = True
        gelir_ozeti_silme = request.POST.get("gelir_ozeti_silme")
        if gelir_ozeti_silme:
            izinler.gelir_ozeti_silme = True
        #
        gider_ozeti_gorme = request.POST.get("gider_ozeti_gorme")
        if gider_ozeti_gorme:
            izinler.gider_ozeti_gorme = True
        gider_ozeti_olusturma = request.POST.get("gider_ozeti_olusturma")
        if gider_ozeti_olusturma:
            izinler.gider_ozeti_olusturma = True
        gider_ozeti_duzenleme = request.POST.get("gider_ozeti_duzenleme")
        if gider_ozeti_duzenleme:
            izinler.gider_ozeti_duzenleme = True
        gider_ozeti_silme = request.POST.get("gider_ozeti_silme")
        if gider_ozeti_silme:
            izinler.gider_ozeti_silme = True
        #
        hesap_ekstra_gorme = request.POST.get("hesap_ekstra_gorme")
        if hesap_ekstra_gorme:
            izinler.hesap_ekstra_gorme = True
        hesap_ekstra_olusturma = request.POST.get("hesap_ekstra_olusturma")
        if hesap_ekstra_olusturma:
            izinler.hesap_ekstra_olusturma = True
        hesap_ekstra_duzenleme = request.POST.get("hesap_ekstra_duzenleme")
        if hesap_ekstra_duzenleme:
            izinler.hesap_ekstra_duzenleme = True
        hesap_ekstra_silme = request.POST.get("hesap_ekstra_silme")
        if hesap_ekstra_silme:
            izinler.hesap_ekstra_silme = True
        #
        kasa_virman_olusturma_izni = request.POST.get("kasa_virman_olusturma_izni")
        if kasa_virman_olusturma_izni:
            izinler.kasa_virman_olusturma_izni = True
        kasa_virman_gorme_izni = request.POST.get("kasa_virman_gorme_izni")
        if kasa_virman_gorme_izni:
            izinler.kasa_virman_gorme_izni = True
        #
        ilerleme_takibi_gorme = request.POST.get("ilerleme_takibi_gorme")
        if ilerleme_takibi_gorme:
            izinler.ilerleme_takibi_gorme = True
        ilerleme_takibi_olusturma = request.POST.get("ilerleme_takibi_olusturma")
        if ilerleme_takibi_olusturma:
            izinler.ilerleme_takibi_olusturma = True
        ilerleme_takibi_duzenleme = request.POST.get("ilerleme_takibi_duzenleme")
        if ilerleme_takibi_duzenleme:
            izinler.ilerleme_takibi_duzenleme = True
        ilerleme_takibi_silme = request.POST.get("ilerleme_takibi_silme")
        if ilerleme_takibi_silme:
            izinler.ilerleme_takibi_silme = True
        #
        sozlesmeler_gorme = request.POST.get("sozlesmeler_gorme")
        if sozlesmeler_gorme:
            izinler.sozlesmeler_gorme = True
        sozlesmeler_olusturma = request.POST.get("sozlesmeler_olusturma")
        if sozlesmeler_olusturma:
            izinler.sozlesmeler_olusturma = True
        sozlesmeler_duzenleme = request.POST.get("sozlesmeler_duzenleme")
        if sozlesmeler_duzenleme:
            izinler.sozlesmeler_duzenleme = True
        sozlesmeler_silme = request.POST.get("sozlesmeler_silme")
        if sozlesmeler_silme:
            izinler.sozlesmeler_silme = True
        #
        yapilacaklar_gorme = request.POST.get("yapilacaklar_gorme")
        if yapilacaklar_gorme:
            izinler.yapilacaklar_gorme = True
        yapilacaklar_olusturma = request.POST.get("yapilacaklar_olusturma")
        if yapilacaklar_olusturma:
            izinler.yapilacaklar_olusturma = True
        yapilacaklar_duzenleme = request.POST.get("yapilacaklar_duzenleme")
        if yapilacaklar_duzenleme:
            izinler.yapilacaklar_duzenleme = True
        yapilacaklar_silme = request.POST.get("yapilacaklar_silme")
        if yapilacaklar_silme:
            izinler.yapilacaklar_silme = True
        #
        dosya_yoneticisi_gorme = request.POST.get("dosya_yoneticisi_gorme")
        if dosya_yoneticisi_gorme:
            izinler.dosya_yoneticisi_gorme = True
        dosya_yoneticisi_olusturma = request.POST.get("dosya_yoneticisi_olusturma")
        if dosya_yoneticisi_olusturma:
            izinler.dosya_yoneticisi_olusturma = True
        dosya_yoneticisi_duzenleme = request.POST.get("dosya_yoneticisi_duzenleme")
        if dosya_yoneticisi_duzenleme:
            izinler.dosya_yoneticisi_duzenleme = True
        dosya_yoneticisi_silme = request.POST.get("dosya_yoneticisi_silme")
        if dosya_yoneticisi_silme:
            izinler.dosya_yoneticisi_silme = True
        #
        projeler_gorme = request.POST.get("projeler_gorme")
        if projeler_gorme:
            izinler.projeler_gorme = True
        projeler_olusturma = request.POST.get("projeler_olusturma")
        if projeler_olusturma:
            izinler.projeler_olusturma = True
        projeler_duzenleme = request.POST.get("projeler_duzenleme")
        if projeler_duzenleme:
            izinler.projeler_duzenleme = True
        projeler_silme = request.POST.get("projeler_silme")
        if projeler_silme:
            izinler.projeler_silme = True

        #
        personeller_gorme = request.POST.get("personeller_gorme")
        if personeller_gorme:
            izinler.personeller_gorme = True
        personeller_olusturma = request.POST.get("personeller_olusturma")
        if personeller_olusturma:
            izinler.personeller_olusturma = True
        personeller_duzenleme = request.POST.get("personeller_duzenleme")
        if personeller_duzenleme:
            izinler.personeller_duzenleme = True
        personeller_silme = request.POST.get("personeller_silme")
        if personeller_silme:
            izinler.personeller_silme = True
        #
        gelir_faturasi_gorme_izni = request.POST.get("gelir_faturasi_gorme_izni")
        if gelir_faturasi_gorme_izni:
            izinler.gelir_faturasi_gorme_izni = True
        gelir_faturasi_kesme_izni = request.POST.get("gelir_faturasi_kesme_izni")
        if gelir_faturasi_kesme_izni:
            izinler.gelir_faturasi_kesme_izni = True
        gelir_faturasi_duzenleme_izni = request.POST.get("gelir_faturasi_duzenleme_izni")
        if gelir_faturasi_duzenleme_izni:
            izinler.gelir_faturasi_duzenleme_izni = True
        gelir_faturasi_silme_izni = request.POST.get("gelir_faturasi_silme_izni")
        if gelir_faturasi_silme_izni:
            izinler.gelir_faturasi_silme_izni = True
        #
        gider_faturasi_gorme_izni = request.POST.get("gider_faturasi_gorme_izni")
        if gider_faturasi_gorme_izni:
            izinler.gider_faturasi_gorme_izni = True
        gider_faturasi_kesme_izni = request.POST.get("gider_faturasi_kesme_izni")
        if gider_faturasi_kesme_izni:
            izinler.gider_faturasi_kesme_izni = True
        gider_faturasi_duzenleme_izni = request.POST.get("gider_faturasi_duzenleme_izni")
        if gider_faturasi_duzenleme_izni:
            izinler.gider_faturasi_duzenleme_izni = True
        gider_faturasi_silme_izni = request.POST.get("gider_faturasi_silme_izni")
        if gider_faturasi_silme_izni:
            izinler.gider_faturasi_silme_izni = True
        #
        
        kasa_gosterme_izni = request.POST.get("kasa_gosterme_izni")
        if kasa_gosterme_izni:
            izinler.kasa_gosterme_izni = True
        kasa_olusturma_izni = request.POST.get("kasa_olusturma_izni")
        if kasa_olusturma_izni:
            izinler.kasa_olusturma_izni = True
        kasa_guncelleme_izni = request.POST.get("kasa_guncelleme_izni")
        if kasa_guncelleme_izni:
            izinler.kasa_guncelleme_izni = True
        Kasa_silme_izni = request.POST.get("Kasa_silme_izni")
        if Kasa_silme_izni:
            izinler.Kasa_silme_izni = True
        #
        cari_gosterme_izni = request.POST.get("cari_gosterme_izni")
        if cari_gosterme_izni:
            izinler.cari_gosterme_izni = True
        cari_olusturma = request.POST.get("cari_olusturma")
        if cari_olusturma:
            izinler.cari_olusturma = True
        cari_guncelleme_izni = request.POST.get("cari_guncelleme_izni")
        if cari_guncelleme_izni:
            izinler.cari_guncelleme_izni = True
        cari_silme_izni = request.POST.get("cari_silme_izni")
        if cari_silme_izni:
            izinler.cari_silme_izni = True
        #
        personeller_gorme = request.POST.get("personeller_gorme")
        if personeller_gorme:
            izinler.personeller_gorme = True
        personeller_olusturma = request.POST.get("personeller_olusturma")
        if personeller_olusturma:
            izinler.personeller_olusturma = True
        personeller_duzenleme = request.POST.get("personeller_duzenleme")
        if personeller_duzenleme:
            izinler.personeller_duzenleme = True
        personeller_silme = request.POST.get("personeller_silme")
        if personeller_silme:
            izinler.personeller_silme = True
        #
        santiye_gorme = request.POST.get("santiye_gorme")
        if santiye_gorme:
            izinler.santiye_gorme = True
        santiye_olusturma  = request.POST.get("santiye_olusturma")
        if santiye_olusturma:
            izinler.santiye_olusturma = True
        santiye_duzenleme = request.POST.get("santiye_duzenleme")
        if santiye_duzenleme:
            izinler.santiye_duzenleme = True
        santiye_silme = request.POST.get("santiye_silme")
        if santiye_silme:
            izinler.santiye_silme = True
        #
        blog_gorme = request.POST.get("blog_gorme")
        if blog_gorme:
            izinler.blog_gorme = True
        blog_olusturma  = request.POST.get("blog_olusturma")
        if blog_olusturma:
            izinler.blog_olusturma = True
        blog_duzenleme = request.POST.get("blog_duzenleme")
        if blog_duzenleme:
            izinler.blog_duzenleme = True
        blog_silme = request.POST.get("blog_silme")
        if blog_silme:
            izinler.blog_silme = True
        #
        kalemleri_gorme = request.POST.get("kalemleri_gorme")
        if kalemleri_gorme:
            izinler.kalemleri_gorme = True
        kalemleri_olusturma  = request.POST.get("kalemleri_olusturma")
        if kalemleri_olusturma:
            izinler.kalemleri_olusturma = True
        kalemleri_duzenleme = request.POST.get("kalemleri_duzenleme")
        if kalemleri_duzenleme:
            izinler.kalemleri_duzenleme = True
        kalemleri_silme = request.POST.get("kalemleri_silme")
        if kalemleri_silme:
            izinler.kalemleri_silme = True
        #
        santiye_raporu_gorme = request.POST.get("santiye_raporu_gorme")
        if santiye_raporu_gorme:
            izinler.santiye_raporu_gorme = True
        santiye_raporu_olusturma  = request.POST.get("santiye_raporu_olusturma")
        if santiye_raporu_olusturma:
            izinler.santiye_raporu_olusturma = True
        santiye_raporu_duzenleme = request.POST.get("santiye_raporu_duzenleme")
        if santiye_raporu_duzenleme:
            izinler.santiye_raporu_duzenleme = True
        santiye_raporu_silme = request.POST.get("santiye_raporu_silme")
        if santiye_raporu_silme:
            izinler.santiye_raporu_silme = True
        #
        taseronlar_gorme = request.POST.get("taseronlar_gorme")
        if taseronlar_gorme:
            izinler.taseronlar_gorme = True
        taseronlar_olusturma = request.POST.get("taseronlar_olusturma")
        if taseronlar_olusturma:
            izinler.taseronlar_olusturma = True
        taseronlar_duzenleme = request.POST.get("taseronlar_duzenleme")
        if taseronlar_duzenleme:
            izinler.taseronlar_duzenleme = True
        taseronlar_silme = request.POST.get("taseronlar_silme")
        if taseronlar_silme:
            izinler.taseronlar_silme = True
        #
        hakedisler_gorme = request.POST.get("hakedisler_gorme")
        if hakedisler_gorme:
            izinler.hakedisler_gorme = True
        hakedisler_olusturma = request.POST.get("hakedisler_olusturma")
        if hakedisler_olusturma:
            izinler.hakedisler_olusturma = True
        hakedisler_duzenleme = request.POST.get("hakedisler_duzenleme")
        if hakedisler_duzenleme:
            izinler.hakedisler_duzenleme = True
        hakedisler_silme = request.POST.get("hakedisler_silme")
        if hakedisler_silme:
            izinler.hakedisler_silme = True
        #
        kasa_detay_izni = request.POST.get("kasa_detay_izni")
        if kasa_detay_izni:
            izinler.kasa_detay_izni = True
        cari_detay_izni = request.POST.get("cari_detay_izni")
        if cari_detay_izni:
            izinler.cari_detay_izni = True
        
        #
        gelir_kategorisi_gorme = request.POST.get("gelir_kategorisi_gorme")
        if gelir_kategorisi_gorme:  
            izinler.gelir_kategorisi_gorme = True
        gelir_kategorisi_olusturma = request.POST.get("gelir_kategorisi_olusturma")
        if gelir_kategorisi_olusturma : 
            izinler.gelir_kategorisi_olusturma = True
        gelir_kategorisi_guncelleme = request.POST.get("gelir_kategorisi_guncelleme")
        if gelir_kategorisi_guncelleme : 
            izinler.gelir_kategorisi_guncelleme = True
        gelir_kategorisi_silme = request.POST.get("gelir_kategorisi_silme")
        if gelir_kategorisi_silme : 
            izinler.gelir_kategorisi_silme = True
        
        #
        gider_kategorisi_gorme = request.POST.get("gider_kategorisi_gorme")
        if gider_kategorisi_gorme:  
            izinler.gider_kategorisi_gorme = True
        gider_kategorisi_olusturma = request.POST.get("gider_kategorisi_olusturma")
        if gider_kategorisi_olusturma : 
            izinler.gider_kategorisi_olusturma = True
        gider_kategorisi_guncelleme = request.POST.get("gider_kategorisi_guncelleme")
        if gider_kategorisi_guncelleme : 
            izinler.gider_kategorisi_guncelleme = True
        gider_kategorisi_silme = request.POST.get("gider_kategorisi_silme")
        if gider_kategorisi_silme : 
            izinler.gider_kategorisi_silme = True
        #
        #
        gelir_etiketi_gorme = request.POST.get("gelir_etiketi_gorme")
        if gelir_etiketi_gorme:  
            izinler.gelir_etiketi_gorme = True
        gelir_etiketi_olusturma = request.POST.get("gelir_etiketi_olusturma")
        if gelir_etiketi_olusturma : 
            izinler.gelir_etiketi_olusturma = True
        gelir_etiketi_guncelleme = request.POST.get("gelir_etiketi_guncelleme")
        if gelir_etiketi_guncelleme : 
            izinler.gelir_etiketi_guncelleme = True
        gelir_etiketi_silme = request.POST.get("gelir_etiketi_silme")
        if gelir_etiketi_silme : 
            izinler.gelir_etiketi_silme = True

        #
        gider_etiketi_gorme = request.POST.get("gider_etiketi_gorme")
        if gider_etiketi_gorme:  
            izinler.gider_etiketi_gorme = True
        gider_etiketi_olusturma = request.POST.get("gider_etiketi_olusturma")
        if gider_etiketi_olusturma : 
            izinler.gider_etiketi_olusturma = True
        gider_etiketi_guncelleme = request.POST.get("gider_etiketi_guncelleme")
        if gider_etiketi_guncelleme : 
            izinler.gider_etiketi_guncelleme = True
        gider_etiketi_silme = request.POST.get("gider_etiketi_silme")
        if gider_etiketi_silme : 
            izinler.gider_etiketi_silme = True
        #
        #
        urun_gorme = request.POST.get("urun_gorme")
        if urun_gorme:  
            izinler.urun_gorme = True
        urun_olusturma = request.POST.get("urun_olusturma")
        if urun_olusturma : 
            izinler.urun_olusturma = True
        urun_guncelleme = request.POST.get("urun_guncelleme")
        if urun_guncelleme : 
            izinler.urun_guncelleme = True
        urun_silme = request.POST.get("urun_silme")
        if urun_silme : 
            izinler.urun_silme = True

        #
        muhasabe_ayarlari_gorme = request.POST.get("muhasabe_ayarlari_gorme")
        if muhasabe_ayarlari_gorme : 
            izinler.muhasabe_ayarlari_gorme = True
        muhasabe_ayarlari_guncelleme = request.POST.get("muhasabe_ayarlari_guncelleme")
        if muhasabe_ayarlari_guncelleme : 
            izinler.muhasabe_ayarlari_guncelleme = True

        #
        gelir_faturasi_makbuz_gorme_izni = request.POST.get("gelir_faturasi_makbuz_gorme_izni")
        if gelir_faturasi_makbuz_gorme_izni:  
            izinler.gelir_faturasi_makbuz_gorme_izni = True
        gelir_faturasi_makbuz_kesme_izni = request.POST.get("gelir_faturasi_makbuz_kesme_izni")
        if gelir_faturasi_makbuz_kesme_izni : 
            izinler.gelir_faturasi_makbuz_kesme_izni = True
        gelir_faturasi_makbuz_duzenleme_izni = request.POST.get("gelir_faturasi_makbuz_duzenleme_izni")
        if gelir_faturasi_makbuz_duzenleme_izni : 
            izinler.gelir_faturasi_makbuz_duzenleme_izni = True
        gelir_faturasi_makbuz_silme_izni = request.POST.get("gelir_faturasi_makbuz_silme_izni")
        if gelir_faturasi_makbuz_silme_izni : 
            izinler.gelir_faturasi_makbuz_silme_izni = True
        #
        gider_faturasi_makbuz_gorme_izni = request.POST.get("gider_faturasi_makbuz_gorme_izni")
        if gider_faturasi_makbuz_gorme_izni:  
            izinler.gider_faturasi_makbuz_gorme_izni = True
        gider_faturasi_makbuz_kesme_izni = request.POST.get("gider_faturasi_makbuz_kesme_izni")
        if gider_faturasi_makbuz_kesme_izni : 
            izinler.gider_faturasi_makbuz_kesme_izni = True
        gider_faturasi_makbuz_duzenleme_izni = request.POST.get("gider_faturasi_makbuz_duzenleme_izni")
        if gider_faturasi_makbuz_duzenleme_izni : 
            izinler.gider_faturasi_makbuz_duzenleme_izni = True
        gider_faturasi_makbuz_silme_izni = request.POST.get("gider_faturasi_makbuz_silme_izni")
        if gider_faturasi_makbuz_silme_izni : 
            izinler.gider_faturasi_makbuz_silme_izni = True
        izinler.save()
    return redirect_with_language("main:kullanici_yetkileri_2",hash)

def proje_ekleme_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        #yetkili_adi
        if super_admin_kontrolu(request):
            kullanici_bilgisi  = request.POST.get("kullanici")
            proje_tip_adi   = request.POST.get("yetkili_adi")
            proje_tipi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = users ,Proje_tipi_adi = proje_tip_adi)
        
    return redirect_with_language("main:proje_tipi_2",hash)
#Proje Adı Silme
def proje_Adi_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        proje_tipi.objects.filter(id = id).update(silinme_bilgisi = True)
    
    return redirect_with_language("main:proje_tipi_2",hash)
#Proje Düzenlme
import folium
def proje_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        id = request.POST.get("buttonId")
    if super_admin_kontrolu(request):
        kullanici_bilgisi  = request.POST.get("kullanici")
        proje_tip_adi   = request.POST.get("yetkili_adi")
        silinmedurumu = request.POST.get("silinmedurumu")
        
        proje_tipi.objects.filter(id = id).update(proje_ait_bilgisi = users ,Proje_tipi_adi = proje_tip_adi )

    
    return redirect_with_language("main:proje_tipi_2",hash)
#Şantiye Projesi Ekleme
############################


def santiyeye_kalem_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            if True:
                projetipi = request.POST.getlist("projetipi")
                yetkili_adi = request.POST.get("yetkili_adi")
                santiye_agirligi = request.POST.get("katsayisi")
                finansal_agirlik = request.POST.get("blogsayisi")
                metraj = request.POST.get("metraj")
                tutar = request.POST.get("tutar")
                birim_bilgisi = request.POST.get("birim_bilgisi")
                kata_veya_binaya_daihil = request.POST.get("kata_veya_binaya_daihil")
                id = bloglar.objects.filter(id__in = projetipi).first()
                kalem = santiye_kalemleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                    proje_ait_bilgisi = users,
                    proje_santiye_Ait =id.proje_santiye_Ait,
                    kalem_adi = yetkili_adi,santiye_agirligi = santiye_agirligi,
                    santiye_finansal_agirligi = finansal_agirlik,
                    birimi = get_object_or_404(birimler,id =birim_bilgisi ),metraj = metraj,
                    tutari = tutar
                )
                if kata_veya_binaya_daihil == "0":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,int(i.kat_sayisi)):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = users,
                                proje_santiye_Ait = id.proje_santiye_Ait,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =kalem.id ),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
                elif kata_veya_binaya_daihil == "1":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,int(i.kat_sayisi)):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = users,
                                proje_santiye_Ait = id.proje_santiye_Ait,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =kalem.id ),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
                            break
                elif kata_veya_binaya_daihil == "2":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,4):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = users,
                                proje_santiye_Ait = id.proje_santiye_Ait,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =kalem.id ),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
                        
    return redirect_with_language("main:santiye_projesi_ekle_2",hash)

def blog_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        santiye_bilgisi = request.POST.get("santiye_bilgisi")
        blok_adi = request.POST.get("blok_adi")
        kat_sayisi = request.POST.get("kat_sayisi")
        baslangictarihi = request.POST.get("baslangictarihi")
        bitistarihi =request.POST.get("bitistarihi")
        bloglar.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
            proje_ait_bilgisi = get_object_or_404(santiye,id = santiye_bilgisi).proje_ait_bilgisi,
            proje_santiye_Ait = get_object_or_404(santiye,id = santiye_bilgisi),
            blog_adi = blok_adi,kat_sayisi = kat_sayisi,
            baslangic_tarihi = baslangictarihi,bitis_tarihi = bitistarihi
        )
    return redirect_with_language("main:santiye_projesi_ekle_2",hash)
def santiye_kalemleri_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        buttonId = request.POST.get("buttonId")
        santiye_kalemlerin_dagilisi.objects.filter(kalem_bilgisi__id =buttonId).delete()
        yetkili_adi = request.POST.get("yetkili_adi")
        santiye_agirligi = request.POST.get("katsayisi")
        finansal_agirlik = request.POST.get("blogsayisi")
        geri_don = request.POST.get("geri_don")
        metraj = request.POST.get("metraj")
        tutar = request.POST.get("tutar")
        birim_bilgisi = request.POST.get("birim_bilgisi")
        kata_veya_binaya_daihil = request.POST.get("kata_veya_binaya_daihil")
        projetipi = request.POST.getlist("projetipi")
        if request.user.is_superuser:
            
            santiye_kalemleri.objects.filter(id  = buttonId).update(
                            proje_ait_bilgisi = users,
                            kalem_adi = yetkili_adi,santiye_agirligi = santiye_agirligi,
                            santiye_finansal_agirligi = finansal_agirlik,birimi = get_object_or_404(birimler,id =birim_bilgisi ),metraj = metraj,
                        tutari = tutar
                        )
            if True:
                if kata_veya_binaya_daihil == "0":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,int(i.kat_sayisi)):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = users,
                                proje_santiye_Ait_id = geri_don,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =buttonId ),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
                elif kata_veya_binaya_daihil == "1":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,int(i.kat_sayisi)):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = users,
                                proje_santiye_Ait_id = geri_don,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =buttonId),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
                            break
                elif kata_veya_binaya_daihil == "2":
                    blog_lar = bloglar.objects.filter(id__in = projetipi)
                    for i in blog_lar:
                        for j in range(0,4):
                            santiye_kalemlerin_dagilisi.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                proje_ait_bilgisi = users,
                                proje_santiye_Ait_id = geri_don,
                                kalem_bilgisi = get_object_or_404(santiye_kalemleri,id =buttonId ),
                                kat = j,blog_bilgisi = get_object_or_404(bloglar,id =i.id ),
                            )
    return redirect_with_language("main:santiye_projesi_ekle_2",hash)
def kalem_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.kalemleri_silme:
                    buttonId = request.POST.get("buttonId")
                    geri_don = request.POST.get("geri_don")
                    santiye_kalemleri.objects.filter(id = buttonId).update(
                        silinme_bilgisi = True
                    )
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            buttonId = request.POST.get("buttonId")
            geri_don = request.POST.get("geri_don")
            santiye_kalemleri.objects.filter(id = buttonId).update(
                silinme_bilgisi = True
            )
    return redirect_with_language("main:santiye_projesi_ekle_2",hash)

def blog_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.blog_duzenleme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    if request.POST:
        santiye_bilgisi = request.POST.get("santiye_bilgisi")
        blog = request.POST.get("blog")
        blok_adi = request.POST.get("blok_adi")
        kat_sayisi = request.POST.get("kat_sayisi")
        baslangictarihi = request.POST.get("baslangictarihi")
        bitistarihi =request.POST.get("bitistarihi")
        bloglar.objects.filter(id = blog).update(
            proje_ait_bilgisi = get_object_or_404(santiye,id = santiye_bilgisi).proje_ait_bilgisi,
            proje_santiye_Ait = get_object_or_404(santiye,id = santiye_bilgisi),
            blog_adi = blok_adi,kat_sayisi = kat_sayisi,
            baslangic_tarihi = baslangictarihi,bitis_tarihi = bitistarihi
        )
       
    return redirect_with_language("main:santiye_projesi_ekle_2",hash)
def blog_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.user.kullanicilar_db:
        a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
        if a:
            if a.izinler.blog_silme:
                pass
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            return redirect_with_language("main:yetkisiz")
    if request.POST:
        buttonId = request.POST.get("buttonId")
        geri = request.POST.get("geri")
        blog_bilgisi = get_object_or_404(bloglar,id = buttonId)
        bloglar.objects.filter(id = buttonId).delete()
    return redirect_with_language("main:santiye_projesi_ekle_2",hash)

"""
content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
"""
def katman_sayfasi_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if super_admin_kontrolu(request):
        profile =katman.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        profile = katman.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = users)
        content["insaatlar"] = santiye.objects.filter(proje_ait_bilgisi =  users,silinme_bilgisi = False)
    content["santiyeler"] = profile
    content["top"]  = profile
    content["medya"] = profile
    return render(request,"santiye_yonetimi/katman.html",content)

def katman_ekle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            kullanici = users
            
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a.izinler.katman_olusturma:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                kullanici = request.user
        katman_adi = request.POST.get("taseron_adi")
        santiye_al = request.POST.get("blogbilgisi")
        dosya = request.FILES.get("file")
        katman.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
            proje_ait_bilgisi = kullanici,
            proje_santiye_Ait = get_object_or_none(santiye,id = santiye_al),
            katman_adi  = katman_adi,
            katman_dosyasi = dosya
        )
    return redirect_with_language("main:katman_sayfasi_2",hash)

def katman_sil_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            kullanici = request.POST.get("kullanici")
            buttonIdInput = request.POST.get("buttonId")
            katman.objects.filter(id = buttonIdInput).update(
                    silinme_bilgisi = True
                )
            return redirect_with_language("main:katman_sayfasi_2",hash)

    

def katman_duzenle_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    users = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = users
    if request.POST:
        if request.user.is_superuser:
            
            kullanici = users
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a.izinler.katman_olusturma:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                kullanici = request.user
        katman_adi = request.POST.get("taseron_adi")
        santiye_al = request.POST.get("blogbilgisi")
        dosya = request.FILES.get("file")
        buttonId = request.POST.get("buttonId")
        if dosya:
            katman.objects.filter(id = buttonId).update(
                proje_ait_bilgisi = kullanici,
                proje_santiye_Ait = get_object_or_none(santiye,id = santiye_al),
                katman_adi  = katman_adi,
                katman_dosyasi = dosya
            )
        else:
            katman.objects.filter(id = buttonId).update(
            proje_ait_bilgisi = kullanici,
            proje_santiye_Ait = get_object_or_none(santiye,id = santiye_al),
            katman_adi  = katman_adi
        )
    return redirect_with_language("main:katman_sayfasi_2",hash)


#taseron olaylari
def genel_rapor_sayfasi(request):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.genel_rapor_gorme:
                   kullanici =  request.user.kullanicilar_db
                   
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici =  request.user
    content["projeler"] = santiye.objects.filter(proje_ait_bilgisi = kullanici,silinme_bilgisi = False)
    content["personel_depertmani"] = calisanlar_kategorisi.objects.filter(kategori_kime_ait = kullanici,silinme_bilgisi = False)
    content["urunler"] =  urunler.objects.filter(urun_ait_oldugu = kullanici,silinme_bilgisi = False)
    content["imalat_kalemleri"] =  santiye_kalemleri.objects.filter(proje_ait_bilgisi = kullanici,silinme_bilgisi = False)
    content["santiyeler"] = genel_rapor.objects.filter(proje_ait_bilgisi = kullanici,silinme_bilgisi = False).order_by("tarih")
    from django.db.models import Sum
    # Kaybedilen gün sayısını hesapla ve content sözlüğüne ekle
    content["rapor_sayisi"] = genel_rapor.objects.filter(proje_ait_bilgisi=kullanici, silinme_bilgisi=False).aggregate(toplam=Sum('kayip_gun_sayisi'))['toplam'] or 0
    
    content["hava_durumu_kaynakli"] = genel_rapor.objects.filter(
        kayip_gun_sebebi="1", proje_ait_bilgisi=kullanici, silinme_bilgisi=False
    ).aggregate(toplam=Sum('kayip_gun_sayisi'))['toplam'] or 0

    content["ust_yuklenici"] = genel_rapor.objects.filter(
        kayip_gun_sebebi="0", proje_ait_bilgisi=kullanici, silinme_bilgisi=False
    ).aggregate(toplam=Sum('kayip_gun_sayisi'))['toplam'] or 0

    content["elektirik_kesintisi"] = genel_rapor.objects.filter(
        kayip_gun_sebebi="3", proje_ait_bilgisi=kullanici, silinme_bilgisi=False
    ).aggregate(toplam=Sum('kayip_gun_sayisi'))['toplam'] or 0
    return render(request,"santiye_yonetimi/genel_rapor.html",content)
#taseron olaylari
def genel_rapor_sayfasi_2(request,hash):
    content = sozluk_yapisi()
    d = decode_id(hash)
    content["hashler"] = hash
    kullanici = get_object_or_404(CustomUser,id = d)
    content["hash_bilgi"] = kullanici
    if super_admin_kontrolu(request):
        
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    
    content["projeler"] = santiye.objects.filter(proje_ait_bilgisi = kullanici,silinme_bilgisi = False)
    content["personel_depertmani"] = calisanlar_kategorisi.objects.filter(kategori_kime_ait = kullanici,silinme_bilgisi = False)
    content["urunler"] =  urunler.objects.filter(urun_ait_oldugu = kullanici,silinme_bilgisi = False)
    content["imalat_kalemleri"] =  santiye_kalemleri.objects.filter(proje_ait_bilgisi = kullanici,silinme_bilgisi = False)
    content["santiyeler"] = genel_rapor.objects.filter(proje_ait_bilgisi = kullanici,silinme_bilgisi = False).order_by("tarih")
    from django.db.models import Sum
    # Kaybedilen gün sayısını hesapla ve content sözlüğüne ekle
    content["rapor_sayisi"] = genel_rapor.objects.filter(proje_ait_bilgisi=kullanici, silinme_bilgisi=False).aggregate(toplam=Sum('kayip_gun_sayisi'))['toplam'] or 0
    
    content["hava_durumu_kaynakli"] = genel_rapor.objects.filter(
        kayip_gun_sebebi="1", proje_ait_bilgisi=kullanici, silinme_bilgisi=False
    ).aggregate(toplam=Sum('kayip_gun_sayisi'))['toplam'] or 0

    content["ust_yuklenici"] = genel_rapor.objects.filter(
        kayip_gun_sebebi="0", proje_ait_bilgisi=kullanici, silinme_bilgisi=False
    ).aggregate(toplam=Sum('kayip_gun_sayisi'))['toplam'] or 0

    content["elektirik_kesintisi"] = genel_rapor.objects.filter(
        kayip_gun_sebebi="3", proje_ait_bilgisi=kullanici, silinme_bilgisi=False
    ).aggregate(toplam=Sum('kayip_gun_sayisi'))['toplam'] or 0
    return render(request,"santiye_yonetimi/genel_rapor.html",content)
#taseron olaylari
def genel_rapor_olustur(request):
    content = sozluk_yapisi()
    if request.POST:
        rapor_tarihi = request.POST.get("rapor_tarihi")
        secili_santiye = request.POST.get("santiye")
        depertman = request.POST.getlist("depert")
        personel = request.POST.getlist("depertsayisi")
        hava_durumu_sicaklik = request.POST.getlist("hava_durumu_sicaklik")
        hava_durumu_ruzgar = request.POST.getlist("hava_durumu_ruzgar")
        malzeme = request.POST.getlist("malzeme")
        malzemesayisi = request.POST.getlist("malzemesayisi")
        imalat = request.POST.getlist("imalat")
        aciklama = request.POST.getlist("aciklama")
        aciklamalar = request.POST.getlist("aciklamalar")
        rapor_bitis_tarihi =request.POST.get("rapor_bitis_tarihi")
        kayipgun = request.POST.get("kayipgun")
        kayipsebebi = request.POST.get("kayipsebebi")
        otherReason = request.POST.get("otherReason")
        if kayipgun:
            pass
        else:
            kayipgun = 0
        if super_admin_kontrolu(request):
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.genel_rapor_gorme:
                        kullanici =  request.user.kullanicilar_db
                    
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                kullanici =  request.user
        veri = genel_rapor.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,raporu_olusturan = request.user,
                                   proje_santiye_Ait = get_object_or_none(santiye,id=secili_santiye),
                                   tarih =rapor_tarihi,bitis_tarih = rapor_bitis_tarihi,kayip_gun_sayisi = float(kayipgun),
                                   kayip_gun_aciklamasi =otherReason , kayip_gun_sebebi = kayipsebebi )
        for i in range(len(depertman)):
            genel_personel.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),hangi_rapor = kullanici,
                                          proje_ait_bilgisi = get_object_or_none(genel_rapor,id = veri.id),
                                          personel_departmani = get_object_or_none(calisanlar_kategorisi,id = depertman[i]),
                                          personel_sayisi = float(personel[i]) )
        for i in range(len(hava_durumu_sicaklik)):
            genel_hava_durumu.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),hangi_rapor = kullanici,
                                          proje_ait_bilgisi = get_object_or_none(genel_rapor,id = veri.id),
                                          hava_durumu_sicaklik =  hava_durumu_sicaklik[i],
                                          hava_durumu_ruzgar = float(hava_durumu_ruzgar[i]) )
        for i in range(len(malzeme)):
            gelen_malzeme.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),hangi_rapor = kullanici,
                                          proje_ait_bilgisi = get_object_or_none(genel_rapor,id = veri.id),
                                          urun = get_object_or_none(urunler,id = malzeme[i]),
                                          urun_adeti = float(malzemesayisi[i]) )
        for i in range(len(imalat)):
            genel_imalat.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),hangi_rapor = kullanici,
                                          proje_ait_bilgisi = get_object_or_none(genel_rapor,id = veri.id),
                                          imalet_kalemi = get_object_or_none(santiye_kalemleri,id = imalat[i]),
                                          imalat_aciklama = aciklama[i] )
        for i in range(len(aciklamalar)):
            genel_aciklamalar.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),hangi_rapor = kullanici,
                                          proje_ait_bilgisi = get_object_or_none(genel_rapor,id = veri.id),
                                          genel_aciklama = aciklamalar[i] )
    return redirect_with_language("main:genel_rapor_sayfasi")
#taseron olaylari
def genel_rapor_olustur_2(request,hash):
    content = sozluk_yapisi()
    if request.POST:
        rapor_tarihi = request.POST.get("rapor_tarihi")
        secili_santiye = request.POST.get("santiye")
        depertman = request.POST.getlist("depert")
        personel = request.POST.getlist("depertsayisi")
        hava_durumu_sicaklik = request.POST.getlist("hava_durumu_sicaklik")
        hava_durumu_ruzgar = request.POST.getlist("hava_durumu_ruzgar")
        malzeme = request.POST.getlist("malzeme")
        malzemesayisi = request.POST.getlist("malzemesayisi")
        imalat = request.POST.getlist("imalat")
        aciklama = request.POST.getlist("aciklama")
        aciklamalar = request.POST.getlist("aciklamalar")
        rapor_bitis_tarihi =request.POST.get("rapor_bitis_tarihi")
        kayipgun = request.POST.get("kayipgun")
        kayipsebebi = request.POST.get("kayipsebebi")
        otherReason = request.POST.get("otherReason")
        if kayipgun:
            pass
        else:
            kayipgun = 0
        if super_admin_kontrolu(request):
            d = decode_id(hash)
            content["hashler"] = hash
            kullanici = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = kullanici
        
        veri = genel_rapor.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,raporu_olusturan = request.user,
                                   proje_santiye_Ait = get_object_or_none(santiye,id=secili_santiye),
                                   tarih =rapor_tarihi,bitis_tarih = rapor_bitis_tarihi,kayip_gun_sayisi = float(kayipgun),
                                   kayip_gun_aciklamasi =otherReason , kayip_gun_sebebi = kayipsebebi )
        for i in range(len(depertman)):
            genel_personel.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),hangi_rapor = kullanici,
                                          proje_ait_bilgisi = get_object_or_none(genel_rapor,id = veri.id),
                                          personel_departmani = get_object_or_none(calisanlar_kategorisi,id = depertman[i]),
                                          personel_sayisi = float(personel[i]) )
        for i in range(len(hava_durumu_sicaklik)):
            genel_hava_durumu.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),hangi_rapor = kullanici,
                                          proje_ait_bilgisi = get_object_or_none(genel_rapor,id = veri.id),
                                          hava_durumu_sicaklik =  hava_durumu_sicaklik[i],
                                          hava_durumu_ruzgar = float(hava_durumu_ruzgar[i]) )
        for i in range(len(malzeme)):
            gelen_malzeme.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),hangi_rapor = kullanici,
                                          proje_ait_bilgisi = get_object_or_none(genel_rapor,id = veri.id),
                                          urun = get_object_or_none(urunler,id = malzeme[i]),
                                          urun_adeti = float(malzemesayisi[i]) )
        for i in range(len(imalat)):
            genel_imalat.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),hangi_rapor = kullanici,
                                          proje_ait_bilgisi = get_object_or_none(genel_rapor,id = veri.id),
                                          imalet_kalemi = get_object_or_none(santiye_kalemleri,id = imalat[i]),
                                          imalat_aciklama = aciklama[i] )
        for i in range(len(aciklamalar)):
            genel_aciklamalar.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),hangi_rapor = kullanici,
                                          proje_ait_bilgisi = get_object_or_none(genel_rapor,id = veri.id),
                                          genel_aciklama = aciklamalar[i] )
    return redirect_with_language("main:genel_rapor_sayfasi_2",hash)

def genel_rapor_onaylama(request,id):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.genel_rapor_gorme:
                   kullanici =  request.user.kullanicilar_db
                   
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici =  request.user
    content["logosu"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu =kullanici).last()
    content["santiyeler"] = get_object_or_404(genel_rapor,proje_ait_bilgisi = kullanici,silinme_bilgisi = False,id = id) 
    content["gelen_malzeme"] = gelen_malzeme.objects.filter(hangi_rapor = kullanici, proje_ait_bilgisi__id = id )
    content["genel_personel"] = genel_personel.objects.filter(hangi_rapor = kullanici, proje_ait_bilgisi__id = id )
    content["genel_imalat"] = genel_imalat.objects.filter(hangi_rapor = kullanici, proje_ait_bilgisi__id = id )
    content["genel_aciklamalar"] = genel_aciklamalar.objects.filter(hangi_rapor = kullanici, proje_ait_bilgisi__id = id )
    content["genel_hava_durumu"] = genel_hava_durumu.objects.filter(hangi_rapor = kullanici, proje_ait_bilgisi__id = id ).last()
    content["perseonel_iznleri_gonder"] =bagli_kullanicilar.objects.filter(izinler__izinlerin_sahibi_kullanici = kullanici,izinler__genel_rapor_onaylama = True)
    return render(request,"santiye_yonetimi/genel_rapor_onaylama.html",content)
def genel_rapor_onaylama_2(request,id,hash):
    content = sozluk_yapisi()

    if super_admin_kontrolu(request):
        d = decode_id(hash)
        content["hashler"] = hash
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hash_bilgi"] = kullanici

        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    
    content["santiyeler"] = get_object_or_404(genel_rapor,proje_ait_bilgisi = kullanici,silinme_bilgisi = False,id = id) 
    content["gelen_malzeme"] = gelen_malzeme.objects.filter(hangi_rapor = kullanici, proje_ait_bilgisi__id = id )
    content["genel_personel"] = genel_personel.objects.filter(hangi_rapor = kullanici, proje_ait_bilgisi__id = id )
    content["genel_imalat"] = genel_imalat.objects.filter(hangi_rapor = kullanici, proje_ait_bilgisi__id = id )
    content["genel_aciklamalar"] = genel_aciklamalar.objects.filter(hangi_rapor = kullanici, proje_ait_bilgisi__id = id )
    content["genel_hava_durumu"] = genel_hava_durumu.objects.filter(hangi_rapor = kullanici, proje_ait_bilgisi__id = id ).last()
    content["perseonel_iznleri_gonder"] =bagli_kullanicilar.objects.filter(izinler__izinlerin_sahibi_kullanici = kullanici,izinler__genel_rapor_onaylama = True)
    return render(request,"santiye_yonetimi/genel_rapor_onaylama.html",content)
def rapor_onaylama(request):
    if request.POST:
        buttonId = request.POST.get("buttonId")
        kullanici_bilgisi = request.POST.get("kullanici_bilgisi")
        sifre_bilgisi = request.POST.get("sifre_bilgisi")
        if super_admin_kontrolu(request):
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.genel_rapor_gorme:
                        kullanici =  request.user.kullanicilar_db
                        
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                kullanici =  request.user
        users = get_object_or_404(CustomUser,id = kullanici_bilgisi)
        if users.imza_sifresi == sifre_bilgisi:
            genel_rapor.objects.filter(id = buttonId).update(raporu_onaylayan = users,onaylama_tarihi =datetime.now())
        else:
            pass
        return redirect_with_language("main:genel_rapor_onaylama",buttonId)
def rapor_onaylama_2(request,hash):
    content = sozluk_yapisi()
    


    if request.POST:
        buttonId = request.POST.get("buttonId")
        kullanici_bilgisi = request.POST.get("kullanici_bilgisi")
        sifre_bilgisi = request.POST.get("sifre_bilgisi")
        if super_admin_kontrolu(request):
            d = decode_id(hash)
            content["hashler"] = hash
            kullanici = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = kullanici
        
        users = get_object_or_404(CustomUser,id = kullanici_bilgisi)
        if users.imza_sifresi == sifre_bilgisi:
            genel_rapor.objects.filter(id = buttonId).update(raporu_onaylayan = users,onaylama_tarihi =datetime.now())
        else:
            pass
        return redirect_with_language("main:genel_rapor_sayfasi_2",hash)
def rapor_sil(request):
    if request.POST:
        buttonId = request.POST.get("buttonId")
        if super_admin_kontrolu(request):
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.genel_rapor_onaylama:
                        kullanici =  request.user.kullanicilar_db
                    
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                kullanici =  request.user
        genel_rapor.objects.filter(id = buttonId).update(silinme_bilgisi = True)
        return redirect_with_language("main:genel_rapor_sayfasi")
def rapor_sil_2(request,hash):
    content = sozluk_yapisi()
    if request.POST:
        buttonId = request.POST.get("buttonId")
        if super_admin_kontrolu(request):
            d = decode_id(hash)
            content["hashler"] = hash
            kullanici = get_object_or_404(CustomUser,id = d)
            content["hash_bilgi"] = kullanici

        
        genel_rapor.objects.filter(id = buttonId).update(silinme_bilgisi = True)
        return redirect_with_language("main:genel_rapor_sayfasi_2",hash)
def rapor_gonder(request, rapor_id):
    # Belirtilen ID'ye sahip genel rapor bilgisi yoksa hata döner
    genel_rapor_bilgisi = get_object_or_404(genel_rapor, id=rapor_id)

    # İlgili diğer bilgiler, genel rapora bağlı nesnelerden alınır
    gelen_malzemeler_bilgisi = gelen_malzeme.objects.filter(proje_ait_bilgisi=genel_rapor_bilgisi)
    genel_personel_bilgisi = genel_personel.objects.filter(proje_ait_bilgisi=genel_rapor_bilgisi)
    genel_imalat_bilgisi = genel_imalat.objects.filter(proje_ait_bilgisi=genel_rapor_bilgisi)
    genel_aciklama_bilgisi = genel_aciklamalar.objects.filter(proje_ait_bilgisi=genel_rapor_bilgisi)
    genel_hava_bilgisi = genel_hava_durumu.objects.filter(proje_ait_bilgisi=genel_rapor_bilgisi).last()
    print(gelen_malzemeler_bilgisi,genel_personel_bilgisi,genel_imalat_bilgisi,genel_aciklama_bilgisi,genel_hava_bilgisi)
    # Rapor verilerini toplama
    fatura_data = {
        'id': genel_rapor_bilgisi.id if genel_rapor_bilgisi else None,
        'santiye_id': genel_rapor_bilgisi.proje_santiye_Ait.id if genel_rapor_bilgisi and genel_rapor_bilgisi.proje_santiye_Ait else None,
        'santiye_Adi': genel_rapor_bilgisi.proje_santiye_Ait.proje_adi if genel_rapor_bilgisi and genel_rapor_bilgisi.proje_santiye_Ait else None,
        'rapor_tarihi': genel_rapor_bilgisi.tarih.strftime("%d.%m.%Y") if genel_rapor_bilgisi and genel_rapor_bilgisi.tarih else None,
        "bitis_tarih" : genel_rapor_bilgisi.bitis_tarih.strftime("%d.%m.%Y") if genel_rapor_bilgisi and genel_rapor_bilgisi.bitis_tarih else None,
        "kayip_gun_aciklamasi" : genel_rapor_bilgisi.kayip_gun_aciklamasi if genel_rapor_bilgisi and genel_rapor_bilgisi.kayip_gun_aciklamasi else None,
        "kayip_gun_sebebi": genel_rapor_bilgisi.kayip_gun_sebebi if genel_rapor_bilgisi and genel_rapor_bilgisi.kayip_gun_sebebi else None,
        'raporu_olusturan': genel_rapor_bilgisi.raporu_olusturan.last_name if genel_rapor_bilgisi and genel_rapor_bilgisi.raporu_olusturan else None,
        'rapor_sahibi': genel_rapor_bilgisi.raporu_olusturan.first_name if genel_rapor_bilgisi and genel_rapor_bilgisi.raporu_olusturan else None,
        'hava_durumu_sicaklik': genel_hava_bilgisi.hava_durumu_sicaklik if genel_hava_bilgisi else None,
        'hava_durumu_ruzgar': genel_hava_bilgisi.hava_durumu_ruzgar if genel_hava_bilgisi else None,
        "gelen_malzemeler_bilgisi": [
            {
                "id": malzeme.id ,
                "malzeme_id": malzeme.urun.id if malzeme.urun else None,
                "malzeme_adi": malzeme.urun.urun_adi if malzeme.urun else None,
                "malzeme_adedi": malzeme.urun_adeti
            } for malzeme in gelen_malzemeler_bilgisi
        ],
        "genel_personel_bilgisi": [
            {
                "id": departman.id,
                "departman_id": departman.personel_departmani.id,
                "departman_adi": departman.personel_departmani.kategori_isimi,
                "personel_adedi": departman.personel_sayisi
            } for departman in genel_personel_bilgisi
        ],
        "genel_imalat_bilgisi": [
            {
                "id": imalat.id,
                "imalat_id": imalat.imalet_kalemi.id,
                "imalat_adi": imalat.imalet_kalemi.kalem_adi,
                "imalat_aciklama": imalat.imalat_aciklama
            } for imalat in genel_imalat_bilgisi
        ],
        "genel_aciklama_bilgisi": [
            {
                "id": aciklama.id,
                "aciklama": aciklama.genel_aciklama
            } for aciklama in genel_aciklama_bilgisi
        ],
    }
    print(fatura_data)
    # Veriyi JSON olarak döndür
    return JsonResponse(fatura_data)




def santiye_projeleri(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici =request.user
    content["blok"] = get_object_or_none(bloglar,id = id)
    #content["projeler"] = santiye.objects.filter(proje_ait_bilgisi = kullanici,silinme_bilgisi = False)
    return render(request,"checklist/santiye_projeleri.html",content)
def santiye_proje_olustur(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici =request.user
    if request.POST:
        blok = request.POST.get("blok")
        kat_basina_daire = request.POST.get("kat_basina_daire")
        daire_numarai = request.POST.getlist("daire_numarai")
        ortak_alanadi = request.POST.getlist("ortak_alanadi")
        verisi = request.POST.getlist("verisi")
        cepheaciklmasi = request.POST.getlist("cepheaciklmasi")
        cephe_verisi = request.POST.getlist("cephe_verisi")
        kat_bilgisi = 1
        block  = get_object_or_none(bloglar,id = blok)
        imalatlari = santiye_imalat_kalemleri.objects.filter(proje_santiye_Ait = block.proje_santiye_Ait,silinme_bilgisi = False)
        for i in range(1,len(daire_numarai)+1):
            a = checkdaireleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),daire_no = daire_numarai[i-1] ,kat_daire_sayisi = kat_basina_daire,kat = kat_bilgisi,proje_ait_bilgisi = kullanici,blog_bilgisi = block,proje_santiye_Ait= block.proje_santiye_Ait)
            if i / int(kat_basina_daire) == 1:
                kat_bilgisi += 1
            for imalat in imalatlari:
                imalat_daire_balama.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,blog_bilgisi = block,proje_santiye_Ait= block.proje_santiye_Ait,daire_bilgisi = a,imalat_detayi = imalat)
            print(i)
        for i in range(len(ortak_alanadi)):
            blog_ortak_alan_ve_cepheleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,blog_ait_bilgisi = block,proje_santiye_Ait= block.proje_santiye_Ait,aciklama_adi = ortak_alanadi[i],bolum_icerigi = verisi[i])
        for i in range(len(cepheaciklmasi)):
            blog_ortak_alan_ve_cepheleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,blog_ait_bilgisi = block,proje_santiye_Ait= block.proje_santiye_Ait,aciklama_adi = cepheaciklmasi[i],bolum_icerigi = cephe_verisi[i])
    return redirect_with_language("main:yapilarim",block.proje_santiye_Ait.id)
    

def santiye_onay_listesi(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:       
            kullanici =request.user
    profile = checkdaireleri.objects.filter(proje_ait_bilgisi = kullanici) 
    content["santiyeler"] = profile
    return render(request,"checklist/santiye_onay_listesi.html",content)
def santiye_sablonu(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:       
            kullanici =request.user
    content["santiye"] = get_object_or_404(santiye,id = id)
    return render(request,"checklist/santiye_sablonu.html",content)
def santiye_kontrolculeri_isle(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:   
            kullanici =request.user
    content["kullanicilar"] = CustomUser.objects.filter(kullanicilar_db = kullanici,is_superuser = False)
    content["santiyeler"] = get_object_or_none(santiye,id = id)
    kontroller = []

    a = santiye_imalat_kalemleri.objects.filter(silinme_bilgisi = False,proje_santiye_Ait__id =id)
    for i in a:
        if i.is_grubu in kontroller:
            pass
        else:
            kontroller.append(i.is_grubu)
    content["is_gruplari"] = kontroller
    #content["kontrolculer"] = bagli_kullanicilar.objects.filter(izinler__santiye_kontrol = True)
    return render(request,"checklist/santiye_kontrolculeri.html",content)

def santiye_onay_listesi_kontrol(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:       
            kullanici =request.user
    profile = checkdaireleri.objects.filter(proje_ait_bilgisi = kullanici,proje_santiye_Ait__id = id) 
    content["santiyeler"] = profile
    content["kullanicilar"] = CustomUser.objects.filter(kullanicilar_db = kullanici,is_superuser = False)
    return render(request,"checklist/santiye_onay_listesi.html",content)
def santiye_kontrolculeri_ekle(request,id,slug):
    content = sozluk_yapisi()
    
    if True:
        if super_admin_kontrolu(request):
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.santiye_kontrol:
                        kullanici =request.user.kullanicilar_db 
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                kullanici =request.user
    content["santiye"] = id
    content["is_grubu"] = slug
    content["users"] =  CustomUser.objects.filter(kullanicilar_db = kullanici,is_superuser = False)
    content["kontrolculer"] = check_liste_onaylama_gruplari.objects.filter(imalat_kalemi_ait__proje_santiye_Ait__id = id,imalat_kalemi_ait__imalat_detayi__is_grubu = slug).values("onaylayan").distinct()

    return render(request,"checklist/kontrolculeri_isleme_sayfasi.html",content)

def kontrolculeri_kaydet(request):
    if request.POST:
        kullanicilarim = request.POST.getlist("kullanici")
        santiye = request.POST.get("santiye")
        is_grubu = request.POST.get("is_grubu")
        if super_admin_kontrolu(request):
            pass
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.santiye_kontrol:
                        kullanici =request.user.kullanicilar_db
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                kullanici =request.user
        
        daireler_kontrol_list = imalat_daire_balama.objects.filter(proje_ait_bilgisi = kullanici,proje_santiye_Ait__id = santiye,imalat_detayi__is_grubu = is_grubu)
        for i in daireler_kontrol_list:
            for k in kullanicilarim:
                check_liste_onaylama_gruplari.objects.filter(imalat_kalemi_ait = i).delete()
        for i in daireler_kontrol_list:
            for k in kullanicilarim:
                check_liste_onaylama_gruplari.objects.create(imalat_kalemi_ait = i,onaylayan = get_object_or_none(CustomUser,id = k))
        return redirect_with_language("main:santiye_kontrolculeri_isle",santiye)


def santiye_sablonu_duzenle(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:       
            kullanici =request.user
    content["santiye"] = get_object_or_404(santiye,id = id)
    content["santiye_sablonlari"]  = get_object_or_none(santiye_sablonlari,proje_santiye_Ait = get_object_or_404(santiye,id = id),silinme_bilgisi = False)
    content["sanytiye_sablon_bolumleri"] = sanytiye_sablon_bolumleri.objects.filter(proje_santiye_Ait = get_object_or_404(santiye,id = id),silinme_bilgisi = False)
    content["santiye_imalat_kalemleri"] = santiye_imalat_kalemleri.objects.filter(proje_santiye_Ait = get_object_or_404(santiye,id = id),silinme_bilgisi = False)
    content["imalat_kalemleri_imalat_detaylari"] = imalat_kalemleri_imalat_detaylari.objects.filter(proje_santiye_Ait = get_object_or_404(santiye,id = id),silinme_bilgisi = False)
    return render(request,"checklist/sablon_duzeltme.html",content)

import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt


@csrf_exempt
def save_template(request):
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
    else:       
        kullanici =request.user
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print(data)
            #şablonu oluştur
            sablon = santiye_sablonlari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,
                    sablon_adi =data.get('name', 'Unnamed Template'),sablon_durumu = data.get('projectType', 'Unknown') ,
                    proje_santiye_Ait = get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')))
            
            for section_data in data.get('sections', []):
                sablon_bolumleri = sanytiye_sablon_bolumleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,proje_santiye_Ait = get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')) , sablon_adi = get_object_or_none(santiye_sablonlari,id = sablon.id),bolum =section_data.get('type', 'Unknown')  )
                for category_data in section_data.get('categories', []):
                    sablon_imalat_olayi = santiye_imalat_kalemleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                        proje_ait_bilgisi = kullanici ,proje_santiye_Ait =get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')),
                        detay = get_object_or_none(sanytiye_sablon_bolumleri,id = sablon_bolumleri.id), 
                        icerik = category_data.get('name', 'Unnamed Category'),is_grubu = category_data.get('workGroup', 'Unknown Work Group') )
                    for item_name in category_data.get('checklistItems', []):
                        imalat_kalemleri_imalat_detaylari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,
                        proje_santiye_Ait= get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')),
                        icerik = get_object_or_none(santiye_imalat_kalemleri,id = sablon_imalat_olayi.id),
                        imalat_detayi = item_name)
            return JsonResponse({"status": "success", "message": "Şablon başarıyla kaydedildi."})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    else:
        return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)



@csrf_exempt
def save_template_duzenle(request):
    if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
    else:       
        kullanici =request.user
    if request.method == 'POST':
        if True:
            data = json.loads(request.body)
            print(data)
            #şablonu oluştur
            sablon = santiye_sablonlari.objects.filter(proje_santiye_Ait = get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown'))).update(proje_ait_bilgisi = kullanici,
                    sablon_adi =data.get('name', 'Unnamed Template'),sablon_durumu = data.get('projectType', 'Unknown') ,
                    proje_santiye_Ait = get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')))
            sablon = santiye_sablonlari.objects.filter(proje_santiye_Ait = get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown'))).last()
            sanytiye_sablon_bolumleri.objects.filter(proje_santiye_Ait = get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown'))).update(silinme_bilgisi = True  )
            for section_data in data.get('sections', []):
                if section_data.get('id') :
                    sablon_bolumleri = sanytiye_sablon_bolumleri.objects.filter(id = int(section_data.get('id'))).update(proje_ait_bilgisi = kullanici,proje_santiye_Ait = get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')) , sablon_adi = get_object_or_none(santiye_sablonlari,id = sablon.id),bolum =section_data.get('type', 'Unknown'),silinme_bilgisi = False  )
                    sablon_bolumleri = sanytiye_sablon_bolumleri.objects.filter(id = section_data.get('id', 'Unknown')).last()
                else:
                    sablon_bolumleri = sanytiye_sablon_bolumleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,proje_santiye_Ait = get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')) , sablon_adi = get_object_or_none(santiye_sablonlari,id = sablon.id),bolum =section_data.get('type', 'Unknown')  )
                for category_data in section_data.get('categories', []):
                    if category_data.get('last_name'):
                        sablon_imalat_olayi = santiye_imalat_kalemleri.objects.filter(id = category_data.get('last_name')).update(
                        proje_ait_bilgisi = kullanici ,proje_santiye_Ait =get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')),
                        detay = get_object_or_none(sanytiye_sablon_bolumleri,id = section_data.get('id', 'Unknown')), 
                        icerik = category_data.get('name'),is_grubu = category_data.get('workGroup', 'Unknown Work Group') )
                        imalat_kalemleri_imalat_detaylari.objects.filter(icerik__id = category_data.get('last_name', 'Unnamed Category')).update(silinme_bilgisi = True)
                        for item_name in category_data.get('checklistItems', []):
                            imalat_kalemleri_imalat_detaylari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,
                            proje_santiye_Ait= get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')),
                            icerik = get_object_or_none(santiye_imalat_kalemleri,id = category_data.get('last_name', 'Unnamed Category')),
                            imalat_detayi = item_name)
                    else:
                        sablon_imalat_olayi = santiye_imalat_kalemleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                        proje_ait_bilgisi = kullanici ,proje_santiye_Ait =get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')),
                        detay = get_object_or_none(sanytiye_sablon_bolumleri,id = sablon_bolumleri.id), 
                        icerik = category_data.get('name', 'Unnamed Category'),is_grubu = category_data.get('workGroup', 'Unknown Work Group') )
                        for item_name in category_data.get('checklistItems', []):
                            imalat_kalemleri_imalat_detaylari.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),proje_ait_bilgisi = kullanici,
                            proje_santiye_Ait= get_object_or_none(santiye,id = data.get('santiyeId', 'Unknown')),
                            icerik = get_object_or_none(santiye_imalat_kalemleri,id = sablon_imalat_olayi.id),
                            imalat_detayi = item_name)
            return JsonResponse({"status": "success", "message": "Şablon başarıyla kaydedildi."})
        
    else:
        return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)

def santiyelerim(request):
    content = sozluk_yapisi()
    m = folium.Map(location=[20, 0], zoom_start=2)
    # Haritayı HTML'ye dönüştürme
    map_html = m._repr_html_()
    content['map'] = map_html
    content["birim_bilgisi"] = birimler.objects.filter(silinme_bilgisi = False)
    content["proje_tipleri"] = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi =  request.user)
    if super_admin_kontrolu(request):
        profile =santiye.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_gorme:
                    profile = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user.kullanicilar_db)
                    content["proje_tipleri"] = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi =  request.user.kullanicilar_db)
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            profile = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = request.user)
            content["proje_tipleri"] = proje_tipi.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi =  request.user)
    if request.GET.get("search"):
        search = request.GET.get("search")
        if super_admin_kontrolu(request):
            profile =santiye.objects.filter(Q(proje_ait_bilgisi__last_name__icontains = search)|Q(Proje_tipi_adi__icontains = search))
            kullanicilar = CustomUser.objects.filter( kullanicilar_db = None,is_superuser = False).order_by("-id")
            content["kullanicilar"] =kullanicilar
        else:
            if request.user.kullanicilar_db:
                a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
                if a:
                    if a.izinler.santiye_gorme:
                        profile = santiye.objects.filter(Q(proje_ait_bilgisi = request.user.kullanicilar_db) & Q(Proje_tipi_adi__icontains = search)& Q(silinme_bilgisi = False))
                    else:
                        return redirect_with_language("main:yetkisiz")
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                profile = santiye.objects.filter(Q(proje_ait_bilgisi = request.user) & Q(Proje_tipi_adi__icontains = search)& Q(silinme_bilgisi = False))
    page_num = request.GET.get('page', 1)
    paginator = Paginator(profile, 10) # 6 employees per page

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
            # if page is not an integer, deliver the first page
        page_obj = paginator.page(1)
    except EmptyPage:
            # if the page is out of range, deliver the last page
        page_obj = paginator.page(paginator.num_pages)
    content["santiyeler"] = profile

    return render(request,"checklist/santiyeler.html",content)


def yapilarim(request,id):
    content = sozluk_yapisi()
    m = folium.Map(location=[20, 0], zoom_start=2)
    # Haritayı HTML'ye dönüştürme
    if super_admin_kontrolu(request):
        profile =santiye.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_gorme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
        
    profile = bloglar.objects.filter(proje_santiye_Ait = get_object_or_none(santiye,id = id),proje_ait_bilgisi = kullanici) 
    content["santiyeler"] = profile

    return render(request,"checklist/yapilar.html",content)

def daireleri_gor(request,id):
    content = sozluk_yapisi()
    # Haritayı HTML'ye dönüştürme
    if super_admin_kontrolu(request):
        profile =santiye.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_gorme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
        
    profile = checkdaireleri.objects.filter(blog_bilgisi__id = id,proje_ait_bilgisi = kullanici) 
    content["santiyeler"] = profile

    return render(request,"checklist/daireler.html",content)



def santiye_kontrol_detayi(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")   
        else:
            kullanici =request.user  
    daire = get_object_or_404(checkdaireleri,id = id)
    content["santiye"] = daire
    content["ortak_alanlar"] = blog_ortak_alan_ve_cepheleri.objects.filter(blog_ait_bilgisi = daire.blog_bilgisi)
    content["imalatlar"] = imalat_daire_balama.objects.filter(daire_bilgisi = daire,tamamlanma_bilgisi = False)
    content["sanytiye_sablon_bolumleri"] = sanytiye_sablon_bolumleri.objects.filter(proje_santiye_Ait = daire.proje_santiye_Ait)
    print(content)
    return render(request,"checklist/santiye_kontrol_detayi.html",content)

def daire_imalat_checklist(request):
    if request.POST:
        check = request.POST.getlist("check")
        aciklma = request.POST.get("aciklma")
        daire = request.POST.get("daire")
        fileGeneral = request.FILES.get("fileGeneral")  
        print(check)
        for i in check:
            note = request.POST.get("note"+i)
            file = request.FILES.get("file"+i)
            if file:
                imalat_daire_balama.objects.filter(id = i).update(tamamlanma_bilgisi = True,tamamlamayi_yapan = request.user,tarih = datetime.now(),genel_notlar = note,dosya = file)
            else:
                imalat_daire_balama.objects.filter(id = i).update(tamamlanma_bilgisi = True,tamamlamayi_yapan = request.user,tarih = datetime.now(),genel_notlar = note)
        if fileGeneral:
            checkdaireleri.objects.filter(id = daire).update(genel_notlar = aciklma,dosya = fileGeneral)
        else:
            checkdaireleri.objects.filter(id = daire).update(genel_notlar = aciklma)
    return redirect_with_language("main:santiyelerim")
    

def katlara_gore_gor(request,id):
    content = sozluk_yapisi()
    content = sozluk_yapisi()
    m = folium.Map(location=[20, 0], zoom_start=2)
    # Haritayı HTML'ye dönüştürme
    if super_admin_kontrolu(request):
        profile =santiye.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_gorme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
        
    profile =get_object_or_none(bloglar,id = id,proje_ait_bilgisi = kullanici)
    katlar = []
    for i in range(1,int(profile.kat_sayisi)+1):
        katlar.append(i)
    content["santiyeler"] = katlar
    content["blok"] = profile
    return render(request,"checklist/katlara_gore_daireler.html",content)

def kat_daire_bilgisi(request,id,kat):
    content = sozluk_yapisi()
    content = sozluk_yapisi()
    m = folium.Map(location=[20, 0], zoom_start=2)
    # Haritayı HTML'ye dönüştürme
    if super_admin_kontrolu(request):
        profile =santiye.objects.all()
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_gorme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
        
    profile =get_object_or_none(bloglar,id = id,proje_ait_bilgisi = kullanici)
    
    content["santiyeler"] = profile
    content["daireler"]  = checkdaireleri.objects.filter(blog_bilgisi = profile,kat = kat)
    content["imalat"] = santiye_imalat_kalemleri.objects.filter(proje_santiye_Ait = profile.proje_santiye_Ait)
    content["kat"] = kat
    return render(request,"checklist/santiye_kontrol.html",content)


def santiye_kontrol_detayi_ust_yonetici(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")   
        else:
            kullanici =request.user  
    daire = get_object_or_404(checkdaireleri,id = id)
    content["santiye"] = daire
    content["ortak_alanlar"] = blog_ortak_alan_ve_cepheleri.objects.filter(blog_ait_bilgisi = daire.blog_bilgisi)
    content["imalatlar"] = imalat_daire_balama.objects.filter(daire_bilgisi = daire,tamamlanma_bilgisi = True)
    content["sanytiye_sablon_bolumleri"] = sanytiye_sablon_bolumleri.objects.filter(proje_santiye_Ait = daire.proje_santiye_Ait)
    print(content)
    return render(request,"checklist/santiye_kontrol_detayi_onaylama.html",content)

def santiye_kontrol_detayi_ust_yonetici(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")   
        else:
            kullanici =request.user  
    daire = get_object_or_404(checkdaireleri,id = id)
    content["santiye"] = daire
    content["ortak_alanlar"] = blog_ortak_alan_ve_cepheleri.objects.filter(blog_ait_bilgisi = daire.blog_bilgisi)
    content["imalatlar"] = imalat_daire_balama.objects.filter(daire_bilgisi = daire,tamamlanma_bilgisi = True)
    content["sanytiye_sablon_bolumleri"] = sanytiye_sablon_bolumleri.objects.filter(proje_santiye_Ait = daire.proje_santiye_Ait)
    print(content)
    return render(request,"checklist/santiye_kontrol_detayi_onaylama.html",content)



def daire_imalat_checklist_onaylama(request):
    if request.POST:
        check = request.POST.getlist("check")
        aciklma = request.POST.get("aciklma")
        onayla = request.POST.get("onayla")
        if onayla == "onayla":
            for i in check:
                #imalat_daire_balama.objects.filter(id = i).update(onaylma_tarihi = datetime.now(),onaylayan = request.user)
                a = imalat_daire_balama.objects.filter(id = i)
                print(a)
                for i in a:
                    check_liste_onaylama_gruplari.objects.filter(imalat_kalemi_ait = i,onaylayan = request.user).update(onaylama_notu = aciklma,tamamlanma_bilgisi = True,onaylma_tarihi = datetime.now())
        else:
           for i in check:
                imalat_daire_balama.objects.filter(id = i).update(tamamlanma_bilgisi = False,tamamlamayi_yapan = None,tarih = datetime.now(), ) 
    return redirect_with_language("main:santiyelerim")


def santiye_kontrol_detayi_ust_yoneticii(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")   
        else:
            kullanici =request.user  
    daire = get_object_or_404(checkdaireleri,id = id)
    content["santiye"] = daire
    content["ortak_alanlar"] = blog_ortak_alan_ve_cepheleri.objects.filter(blog_ait_bilgisi = daire.blog_bilgisi)
    content["imalatlar"] = imalat_daire_balama.objects.filter(daire_bilgisi = daire)
    content["sanytiye_sablon_bolumleri"] = sanytiye_sablon_bolumleri.objects.filter(proje_santiye_Ait = daire.proje_santiye_Ait)
    print(content)
    return render(request,"checklist/santiye_kontrol_de_onaylama_goster.html",content)

def rfi_Olustur(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")   
        else:
            kullanici =request.user  

    content["santiyeler"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
    if request.POST:
        santiye_bigisi = request.POST.get("santiye_bigisi")
        rfi_adi = request.POST.get("rfi_adi")
        rfi_kategorisi = request.POST.get("rfi_kategorisi")
        rfi_aciklama = request.POST.get("rfi_aciklama")
        ana_imalat_adi = request.POST.get("ana_imalat_adi")
        kontrol = request.POST.getlist("kontrol")
        sablon_bilgileri = rfi_sablonlar.objects.create(kayit_tarihi = get_kayit_tarihi_from_request(request),rfi_kime_ait = kullanici,
        rfi_santiye = get_object_or_none(santiye,id = santiye_bigisi,proje_ait_bilgisi = kullanici),rfi_baslik = rfi_adi,rfi_kategorisi = rfi_kategorisi,rfi_aciklama = rfi_aciklama,olusturan = request.user)
        if len(kontrol) > 0:
            for i in kontrol:
                rfi_kontrol = rfi_sablon_kalemleri.objects.create(kayit_tarihi = get_kayit_tarihi_from_request(request),sablon_bilgisi = get_object_or_none(rfi_sablonlar,id = sablon_bilgileri.id),
                kalem_baslik = i)
        else:
            for i in kontrol:
                rfi_kontrol = rfi_sablon_kalemleri.objects.create(kayit_tarihi = get_kayit_tarihi_from_request(request),sablon_bilgisi = get_object_or_none(rfi_sablonlar,id = sablon_bilgileri.id),
                kalem_baslik = ana_imalat_adi)
        return redirect_with_language("main:rfi_listesi")
    content["logosu"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu =kullanici).last()
    return render(request,"checklist/rfi_olustur.html",content)

def rfi_listesi(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar, kullanicilar=request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user

    if request.method == "POST":
        rfi_sablonu = request.POST.get("rfi_sablonu")
        yapi_gonder = request.POST.get("yapi_gonder")
        kat = request.POST.get("kat")
        apartmentNo = request.POST.get("apartmentNo")
        location = request.POST.get("location")
        file = request.FILES.get("file")
        notlar = request.POST.get("notlar")

        # Validate required fields
        if not (rfi_sablonu and yapi_gonder and kat and apartmentNo and location and file):
            return JsonResponse({"status": "error", "message": "Eksik bilgi gönderildi."}, status=400)
        print(get_kayit_tarihi_from_request(request),"saat")
        rfi_kontrol.objects.create(
            kayit_tarihi=get_kayit_tarihi_from_request(request),
            sablon_bilgisi=get_object_or_none(rfi_sablonlar, id=rfi_sablonu),
            blok=get_object_or_none(bloglar, id=yapi_gonder),
            kat_bilgisi=kat,
            daire_no=apartmentNo,
            mahal=location,
            file=file,
            notlar=notlar,
            kontrol_ekleyen=request.user
        )
        return JsonResponse({"status": "success", "message": "RFI başarıyla oluşturuldu."})

    content["rfi_sablonlari"] = rfi_sablonlar.objects.filter(rfi_kime_ait=kullanici)
    content["rfi_listesi_onay_bekleyen"] = rfi_kontrol.objects.filter(
        sablon_bilgisi__rfi_kime_ait=kullanici, onaylama_bilgisi=False, onaylayan_bilgisi=None
    ).order_by("kayit_tarihi")
    content["rfi_listesi_onay_bekleyen_sayisi"] = content["rfi_listesi_onay_bekleyen"].count()
    content["rfi_listesi_red_yiyenler"] = rfi_kontrol.objects.filter(
        sablon_bilgisi__rfi_kime_ait=kullanici, onaylama_bilgisi=False
    ).exclude(onaylayan_bilgisi=None).order_by("kayit_tarihi")
    content["rfi_listesi_onay_alanlar"] = rfi_kontrol.objects.filter(
        sablon_bilgisi__rfi_kime_ait=kullanici, onaylama_bilgisi=True
    ).exclude(onaylayan_bilgisi=None).order_by("kayit_tarihi")
    return render(request, "checklist/rfi_listesi.html", content)


def rfi_approve(request):
    if True:
        id = request.GET.get("id")
        if not id:
            return JsonResponse({"status": "error", "message": "RFI ID eksik."}, status=400)

        rfi = rfi_kontrol.objects.filter(id=id).first()
        if not rfi:
            return JsonResponse({"status": "error", "message": "RFI bulunamadı."}, status=404)

        rfi.onaylama_bilgisi = True
        rfi.onaylayan_bilgisi = request.user
        rfi.onaylayan_tarih = datetime.now()
        rfi.save()

        return JsonResponse({"status": "success", "message": "RFI başarıyla onaylandı."})
    return JsonResponse({"status": "error", "message": "Geçersiz istek."}, status=405)


def rfi_reject(request):
    
    if True:
        id = request.GET.get("id")
        red_sebebi = request.GET.get("red_sebebi")
        
        if not id or not red_sebebi:
            return JsonResponse({"status": "error", "message": "Eksik bilgi gönderildi."}, status=400)

        rfi = rfi_kontrol.objects.filter(id=id).first()
        if not rfi:
            return JsonResponse({"status": "error", "message": "RFI bulunamadı."}, status=404)

        rfi.onaylama_bilgisi = False
        rfi.onaylayan_bilgisi = request.user
        rfi.onaylayan_tarih = datetime.now()
        rfi.red_sebebi = red_sebebi
        rfi.save()

        return JsonResponse({"status": "success", "message": "RFI başarıyla reddedildi."})
    return JsonResponse({"status": "error", "message": "Geçersiz istek."}, status=405)

def rfi_template(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")   
        else:
            kullanici =request.user  
    content["rfi_sablonlari"] = rfi_sablonlar.objects.filter(rfi_kime_ait = kullanici)
    content["rfi_kategorileri"] = rfi_sablonlar.objects.filter(rfi_kime_ait = kullanici).values("rfi_kategorisi").distinct() 
    return render(request,"checklist/rfi_template.html",content)


def rfi_detail(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")   
        else:
            kullanici =request.user  
        content["onayli"] = get_object_or_none(rfi_kontrol,sablon_bilgisi__rfi_kime_ait = kullanici,id = id)
        
        #rfi_kontrol.objects.filter(sablon_bilgisi__rfi_kime_ait = kullanici,onaylama_bilgisi = False,onaylayan_bilgisi = None).order_by("kayit_tarihi")
    return render(request,"checklist/rfi_detail.html",content)

def rfi_duzenleme(request, id):
    content = sozluk_yapisi()

    # Check if the user has super admin privileges
    if not super_admin_kontrolu(request):
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar, kullanicilar=request.user)
            if not a or not a.izinler.santiye_kontrol:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user

    # Fetch the RFI template and its related controls
    rfi_sablon = get_object_or_none(rfi_sablonlar, id=id)
    if not rfi_sablon:
        return redirect_with_language("main:yetkisiz")  # Redirect if the RFI template does not exist

    content["rfi_sablonlari"] = rfi_sablon
    content["rfi_kontrolleri"] = rfi_sablon_kalemleri.objects.filter(sablon_bilgisi=rfi_sablon)
    if request.POST:
        # Update the RFI template
        rfi_adi = request.POST.get("rfi_baslik")
        rfi_kategorisi = request.POST.get("rfi_kategorisi")
        rfi_aciklama = request.POST.get("rfi_aciklama")
        ana_imalat_adi = request.POST.get("ana_imalat_adi")
        kontrol = request.POST.getlist("kontroll")
        rfi_sablonlar.objects.filter(id=id).update(rfi_baslik=rfi_adi, rfi_kategorisi=rfi_kategorisi, rfi_aciklama=rfi_aciklama)
        rfi_sablon_kalemleri.objects.filter(sablon_bilgisi=rfi_sablon).delete()
        
        if len(kontrol) > 0:
            for i in kontrol:
                rfi_kontrol = rfi_sablon_kalemleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                                                    sablon_bilgisi=get_object_or_none(rfi_sablonlar, id=id),kalem_baslik = i)
        else:
            for i in kontrol:
                rfi_kontrol = rfi_sablon_kalemleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                                                    sablon_bilgisi=get_object_or_none(rfi_sablonlar, id=id))
        print(request.POST)
        return redirect_with_language("main:rfi_template")
    return render(request, "checklist/rfi_duzenleme.html", content)



def rfi_show(request,id):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar,kullanicilar = request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici =request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")   
        else:
            kullanici =request.user  
        content["onayli"] = get_object_or_none(rfi_sablonlar,rfi_kime_ait = kullanici,id = id)
        content["kalemler"] = rfi_sablon_kalemleri.objects.filter(sablon_bilgisi = get_object_or_none(rfi_sablonlar,rfi_kime_ait = kullanici,id = id))
        content["kalemler_count"] = rfi_sablon_kalemleri.objects.filter(sablon_bilgisi = get_object_or_none(rfi_sablonlar,rfi_kime_ait = kullanici,id = id)).count()
        content["logosu"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu =kullanici).last()
        #rfi_kontrol.objects.filter(sablon_bilgisi__rfi_kime_ait = kullanici,onaylama_bilgisi = False,onaylayan_bilgisi = None).order_by("kayit_tarihi")
    return render(request,"checklist/rfi_onizle.html",content)
def rapor_olusturma(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar, kullanicilar=request.user)
            if a:
                if a.izinler.rapor_olusturucu_gorme and a.izinler.rapor_olusturucu_olusturma:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = kullanici)     
        content["cariler"] = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = kullanici)
        content["santiyeler"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        content["bloklar"] = bloglar.objects.filter(proje_santiye_Ait__silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        content["is_planlari"] = IsplaniPlanlari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici,status = "Completed")
        content["is_planlarii"] = IsplaniPlanlari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        content["personeller_listesi"] = calisanlar.objects.filter(status = "0",silinme_bilgisi = False,calisan_kime_ait = kullanici)
        content["logo_islemi"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu = kullanici).last()
    return render(request, "santiye_yonetimi/rapor_olusturucu.html", content)

def raporlari_gor_sayfasi(request):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar, kullanicilar=request.user)
            if a:
                if a.izinler.rapor_olusturucu_gorme:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user
    content["raporlar"] = rapor_bilgisi.objects.filter(silinme_bilgisi = False,rapor_kime_ait = kullanici)
    return render(request, "santiye_yonetimi/raporlari_goster.html", content)

import json
import base64
import os
import uuid
from django.http import JsonResponse
from django.conf import settings
from django.shortcuts import redirect

def rapor_kaydedici(request):
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar, kullanicilar=request.user)
            if a:
                if a.izinler.rapor_olusturucu_olusturma:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user

    if request.method == 'POST':
        data = json.loads(request.body)
        description = data["description"]
        tarih_baslangic = data["dateRange"]["start"]
        tarih_bitis = data["dateRange"]["end"]
        veri = data["content"]
        proje_getir = data["project"]
        pdf = data["pdf"] 

        # PDF'yi kaydet
        try:
            pdf_base64 = pdf # "data:application/pdf;base64," kısmını at
            pdf_icerik = base64.b64decode(pdf_base64)

            # Dosya adını oluştur
            dosya_adi = f"{uuid.uuid4().hex}.pdf"
            klasor_yolu = os.path.join(settings.MEDIA_ROOT, "rapor_dosyalari")
            os.makedirs(klasor_yolu, exist_ok=True)
            pdf_yolu = os.path.join(klasor_yolu, dosya_adi)

            # Dosyayı kaydet
            with open(pdf_yolu, "wb") as f:
                f.write(pdf_icerik)

            pdf_url = f"/media/rapor_dosyalari/{dosya_adi}"
            c = f"rapor_dosyalari/{dosya_adi}"
            rapor_bilgisi.objects.create(
                rapor_kime_ait = kullanici,
                rapor_basligi = name,
                rapor_aciklama = description,
                rapor_icerigi = veri,
                olusturan = request.user,
                baslangic_tarihi = tarih_baslangic,
                bitis_tarihi = tarih_bitis,
                rapor_dosyalari = c)
        except Exception as e:
            return JsonResponse({
                "status": "error",
                "message": f"PDF kaydedilirken hata oluştu: {str(e)}"
            })

        return JsonResponse({
            "status": "success",
            "message": "Rapor başarıyla kaydedildi.",
            "pdf_url": pdf_url
        })
#####################RFİ
def rfi_Olustur_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = kullanici
        #profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    

    content["santiyeler"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
    if request.POST:
        santiye_bigisi = request.POST.get("santiye_bigisi")
        rfi_adi = request.POST.get("rfi_adi")
        rfi_kategorisi = request.POST.get("rfi_kategorisi")
        rfi_aciklama = request.POST.get("rfi_aciklama")
        ana_imalat_adi = request.POST.get("ana_imalat_adi")
        kontrol = request.POST.getlist("kontrol")
        sablon_bilgileri = rfi_sablonlar.objects.create(kayit_tarihi = get_kayit_tarihi_from_request(request),rfi_kime_ait = kullanici,
        rfi_santiye = get_object_or_none(santiye,id = santiye_bigisi,proje_ait_bilgisi = kullanici),rfi_baslik = rfi_adi,rfi_kategorisi = rfi_kategorisi,rfi_aciklama = rfi_aciklama,olusturan = request.user)
        if len(kontrol) > 0:
            for i in kontrol:
                rfi_kontrol = rfi_sablon_kalemleri.objects.create(kayit_tarihi = get_kayit_tarihi_from_request(request),sablon_bilgisi = get_object_or_none(rfi_sablonlar,id = sablon_bilgileri.id),
                kalem_baslik = i)
        else:
            for i in kontrol:
                rfi_kontrol = rfi_sablon_kalemleri.objects.create(kayit_tarihi = get_kayit_tarihi_from_request(request),sablon_bilgisi = get_object_or_none(rfi_sablonlar,id = sablon_bilgileri.id),
                kalem_baslik = ana_imalat_adi)
        return redirect_with_language("main:rfi_listesi_2",hash)
    content["logosu"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu =kullanici).last()
    return render(request,"checklist/rfi_olustur.html",content)

def rfi_listesi_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = kullanici
        #profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    if request.method == "POST":
        rfi_sablonu = request.POST.get("rfi_sablonu")
        yapi_gonder = request.POST.get("yapi_gonder")
        kat = request.POST.get("kat")
        apartmentNo = request.POST.get("apartmentNo")
        location = request.POST.get("location")
        file = request.FILES.get("file")
        notlar = request.POST.get("notlar")

        # Validate required fields
        if not (rfi_sablonu and yapi_gonder and kat and apartmentNo and location and file):
            return JsonResponse({"status": "error", "message": "Eksik bilgi gönderildi."}, status=400)
        print(get_kayit_tarihi_from_request(request),"saat")
        rfi_kontrol.objects.create(
            kayit_tarihi=get_kayit_tarihi_from_request(request),
            sablon_bilgisi=get_object_or_none(rfi_sablonlar, id=rfi_sablonu),
            blok=get_object_or_none(bloglar, id=yapi_gonder),
            kat_bilgisi=kat,
            daire_no=apartmentNo,
            mahal=location,
            file=file,
            notlar=notlar,
            kontrol_ekleyen=request.user
        )
        return JsonResponse({"status": "success", "message": "RFI başarıyla oluşturuldu."})

    content["rfi_sablonlari"] = rfi_sablonlar.objects.filter(rfi_kime_ait=kullanici)
    content["rfi_listesi_onay_bekleyen"] = rfi_kontrol.objects.filter(
        sablon_bilgisi__rfi_kime_ait=kullanici, onaylama_bilgisi=False, onaylayan_bilgisi=None
    ).order_by("kayit_tarihi")
    content["rfi_listesi_onay_bekleyen_sayisi"] = content["rfi_listesi_onay_bekleyen"].count()
    content["rfi_listesi_red_yiyenler"] = rfi_kontrol.objects.filter(
        sablon_bilgisi__rfi_kime_ait=kullanici, onaylama_bilgisi=False
    ).exclude(onaylayan_bilgisi=None).order_by("kayit_tarihi")
    content["rfi_listesi_onay_alanlar"] = rfi_kontrol.objects.filter(
        sablon_bilgisi__rfi_kime_ait=kullanici, onaylama_bilgisi=True
    ).exclude(onaylayan_bilgisi=None).order_by("kayit_tarihi")
    return render(request, "checklist/rfi_listesi.html", content)


def rfi_approve_2(request,hash):
    
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = kullanici
        #profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
        
    if True:
        id = request.GET.get("id")
        if not id:
            return JsonResponse({"status": "error", "message": "RFI ID eksik."}, status=400)

        rfi = rfi_kontrol.objects.filter(id=id).first()
        if not rfi:
            return JsonResponse({"status": "error", "message": "RFI bulunamadı."}, status=404)

        rfi.onaylama_bilgisi = True
        rfi.onaylayan_bilgisi = request.user
        rfi.onaylayan_tarih = datetime.now()
        rfi.save()

        return JsonResponse({"status": "success", "message": "RFI başarıyla onaylandı."})
    return JsonResponse({"status": "error", "message": "Geçersiz istek."}, status=405)


def rfi_reject_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = kullanici
        #profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    if True:
        id = request.GET.get("id")
        red_sebebi = request.GET.get("red_sebebi")
        
        if not id or not red_sebebi:
            return JsonResponse({"status": "error", "message": "Eksik bilgi gönderildi."}, status=400)

        rfi = rfi_kontrol.objects.filter(id=id).first()
        if not rfi:
            return JsonResponse({"status": "error", "message": "RFI bulunamadı."}, status=404)

        rfi.onaylama_bilgisi = False
        rfi.onaylayan_bilgisi = request.user
        rfi.onaylayan_tarih = datetime.now()
        rfi.red_sebebi = red_sebebi
        rfi.save()

        return JsonResponse({"status": "success", "message": "RFI başarıyla reddedildi."})
    return JsonResponse({"status": "error", "message": "Geçersiz istek."}, status=405)

def rfi_template_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = kullanici
        #profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    content["rfi_sablonlari"] = rfi_sablonlar.objects.filter(rfi_kime_ait = kullanici)
    content["rfi_kategorileri"] = rfi_sablonlar.objects.filter(rfi_kime_ait = kullanici).values("rfi_kategorisi").distinct() 
    return render(request,"checklist/rfi_template.html",content)


def rfi_detail_2(request,id,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = kullanici
        #profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
     
        content["onayli"] = get_object_or_none(rfi_kontrol,sablon_bilgisi__rfi_kime_ait = kullanici,id = id)
        
        #rfi_kontrol.objects.filter(sablon_bilgisi__rfi_kime_ait = kullanici,onaylama_bilgisi = False,onaylayan_bilgisi = None).order_by("kayit_tarihi")
    return render(request,"checklist/rfi_detail.html",content)

def rfi_duzenleme_2(request, id,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = kullanici
        #profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar

    

    # Fetch the RFI template and its related controls
    rfi_sablon = get_object_or_none(rfi_sablonlar, id=id)
    if not rfi_sablon:
        return redirect_with_language("main:yetkisiz")  # Redirect if the RFI template does not exist

    content["rfi_sablonlari"] = rfi_sablon
    content["rfi_kontrolleri"] = rfi_sablon_kalemleri.objects.filter(sablon_bilgisi=rfi_sablon)
    if request.POST:
        # Update the RFI template
        rfi_adi = request.POST.get("rfi_baslik")
        rfi_kategorisi = request.POST.get("rfi_kategorisi")
        rfi_aciklama = request.POST.get("rfi_aciklama")
        ana_imalat_adi = request.POST.get("ana_imalat_adi")
        kontrol = request.POST.getlist("kontroll")
        rfi_sablonlar.objects.filter(id=id).update(rfi_baslik=rfi_adi, rfi_kategorisi=rfi_kategorisi, rfi_aciklama=rfi_aciklama)
        rfi_sablon_kalemleri.objects.filter(sablon_bilgisi=rfi_sablon).delete()
        
        if len(kontrol) > 0:
            for i in kontrol:
                rfi_kontrol = rfi_sablon_kalemleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                                                    sablon_bilgisi=get_object_or_none(rfi_sablonlar, id=id),kalem_baslik = i)
        else:
            for i in kontrol:
                rfi_kontrol = rfi_sablon_kalemleri.objects.create(kayit_tarihi=get_kayit_tarihi_from_request(request),
                                                                    sablon_bilgisi=get_object_or_none(rfi_sablonlar, id=id))
        print(request.POST)
        return redirect_with_language("main:rfi_template_2",hash)
    return render(request, "checklist/rfi_duzenleme.html", content)



def rfi_show_2(request,id,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = kullanici
        #profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar 
        content["onayli"] = get_object_or_none(rfi_sablonlar,rfi_kime_ait = kullanici,id = id)
        content["kalemler"] = rfi_sablon_kalemleri.objects.filter(sablon_bilgisi = get_object_or_none(rfi_sablonlar,rfi_kime_ait = kullanici,id = id))
        content["logosu"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu =kullanici).last()
        #rfi_kontrol.objects.filter(sablon_bilgisi__rfi_kime_ait = kullanici,onaylama_bilgisi = False,onaylayan_bilgisi = None).order_by("kayit_tarihi")
    return render(request,"checklist/rfi_onizle.html",content)
def rapor_olusturma_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = kullanici
        #profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    
        content["kasa"] = Kasa.objects.filter(silinme_bilgisi = False,kasa_kart_ait_bilgisi = kullanici)     
        content["cariler"] = cari.objects.filter(silinme_bilgisi = False,cari_kart_ait_bilgisi = kullanici)
        content["santiyeler"] = santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        content["bloklar"] = bloglar.objects.filter(proje_santiye_Ait__silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        content["is_planlari"] = IsplaniPlanlari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici,status = "Completed")
        content["is_planlarii"] = IsplaniPlanlari.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        content["personeller_listesi"] = calisanlar.objects.filter(status = "0",silinme_bilgisi = False,calisan_kime_ait = kullanici)
        content["logo_islemi"] = faturalar_icin_logo.objects.filter(gelir_kime_ait_oldugu = kullanici).last()
    return render(request, "santiye_yonetimi/rapor_olusturucu.html", content)

def raporlari_gor_sayfasi_2(request,hash):
    content = sozluk_yapisi()
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        content["hashler"] = hash
        content["hash_bilgi"] = kullanici
        #profile =santiye.objects.filter(silinme_bilgisi = False,proje_ait_bilgisi = kullanici)
        kullanicilar = CustomUser.objects.filter(kullanicilar_db = None,is_superuser = False).order_by("-id")
        content["kullanicilar"] =kullanicilar
    
    content["raporlar"] = rapor_bilgisi.objects.filter(silinme_bilgisi = False,rapor_kime_ait = kullanici)
    return render(request, "santiye_yonetimi/raporlari_goster.html", content)

import json
import base64
import os
import uuid
from django.http import JsonResponse
from django.conf import settings
from django.shortcuts import redirect

def rapor_kaydedici(request):
    if super_admin_kontrolu(request):
        pass
    else:
        if request.user.kullanicilar_db:
            a = get_object_or_none(bagli_kullanicilar, kullanicilar=request.user)
            if a:
                if a.izinler.santiye_kontrol:
                    kullanici = request.user.kullanicilar_db
                else:
                    return redirect_with_language("main:yetkisiz")
            else:
                return redirect_with_language("main:yetkisiz")
        else:
            kullanici = request.user

    if request.method == 'POST':
        data = json.loads(request.body)
        description = data["description"]
        tarih_baslangic = data["dateRange"]["start"]
        tarih_bitis = data["dateRange"]["end"]
        veri = data["content"]
        proje_getir = data["project"]
        pdf = data["pdf"] 

        # PDF'yi kaydet
        try:
            pdf_base64 = pdf # "data:application/pdf;base64," kısmını at
            pdf_icerik = base64.b64decode(pdf_base64)

            # Dosya adını oluştur
            dosya_adi = f"{uuid.uuid4().hex}.pdf"
            klasor_yolu = os.path.join(settings.MEDIA_ROOT, "rapor_dosyalari")
            os.makedirs(klasor_yolu, exist_ok=True)
            pdf_yolu = os.path.join(klasor_yolu, dosya_adi)

            # Dosyayı kaydet
            with open(pdf_yolu, "wb") as f:
                f.write(pdf_icerik)

            pdf_url = f"/media/rapor_dosyalari/{dosya_adi}"
            c = f"rapor_dosyalari/{dosya_adi}"
            rapor_bilgisi.objects.create(
                rapor_kime_ait = kullanici,
                rapor_basligi = name,
                rapor_aciklama = description,
                rapor_icerigi = veri,
                olusturan = request.user,
                baslangic_tarihi = tarih_baslangic,
                bitis_tarihi = tarih_bitis,
                rapor_dosyalari = c)
        except Exception as e:
            return JsonResponse({
                "status": "error",
                "message": f"PDF kaydedilirken hata oluştu: {str(e)}"
            })

        return JsonResponse({
            "status": "success",
            "message": "Rapor başarıyla kaydedildi.",
            "pdf_url": pdf_url
        })
def rapor_kaydedici_2(request,hash):
    if super_admin_kontrolu(request):
        d = decode_id(hash)
        kullanici = get_object_or_404(CustomUser,id = d)
        

    if request.method == 'POST':
        data = json.loads(request.body)
        description = data["description"]
        tarih_baslangic = data["dateRange"]["start"]
        tarih_bitis = data["dateRange"]["end"]
        veri = data["content"]
        proje_getir = data["project"]
        pdf = data["pdf"] 

        # PDF'yi kaydet
        try:
            pdf_base64 = pdf # "data:application/pdf;base64," kısmını at
            pdf_icerik = base64.b64decode(pdf_base64)

            # Dosya adını oluştur
            dosya_adi = f"{uuid.uuid4().hex}.pdf"
            klasor_yolu = os.path.join(settings.MEDIA_ROOT, "rapor_dosyalari")
            os.makedirs(klasor_yolu, exist_ok=True)
            pdf_yolu = os.path.join(klasor_yolu, dosya_adi)

            # Dosyayı kaydet
            with open(pdf_yolu, "wb") as f:
                f.write(pdf_icerik)

            pdf_url = f"/media/rapor_dosyalari/{dosya_adi}"
            c = f"rapor_dosyalari/{dosya_adi}"
            rapor_bilgisi.objects.create(
                rapor_kime_ait = kullanici,
                rapor_basligi = name,
                rapor_aciklama = description,
                rapor_icerigi = veri,
                olusturan = request.user,
                baslangic_tarihi = tarih_baslangic,
                bitis_tarihi = tarih_bitis,
                rapor_dosyalari = c)
        except Exception as e:
            return JsonResponse({
                "status": "error",
                "message": f"PDF kaydedilirken hata oluştu: {str(e)}"
            })

        return JsonResponse({
            "status": "success",
            "message": "Rapor başarıyla kaydedildi.",
            "pdf_url": pdf_url
        })

####################### RFİ
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.utils import timezone


@login_required
@user_passes_test(lambda u: u.is_superuser)
@transaction.atomic
def kullanici_verilerini_klonla_view(request, kaynak_kullanici_id, hedef_kullanici_id):
    kaynak_kullanici = get_object_or_404(CustomUser, pk=kaynak_kullanici_id)
    hedef_kullanici = get_object_or_404(CustomUser, pk=hedef_kullanici_id)

    try:
        eski_yeni_proje_tipleri = {}
        for pt in proje_tipi.objects.filter(proje_ait_bilgisi=kaynak_kullanici, silinme_bilgisi=False):
            eski_id = pt.pk
            pt.pk = None
            pt.proje_ait_bilgisi = hedef_kullanici
            pt.kayit_tarihi = timezone.now()
            pt.save()
            eski_yeni_proje_tipleri[eski_id] = pt

        eski_yeni_santiyeler = {}
        for s in santiye.objects.filter(proje_ait_bilgisi=kaynak_kullanici, silinme_bilgisi=False):
            eski_id = s.pk
            s.pk = None
            s.proje_ait_bilgisi = hedef_kullanici
            s.proje_tipi = eski_yeni_proje_tipleri.get(s.proje_tipi.pk) if s.proje_tipi else None
            s.kayit_tarihi = timezone.now()
            s.save()
            eski_yeni_santiyeler[eski_id] = s

        eski_yeni_bloglar = {}
        for b in bloglar.objects.filter(proje_ait_bilgisi=kaynak_kullanici):
            eski_id = b.pk
            b.pk = None
            b.proje_ait_bilgisi = hedef_kullanici
            b.proje_santiye_Ait = eski_yeni_santiyeler.get(b.proje_santiye_Ait.pk) if b.proje_santiye_Ait else None
            b.kayit_tarihi = timezone.now()
            b.save()
            eski_yeni_bloglar[eski_id] = b

        eski_yeni_kalemler = {}
        for k in santiye_kalemleri.objects.filter(proje_ait_bilgisi=kaynak_kullanici):
            eski_id = k.pk
            k.pk = None
            k.proje_ait_bilgisi = hedef_kullanici
            k.proje_santiye_Ait = eski_yeni_santiyeler.get(k.proje_santiye_Ait.pk) if k.proje_santiye_Ait else None
            k.kayit_tarihi = timezone.now()
            k.save()
            eski_yeni_kalemler[eski_id] = k

        for d in santiye_kalemlerin_dagilisi.objects.filter(proje_ait_bilgisi=kaynak_kullanici):
            d.pk = None
            d.proje_ait_bilgisi = hedef_kullanici
            d.proje_santiye_Ait = eski_yeni_santiyeler.get(d.proje_santiye_Ait.pk) if d.proje_santiye_Ait else None
            d.kalem_bilgisi = eski_yeni_kalemler.get(d.kalem_bilgisi.pk) if d.kalem_bilgisi else None
            d.blog_bilgisi = eski_yeni_bloglar.get(d.blog_bilgisi.pk) if d.blog_bilgisi else None
            d.kayit_tarihi = timezone.now()
            d.save()

        messages.success(request, "Veriler başarıyla kopyalandı.")
    except Exception as e:
        messages.error(request, f"Hata oluştu: {str(e)}")
        raise

    return redirect("admin:index")
